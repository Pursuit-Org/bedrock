"""Jobs pipeline API.

  GET    /api/jobs/opportunities              — list (filterable by stage, owner, account)
  POST   /api/jobs/opportunities              — create
  GET    /api/jobs/opportunities/:id          — get one with stage history + activity
  PATCH  /api/jobs/opportunities/:id          — update (stage change auto-logs history)
  DELETE /api/jobs/opportunities/:id          — soft delete
  GET    /api/jobs/opportunities/pipeline     — grouped stage counts for pipeline view
"""

import asyncio
import asyncpg
import io
import json
import logging
import re
from datetime import date, datetime, timedelta, timezone
from typing import Optional
from uuid import UUID

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from auth import require_auth
from db import get_db, get_pool
from dependencies import get_mcp_client, require_sf_mcp_client
from sf_errors import sf_http_error
from services.placement_sf import sync_placement_to_sf, record_sync_error, NotEligible, AccountAmbiguous
from services.outreach_targets import user_pipeline_target, activity_pipeline_target
from services.jobs_activity_link import has_membership_history

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/jobs", tags=["jobs"])

# Every stage the app UNDERSTANDS, old and new. The 2026-08-05 simplification
# drops initial_outreach and the three on_hold_* values and renames
# active_builder_interview -> reviewing_builders, but the legacy values stay
# listed here so the app can still read and label rows written before the
# migration (and so it keeps working if the migration hasn't been applied yet).
# What can actually be WRITTEN is decided by the database CHECK constraint —
# see _writable_stages(), which is what the UI's pickers are built from.
VALID_STAGES = {
    "lead_submitted",
    "active_in_discussions", "active_opportunity_confirmed", "reviewing_builders",
    "closed_won", "closed_lost",
    # legacy, pre-2026-08-05
    "initial_outreach", "active_builder_interview",
    "on_hold_not_selected", "on_hold_not_interested", "on_hold_not_responsive",
}

# Post-migration stage lists, in funnel order. These drive the UI once the
# database accepts them.
OPPORTUNITY_STAGES_NEW = [
    "lead_submitted", "active_in_discussions", "active_opportunity_confirmed",
    "reviewing_builders", "closed_won", "closed_lost",
]
MEMBERSHIP_STAGES_NEW = [
    "assigned", "initial_outreach", "call_booked",
    "converted_to_opportunity", "revisit", "not_a_fit",
]
MEMBERSHIP_STAGES_LEGACY = [
    "assigned", "initial_outreach", "converted_to_opportunity", "on_hold", "not_a_fit",
]

CLOSED_LOST_REASONS = [
    ("budget", "No budget"), ("timing", "Timing / not now"),
    ("hired_elsewhere", "Hired elsewhere"), ("not_a_fit", "Not a fit"),
    ("no_response", "Went cold"), ("role_cancelled", "Role cancelled"),
    ("not_interested", "Not interested"), ("not_selected", "Not selected"),
    ("not_responsive", "Not responsive"), ("revisit", "Revisit later"),
    ("other", "Other"),
]

VALID_DEAL_TYPES = {"ft", "pt_contract", "capstone", "volunteer", "workshop", "pilot"}

# The jobs team's mailboxes. The Outreach and Calls/Mtgs dashboard metrics count
# FIRST TOUCHES by these senders only: each external contact counts once, ever,
# across the whole team (3 emails to the same person in a week = 1; emailing
# someone the team already reached before = 0). Synced gmail/calendar only —
# manual deal-logs have no counterpart email to dedupe on.
# Core jobs team — default scope for outreach/activation metrics. Per the
# 2026-07-06 review ("remove everyone except Damon, Avni, Devika"); other
# staff remain reachable via the per-person owner filter (staff drill).
JOBS_TEAM_EMAILS = ["avni@pursuit.org", "damon.kornhauser@pursuit.org", "devika@pursuit.org"]

_FT_EXTERNAL = """
        WHERE counterpart NOT LIKE '%@pursuit.org' AND counterpart NOT LIKE '%@pursuit.com'
          AND counterpart NOT LIKE '%.calendar.google.com' AND counterpart <> ''
"""


def _first_touch_email_cte() -> str:
    """CTE `ext(counterpart, first_touch)`: external recipients of outbound
    emails sent by the jobs team, with each contact's first-ever touch date."""
    sender = " OR ".join(f"a.email_from ILIKE '%{e}%'" for e in JOBS_TEAM_EMAILS)
    return f"""
    WITH outbound AS (
      SELECT lower(e) AS counterpart, a.activity_date
      FROM bedrock.activity a,
           unnest(coalesce(a.email_to,'{{}}') || coalesce(a.email_cc,'{{}}')) e
      WHERE a.source = 'gmail-sync' AND a.deleted_at IS NULL AND ({sender})
        AND {_not_autoreply('a')} AND {_jobs_relevant('a')}
    ),
    ext AS (
      SELECT counterpart, min(activity_date) AS first_touch
      FROM outbound {_FT_EXTERNAL}
      GROUP BY counterpart
    )
    """


def _first_touch_meeting_cte() -> str:
    """CTE `ext(counterpart, first_touch)`: external attendees of meetings on
    the jobs team's calendars, with each person's first-ever meeting date."""
    team = ", ".join(f"'{e}'" for e in JOBS_TEAM_EMAILS)
    return f"""
    WITH mtg AS (
      SELECT lower(att->>'email') AS counterpart, a.activity_date
      FROM bedrock.activity a, jsonb_array_elements(a.meeting_attendees) att
      WHERE a.source = 'calendar-sync' AND a.deleted_at IS NULL
        AND lower(a.logged_by) IN ({team}) AND {_jobs_relevant('a')}
    ),
    ext AS (
      SELECT counterpart, min(activity_date) AS first_touch
      FROM mtg {_FT_EXTERNAL}
      GROUP BY counterpart
    )
    """

def _jobs_activity_flag(alias: str = "a") -> str:
    """SQL boolean: is this activity row 'jobs activity'? True when it's tied to
    a jobs opportunity, is a manual jobs channel (call/text/linkedin), OR is an
    email/meeting in a jobs-team mailbox (Avni/Damon) — so their synced email
    lands in the Jobs section, not the generic comms bucket.

    The single source of truth for the per-row is_jobs split in the
    account/contact/opp activity rollups. It is INTENTIONALLY narrower than a
    full "in the jobs pipeline" check: those rollups already filter to the
    account/contact/opp, so matching on account_id/participant here would mark
    *every* row jobs and collapse the Jobs-vs-Comms distinction. This flag means
    "a structured or team-driven touch" within that already-scoped feed."""
    team = " OR ".join(
        f"{alias}.email_from ILIKE '%{e}%' OR {alias}.logged_by ILIKE '%{e}%'"
        for e in JOBS_TEAM_EMAILS
    )
    return (
        f"({alias}.jobs_opportunity_id IS NOT NULL "
        f"OR {alias}.source = 'manual' "        # a hand-logged touch is always a jobs touch
        f"OR {alias}.type IN ('call','text','linkedin') "
        f"OR {team})"
    )


STAGE_LABELS = {
    "lead_submitted":               "Lead Submitted",
    "initial_outreach":             "Initial Outreach",
    "active_in_discussions":        "In Discussions",
    "active_opportunity_confirmed": "Opportunity Confirmed",
    "active_builder_interview":     "Builder Interview",
    "closed_won":                   "Closed — Won",
    "closed_lost":                  "Closed — Lost",
    "reviewing_builders":           "Reviewing Builders",
    # Legacy values — kept so history and un-migrated rows still render a name
    # rather than a raw slug.
    "on_hold_not_selected":         "On Hold: Not Selected",
    "on_hold_not_interested":       "On Hold: Not Interested",
    "on_hold_not_responsive":       "On Hold: Not Responsive",
}

MEMBERSHIP_STAGE_LABELS = {
    "assigned":                 "Assigned",
    "initial_outreach":         "Initial Outreach",
    "call_booked":              "Call Booked",
    "converted_to_opportunity": "Converted to Opportunity",
    "revisit":                  "Revisit",
    "not_a_fit":                "Not a Fit",
    "on_hold":                  "On Hold",   # legacy, pre-2026-08-05
}


VALID_LIKELIHOODS = {"low", "medium", "high"}


class OpportunityCreate(BaseModel):
    account_id: str
    account_name: Optional[str] = None
    stage: str = "lead_submitted"
    deal_type: Optional[str] = None
    title: Optional[str] = None
    description: Optional[str] = None
    salary_expected: Optional[int] = None
    num_roles: Optional[int] = None
    likelihood: Optional[str] = None
    source: Optional[str] = None
    owner_email: Optional[str] = None
    relationship_owner: Optional[str] = None
    sf_contact_ids: list[str] = []
    builder_ids: list[str] = []
    follow_up_date: Optional[datetime] = None
    note: Optional[str] = None  # note for initial stage history entry


class OpportunityUpdate(BaseModel):
    stage: Optional[str] = None
    deal_type: Optional[str] = None
    title: Optional[str] = None
    description: Optional[str] = None
    salary_expected: Optional[int] = None
    num_roles: Optional[int] = None
    likelihood: Optional[str] = None
    source: Optional[str] = None
    owner_email: Optional[str] = None
    relationship_owner: Optional[str] = None
    sf_contact_ids: Optional[list[str]] = None
    builder_ids: Optional[list[str]] = None
    follow_up_date: Optional[datetime] = None
    target_close_date: Optional[date] = None
    touch_count: Optional[int] = None
    sf_opportunity_id: Optional[str] = None
    note: Optional[str] = None  # optional note when changing stage
    # Call-feedback round
    closed_lost_reason: Optional[str] = None
    closed_lost_note: Optional[str] = None
    priority: Optional[int] = None
    segment: Optional[str] = None
    intro_by: Optional[str] = None
    # Campaign tags, same vocabulary as contacts (bedrock.contact_tag_catalog).
    # Column arrives with the 2026-08-05 migration; writes are skipped until then
    # rather than 42703-ing the whole update.
    tags: Optional[list[str]] = None


@router.get("/metrics/{key}")
async def metric_drilldown(
    key: str,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Return the underlying records behind a leadership-dashboard metric.

    Generic shape: {title, columns:[{key,label}], rows:[{...}], count}.
    Lets the frontend render any metric's detail in one table component.
    """
    ENGAGED = ("initial_outreach", "active", "on_hold")

    # ---- contact-based metrics ----
    contact_cols = [
        {"key": "full_name", "label": "Name"},
        {"key": "current_company", "label": "Company"},
        {"key": "current_title", "label": "Title"},
        {"key": "contact_stage", "label": "Stage"},
        {"key": "email", "label": "Email"},
    ]
    # ---- deal-based metrics ----
    deal_cols = [
        {"key": "account_name", "label": "Company"},
        {"key": "stage", "label": "Stage"},
        {"key": "deal_type", "label": "Type"},
        {"key": "owner_email", "label": "Owner"},
        {"key": "title", "label": "Role"},
    ]
    # ---- application-based metrics ----
    app_cols = [
        {"key": "company_name", "label": "Company"},
        {"key": "role_title", "label": "Role"},
        {"key": "stage", "label": "Stage"},
        {"key": "date_applied", "label": "Applied"},
    ]

    async def contacts(where: str):
        rows = await conn.fetch(
            f"SELECT contact_id AS id, full_name, current_company, current_title, contact_stage, email "
            f"FROM public.contacts WHERE is_jobs_contact=true AND {where} "
            f"ORDER BY full_name"
        )
        return contact_cols, [dict(r) for r in rows], "contact"

    async def deals(where: str):
        rows = await conn.fetch(
            f"SELECT id, account_name, stage, deal_type, owner_email, title "
            f"FROM bedrock.jobs_opportunity WHERE deleted_at IS NULL AND {where} "
            f"ORDER BY account_name"
        )
        return deal_cols, [dict(r) for r in rows], "deal"

    async def engaged_prospects(_where: str):
        # Distinct jobs prospects we've actually had activity with (linked via
        # participant_public_contact_id by the jobs-activity-link pass).
        rows = await conn.fetch(
            "SELECT contact_id AS id, full_name, current_company, current_title, contact_stage, email "
            "FROM public.contacts ct WHERE ct.is_jobs_contact=true AND EXISTS("
            "  SELECT 1 FROM bedrock.activity a "
            "  WHERE a.participant_public_contact_id = ct.contact_id AND a.deleted_at IS NULL "
            f"    AND {_jobs_relevant('a')}) "
            "ORDER BY full_name"
        )
        return contact_cols, [dict(r) for r in rows], "contact"

    # First-touch drill: one row per external contact, with their first-ever
    # touch by the jobs team and (when their email matches a jobs prospect)
    # the known prospect name/company.
    ft_cols = [
        {"key": "email", "label": "Contact"},
        {"key": "full_name", "label": "Known Prospect"},
        {"key": "current_company", "label": "Company"},
        {"key": "first_touch", "label": "First Touch"},
    ]

    async def first_touch(kind: str, where: str):
        cte = _first_touch_email_cte() if kind == "email" else _first_touch_meeting_cte()
        rows = await conn.fetch(f"""
            {cte}
            SELECT ext.counterpart AS email, ext.first_touch,
                   p.full_name, p.current_company
            FROM ext
            LEFT JOIN LATERAL (
                SELECT full_name, current_company FROM public.contacts c
                WHERE lower(c.email) = ext.counterpart AND c.is_jobs_contact = true
                LIMIT 1
            ) p ON true
            WHERE {where}
            ORDER BY ext.first_touch DESC
            LIMIT 500
        """)
        return ft_cols, [dict(r) for r in rows], "activity"

    # Company-level rollup for the "Companies with…" metrics — one row per company.
    company_cols = [
        {"key": "company_name", "label": "Company"},
        {"key": "candidates", "label": "# Candidates"},
        {"key": "roles", "label": "Roles"},
    ]

    async def companies(where: str):
        rows = await conn.fetch(
            f"""
            SELECT company_name,
                   count(*)                              AS candidates,
                   string_agg(DISTINCT role_title, ', ') AS roles
            FROM public.job_applications
            WHERE source_type='Pursuit_referred' AND {where}
            GROUP BY company_name
            ORDER BY count(*) DESC, company_name
            """
        )
        return company_cols, [dict(r) for r in rows], "company"

    async def placements(where: str):
        # Drill for the "FT Roles Secured" headline. Flat table, one row per role:
        # Company | Builder | Status | Role. Three kinds —
        #   (1) FT-placed  → builder filled the full-time seat
        #   (2) trial active → builder is in a committed trial (converts to a separate
        #        open FT req, so it never inflates the FT number; Fowler/Ethan)
        #   (3) committed open req → seat locked in, no builder placed yet (Builder = —)
        # The headline count = (1)+(3); trials are shown here but not counted.
        out = []
        # (1) FT-placed builders (full_time employment_records)
        placed = await conn.fetch(
            "SELECT s.*, (SELECT r2.opportunity_id::text FROM bedrock.jobs_role r2 "
            "             WHERE r2.employment_record_id = s.id LIMIT 1) AS opp_id "
            "FROM bedrock.secured_jobs() s "
            f"WHERE s.payment_amount > 0 AND s.employment_type = 'full_time' AND {where} AND {_live_placement('s')} ORDER BY s.builder"
        )
        for r in placed:
            out.append({
                "id": str(r["id"]), "kind": "placed",
                "opportunity_id": r["opp_id"],
                "company": r["company_name"] or "—",
                "builder": r["builder"] or "—",
                "status": "Full-time placed",
                "role": r["role_title"] or "—",
                "salary": str(int(r["payment_amount"])) if r["payment_amount"] else "",
                "counted": "✓",
            })
        # (2) committed active trials — builder in a trial (converts to the open FT req below)
        trials = await conn.fetch("""
            SELECT r.id, r.approx_salary, r.opportunity_id::text AS opp_id, o.account_name, r.title, s.builder
            FROM bedrock.jobs_role r
            JOIN bedrock.jobs_opportunity o ON o.id = r.opportunity_id
            LEFT JOIN bedrock.secured_jobs() s ON s.id = r.employment_record_id
            WHERE o.deleted_at IS NULL AND r.commitment = 'committed' AND r.is_trial = true
              AND r.filled_by_user_id IS NOT NULL AND r.status <> 'cancelled'
            ORDER BY o.account_name
        """)
        for tr in trials:
            out.append({
                "id": str(tr["id"]), "kind": "role",
                "opportunity_id": tr["opp_id"],
                "company": tr["account_name"] or "—",
                "builder": tr["builder"] or "—",
                "status": "Trial active — counts on conversion",
                "role": tr["title"] or "Trial",
                "salary": str(int(tr["approx_salary"])) if tr["approx_salary"] else "",
                "counted": "—",
            })
        # (3) committed FT roles still open — no builder placed yet
        committed = await conn.fetch("""
            SELECT r.id, r.approx_salary, r.opportunity_id::text AS opp_id, o.account_name, r.title
            FROM bedrock.jobs_role r
            JOIN bedrock.jobs_opportunity o ON o.id = r.opportunity_id
            WHERE r.status = 'open' AND o.deleted_at IS NULL
              AND r.commitment = 'committed' AND r.is_trial = false
              AND (r.employment_type = 'full_time' OR (r.employment_type IS NULL AND o.deal_type = 'ft'))
            ORDER BY o.account_name
        """)
        for cr in committed:
            out.append({
                "id": str(cr["id"]), "kind": "role",
                "opportunity_id": cr["opp_id"],
                "company": cr["account_name"] or "—",
                "builder": "—",
                "status": "Committed – open req",
                "role": cr["title"] or "FT role",
                "salary": str(int(cr["approx_salary"])) if cr["approx_salary"] else "",
                "counted": "✓",
            })
        # Sort so one company's rows sit together — a trial + its conversion seat
        # read as one story instead of a confusing duplicate (JPMC, Fowler).
        out.sort(key=lambda r: (r["company"].lower(), r["counted"] != "✓"))
        cols = [
            {"key": "company", "label": "Company"},
            {"key": "builder", "label": "Builder"},
            {"key": "status",  "label": "Status"},
            {"key": "role",    "label": "Role"},
            {"key": "salary",  "label": "Pay"},
            {"key": "counted", "label": "In FT #"},
        ]
        return cols, out, "placement"

    async def salaries(_where: str):
        # Flat, EDITABLE breakdown of everything feeding Avg FT Salary: each FT
        # placement (employment_record) + each committed open FT role, with the
        # id needed to edit it inline. Placed rows edit via the placement; committed
        # via the role (both stay in sync once filled).
        placed = await conn.fetch(
            "SELECT id, builder, company_name, role_title, payment_amount "
            f"FROM bedrock.secured_jobs() WHERE payment_amount > 0 AND employment_type='full_time' AND {_live_placement()} ORDER BY builder")
        committed = await conn.fetch("""
            SELECT r.id, o.account_name, r.title, r.approx_salary
            FROM bedrock.jobs_role r JOIN bedrock.jobs_opportunity o ON o.id = r.opportunity_id
            WHERE r.status='open' AND o.deleted_at IS NULL AND r.commitment='committed' AND r.is_trial=false
              AND (r.employment_type='full_time' OR (r.employment_type IS NULL AND o.deal_type='ft'))
            ORDER BY o.account_name""")
        out = []
        for r in placed:
            out.append({"id": str(r["id"]), "kind": "placed", "name": r["builder"],
                        "where": r["company_name"] or "—", "role": r["role_title"] or "—",
                        "status": "Placed", "salary": str(int(r["payment_amount"])) if r["payment_amount"] else ""})
        for r in committed:
            out.append({"id": str(r["id"]), "kind": "committed", "name": r["account_name"] or "—",
                        "where": r["account_name"] or "—", "role": r["title"] or "FT role",
                        "status": "Committed", "salary": str(int(r["approx_salary"])) if r["approx_salary"] else ""})
        cols = [
            {"key": "name", "label": "Builder / Account"},
            {"key": "role", "label": "Role"},
            {"key": "status", "label": "Status"},
            {"key": "salary", "label": "Salary"},
        ]
        return cols, out, "salary"

    async def any_paid(_where: str):
        # Paid work = recorded pay OR paid-type work whose pay wasn't recorded
        # (contract/freelance/part-time from the Pathfinder era often has no
        # amount — TKT-129: 'sometimes counted, sometimes not' depended on
        # whether someone filled the pay field). pro_bono stays excluded.
        rows = await conn.fetch(f"""
            SELECT * FROM bedrock.secured_jobs()
            WHERE (payment_amount > 0
               OR (coalesce(payment_amount, 0) = 0 AND employment_type IN ('contract','freelance','part_time')))
              AND {_live_placement()}
            ORDER BY builder""")
        # One row PER PAID ROLE: builders with several paid engagements show each
        # (the headline still counts distinct builders). Grouped by builder.
        rows = sorted(rows, key=lambda r: ((r["builder"] or "").lower(),
                                           r["employment_type"] != "full_time"))
        out = [{
            "builder": r["builder"],
            "company": r["company_name"] or "—",
            "role": r["role_title"] or "—",
            "type": ("Full-Time" if r["employment_type"] == "full_time"
                     else (r["employment_type"] or "—").replace("_", " ").title()),
            "salary": f"${int(r['payment_amount']):,}" if r["payment_amount"] else "— (pay unrecorded)",
        } for r in rows]
        cols = [{"key": "builder", "label": "Builder"}, {"key": "company", "label": "Company"},
                {"key": "role", "label": "Role"}, {"key": "type", "label": "Type"}, {"key": "salary", "label": "Pay"}]
        return cols, out, "builder"

    async def committed_roles(_where: str):
        rows = await conn.fetch("""
            SELECT o.account_name, r.title, r.approx_salary
            FROM bedrock.jobs_role r JOIN bedrock.jobs_opportunity o ON o.id = r.opportunity_id
            WHERE r.status='open' AND o.deleted_at IS NULL AND r.commitment='committed' AND r.is_trial=false
              AND (r.employment_type='full_time' OR (r.employment_type IS NULL AND o.deal_type='ft'))
            ORDER BY o.account_name
        """)
        out = [{"company": r["account_name"] or "—", "role": r["title"] or "FT role",
                "salary": f"${int(r['approx_salary']):,}" if r["approx_salary"] else "—"} for r in rows]
        cols = [{"key": "company", "label": "Company"}, {"key": "role", "label": "Role"}, {"key": "salary", "label": "Expected Pay"}]
        return cols, out, "role"

    async def builders_interviewing(_where: str):
        rows = await conn.fetch("""
            SELECT trim(split_part(notes, ':', 1)) AS builder, company_name, role_title
            FROM public.job_applications
            WHERE stage='interview' AND source_type='Pursuit_referred'
            ORDER BY company_name
        """)
        out = [{"builder": r["builder"] or "—", "company": r["company_name"] or "—",
                "role": r["role_title"] or "—", "stage": "Interviewing"} for r in rows]
        cols = [{"key": "builder", "label": "Builder"}, {"key": "company", "label": "Company"},
                {"key": "role", "label": "Role"}, {"key": "stage", "label": "Stage"}]
        return cols, out, "builder"

    DISPATCH = {
        "any_paid":             ("Builders With Paid Work",   lambda: any_paid("true")),
        "committed_roles":      ("Committed FT Roles (unfilled)", lambda: committed_roles("true")),
        "interviewing_builders": ("Builders Interviewing",    lambda: builders_interviewing("true")),
        "total_leads":          ("Total Leads",              lambda: contacts("true")),
        "engaged_leads":        ("Engaged Prospects",        lambda: engaged_prospects("true")),
        "outreach_week":        ("New Contacts Emailed — Last 7 Days", lambda: first_touch("email", "ext.first_touch >= now() - interval '7 days'")),
        "outreach_total":       ("Contacts Emailed — All Time",        lambda: first_touch("email", "true")),
        "calls_total":          ("Contacts Met — All Time",            lambda: first_touch("meeting", "true")),
        "calls_week":           ("New Contacts Met — Last 7 Days",     lambda: first_touch("meeting", "ext.first_touch >= now() - interval '7 days'")),
        "active_orgs":          ("Active Orgs",              lambda: deals("stage LIKE 'active_%'")),
        "active_companies":     ("Active Companies",         lambda: deals("stage LIKE 'active_%'")),
        "in_discussion":        ("In Discussion",            lambda: deals("stage='active_in_discussions'")),
        "builder_interviews":   ("Reviewing Builders",        lambda: deals("stage IN ('reviewing_builders','active_builder_interview')")),
        "placements":           ("FT Roles Secured", lambda: placements("true")),
        "ft_salaries":          ("FT Salaries", lambda: salaries("true")),
        "candidates_submitted": ("Companies w/ Candidates Submitted", lambda: companies("stage IN ('applied','interview','accepted')")),
        "interviewing":         ("Companies Interviewing Builders",   lambda: companies("stage='interview'")),
    }

    if key not in DISPATCH:
        raise HTTPException(404, f"Unknown metric: {key}")

    title, fn = DISPATCH[key]
    columns, rows, entity = await fn()
    # columns may be a list (flat table) or a dict {columns, child_columns} (expandable)
    child_columns = None
    if isinstance(columns, dict):
        child_columns = columns.get("child_columns")
        columns = columns["columns"]
    # serialize dates + ids
    for r in rows:
        for k, v in list(r.items()):
            if hasattr(v, "isoformat"):
                r[k] = v.isoformat()
            elif k == "id" and v is not None:
                r[k] = str(v)
    return {
        "success": True,
        "data": {
            "title": title, "columns": columns, "rows": rows,
            "count": len(rows), "entity": entity, "child_columns": child_columns,
        },
    }


@router.get("/placements")
async def get_placements(
    segment: Optional[str] = Query(None),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Secured jobs — single source of truth = public.employment_records.

    Counts ALL placements (incl. builder self-sourced, no deal link) and
    separates by `influenced`. `segment` (an L3 cohort) scopes the builder-side
    numbers to that segment of the L3+ pool; committed roles stay global (demand).
    """
    seg = segment if segment and segment != "all" else None
    seg_uids: Optional[set] = None
    if seg:
        prows = await conn.fetch(f"WITH {_L3PLUS_POOL} SELECT user_id, segment FROM pool")
        seg_uids = {r["user_id"] for r in prows if r["segment"] == seg}

    # Same inclusion rule as the any_paid drill: typed paid work counts even
    # when the pay amount wasn't recorded (TKT-129); pro_bono stays excluded.
    rows = await conn.fetch(f"""
        SELECT * FROM bedrock.secured_jobs()
        WHERE (payment_amount > 0
           OR (coalesce(payment_amount, 0) = 0 AND employment_type IN ('contract','freelance','part_time')))
          AND {_live_placement()}""")
    if seg_uids is not None:
        rows = [r for r in rows if r["user_id"] in seg_uids]

    # Metric is DISTINCT BUILDERS placed (a builder with 2 PT jobs counts once).
    # Two tracked numbers: any paid work, and full-time.
    def best(a, b):
        """Pick the more representative placement for a builder: FT > influenced > higher pay."""
        ka = (a["employment_type"] == "full_time", a["influenced"] is True, a["payment_amount"] or 0)
        kb = (b["employment_type"] == "full_time", b["influenced"] is True, b["payment_amount"] or 0)
        return a if ka >= kb else b

    by_builder: dict = {}
    for r in rows:
        uid = r["user_id"]
        by_builder[uid] = best(by_builder[uid], r) if uid in by_builder else r

    # A builder is FT-placed if ANY of their paid records is full_time.
    ft_uids   = {r["user_id"] for r in rows if r["employment_type"] == "full_time"}
    any_uids  = set(by_builder.keys())
    infl_uids = {r["user_id"] for r in rows if r["influenced"] is True}

    ft_builders   = len(ft_uids)
    any_builders  = len(any_uids)
    infl_ft       = len(ft_uids & infl_uids)
    infl_any      = len(any_uids & infl_uids)

    # Committed FT roles still OPEN (unfilled reqs) on full-time opportunities —
    # demand the team has locked in but not yet placed a builder into. Excludes
    # open-market roles (CVs welcome but no hiring commitment) and trials (those
    # convert into a separate FT role; the FT role is what counts as committed).
    committed_ft_roles = await conn.fetchval("""
        SELECT count(*) FROM bedrock.jobs_role r
        JOIN bedrock.jobs_opportunity o ON o.id = r.opportunity_id
        WHERE r.status = 'open' AND o.deleted_at IS NULL
          AND r.commitment = 'committed' AND r.is_trial = false
          AND (r.employment_type = 'full_time' OR (r.employment_type IS NULL AND o.deal_type = 'ft'))
    """) or 0

    # Avg salaries: per FT-placed builder's representative FT pay (placed), and
    # that pool blended with committed FT roles' expected salary (secured).
    ft_pay = {}
    for r in rows:
        if r["employment_type"] == "full_time" and r["payment_amount"]:
            uid = r["user_id"]
            ft_pay[uid] = max(ft_pay.get(uid, 0), float(r["payment_amount"]))
    placed_salaries = list(ft_pay.values())
    avg_salary_ft_placed = round(sum(placed_salaries) / len(placed_salaries)) if placed_salaries else None
    crows = await conn.fetch("""
        SELECT r.approx_salary FROM bedrock.jobs_role r
        JOIN bedrock.jobs_opportunity o ON o.id = r.opportunity_id
        WHERE r.status='open' AND o.deleted_at IS NULL AND r.commitment='committed' AND r.is_trial=false
          AND (r.employment_type='full_time' OR (r.employment_type IS NULL AND o.deal_type='ft'))
          AND r.approx_salary IS NOT NULL
    """)
    committed_salaries = [float(r["approx_salary"]) for r in crows if r["approx_salary"]]
    secured_salaries = placed_salaries + committed_salaries
    avg_salary_ft_secured = round(sum(secured_salaries) / len(secured_salaries)) if secured_salaries else None

    # Builders currently in a committed, active paid trial (someone IS in the trial;
    # its FT conversion is a separate role that stays open until they convert). This
    # is the "Committed: trial active" status — surfaced so a trial like Fowler/Ethan
    # is neither counted as FT-placed nor invisible (fixes the Home vs Accounts gap).
    trial_rows = await conn.fetch("""
        SELECT DISTINCT r.filled_by_user_id AS uid FROM bedrock.jobs_role r
        JOIN bedrock.jobs_opportunity o ON o.id = r.opportunity_id
        WHERE o.deleted_at IS NULL AND r.commitment = 'committed' AND r.is_trial = true
          AND r.filled_by_user_id IS NOT NULL AND r.status <> 'cancelled'
    """)
    trial_uids = {r["uid"] for r in trial_rows}
    if seg_uids is not None:
        trial_uids &= seg_uids
    committed_trial_active = len(trial_uids)

    # Builders currently interviewing (job_applications), optionally segment-scoped.
    iv_rows = await conn.fetch("""
        SELECT DISTINCT ja.builder_id FROM public.job_applications ja
        WHERE ja.stage = 'interview' AND ja.source_type = 'Pursuit_referred' AND ja.builder_id IS NOT NULL
    """)
    iv_uids = {r["builder_id"] for r in iv_rows}
    if seg_uids is not None:
        iv_uids &= seg_uids
    interviewing = len(iv_uids)

    # One row per builder for the drill list (their representative placement)
    out = []
    for uid, r in sorted(by_builder.items(),
                         key=lambda kv: (kv[1]["employment_type"] != "full_time", kv[1]["builder"] or "")):
        et = r["employment_type"]
        type_label = ("Full-Time" if et == "full_time"
                      else "PT / Contract" if et in ("contract", "freelance")
                      else (et or "—").replace("_", " ").title())
        out.append({
            "id": str(r["id"]),
            "builder": r["builder"],
            "role_title": r["role_title"] or "—",
            "company_name": r["company_name"] or "—",
            "employment_type": type_label,
            "ft_placed": uid in ft_uids,
            "influenced": r["influenced"],
            "salary": int(r["payment_amount"]) if r["payment_amount"] else None,
            "source": r["source"],
        })

    return {
        "success": True,
        "data": {
            "ft_builders":  ft_builders,
            "any_builders": any_builders,
            "influenced_ft":  infl_ft,
            "influenced_any": infl_any,
            "committed_ft_roles": committed_ft_roles,
            "committed_trial_active": committed_trial_active,
            # Committed roles have no builder, so no cohort — including them in
            # a cohort-scoped headline repeated the same roles under every
            # cohort (TKT-127). Under a segment they're reported separately
            # (committed_ft_roles) and excluded from the additive number.
            "ft_roles_secured": ft_builders + (0 if seg else committed_ft_roles),
            "committed_is_global": True,
            "avg_salary_ft_placed": avg_salary_ft_placed,
            "avg_salary_ft_secured": avg_salary_ft_secured,
            "interviewing": interviewing,
            "rows": out,
        },
    }


@router.get("/staff")
async def list_staff(
    q: Optional[str] = Query(None),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Active Pursuit staff (for owner pickers). Deduped by PERSON — several
    staff have two active org_users rows (a short alias like greg@ and a
    full-name greg.hogue@), which showed the same name twice. Keep one row per
    person, valued by the email that actually appears in sent activity (so the
    owner filter, which matches email_from, still works), else the shorter."""
    rows = await conn.fetch("""
        WITH staff AS (
          SELECT o.email, o.display_name,
                 (SELECT count(*) FROM bedrock.activity a
                  WHERE a.email_from ILIKE '%'||o.email||'%') AS sent
          FROM public.org_users o
          WHERE o.is_active = true
            -- real Pursuit people only: drop bots/system accounts (bug-fix-agent@
            -- pursuit-factory.local 'Brainiac', systems@ 'Systems Admin')
            AND o.email LIKE '%@pursuit.org'
            AND o.email NOT IN ('systems@pursuit.org')
        ),
        deduped AS (
          SELECT DISTINCT ON (lower(display_name)) email, display_name, sent
          FROM staff
          ORDER BY lower(display_name), sent DESC, length(email)
        )
        -- an owner picker only needs people who actually do outreach; this also
        -- drops never-active / departed rows that linger with is_active=true.
        SELECT email, display_name FROM deduped WHERE sent > 0 ORDER BY lower(display_name)
    """)
    out = [{"email": r["email"], "name": r["display_name"] or r["email"]} for r in rows]
    if q:
        ql = q.lower()
        out = [s for s in out if ql in s["email"].lower() or ql in (s["name"] or "").lower()]
    out.sort(key=lambda s: s["name"])
    return {"success": True, "data": out[:50]}


@router.get("/placements/unlinked")
async def search_unlinked_placements(
    q: Optional[str] = Query(None),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Existing employment_records not yet tied to an opportunity (for linking on win)."""
    rows = await conn.fetch("""
        SELECT er.id, er.role_title, er.company_name, er.employment_type,
               COALESCE(NULLIF(trim(u.first_name||' '||u.last_name),''),
                        'Builder #'||er.user_id) AS builder
        FROM public.employment_records er
        LEFT JOIN public.users u ON u.user_id = er.user_id
        WHERE er.opportunity_id IS NULL
        ORDER BY er.id DESC LIMIT 200
    """)
    out = [
        {"id": str(r["id"]), "builder": r["builder"], "role_title": r["role_title"],
         "company_name": r["company_name"], "employment_type": r["employment_type"]}
        for r in rows
    ]
    if q:
        ql = q.lower()
        out = [r for r in out if ql in (r["builder"] or "").lower()
               or ql in (r["company_name"] or "").lower()
               or ql in (r["role_title"] or "").lower()]
    return {"success": True, "data": out[:30]}


@router.get("/opportunities/{opp_id}/placements")
async def list_opp_placements(
    opp_id: UUID,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Placements already linked to this opportunity."""
    rows = await conn.fetch("""
        SELECT er.id, er.role_title, er.company_name, er.employment_type,
               er.payment_amount, er.user_id,
               COALESCE(NULLIF(trim(u.first_name||' '||u.last_name),''),
                        'Builder #'||er.user_id) AS builder
        FROM public.employment_records er
        LEFT JOIN public.users u ON u.user_id = er.user_id
        WHERE er.opportunity_id = $1
        ORDER BY er.id
    """, opp_id)
    return {"success": True, "data": [
        {"id": str(r["id"]), "builder": r["builder"], "role_title": r["role_title"],
         "company_name": r["company_name"], "employment_type": r["employment_type"],
         "salary": int(r["payment_amount"]) if r["payment_amount"] else None}
        for r in rows
    ]}


class PlacementCreate(BaseModel):
    builder_user_id: Optional[int] = None   # link to a platform builder if known
    builder_name:    Optional[str] = None   # free-text fallback
    role_title:      Optional[str] = None
    # full_time | contract | freelance | pro_bono. None → derived from the
    # opportunity's deal_type (pt_contract → contract). A hardcoded full_time
    # default here mis-typed contract placements as FT (Sam MacFarlane / RBM).
    employment_type: Optional[str] = None
    salary:          Optional[int] = None


def _default_employment_type(deal_type: Optional[str]) -> str:
    """Deal-type-aware fallback for placements missing an explicit type."""
    return "contract" if deal_type == "pt_contract" else "full_time"


@router.post("/opportunities/{opp_id}/placements")
async def create_opp_placement(
    opp_id: UUID,
    body: PlacementCreate,
    user=Depends(require_auth),
    conn=Depends(get_db),
    client=Depends(get_mcp_client),
):
    """Create a new secured-job placement under this opportunity (jobs-team influenced)."""
    opp = await conn.fetchrow(
        "SELECT account_id, account_name, deal_type FROM bedrock.jobs_opportunity WHERE id=$1 AND deleted_at IS NULL",
        opp_id,
    )
    if not opp:
        raise HTTPException(404, "Opportunity not found")
    emp_type = body.employment_type or _default_employment_type(opp["deal_type"])

    user_email = user.get("email") if isinstance(user, dict) else getattr(user, "email", None)
    # employment_records.user_id is NOT NULL — require a builder_user_id, or store name in story/notes
    if not body.builder_user_id and not body.builder_name:
        raise HTTPException(400, "Provide builder_user_id or builder_name")

    role = body.role_title or "(role TBD)"
    notes = f"{body.builder_name}: jobs-team placement" if body.builder_name else "jobs-team placement"

    if body.builder_user_id:
        # Dedup guard: if this builder already has a placement at this company,
        # enrich it with jobs-team attribution instead of creating a duplicate.
        existing = await conn.fetchrow("""
            SELECT id FROM public.employment_records
            WHERE user_id = $1 AND lower(company_name) = lower($2)
            ORDER BY id LIMIT 1
        """, body.builder_user_id, opp["account_name"])
        if existing:
            await conn.execute("""
                UPDATE public.employment_records
                SET opportunity_id=$1, influenced=true,
                    role_title=COALESCE(role_title, $2),
                    employment_type=COALESCE(NULLIF(employment_type,''), $3),
                    payment_amount=COALESCE(payment_amount, $4),
                    updated_at=now()
                WHERE id=$5
            """, opp_id, role, emp_type, body.salary, existing["id"])
            rec_id, merged = existing["id"], True
        else:
            rec_id = await conn.fetchval("""
                INSERT INTO public.employment_records
                    (user_id, role_title, company_name, employment_type, engagement_stage,
                     payment_amount, source, opportunity_id, influenced, notes)
                VALUES ($1,$2,$3,$4,'completed',$5,'staff_created',$6,true,$7)
                RETURNING id
            """, body.builder_user_id, role, opp["account_name"], emp_type,
                body.salary, opp_id, notes)
            merged = False
    else:
        # No platform user — store as a name-only record (user_id required, use 0 sentinel won't work
        # if FK; instead reject). Most placements should link a real builder.
        raise HTTPException(400, "builder_user_id required to create a placement (name-only not supported)")

    # Mirror to Salesforce like the hire path — previously this endpoint never
    # synced, so placements made here silently never reached SF (and never
    # showed in sf-sync-status). Best-effort; failures are recorded + retryable.
    sf_sync: dict = {"status": "skipped", "message": "Salesforce not connected — sync from the placement later."}
    try:
        if client and "salesforce" in (client.connected_services or []):
            result = await sync_placement_to_sf(conn, client.salesforce, rec_id)
            sf_sync = {"status": "synced", **result}
        else:
            await record_sync_error(conn, rec_id, "Salesforce not connected at placement create time")
    except NotEligible as ne:
        await record_sync_error(conn, rec_id, str(ne), status="skipped")
        sf_sync = {"status": "skipped", "message": str(ne)}
    except AccountAmbiguous as aa:
        await record_sync_error(conn, rec_id, str(aa), status="needs_choice")
        sf_sync = {"status": "needs_choice", "message": str(aa)}
    except Exception as e:
        await record_sync_error(conn, rec_id, str(e)[:200])
        sf_sync = {"status": "error", "message": str(e)[:200]}

    return {"success": True, "data": {"id": str(rec_id), "merged": merged, "sf_sync": sf_sync}}


@router.post("/opportunities/{opp_id}/placements/{placement_id}/link")
async def link_opp_placement(
    opp_id: UUID,
    placement_id: int,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Link an existing employment_record to this opportunity (mark jobs-team influenced)."""
    result = await conn.execute(
        "UPDATE public.employment_records SET opportunity_id=$1, influenced=true, updated_at=now() WHERE id=$2",
        opp_id, placement_id,
    )
    if result == "UPDATE 0":
        raise HTTPException(404, "Placement not found")
    return {"success": True, "data": {"id": str(placement_id), "opportunity_id": str(opp_id)}}


class PlacementUpdate(BaseModel):
    influenced: Optional[bool] = None  # true / false / null (unclassify)
    salary: Optional[int] = None       # edit payment_amount (the secured-jobs SoT)
    role_title: Optional[str] = None   # edit the title (syncs the linked role)


@router.patch("/placements/{placement_id}")
async def update_placement(
    placement_id: int,
    body: PlacementUpdate,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Edit a secured job: influence attribution and/or salary (payment_amount)."""
    sets, params, i = [], [], 1
    fields = body.model_dump(exclude_unset=True)
    if "influenced" in fields:
        sets.append(f"influenced=${i}"); params.append(body.influenced); i += 1
    if "salary" in fields:
        sets.append(f"payment_amount=${i}"); params.append(body.salary); i += 1
    if "role_title" in fields and (body.role_title or "").strip():
        sets.append(f"role_title=${i}"); params.append(body.role_title.strip()); i += 1
    if not sets:
        return {"success": True, "data": {"id": placement_id}}
    params.append(placement_id)
    result = await conn.execute(
        f"UPDATE public.employment_records SET {', '.join(sets)}, updated_at=now() WHERE id=${i}", *params)
    if result == "UPDATE 0":
        raise HTTPException(404, "Placement not found")
    # Keep the linked role in sync (the other half of the filled-role pair).
    if "salary" in fields:
        await conn.execute(
            "UPDATE bedrock.jobs_role SET approx_salary=$1, updated_at=now() WHERE employment_record_id=$2",
            body.salary, placement_id)
    if "role_title" in fields and (body.role_title or "").strip():
        await conn.execute(
            "UPDATE bedrock.jobs_role SET title=$1, updated_at=now() WHERE employment_record_id=$2",
            body.role_title.strip(), placement_id)
    return {"success": True, "data": {"id": placement_id, **fields}}


@router.delete("/placements/{placement_id}")
async def delete_placement(placement_id: int, user=Depends(require_auth), conn=Depends(get_db)):
    """Hard-delete a placement recorded in error (TKT-161). Any role that was
    filled by it reopens. Requires the bedrock_user DELETE grant
    (db/migrations/2026-07-30-employment-records-delete.sql)."""
    row = await conn.fetchrow("SELECT id, company_name FROM public.employment_records WHERE id=$1", placement_id)
    if not row:
        raise HTTPException(404, "Placement not found")
    try:
        async with conn.transaction():
            await conn.execute(
                "UPDATE bedrock.jobs_role SET status='open', filled_by_user_id=NULL, "
                "employment_record_id=NULL, updated_at=now() WHERE employment_record_id=$1",
                placement_id)
            await conn.execute("DELETE FROM public.employment_records WHERE id=$1", placement_id)
    except asyncpg.exceptions.InsufficientPrivilegeError:
        raise HTTPException(503, "Placement deletion needs a DB grant that isn't applied yet — ask Jac to run the 2026-07-30 employment-records-delete migration.")
    return {"success": True, "data": {"deleted": placement_id}}


class RelevanceOverride(BaseModel):
    override: Optional[str] = None  # 'jobs' | 'not_jobs' | None (clear → back to AI verdict)


@router.patch("/activity/{activity_id}/relevance")
async def set_activity_relevance(
    activity_id: str,
    body: RelevanceOverride,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Human override of the jobs-relevance classifier. Open to ALL staff —
    the person who had the conversation knows better than the model. Metrics
    gate on coalesce(override, classifier verdict); clearing restores the AI
    verdict. Audit-trailed (who/when)."""
    if body.override not in (None, "jobs", "not_jobs"):
        raise HTTPException(400, "override must be 'jobs', 'not_jobs', or null")
    who = user.get("email") if isinstance(user, dict) else getattr(user, "email", None)
    result = await conn.execute(
        """UPDATE bedrock.activity
           SET jobs_relevance_override = $1::text,
               jobs_relevance_override_by = CASE WHEN $1::text IS NULL THEN NULL ELSE $2::text END,
               jobs_relevance_override_at = CASE WHEN $1::text IS NULL THEN NULL ELSE now() END,
               updated_at = now()
           WHERE id = $3::uuid AND deleted_at IS NULL""",
        body.override, who, activity_id)
    if result == "UPDATE 0":
        raise HTTPException(404, "Activity not found")
    return {"success": True, "data": {"id": activity_id, "override": body.override}}


# ── Roles (jobs_role) — open roles on an opportunity ──────────────────────────

VALID_ROLE_STATUSES = {"open", "filled", "cancelled"}


VALID_COMMITMENTS = {"committed", "open_market"}
VALID_RATE_PERIODS = {"annual", "monthly", "weekly", "daily", "hourly"}


def _placement_status_py(r) -> str:
    """Python mirror of _placement_status_sql — same derivation, for serialized rows."""
    if r.get("status") == "cancelled":
        return "cancelled"
    filled = r.get("filled_by_user_id") is not None
    if filled and not bool(r.get("is_trial")):
        return "ft_placed"
    if filled and r.get("is_trial"):
        return "trial_active"
    if r.get("commitment") == "committed":
        return "committed_open"
    return "open_market"


def _role_dict(r) -> dict:
    """Serialize a bedrock.jobs_role row for the API."""
    d = dict(r)
    d["id"] = str(d["id"])
    _st = _placement_status_py(d)
    d["placement_status"] = _st
    d["placement_status_label"] = PLACEMENT_STATUS_LABELS.get(_st, _st)
    if d.get("opportunity_id") is not None:
        d["opportunity_id"] = str(d["opportunity_id"])
    if d.get("converts_to_role_id") is not None:
        d["converts_to_role_id"] = str(d["converts_to_role_id"])
    if d.get("pay_rate") is not None:
        d["pay_rate"] = float(d["pay_rate"])
    for k in ("start_date", "end_date", "created_at", "updated_at"):
        v = d.get(k)
        if v is not None and hasattr(v, "isoformat"):
            d[k] = v.isoformat()
    return d


class RoleCreate(BaseModel):
    title:               str
    allow_duplicate:     bool = False   # bypass the rapid-duplicate guard
    approx_salary:       Optional[int] = None
    employment_type:     Optional[str] = None
    start_date:          Optional[date] = None
    notes:               Optional[str] = None
    commitment:          Optional[str] = None          # committed | open_market (default committed)
    is_trial:            Optional[bool] = None
    converts_to_role_id: Optional[str] = None
    pay_rate:            Optional[float] = None
    rate_period:         Optional[str] = None           # annual | monthly | weekly | daily | hourly
    end_date:            Optional[date] = None
    pay_cadence:         Optional[str] = None
    benefits:            Optional[str] = None
    payment_schedule:    Optional[str] = None
    negotiation_notes:   Optional[str] = None
    jd_url:              Optional[str] = None
    pathfinder_visible:  Optional[bool] = None   # publish to the builder-facing Pathfinder feed


class RoleUpdate(BaseModel):
    title:               Optional[str] = None
    approx_salary:       Optional[int] = None
    employment_type:     Optional[str] = None
    start_date:          Optional[date] = None
    status:              Optional[str] = None
    notes:               Optional[str] = None
    commitment:          Optional[str] = None
    is_trial:            Optional[bool] = None
    converts_to_role_id: Optional[str] = None
    pay_rate:            Optional[float] = None
    rate_period:         Optional[str] = None
    end_date:            Optional[date] = None
    pay_cadence:         Optional[str] = None
    benefits:            Optional[str] = None
    payment_schedule:    Optional[str] = None
    negotiation_notes:   Optional[str] = None
    jd_url:              Optional[str] = None
    pathfinder_visible:  Optional[bool] = None   # publish to the builder-facing Pathfinder feed


@router.get("/opportunities/{opp_id}/roles")
async def list_opp_roles(
    opp_id: UUID,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """All roles (any status) for an opportunity, oldest first."""
    rows = await conn.fetch(
        "SELECT * FROM bedrock.jobs_role WHERE opportunity_id=$1 ORDER BY created_at",
        opp_id,
    )
    return {"success": True, "data": [_role_dict(r) for r in rows]}


@router.post("/opportunities/{opp_id}/roles")
async def create_opp_role(
    opp_id: UUID,
    body: RoleCreate,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Create an open role on an opportunity."""
    if body.commitment is not None and body.commitment not in VALID_COMMITMENTS:
        raise HTTPException(400, f"Invalid commitment: {body.commitment}")
    if body.rate_period is not None and body.rate_period not in VALID_RATE_PERIODS:
        raise HTTPException(400, f"Invalid rate_period: {body.rate_period}")
    opp = await conn.fetchrow(
        "SELECT id FROM bedrock.jobs_opportunity WHERE id=$1 AND deleted_at IS NULL",
        opp_id,
    )
    if not opp:
        raise HTTPException(404, "Opportunity not found")

    # Guard against accidental duplicate roles (the Citizens Bank ×5 case: an
    # identical open role re-added seconds apart). Block an exact-title open
    # role created on this opp in the last 5 min unless explicitly forced.
    if body.title and not getattr(body, "allow_duplicate", False):
        dup = await conn.fetchval(
            """SELECT 1 FROM bedrock.jobs_role
               WHERE opportunity_id=$1 AND status='open'
                 AND lower(trim(title)) = lower(trim($2))
                 AND created_at > now() - interval '5 minutes' LIMIT 1""",
            opp_id, body.title)
        if dup:
            raise HTTPException(409, {
                "error": "duplicate_role",
                "message": f"An open '{body.title}' role was just added to this opportunity. "
                           "For multiple seats, set the role's number of seats instead of adding it again.",
            })

    converts_to = UUID(body.converts_to_role_id) if body.converts_to_role_id else None
    row = await conn.fetchrow(
        """
        INSERT INTO bedrock.jobs_role
            (opportunity_id, title, approx_salary, employment_type, start_date, status, notes,
             commitment, is_trial, converts_to_role_id, pay_rate, rate_period, end_date,
             pay_cadence, benefits, payment_schedule, negotiation_notes, jd_url)
        VALUES ($1,$2,$3,$4,$5,'open',$6,
                COALESCE($7,'committed'), COALESCE($8,false), $9,$10,$11,$12,
                $13,$14,$15,$16,$17)
        RETURNING *
        """,
        opp_id, body.title, body.approx_salary, body.employment_type,
        body.start_date, body.notes,
        body.commitment, body.is_trial, converts_to, body.pay_rate, body.rate_period, body.end_date,
        body.pay_cadence, body.benefits, body.payment_schedule, body.negotiation_notes, body.jd_url,
    )
    # Publish straight to Pathfinder if the role was created visible.
    if getattr(body, "pathfinder_visible", None):
        try:
            await conn.execute(
                "UPDATE bedrock.jobs_role SET pathfinder_visible=true WHERE id=$1", row["id"])
            await conn.execute("SELECT bedrock.sync_role_to_pathfinder($1)", row["id"])
            row = await conn.fetchrow("SELECT * FROM bedrock.jobs_role WHERE id=$1", row["id"])
        except Exception as e:
            logger.warning("pathfinder sync failed for new role %s: %s", row["id"], e)
    return {"success": True, "data": _role_dict(row)}


@router.patch("/roles/{role_id}")
async def update_role(
    role_id: UUID,
    body: RoleUpdate,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Update role core fields, commitment/trial/conversion, and compensation."""
    if body.status is not None and body.status not in VALID_ROLE_STATUSES:
        raise HTTPException(400, f"Invalid status: {body.status}")
    if body.commitment is not None and body.commitment not in VALID_COMMITMENTS:
        raise HTTPException(400, f"Invalid commitment: {body.commitment}")
    if body.rate_period is not None and body.rate_period not in VALID_RATE_PERIODS:
        raise HTTPException(400, f"Invalid rate_period: {body.rate_period}")

    existing = await conn.fetchrow("SELECT * FROM bedrock.jobs_role WHERE id=$1", role_id)
    if not existing:
        raise HTTPException(404, "Role not found")

    sets, params = [], []
    i = 1
    for field in (
        "title", "approx_salary", "employment_type", "start_date", "status", "notes",
        "commitment", "is_trial", "pay_rate", "rate_period", "end_date",
        "pay_cadence", "benefits", "payment_schedule", "negotiation_notes", "jd_url",
        "pathfinder_visible",
    ):
        val = getattr(body, field, None)
        if val is not None:
            sets.append(f"{field} = ${i}"); params.append(val); i += 1
    # converts_to_role_id is a uuid column — cast the incoming string.
    if body.converts_to_role_id is not None:
        sets.append(f"converts_to_role_id = ${i}")
        params.append(UUID(body.converts_to_role_id)); i += 1

    if not sets:
        return {"success": True, "data": _role_dict(existing)}

    params.append(role_id)
    row = await conn.fetchrow(
        f"UPDATE bedrock.jobs_role SET {', '.join(sets)}, updated_at=now() WHERE id=${i} RETURNING *",
        *params,
    )
    # Keep a FILLED role's placement record in sync with the role — salary,
    # title, AND employment type. Otherwise editing the role after hiring
    # leaves the placement (which the Builders view + FT-Roles-Secured metric
    # read) on the old values (Acture: stale title; Sam MacFarlane / RBM:
    # role corrected to contract but the placement stayed full_time and kept
    # counting as an FT role secured).
    if row["employment_record_id"]:
        er_sets, er_params = [], []
        if body.approx_salary is not None:
            er_sets.append(f"payment_amount=${len(er_params)+1}"); er_params.append(body.approx_salary)
        if body.title is not None:
            er_sets.append(f"role_title=${len(er_params)+1}"); er_params.append(body.title)
        if body.employment_type is not None:
            er_sets.append(f"employment_type=${len(er_params)+1}"); er_params.append(body.employment_type)
        if er_sets:
            er_params.append(row["employment_record_id"])
            await conn.execute(
                f"UPDATE public.employment_records SET {', '.join(er_sets)}, updated_at=now() "
                f"WHERE id=${len(er_params)}", *er_params)

    # Publish/refresh to Pathfinder when visibility was toggled, or when a role
    # already linked to a posting had its source fields (title/salary/url/notes)
    # edited. Idempotent + RLS-safe via the SECURITY DEFINER function.
    if body.pathfinder_visible is not None or row["job_posting_id"] is not None:
        try:
            await conn.execute("SELECT bedrock.sync_role_to_pathfinder($1)", role_id)
            row = await conn.fetchrow("SELECT * FROM bedrock.jobs_role WHERE id=$1", role_id)
        except Exception as e:
            logger.warning("pathfinder sync failed for role %s: %s", role_id, e)
    return {"success": True, "data": _role_dict(row)}


@router.delete("/roles/{role_id}")
async def delete_role(
    role_id: UUID,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Hard-delete a role (jobs-team owned, low volume)."""
    # If this role was published to Pathfinder, unpublish its posting first —
    # otherwise the builder-facing job_postings row is orphaned (still shared,
    # pointing at a deleted role). bedrock_user can't write public.job_postings
    # directly, so go through the SECURITY DEFINER sync fn.
    posting = await conn.fetchval("SELECT job_posting_id FROM bedrock.jobs_role WHERE id=$1", role_id)
    if posting is not None:
        await conn.execute("UPDATE bedrock.jobs_role SET pathfinder_visible=false WHERE id=$1", role_id)
        try:
            await conn.execute("SELECT bedrock.sync_role_to_pathfinder($1)", role_id)
        except Exception as e:
            logger.warning("pathfinder unpublish on delete failed for %s: %s", role_id, e)
    result = await conn.execute("DELETE FROM bedrock.jobs_role WHERE id=$1", role_id)
    if result == "DELETE 0":
        raise HTTPException(404, "Role not found")
    return {"success": True, "data": {"deleted": True}}


class RoleHire(BaseModel):
    user_id:         int
    start_date:      Optional[date] = None
    salary:          Optional[int] = None
    employment_type: Optional[str] = None


@router.post("/roles/{role_id}/hire")
async def hire_into_role(
    role_id: UUID,
    body: RoleHire,
    user=Depends(require_auth),
    conn=Depends(get_db),
    client=Depends(get_mcp_client),
):
    """Fill a role: create the employment_record placement and mark the role filled."""
    role = await conn.fetchrow("SELECT * FROM bedrock.jobs_role WHERE id=$1", role_id)
    if not role:
        raise HTTPException(404, "Role not found")
    opp = await conn.fetchrow(
        "SELECT id, account_name, deal_type FROM bedrock.jobs_opportunity WHERE id=$1 AND deleted_at IS NULL",
        role["opportunity_id"],
    )
    if not opp:
        raise HTTPException(404, "Opportunity not found")

    # Fallback chain used to bottom out at 'full_time' regardless of deal type,
    # mis-typing hires on contract deals (Sam MacFarlane / RBM Maintenance).
    employment_type = body.employment_type or role["employment_type"] or _default_employment_type(opp["deal_type"])
    payment_amount  = body.salary if body.salary is not None else role["approx_salary"]
    start_date      = body.start_date or role["start_date"]

    async with conn.transaction():
        # Dedup guard (same as create_opp_placement): if this builder already
        # has a placement at this company — e.g. recorded via the Placements
        # modal before the role was filled — enrich THAT record instead of
        # inserting a duplicate. A modal-then-hire sequence used to create two
        # records, one carrying the modal's full_time default (Sam MacFarlane /
        # RBM: dup record kept counting as an FT role secured).
        existing_id = await conn.fetchval("""
            SELECT id FROM public.employment_records
            WHERE user_id = $1 AND lower(company_name) = lower($2)
            ORDER BY id LIMIT 1
        """, body.user_id, opp["account_name"])
        if existing_id:
            await conn.execute(
                """
                UPDATE public.employment_records
                SET role_title=$1, employment_type=$2, engagement_stage='active',
                    payment_amount=COALESCE($3, payment_amount),
                    opportunity_id=$4, influenced=true,
                    start_date=COALESCE($5, start_date), updated_at=now()
                WHERE id=$6
                """,
                role["title"], employment_type, payment_amount, opp["id"],
                start_date, existing_id,
            )
            new_id = existing_id
        else:
            new_id = await conn.fetchval(
                """
                INSERT INTO public.employment_records
                    (user_id, role_title, company_name, employment_type, engagement_stage,
                     payment_amount, source, opportunity_id, influenced, start_date)
                VALUES ($1,$2,$3,$4,'active',$5,'staff_created',$6,true,$7)
                RETURNING id
                """,
                body.user_id, role["title"], opp["account_name"], employment_type,
                payment_amount, opp["id"], start_date,
            )
        updated = await conn.fetchrow(
            """
            UPDATE bedrock.jobs_role
            SET status='filled', filled_by_user_id=$1, employment_record_id=$2, updated_at=now()
            WHERE id=$3
            RETURNING *
            """,
            body.user_id, new_id, role_id,
        )

    # A filled role must stop advertising in Pathfinder (the sync fn now
    # force-unpublishes when status is filled/cancelled).
    if updated["job_posting_id"]:
        try:
            await conn.execute("SELECT bedrock.sync_role_to_pathfinder($1)", role_id)
        except Exception as e:
            logger.warning("pathfinder unpublish on hire failed for %s: %s", role_id, e)

    # Mirror the placement into Salesforce (fellow contact + account +
    # affiliation) — best-effort so a SF hiccup never blocks the hire; the
    # failure is recorded and retryable via POST /placements/{id}/sync-sf.
    sf_sync: dict = {"status": "skipped", "message": "Salesforce not connected — sync from the placement later."}
    try:
        if client and "salesforce" in (client.connected_services or []):
            result = await sync_placement_to_sf(conn, client.salesforce, new_id)
            sf_sync = {"status": "synced", **result}
        else:
            await record_sync_error(conn, new_id, "Salesforce not connected at hire time")
    except NotEligible as ne:
        await record_sync_error(conn, new_id, str(ne), status="skipped")
        sf_sync = {"status": "skipped", "message": str(ne)}
    except AccountAmbiguous as aa:
        await record_sync_error(conn, new_id, str(aa), status="needs_choice")
        sf_sync = {"status": "needs_choice", "message": str(aa), "account_candidates": aa.candidates}
    except ValueError as ve:
        await record_sync_error(conn, new_id, str(ve))
        sf_sync = {"status": "needs_info", "message": str(ve)}
    except Exception as e:  # noqa: BLE001 — never fail the hire on SF errors
        logger.warning(f"placement SF sync failed for er {new_id}: {e}")
        await record_sync_error(conn, new_id, str(e))
        sf_sync = {"status": "error", "message": str(e)}

    return {"success": True, "data": {"role": _role_dict(updated), "employment_record_id": new_id, "sf_sync": sf_sync}}


@router.get("/roles/board")
async def roles_board(
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Every role across every (non-deleted) opportunity, each with its matched
    builder applications — the cross-account Placement > Roles board. Builder
    name is derived from the job_applications.notes prefix (see
    opp_builder_activity) rather than joined from public.users, which isn't
    directly readable under every DB role (e.g. avni_dev)."""
    roles = await conn.fetch(
        """
        SELECT r.*, o.account_name, o.stage AS opp_stage, s.sort_position
        FROM bedrock.jobs_role r
        JOIN bedrock.jobs_opportunity o ON o.id = r.opportunity_id
        LEFT JOIN jobs_analytics.role_sort_order s ON s.jobs_role_id = r.id
        LEFT JOIN jobs_analytics.role_origin ro ON ro.jobs_role_id = r.id
        WHERE o.deleted_at IS NULL AND ro.jobs_role_id IS NULL
        ORDER BY (r.status = 'cancelled'), (s.sort_position IS NULL) DESC, s.sort_position, r.updated_at DESC
        """
    )
    apps = await conn.fetch(
        """
        SELECT ja.job_application_id, ja.jobs_role_id,
               COALESCE(
                   NULLIF(trim(b.full_name), ''),
                   -- Only trust the notes prefix as a name when it actually
                   -- follows the app's "Name: ..." convention (has a colon) —
                   -- some rows have free-text notes with no name at all
                   -- (e.g. "Applied via Google form through Pursuit"), which
                   -- would otherwise get displayed as if it were a builder.
                   CASE WHEN ja.notes LIKE '%:%'
                        THEN NULLIF(trim(split_part(ja.notes, ':', 1)), '') END,
                   'Builder #' || ja.builder_id
               ) AS builder,
               ja.stage, ja.date_applied, ja.updated_at
        FROM public.job_applications ja
        LEFT JOIN LATERAL bedrock.builder_by_id(ja.builder_id) b ON true
        WHERE ja.jobs_role_id IS NOT NULL
        ORDER BY ja.date_applied DESC NULLS LAST
        """
    )
    apps_by_role: dict = {}
    for a in apps:
        apps_by_role.setdefault(str(a["jobs_role_id"]), []).append({
            "job_application_id": a["job_application_id"],
            "builder": a["builder"],
            "stage": a["stage"],
            "date_applied": a["date_applied"].isoformat() if a["date_applied"] else None,
            "updated_at": a["updated_at"].isoformat() if a["updated_at"] else None,
        })

    out = []
    for r in roles:
        d = _role_dict(r)
        if d["placement_status"] == "ft_placed":
            continue  # already filled full-time — nothing left to track here
        d["applications"] = apps_by_role.get(d["id"], [])
        out.append(d)
    return {"success": True, "data": out}


class RolesBoardOrder(BaseModel):
    # full visible-board order, top to bottom. Pydantic validates each as a
    # UUID (a malformed id 422s cleanly instead of raising ValueError mid-loop);
    # max_length is a generous ceiling — no real board is anywhere near this
    # size — just a backstop against an abusive payload.
    role_ids: list[UUID] = Field(..., max_length=2000)


@router.put("/roles/board/order")
async def set_roles_board_order(
    body: RolesBoardOrder,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Persist a manual drag-order for the Roles board. bedrock.jobs_role has
    no sort column of its own (no DDL for this role) — this writes into the
    jobs_analytics scratch schema instead, mirroring the existing tag-campaign
    priority pattern (bedrock.contact_tag_catalog.sort_order) but without an
    FK to jobs_role so a role hard-delete is never blocked by this table.
    SET LOCAL ROLE auto-reverts at transaction end either way, so it can never
    leak the jobs_team role onto a later request via a reused pooled connection."""
    who = user.get("email") if isinstance(user, dict) else getattr(user, "email", None)
    async with conn.transaction():
        await conn.execute("SET LOCAL ROLE jobs_team")
        for i, role_id in enumerate(body.role_ids):
            await conn.execute(
                """
                INSERT INTO jobs_analytics.role_sort_order (jobs_role_id, sort_position, updated_by)
                VALUES ($1, $2, $3)
                ON CONFLICT (jobs_role_id) DO UPDATE
                  SET sort_position = EXCLUDED.sort_position,
                      updated_at = now(),
                      updated_by = EXCLUDED.updated_by
                """,
                role_id, i, who,
            )
    return {"success": True, "data": {"updated": len(body.role_ids)}}


@router.post("/roles/{role_id}/mark-builder-sourced")
async def mark_role_builder_sourced(
    role_id: UUID,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Flag a role as builder-sourced (see jobs_analytics.role_origin) — for
    when a role was created through the app's own Opportunities/Placement
    Roles flow just to track a self-found builder's progress, without
    Pursuit actually having a relationship with the company. Drops the role
    out of the Pursuit-Supported board and into Builder-Sourced instead."""
    who = user.get("email") if isinstance(user, dict) else getattr(user, "email", None)
    async with conn.transaction():
        await conn.execute("SET LOCAL ROLE jobs_team")
        await conn.execute(
            """
            INSERT INTO jobs_analytics.role_origin (jobs_role_id, set_by)
            VALUES ($1, $2)
            ON CONFLICT (jobs_role_id) DO NOTHING
            """,
            role_id, who,
        )
    return {"success": True, "data": {"jobs_role_id": str(role_id), "origin": "builder_sourced"}}


@router.delete("/roles/{role_id}/mark-builder-sourced")
async def unmark_role_builder_sourced(
    role_id: UUID,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Revert a role back to the Pursuit-Supported default."""
    async with conn.transaction():
        await conn.execute("SET LOCAL ROLE jobs_team")
        await conn.execute(
            "DELETE FROM jobs_analytics.role_origin WHERE jobs_role_id = $1", role_id,
        )
    return {"success": True, "data": {"jobs_role_id": str(role_id), "origin": "pursuit_supported"}}


class PlacementSyncChoice(BaseModel):
    sf_account_id: Optional[str] = None        # link this existing SF account
    force_create_account: bool = False         # or explicitly create a new one


@router.post("/placements/{employment_record_id}/sync-sf")
async def sync_placement_sf(
    employment_record_id: int,
    body: Optional[PlacementSyncChoice] = None,
    user=Depends(require_auth),
    conn=Depends(get_db),
    client=Depends(require_sf_mcp_client),
):
    """(Re)sync one placement to Salesforce — creates the fellow contact,
    the employer account, and the affiliation as needed. When SF holds a
    similar-but-not-identical account name the response returns
    status=needs_choice with candidates; retry with sf_account_id (link) or
    force_create_account=true."""
    try:
        result = await sync_placement_to_sf(
            conn, client.salesforce, employment_record_id,
            sf_account_id_override=(body.sf_account_id if body else None),
            force_create_account=(body.force_create_account if body else False))
    except NotEligible as ne:
        await record_sync_error(conn, employment_record_id, str(ne), status="skipped")
        return {"success": True, "data": {"status": "skipped", "reason": str(ne)}}
    except AccountAmbiguous as aa:
        await record_sync_error(conn, employment_record_id, str(aa), status="needs_choice")
        return {"success": True, "data": {"status": "needs_choice", "reason": str(aa),
                                          "account_candidates": aa.candidates}}
    except ValueError as ve:
        await record_sync_error(conn, employment_record_id, str(ve))
        raise HTTPException(400, str(ve))
    except HTTPException:
        raise
    except Exception as e:
        await record_sync_error(conn, employment_record_id, str(e))
        raise sf_http_error(e, "placement sync")
    return {"success": True, "data": result}


@router.get("/placements/sf-sync-status")
async def placements_sf_sync_status(
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Sync state for every placement — deal-linked or not; skipped rows
    carry the paid-work policy reason."""
    rows = await conn.fetch(
        """SELECT er.id AS employment_record_id, er.role_title, er.company_name, er.start_date,
                  er.employment_type, er.payment_amount, (er.opportunity_id IS NOT NULL) AS deal_linked,
                  b.full_name AS builder_name,
                  s.status, s.error, s.sf_contact_id, s.sf_account_id, s.sf_affiliation_id, s.synced_at
           FROM public.employment_records er
           LEFT JOIN bedrock.placement_sf_sync s ON s.employment_record_id = er.id
           LEFT JOIN LATERAL bedrock.builder_by_id(er.user_id) b ON true
           ORDER BY er.created_at DESC""")
    return {"success": True, "data": [{
        **dict(r),
        "start_date": r["start_date"].isoformat() if r["start_date"] else None,
        "synced_at": r["synced_at"].isoformat() if r["synced_at"] else None,
        "status": r["status"] or "pending",
    } for r in rows]}


# ── Builder activity on an opportunity ────────────────────────────────────────

@router.get("/opportunities/{opp_id}/builder-activity")
async def opp_builder_activity(
    opp_id: UUID,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Builder submissions/interviews/hires tied to this opportunity.

    Rows come from public.job_applications where jobs_opportunity_id matches.
    Includes a stage summary {applied, interview, accepted}.
    """
    rows = await conn.fetch(
        """
        SELECT job_application_id,
               trim(split_part(notes, ':', 1)) AS builder,
               company_name, role_title, stage, jobs_role_id, date_applied
        FROM public.job_applications
        WHERE jobs_opportunity_id = $1
        ORDER BY date_applied DESC NULLS LAST
        """,
        opp_id,
    )
    out = []
    summary = {"applied": 0, "interview": 0, "accepted": 0}
    for r in rows:
        if r["stage"] in summary:
            summary[r["stage"]] += 1
        out.append({
            "job_application_id": r["job_application_id"],
            "builder":            r["builder"],
            "company_name":       r["company_name"],
            "role_title":         r["role_title"],
            "stage":              r["stage"],
            "jobs_role_id":       str(r["jobs_role_id"]) if r["jobs_role_id"] else None,
            "date_applied":       r["date_applied"].isoformat() if r["date_applied"] else None,
        })
    return {"success": True, "data": {"rows": out, "summary": summary}}


@router.get("/interview-pipeline")
async def interview_pipeline(
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Confirmed roles across all live opportunities, each with the builders
    progressing through them — the command center's interview tracker.

    An opportunity is included when it has at least one committed role OR at
    least one builder application/interview/hire tied to it. Builders are
    grouped under the role they're applying to (jobs_role_id) when known, else
    listed at the opportunity level. Closed/deleted opps are excluded.
    """
    roles = await conn.fetch(
        """
        SELECT r.id, r.opportunity_id, r.title, r.status, r.employment_type,
               r.approx_salary, r.filled_by_user_id, r.commitment, r.is_trial,
               o.account_name, o.stage AS opp_stage, o.owner_email
        FROM bedrock.jobs_role r
        JOIN bedrock.jobs_opportunity o ON o.id = r.opportunity_id
        WHERE o.deleted_at IS NULL AND o.stage NOT IN ('closed_won', 'closed_lost')
        ORDER BY o.account_name, r.created_at
        """,
    )
    apps = await conn.fetch(
        """
        SELECT ja.job_application_id, ja.jobs_opportunity_id, ja.jobs_role_id,
               trim(split_part(ja.notes, ':', 1)) AS builder,
               ja.role_title, ja.stage, ja.date_applied,
               o.account_name, o.stage AS opp_stage, o.owner_email
        FROM public.job_applications ja
        JOIN bedrock.jobs_opportunity o ON o.id = ja.jobs_opportunity_id
        WHERE o.deleted_at IS NULL AND o.stage NOT IN ('closed_won', 'closed_lost')
          AND ja.stage IN ('applied', 'interview', 'accepted')
        ORDER BY ja.date_applied DESC NULLS LAST
        """,
    )

    opps: dict = {}

    def _opp(opp_id, account_name, stage, owner_email):
        key = str(opp_id)
        if key not in opps:
            opps[key] = {
                "opportunity_id": key, "account_name": account_name,
                "stage": stage, "owner_email": owner_email,
                "roles": [], "builders": [],
                "summary": {"applied": 0, "interview": 0, "accepted": 0, "open_roles": 0},
            }
        return opps[key]

    for r in roles:
        g = _opp(r["opportunity_id"], r["account_name"], r["opp_stage"], r["owner_email"])
        st = _placement_status_py(dict(r))
        g["roles"].append({
            "id": str(r["id"]), "title": r["title"], "status": r["status"],
            "employment_type": r["employment_type"], "approx_salary": r["approx_salary"],
            "filled_by_user_id": r["filled_by_user_id"],
            "commitment": r["commitment"], "is_trial": r["is_trial"],
            "placement_status": st, "placement_status_label": PLACEMENT_STATUS_LABELS.get(st, st),
        })
        if r["status"] == "open":
            g["summary"]["open_roles"] += 1

    for a in apps:
        g = _opp(a["jobs_opportunity_id"], a["account_name"], a["opp_stage"], a["owner_email"])
        g["builders"].append({
            "job_application_id": a["job_application_id"], "builder": a["builder"],
            "role_title": a["role_title"], "jobs_role_id": str(a["jobs_role_id"]) if a["jobs_role_id"] else None,
            "stage": a["stage"], "date_applied": a["date_applied"].isoformat() if a["date_applied"] else None,
        })
        if a["stage"] in g["summary"]:
            g["summary"][a["stage"]] += 1

    # Opps actively interviewing first, then by builder volume.
    out = sorted(
        opps.values(),
        key=lambda g: (-(g["summary"]["interview"]), -len(g["builders"]), g["account_name"] or ""),
    )
    return {"success": True, "data": out}


# public.job_applications.stage vocabulary (builder-side submission funnel) —
# must mirror the job_applications_stage_check DB constraint exactly. This set
# used to omit screen/oa/offer, which are real, actively-used stage values
# (10 existing rows) — a role's stage dropdown silently mislabeled them.
VALID_APP_STAGES = {"prospect", "applied", "screen", "oa", "interview", "offer", "accepted", "rejected", "withdrawn"}


class BuilderActivityCreate(BaseModel):
    user_id:       int                       # builder user_id (from the builder picker)
    builder_name:  Optional[str] = None      # stored in notes prefix for display
    role_title:    Optional[str] = None
    stage:         str = "applied"
    jobs_role_id:  Optional[str] = None
    date_applied:  Optional[date] = None


class BuilderActivityUpdate(BaseModel):
    stage: str


@router.post("/opportunities/{opp_id}/builder-activity")
async def create_builder_activity(
    opp_id: UUID,
    body: BuilderActivityCreate,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Log a builder application / interview against this opportunity.

    Writes to public.job_applications (the builder submission funnel) tagged with
    jobs_opportunity_id so it surfaces in the Builders tab and the funnel. The
    builder's display name is stored in the notes prefix because the read path
    derives it via split_part(notes, ':', 1).
    """
    if body.stage not in VALID_APP_STAGES:
        raise HTTPException(400, f"Invalid stage: {body.stage}")
    opp = await conn.fetchrow(
        "SELECT account_name FROM bedrock.jobs_opportunity WHERE id=$1 AND deleted_at IS NULL",
        opp_id,
    )
    if not opp:
        raise HTTPException(404, "Opportunity not found")

    role_id = UUID(body.jobs_role_id) if body.jobs_role_id else None
    notes = f"{body.builder_name}: logged via Opportunities" if body.builder_name else None
    app_id = await conn.fetchval(
        """
        INSERT INTO public.job_applications
            (builder_id, company_name, role_title, stage, date_applied,
             source_type, jobs_opportunity_id, jobs_role_id, notes)
        VALUES ($1, $2, $3, $4, COALESCE($5, CURRENT_DATE),
                'Pursuit_referred', $6, $7, $8)
        RETURNING job_application_id
        """,
        body.user_id,
        opp["account_name"],
        body.role_title or "Role",
        body.stage,
        body.date_applied,
        opp_id,
        role_id,
        notes,
    )
    return {"success": True, "data": {"job_application_id": app_id, "stage": body.stage}}


@router.patch("/builder-activity/{app_id}")
async def update_builder_activity(
    app_id: int,
    body: BuilderActivityUpdate,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Update an application's stage inline (applied → interview → accepted, …)."""
    if body.stage not in VALID_APP_STAGES:
        raise HTTPException(400, f"Invalid stage: {body.stage}")
    result = await conn.execute(
        """
        UPDATE public.job_applications
        SET stage=$1, updated_at=now()
        WHERE job_application_id=$2 AND jobs_opportunity_id IS NOT NULL
        """,
        body.stage, app_id,
    )
    if result == "UPDATE 0":
        raise HTTPException(404, "Application not found")
    return {"success": True, "data": {"job_application_id": app_id, "stage": body.stage}}


class RoleApplicationCreate(BaseModel):
    user_id:      int                       # builder user_id (from the builder picker)
    builder_name: Optional[str] = None      # stored in notes prefix for display
    stage:        str = "applied"
    date_applied: Optional[date] = None


@router.post("/roles/{role_id}/applications")
async def create_role_application(
    role_id: UUID,
    body: RoleApplicationCreate,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Log a builder application directly against a role from the Roles board
    (rather than needing to open the parent opportunity first). Derives the
    opportunity + company name from the role, then inserts with the same shape
    as create_builder_activity so the existing read paths (opp_builder_activity,
    roles_board) stay consistent."""
    if body.stage not in VALID_APP_STAGES:
        raise HTTPException(400, f"Invalid stage: {body.stage}")
    role = await conn.fetchrow(
        """
        SELECT r.title, r.opportunity_id, o.account_name
        FROM bedrock.jobs_role r
        JOIN bedrock.jobs_opportunity o ON o.id = r.opportunity_id
        WHERE r.id=$1 AND o.deleted_at IS NULL
        """,
        role_id,
    )
    if not role:
        raise HTTPException(404, "Role not found")
    # user_id normally comes from the UI's builder-search picker, never
    # hand-typed — but the API itself didn't verify it, so a bad payload
    # (bypassing the picker, or a stale/mistyped id) would either 500 on the
    # builder_id FK or, worse, silently succeed against some other real user
    # row. Confirm it resolves to an actual builder first.
    builder = await conn.fetchrow("SELECT user_id FROM bedrock.builder_by_id($1)", body.user_id)
    if not builder:
        raise HTTPException(404, "Builder not found")

    notes = f"{body.builder_name}: logged via Placement Roles" if body.builder_name else None
    app_id = await conn.fetchval(
        """
        INSERT INTO public.job_applications
            (builder_id, company_name, role_title, stage, date_applied,
             source_type, jobs_opportunity_id, jobs_role_id, notes)
        VALUES ($1, $2, $3, $4, COALESCE($5, CURRENT_DATE),
                'Pursuit_referred', $6, $7, $8)
        RETURNING job_application_id
        """,
        body.user_id,
        role["account_name"],
        role["title"] or "Role",
        body.stage,
        body.date_applied,
        role["opportunity_id"],
        role_id,
        notes,
    )
    return {"success": True, "data": {"job_application_id": app_id, "stage": body.stage}}


def _title_nkey(title: str) -> str:
    """Normalized role-title key for backfill matching — case/punctuation
    insensitive, same spirit as _acct_nkey but without the legal-suffix strip
    (titles don't have those)."""
    s = (title or "").strip().lower()
    s = re.sub(r"[,\.\&\'\"()]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


@router.get("/job-applications/builder-sourced")
async def builder_sourced_applications(
    days: int = Query(30, ge=1, le=365),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """The Builder-Sourced column: replaces the old separate match-suggestions
    banners (folded in here so there's one place, not a temporary queue plus
    a permanent column showing overlapping data). Two kinds of rows:
    1. Applications not linked to a Bedrock role AND not staff/Pursuit-
       originated. jobs_role_id IS NULL alone isn't enough: plenty of
       staff-logged applications (source_type 'Pursuit_referred' or
       'staff_sourced', e.g. logged via the Opportunities page before a
       formal role existed) also have no role link yet, and those are
       Pursuit-supported, not builder-sourced. 30-day window — this is a
       working queue of recent activity, not a permanent historical archive.
       Each of these gets annotated with `suggested_match` (same matching
       as the old match-suggestions: exact normalized-name match first,
       then a looser normalized-title check within the same account) when
       its company already has an open Bedrock role, so the frontend can
       offer "Confirm match" instead of "Create opportunity" — avoids
       spawning a duplicate opportunity for a company Pursuit already has.
    2. Applications linked to a role explicitly flagged in
       jobs_analytics.role_origin (see mark_role_builder_sourced) — a role
       created through the app's own UI just to track a self-found
       builder's progress, with no real Pursuit-company relationship. No
       date window here: once explicitly marked, it stays surfaced until
       someone unmarks it, regardless of how old the application is. Never
       gets a suggested_match — it already has a role.

    Interview-or-further rows sort first: we only formally track a
    self-sourced lead as a Bedrock opportunity once it reaches that stage,
    to avoid cluttering the pipeline with every speculative application a
    builder happens to log."""
    rows = await conn.fetch(
        """
        WITH combined AS (
            SELECT ja.job_application_id, ja.builder_id, ja.company_name, ja.role_title,
                   ja.date_applied, ja.stage, ja.updated_at,
                   COALESCE(
                       NULLIF(trim(b.full_name), ''),
                       CASE WHEN ja.notes LIKE '%:%'
                            THEN NULLIF(trim(split_part(ja.notes, ':', 1)), '') END,
                       'Builder #' || ja.builder_id
                   ) AS builder,
                   NULL::uuid AS jobs_role_id
            FROM public.job_applications ja
            LEFT JOIN LATERAL bedrock.builder_by_id(ja.builder_id) b ON true
            WHERE ja.jobs_role_id IS NULL
              AND ja.date_applied >= (CURRENT_DATE - $1::int * INTERVAL '1 day')
              AND (ja.source_type IS NULL OR ja.source_type NOT IN ('Pursuit_referred', 'staff_sourced'))
              -- Exclude known seed/test data: rows explicitly marked as seeded, an
              -- old ad-hoc smoke-test row, and anything tied to the synthetic
              -- +auto-testing builder account (regardless of how its notes are
              -- worded — matching by the test account itself is more reliable
              -- than pattern-matching note text).
              -- NULL notes must not exclude a row: `NULL NOT ILIKE x` evaluates to
              -- NULL (not TRUE), which silently dropped every no-notes application
              -- out of both queues regardless of source_type until this COALESCE.
              AND COALESCE(ja.notes, '') NOT ILIKE '%seeded for platform demo%'
              AND COALESCE(ja.notes, '') NOT ILIKE '%smoke test%'
              AND (b.email IS NULL OR b.email NOT ILIKE '%+auto-testing%')

            UNION ALL

            -- Starts from role_origin (not job_applications) and LEFT JOINs
            -- out to applications: a role can be marked builder-sourced
            -- before any builder has applied to it, and it still needs to
            -- show up here — otherwise it disappears from the Pursuit-
            -- Supported board (roles_board excludes flagged roles) without
            -- appearing anywhere on the Builder-Sourced side either.
            SELECT ja.job_application_id, ja.builder_id,
                   COALESCE(o.account_name, ja.company_name) AS company_name,
                   COALESCE(r.title, ja.role_title) AS role_title,
                   ja.date_applied, ja.stage, ja.updated_at,
                   COALESCE(
                       NULLIF(trim(b.full_name), ''),
                       CASE WHEN ja.notes LIKE '%:%'
                            THEN NULLIF(trim(split_part(ja.notes, ':', 1)), '') END,
                       CASE WHEN ja.builder_id IS NOT NULL THEN 'Builder #' || ja.builder_id END
                   ) AS builder,
                   ro.jobs_role_id
            FROM jobs_analytics.role_origin ro
            JOIN bedrock.jobs_role r ON r.id = ro.jobs_role_id
            JOIN bedrock.jobs_opportunity o ON o.id = r.opportunity_id
            LEFT JOIN public.job_applications ja ON ja.jobs_role_id = ro.jobs_role_id
            LEFT JOIN LATERAL bedrock.builder_by_id(ja.builder_id) b ON true
            -- Same two exclusions as roles_board: a soft-deleted opportunity's
            -- role shouldn't linger here forever, and once someone's actually
            -- hired full-time there's nothing left to track (mirrors the
            -- ft_placed skip in roles_board — see _placement_status_sql).
            WHERE o.deleted_at IS NULL
              AND NOT (r.filled_by_user_id IS NOT NULL AND COALESCE(r.is_trial, false) = false)
        )
        SELECT * FROM combined
        ORDER BY
          CASE WHEN stage IN ('interview', 'offer', 'accepted') THEN 0 ELSE 1 END,
          date_applied DESC
        """,
        days,
    )

    return {"success": True, "data": await _annotate_suggested_matches(conn, rows)}


async def _annotate_suggested_matches(conn, rows) -> list[dict]:
    """Shared by builder-sourced and staff-sourced: given fetched application
    rows (each with job_application_id, builder_id, company_name, role_title,
    date_applied, stage, builder, jobs_role_id), annotate the still-unlinked
    ones with a suggested_match when their company already has an open
    Bedrock role — same matching as the old match-suggestions endpoint:
    exact normalized-name match first, then a looser normalized-title check
    within the same account."""
    unlinked = [r for r in rows if r["jobs_role_id"] is None]
    role_index = []
    linked_pairs = set()
    if unlinked:
        roles = await conn.fetch(
            """
            SELECT r.id, r.title, o.account_name
            FROM bedrock.jobs_role r
            JOIN bedrock.jobs_opportunity o ON o.id = r.opportunity_id
            WHERE o.deleted_at IS NULL AND r.status = 'open'
            """
        )
        role_index = [
            {"id": str(r["id"]), "title": r["title"], "account_name": r["account_name"],
             "account_key": _acct_nkey(r["account_name"]), "title_key": _title_nkey(r["title"])}
            for r in roles
        ]
        # A builder already linked to a role via some OTHER application is not
        # a fresh match to suggest — confirming it would just recreate the
        # same duplicate a human already resolved (see unlink_job_application_role).
        already_linked = await conn.fetch(
            "SELECT jobs_role_id, builder_id FROM public.job_applications WHERE jobs_role_id IS NOT NULL"
        )
        linked_pairs = {(str(r["jobs_role_id"]), r["builder_id"]) for r in already_linked}

    out = []
    for r in rows:
        row = {
            "job_application_id": r["job_application_id"],
            "builder": r["builder"],
            "company_name": r["company_name"],
            "role_title": r["role_title"],
            "date_applied": r["date_applied"].isoformat() if r["date_applied"] else None,
            "updated_at": r["updated_at"].isoformat() if r["updated_at"] else None,
            "stage": r["stage"],
            "jobs_role_id": str(r["jobs_role_id"]) if r["jobs_role_id"] else None,
            "suggested_match": None,
        }
        if r["jobs_role_id"] is None:
            app_account_key = _acct_nkey(r["company_name"])
            app_title_key = _title_nkey(r["role_title"])
            best = None
            for ri in role_index:
                if ri["account_key"] != app_account_key:
                    continue
                confidence = "exact" if ri["title_key"] == app_title_key else "normalized"
                if best is None or (confidence == "exact" and best["confidence"] != "exact"):
                    best = {
                        "jobs_role_id": ri["id"], "role_title": ri["title"],
                        "account_name": ri["account_name"], "confidence": confidence,
                    }
                if confidence == "exact":
                    break
            if best and (best["jobs_role_id"], r["builder_id"]) not in linked_pairs:
                row["suggested_match"] = best
        out.append(row)
    return out


@router.get("/job-applications/staff-sourced")
async def staff_sourced_applications(
    days: int = Query(30, ge=1, le=365),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """The Staff-Sourced queue: applications logged by staff (source_type
    'Pursuit_referred' or 'staff_sourced', e.g. via the Opportunities/Activity
    flow) that have no Bedrock role linked yet. These are real Pursuit-
    relationship leads, just not yet formalized into a role/opportunity
    record. Same 30-day working-queue window as Builder-Sourced — someone
    owes each of these a decision: confirm a suggested_match, create the
    opportunity/role from scratch, or — if it turns out staff misattributed
    it and the builder actually found this one on their own — send it to
    Builder-Sourced via mark_application_self_sourced.

    Split out on 2026-08-18: builder-sourced's WHERE clause explicitly
    excludes these source_types (they're not builder-sourced), but nothing
    else was surfacing them, so a staff-logged application with no role
    fell through the cracks entirely. This is the other half of that
    same "jobs_role_id IS NULL" population, not a new concept.

    Grouped by normalized (company, role_title), also 2026-08-18: several
    applicants often share the same not-yet-formalized role (e.g. two
    builders both applying to "Silna — AI Deployment Analyst"), and showing
    each as its own disconnected row obscured that — mirrors the Pursuit-
    Supported board's role-with-nested-applicants shape instead."""
    rows = await conn.fetch(
        """
        SELECT ja.job_application_id, ja.builder_id, ja.company_name, ja.role_title,
               ja.date_applied, ja.stage, ja.updated_at,
               COALESCE(
                   NULLIF(trim(b.full_name), ''),
                   CASE WHEN ja.notes LIKE '%:%'
                        THEN NULLIF(trim(split_part(ja.notes, ':', 1)), '') END,
                   'Builder #' || ja.builder_id
               ) AS builder
        FROM public.job_applications ja
        LEFT JOIN LATERAL bedrock.builder_by_id(ja.builder_id) b ON true
        WHERE ja.jobs_role_id IS NULL
          AND ja.source_type IN ('Pursuit_referred', 'staff_sourced')
          AND ja.date_applied >= (CURRENT_DATE - $1::int * INTERVAL '1 day')
          AND COALESCE(ja.notes, '') NOT ILIKE '%seeded for platform demo%'
          AND COALESCE(ja.notes, '') NOT ILIKE '%smoke test%'
          AND (b.email IS NULL OR b.email NOT ILIKE '%+auto-testing%')
        ORDER BY date_applied DESC
        """,
        days,
    )

    groups: dict = {}
    order = []
    for r in rows:
        key = (_acct_nkey(r["company_name"]), _title_nkey(r["role_title"]))
        if key not in groups:
            groups[key] = {"company_name": r["company_name"], "role_title": r["role_title"], "applications": []}
            order.append(key)
        groups[key]["applications"].append({
            "job_application_id": r["job_application_id"],
            "builder_id": r["builder_id"],
            "builder": r["builder"],
            "stage": r["stage"],
            "date_applied": r["date_applied"].isoformat() if r["date_applied"] else None,
            "updated_at": r["updated_at"].isoformat() if r["updated_at"] else None,
        })

    roles = await conn.fetch(
        """
        SELECT r.id, r.title, o.account_name
        FROM bedrock.jobs_role r
        JOIN bedrock.jobs_opportunity o ON o.id = r.opportunity_id
        WHERE o.deleted_at IS NULL AND r.status = 'open'
        """
    )
    role_index = [
        {"id": str(r["id"]), "title": r["title"], "account_name": r["account_name"],
         "account_key": _acct_nkey(r["account_name"]), "title_key": _title_nkey(r["title"])}
        for r in roles
    ]
    # A builder already linked to the candidate role via some OTHER application
    # shouldn't get re-confirmed onto it — same guard as _annotate_suggested_matches,
    # otherwise "Confirm match (N)" would create a second linked row for them.
    already_linked = await conn.fetch(
        "SELECT jobs_role_id, builder_id FROM public.job_applications WHERE jobs_role_id IS NOT NULL"
    )
    linked_pairs = {(str(r["jobs_role_id"]), r["builder_id"]) for r in already_linked}

    out = []
    for account_key, title_key in order:
        g = groups[(account_key, title_key)]
        # Interview-or-further applicants sort first within the group; stable
        # sort preserves the original date_applied DESC order within each tier.
        g["applications"].sort(key=lambda a: 0 if a["stage"] in ("interview", "offer", "accepted") else 1)
        best = None
        for ri in role_index:
            if ri["account_key"] != account_key:
                continue
            confidence = "exact" if ri["title_key"] == title_key else "normalized"
            if best is None or (confidence == "exact" and best["confidence"] != "exact"):
                best = {"jobs_role_id": ri["id"], "role_title": ri["title"],
                        "account_name": ri["account_name"], "confidence": confidence}
            if confidence == "exact":
                break
        applications = [
            {**{k: v for k, v in a.items() if k != "builder_id"},
             "already_linked": bool(best) and (best["jobs_role_id"], a["builder_id"]) in linked_pairs}
            for a in g["applications"]
        ]
        out.append({
            "company_name": g["company_name"],
            "role_title": g["role_title"],
            "suggested_match": best,
            "applications": applications,
        })
    return {"success": True, "data": out}


@router.post("/job-applications/{app_id}/mark-self-sourced")
async def mark_application_self_sourced(
    app_id: int,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Reclassify a Staff-Sourced application as builder-sourced: clears
    source_type so it falls out of the Staff-Sourced queue and into
    Builder-Sourced on next fetch. One-way — this is a rare correction
    (staff assumed they'd sourced it, but the builder had actually already
    applied on their own), not a routine toggle, so there's no unmark
    endpoint; revert by hand if it's ever wrong."""
    result = await conn.execute(
        """
        UPDATE public.job_applications
        SET source_type = NULL, updated_at = now()
        WHERE job_application_id = $1 AND source_type IN ('Pursuit_referred', 'staff_sourced')
        """,
        app_id,
    )
    if result == "UPDATE 0":
        raise HTTPException(404, "Application not found or not staff-sourced")
    return {"success": True, "data": {"job_application_id": app_id, "source_type": None}}


@router.post("/job-applications/{app_id}/mark-staff-sourced")
async def mark_application_staff_sourced(
    app_id: int,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Reclassify a Builder-Sourced application as staff-sourced: the mirror
    of mark_application_self_sourced. Covers manually-logged applications
    that were never tagged with a source_type at creation (so they defaulted
    to Builder-Sourced) even though staff actually sourced them. One-way,
    same as the other direction."""
    result = await conn.execute(
        """
        UPDATE public.job_applications
        SET source_type = 'staff_sourced', updated_at = now()
        WHERE job_application_id = $1
          AND (source_type IS NULL OR source_type NOT IN ('Pursuit_referred', 'staff_sourced'))
        """,
        app_id,
    )
    if result == "UPDATE 0":
        raise HTTPException(404, "Application not found or already staff-sourced")
    return {"success": True, "data": {"job_application_id": app_id, "source_type": "staff_sourced"}}


class MatchConfirm(BaseModel):
    jobs_role_id: UUID


@router.post("/job-applications/{app_id}/confirm-match")
async def confirm_job_application_match(
    app_id: int,
    body: MatchConfirm,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Link an application to a role. Only fires on an explicit staff confirm
    click in the UI — matches from match-suggestions are suggested, never
    auto-applied."""
    role_id = body.jobs_role_id
    role = await conn.fetchrow("SELECT opportunity_id FROM bedrock.jobs_role WHERE id=$1", role_id)
    if not role:
        raise HTTPException(404, "Role not found")
    result = await conn.execute(
        """
        UPDATE public.job_applications
        SET jobs_role_id=$1, jobs_opportunity_id=$2, updated_at=now()
        WHERE job_application_id=$3
        """,
        role_id, role["opportunity_id"], app_id,
    )
    if result == "UPDATE 0":
        raise HTTPException(404, "Application not found")
    return {"success": True, "data": {"job_application_id": app_id, "jobs_role_id": str(role_id)}}


@router.post("/job-applications/{app_id}/unlink-role")
async def unlink_job_application_role(
    app_id: int,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Remove an application from whatever role/opportunity it's linked to,
    without deleting the row. avni_dev (and jobs_team) has no DELETE grant on
    public.job_applications, so this is the reversible alternative for
    cleaning up a duplicate — e.g. the same builder logged both via the
    Opportunities page and via the Roles board, then a match-suggestion
    confirm linked a third, pre-existing row to the same role."""
    result = await conn.execute(
        """
        UPDATE public.job_applications
        SET jobs_role_id=NULL, jobs_opportunity_id=NULL, updated_at=now()
        WHERE job_application_id=$1
        """,
        app_id,
    )
    if result == "UPDATE 0":
        raise HTTPException(404, "Application not found")
    return {"success": True, "data": {"job_application_id": app_id, "unlinked": True}}


# ── Job-ready (L3+) pool ───────────────────────────────────────────────────────
# Builders who EVER reached L3+ (have an L3+ course enrollment), each tagged with
# the L3 cohort (class) they completed — the dashboard's segment dimension. L3+ is
# one shared pool, so we look back at each builder's L3 enrollment for the segment.
_L3PLUS_POOL = """
  l3plus AS (
    SELECT DISTINCT ue.user_id
    FROM public.user_enrollment ue
    JOIN public.cohort ch ON ch.cohort_id = ue.cohort_id
    JOIN public.course co ON co.course_id = ch.course_id
    WHERE co.level = 'L3+'
  ),
  l3cohort AS (
    SELECT DISTINCT ON (ue.user_id) ue.user_id, ch.name AS segment
    FROM public.user_enrollment ue
    JOIN public.cohort ch ON ch.cohort_id = ue.cohort_id
    JOIN public.course co ON co.course_id = ch.course_id
    WHERE co.level = 'L3' AND ue.user_id IN (SELECT user_id FROM l3plus)
    ORDER BY ue.user_id, ue.enrolled_date DESC
  ),
  pool AS (
    SELECT lp.user_id, COALESCE(lc.segment, 'Other L3+') AS segment
    FROM l3plus lp LEFT JOIN l3cohort lc ON lc.user_id = lp.user_id
  )
"""


@router.get("/builder-segments")
async def builder_segments(user=Depends(require_auth), conn=Depends(get_db)):
    """L3-cohort segments present in the L3+ pool — drives the dashboard filter."""
    rows = await conn.fetch(f"""
        WITH {_L3PLUS_POOL}
        SELECT segment, count(*) AS n FROM pool GROUP BY segment ORDER BY n DESC
    """)
    return {"success": True, "data": {
        "segments": [{"value": r["segment"], "label": r["segment"], "count": r["n"]} for r in rows],
        "total": sum(r["n"] for r in rows),
    }}


# Stages that must not advertise a conversion rate into whatever follows them in
# `stage_order`: the next row isn't a forward step. On Hold is a parking state and
# Converted is the end of the contact funnel.
_NO_CONVERSION_FROM = {"converted_to_opportunity", "revisit", "not_a_fit", "on_hold",
                       "closed_won", "closed_lost"}


@router.get("/funnel/{ftype}")
async def get_funnel(
    ftype: str,
    deal_type: Optional[str] = Query(None),
    segment: Optional[str] = Query(None),
    period_from: Optional[str] = Query(None, description="YYYY-MM-DD, inclusive"),
    period_to: Optional[str] = Query(None, description="YYYY-MM-DD, inclusive"),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Unified funnel for the three pipelines: opportunities | prospects | builders.

    Two modes:

    * **Snapshot** (no period) — stages carry how many records SIT in them right
      now, and conversion is the all-time cohort rate. This is what the Exec view
      uses; it has no period control.
    * **Period flow** (`period_from` + `period_to`) — stages carry how many
      records ENTERED them inside the window, and conversion is
      `next stage entries / this stage entries`. Used by Outreach and Pipeline,
      which both have a period picker.

    Period flow can legitimately exceed 100%: more contacts can enter Initial
    Outreach in a week than were assigned that same week, because they were
    assigned earlier. That's backlog being cleared, so the rate is reported as
    measured rather than clamped.

    `deal_type` (ft | pt_contract | ...) scopes every funnel to that lens:
    opportunities by their own deal_type; prospects to contacts at companies
    that have a deal of that type; builders to applications on such opps.
    """
    dt = deal_type if deal_type and deal_type != "all" else None

    # Period mode is opt-in and only meaningful where we stamp stage entry.
    period: Optional[tuple] = None
    if period_from and period_to and ftype in ("opportunities", "prospects"):
        try:
            p_from = datetime.strptime(period_from, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            # Inclusive end date → exclusive upper bound at the next midnight, so
            # a record stamped 17:04 on the last day still counts.
            p_to = (datetime.strptime(period_to, "%Y-%m-%d").replace(tzinfo=timezone.utc)
                    + timedelta(days=1))
        except ValueError:
            raise HTTPException(400, "period_from / period_to must be YYYY-MM-DD")
        if p_to <= p_from:
            raise HTTPException(400, "period_to must not precede period_from")
        period = (p_from, p_to)
    movement_by_stage: dict = {}  # stage_key -> list of recent transitions touching it

    if ftype == "opportunities":
        # Initial Outreach is retired (2026-08-05), so the funnel starts at In
        # Discussions. Closed Lost is a row of its own rather than an omission:
        # a funnel that only shows the wins overstates the pipeline, and lost is
        # where the on_hold_* deals now land, so leaving it out would have hidden
        # them entirely. It sits after Closed Won as a second terminal row —
        # neither converts into anything, so no rate is drawn between them.
        stage_order = [
            ("active_in_discussions", "In Discussions"),
            ("active_opportunity_confirmed", "Opportunity Confirmed"),
            ("reviewing_builders", "Reviewing Builders"),
            ("closed_won", "Closed — Won"),
            ("closed_lost", "Closed — Lost"),
        ]
        record_columns = [
            {"key": "name", "label": "Company"},
            {"key": "deal_type", "label": "Type"},
            {"key": "owner", "label": "Owner"},
            {"key": "roles", "label": "Roles (commitment)"},
        ]
        # Per-opp roles rollup so the drill shows what "Opportunity Confirmed"
        # actually contains — each role with its status and whether it's a
        # committed seat or open-market (feedback 2026-07-16).
        rows = await conn.fetch("""
            SELECT o.stage, o.account_name AS name, o.deal_type, o.owner_email AS owner,
                   (SELECT string_agg(
                            coalesce(r.title, 'Role') || ' — ' ||
                            CASE WHEN r.status = 'cancelled' THEN 'cancelled'
                                 WHEN r.status = 'filled' AND r.is_trial THEN 'trial active'
                                 WHEN r.status = 'filled' THEN 'filled'
                                 WHEN r.commitment = 'committed' THEN 'committed · open'
                                 ELSE 'open market' END,
                            '  ·  ' ORDER BY r.created_at)
                      FROM bedrock.jobs_role r WHERE r.opportunity_id = o.id) AS roles
            FROM bedrock.jobs_opportunity o
            WHERE o.deleted_at IS NULL AND ($1::text IS NULL OR o.deal_type = $1)
            ORDER BY o.account_name
        """, dt)
        by_stage: dict = {}
        for r in rows:
            # Canonical stage, so a deal still stored as active_builder_interview
            # lands in the Reviewing Builders row instead of no row at all.
            by_stage.setdefault(canon_stage(r["stage"]), []).append(
                {"name": r["name"], "deal_type": r["deal_type"], "owner": r["owner"],
                 "roles": r["roles"] or "—"}
            )

        idx = {k: i for i, (k, _) in enumerate(stage_order)}
        label_of = dict(stage_order)
        # DISTINCT ON keeps only each opp's MOST RECENT transition in the window —
        # so an opp that moved twice shows once (its current stage + where it came
        # from on the latest hop), not a duplicate per hop.
        hist = await conn.fetch("""
            SELECT DISTINCT ON (h.opportunity_id)
                   h.from_stage, h.to_stage, h.changed_at, o.account_name
            FROM bedrock.jobs_stage_history h
            JOIN bedrock.jobs_opportunity o ON o.id = h.opportunity_id
            WHERE h.from_stage IS NOT NULL
              AND h.changed_at >= now() - interval '30 days'
              AND ($1::text IS NULL OR o.deal_type = $1)
            ORDER BY h.opportunity_id, h.changed_at DESC
            LIMIT 100
        """, dt)
        for h in hist:
            h_from, h_to = canon_stage(h["from_stage"]), canon_stage(h["to_stage"])
            fi, ti = idx.get(h_from), idx.get(h_to)
            direction = "advanced" if (fi is not None and ti is not None and ti > fi) else "regressed"
            item = {
                "name": h["account_name"],
                "from_label": label_of.get(h_from, h_from),
                "to_label": label_of.get(h_to, h_to),
                "direction": direction,
                "when": h["changed_at"].isoformat() if h["changed_at"] else None,
            }
            # attach to the destination stage (moved into) and origin stage (moved out of)
            movement_by_stage.setdefault(h_to, []).append({**item, "flow": "in"})
            movement_by_stage.setdefault(h_from, []).append({**item, "flow": "out"})

    elif ftype == "prospects":
        # The Contacts funnel runs on the jobs-pipeline membership stage
        # (bedrock.jobs_contact_membership) — the stage the team actually
        # manages on the Contacts page — not the legacy contacts.contact_stage.
        # on_hold shows as a terminal parking stage; not_a_fit is a dead
        # disposition and stays out of the funnel.
        # Both off-ramps are rows, not omissions: Revisit is work you've parked
        # and Not a Fit is work you've ruled out, and a funnel that shows neither
        # implies every contact is still live. Terminal rows draw no conversion
        # into whatever follows them (see _NO_CONVERSION_FROM).
        stage_order = [
            ("assigned", "Assigned"),
            ("initial_outreach", "Initial Outreach"),
            ("call_booked", "Call Booked"),
            ("converted_to_opportunity", "Converted to Opportunity"),
            ("revisit", "Revisit"),
            ("not_a_fit", "Not a Fit"),
        ]
        record_columns = [
            {"key": "name", "label": "Contact"},
            {"key": "company", "label": "Company"},
        ]
        rows = await conn.fetch("""
            SELECT m.stage, c.full_name AS name, c.current_company AS company
            FROM bedrock.jobs_contact_membership m
            JOIN public.contacts c ON c.contact_id = m.contact_id
            WHERE m.stage <> 'not_a_fit'
              AND ($1::text IS NULL OR lower(c.current_company) IN (
                    SELECT lower(account_name) FROM bedrock.jobs_opportunity
                    WHERE deleted_at IS NULL AND deal_type = $1 AND account_name IS NOT NULL))
            ORDER BY c.full_name
        """, dt)
        by_stage = {}
        for r in rows:
            by_stage.setdefault(canon_membership_stage(r["stage"]), []).append({"name": r["name"], "company": r["company"]})
        # Ever-reached counts (stage-entry stamps), so conversion is a real
        # cohort rate: of everyone who entered a stage, how many got to the next.
        reach = await conn.fetchrow("""
            SELECT count(*) FILTER (WHERE assigned_at IS NOT NULL)       AS reached_assigned,
                   count(*) FILTER (WHERE first_outreach_at IS NOT NULL) AS reached_outreach,
                   count(*) FILTER (WHERE converted_at IS NOT NULL)      AS reached_converted
            FROM bedrock.jobs_contact_membership
        """)
        cohort_conv = {
            "assigned": round(100 * reach["reached_outreach"] / reach["reached_assigned"])
                        if reach["reached_assigned"] else None,
            "initial_outreach": round(100 * reach["reached_converted"] / reach["reached_outreach"])
                        if reach["reached_outreach"] else None,
            # On Hold is a parking state, not the step after Converted — no rate.
            "converted_to_opportunity": None,
        }

    elif ftype == "builders":
        # Job-ready pipeline keyed off the L3+ pool + actual paid placements
        # (employment_records) — so hiring a builder into a role shows them as
        # hired here (the old job_applications-based funnel never updated on hire).
        seg = segment if segment and segment != "all" else None
        stage_order = [
            ("job_ready", "Job Ready (L3+)"),
            ("paid",      "Hired — Any Paid Work"),
            ("ft",        "Hired — Full-Time"),
        ]
        record_columns = [
            {"key": "name", "label": "Builder"},
            {"key": "company", "label": "Company"},
            {"key": "role", "label": "Role"},
        ]
        # bedrock.l3plus_funnel() is SECURITY DEFINER so it can read names from
        # the RLS-protected public.users (the app role can't). Returns the L3+
        # pool with placement flags + the builder's L3-cohort segment.
        prows = await conn.fetch(
            "SELECT name, is_paid, is_ft, company, role FROM bedrock.l3plus_funnel($1)", seg
        )
        by_stage = {"job_ready": [], "paid": [], "ft": []}
        for r in prows:
            rec = {"name": r["name"], "company": r["company"] or "—", "role": r["role"] or "—"}
            by_stage["job_ready"].append(rec)
            if r["is_paid"]:
                by_stage["paid"].append(rec)
            if r["is_ft"]:
                by_stage["ft"].append(rec)
    else:
        raise HTTPException(404, f"Unknown funnel: {ftype}")

    # ── Period flow ────────────────────────────────────────────────────────
    # Swap "how many sit in this stage" for "how many ENTERED it in the window".
    # Stage entry is stamped, so this is read off timestamps rather than inferred.
    period_entries: Optional[dict] = None
    prev_counts: Optional[list] = None

    async def _entries(p_from, p_to) -> dict:
        """Records that entered each stage inside [p_from, p_to).

        Run twice in period mode — once for the selected window and once for the
        window of equal length immediately before it, which is what the volume
        and conversion trends compare against.
        """
        nonlocal record_columns
        period_entries = {k: [] for k, _ in stage_order}

        if ftype == "prospects":
            company_lens = """
                AND ($3::text IS NULL OR lower(c.current_company) IN (
                      SELECT lower(account_name) FROM bedrock.jobs_opportunity
                      WHERE deleted_at IS NULL AND deal_type = $3 AND account_name IS NOT NULL))
            """
            # The three managed stages each have their own entry stamp on the
            # membership row, which is more reliable than the history table
            # (history only goes back to when we started recording it).
            # Two sources, because neither alone is complete: the membership row
            # stamps entry for the three managed stages, and the history table
            # records every transition but only since we started writing it.
            # Keyed by (contact, stage) so a contact with both isn't counted twice;
            # the stamp wins because it's the purpose-built field.
            seen: dict = {}

            prows = await conn.fetch(f"""
                SELECT c.contact_id, c.full_name AS name, c.current_company AS company,
                       m.owner_email AS owner, m.assigned_by,
                       m.assigned_at, m.first_outreach_at, m.converted_at
                FROM bedrock.jobs_contact_membership m
                JOIN public.contacts c ON c.contact_id = m.contact_id
                WHERE m.stage <> 'not_a_fit' AND $1::timestamptz IS NOT NULL
                  AND $2::timestamptz IS NOT NULL
                  {company_lens}
            """, p_from, p_to, dt)
            stamp_for = {
                "assigned": "assigned_at",
                "initial_outreach": "first_outreach_at",
                "converted_to_opportunity": "converted_at",
            }
            for r in prows:
                for stage_key, col in stamp_for.items():
                    ts = r[col]
                    if ts is not None and p_from <= ts < p_to:
                        seen[(r["contact_id"], stage_key)] = {
                            "name": r["name"], "company": r["company"],
                            "owner": r["owner"], "assigned_by": r["assigned_by"],
                            "entered_at": ts.isoformat(),
                        }

            try:
                hrows = await conn.fetch(f"""
                    SELECT DISTINCT ON (h.contact_id, h.to_stage)
                           h.contact_id, h.to_stage, h.changed_at, h.changed_by,
                           c.full_name AS name, c.current_company AS company,
                           m.owner_email AS owner
                    FROM bedrock.jobs_membership_stage_history h
                    JOIN public.contacts c ON c.contact_id = h.contact_id
                    LEFT JOIN bedrock.jobs_contact_membership m ON m.contact_id = h.contact_id
                    WHERE h.changed_at >= $1 AND h.changed_at < $2
                      {company_lens}
                    ORDER BY h.contact_id, h.to_stage, h.changed_at
                """, p_from, p_to, dt)
            except asyncpg.exceptions.InsufficientPrivilegeError:
                # Read-only enhancement: without it the stamps still carry
                # assigned / initial-outreach / converted, so the funnel is
                # complete for those three and only loses on_hold entries and
                # pre-stamp history. Far better than a dead panel.
                logger.warning(
                    "funnel: no SELECT on bedrock.jobs_membership_stage_history for this "
                    "role — falling back to membership stamps only. Apply the "
                    "2026-08-03 membership-stage-history-grant migration.")
                hrows = []
            for r in hrows:
                to_stage = canon_membership_stage(r["to_stage"])
                key = (r["contact_id"], to_stage)
                if to_stage not in period_entries or key in seen:
                    continue
                seen[key] = {
                    "name": r["name"], "company": r["company"], "owner": r["owner"],
                    "assigned_by": r["changed_by"],
                    "entered_at": r["changed_at"].isoformat() if r["changed_at"] else None,
                }

            for (_cid, stage_key), rec in seen.items():
                period_entries[stage_key].append(rec)
            record_columns = [
                {"key": "name", "label": "Contact"},
                {"key": "company", "label": "Company"},
                {"key": "owner", "label": "Owner"},
                {"key": "assigned_by", "label": "Assigned by"},
                {"key": "entered_at", "label": "Entered"},
            ]

        else:  # opportunities
            orows = await conn.fetch("""
                WITH hist AS (
                    SELECT DISTINCT ON (h.opportunity_id, h.to_stage)
                           h.opportunity_id AS oid, h.to_stage AS stage,
                           h.changed_at AS entered_at, h.changed_by
                    FROM bedrock.jobs_stage_history h
                    JOIN bedrock.jobs_opportunity o ON o.id = h.opportunity_id
                    WHERE o.deleted_at IS NULL
                      AND h.changed_at >= $1 AND h.changed_at < $2
                      AND ($3::text IS NULL OR o.deal_type = $3)
                    ORDER BY h.opportunity_id, h.to_stage, h.changed_at
                ),
                created AS (
                    -- An opp created straight into a stage never gets a history
                    -- row, so without this it would vanish from the funnel.
                    SELECT o.id AS oid, o.stage, o.created_at AS entered_at,
                           NULL::text AS changed_by
                    FROM bedrock.jobs_opportunity o
                    WHERE o.deleted_at IS NULL
                      AND o.created_at >= $1 AND o.created_at < $2
                      AND ($3::text IS NULL OR o.deal_type = $3)
                      AND NOT EXISTS (SELECT 1 FROM bedrock.jobs_stage_history h2
                                      WHERE h2.opportunity_id = o.id)
                )
                SELECT e.stage, e.entered_at, e.changed_by, o.account_name AS name,
                       o.deal_type, o.owner_email AS owner
                FROM (SELECT * FROM hist UNION ALL SELECT * FROM created) e
                JOIN bedrock.jobs_opportunity o ON o.id = e.oid
                ORDER BY e.entered_at DESC
            """, p_from, p_to, dt)
            for r in orows:
                if r["stage"] in period_entries:
                    period_entries[r["stage"]].append({
                        "name": r["name"], "deal_type": r["deal_type"],
                        "owner": r["owner"], "assigned_by": r["changed_by"],
                        "entered_at": r["entered_at"].isoformat() if r["entered_at"] else None,
                    })
            record_columns = [
                {"key": "name", "label": "Company"},
                {"key": "deal_type", "label": "Type"},
                {"key": "owner", "label": "Owner"},
                {"key": "assigned_by", "label": "Moved by"},
                {"key": "entered_at", "label": "Entered"},
            ]

        for k in period_entries:
            period_entries[k].sort(key=lambda rec: rec["entered_at"] or "", reverse=True)
        return period_entries

    last_movement_at = None
    if period is not None:
        p_from, p_to = period
        period_entries = await _entries(p_from, p_to)
        # Prior window = same length, immediately before. Only its counts matter.
        span = p_to - p_from
        prev = await _entries(p_from - span, p_from)
        prev_counts = [len(prev.get(k, [])) for k, _ in stage_order]

        # When the window is empty, "nothing happened" is indistinguishable from
        # "you're looking at the wrong week". Report the most recent movement of
        # ANY kind so the empty state can point at where the data actually is.
        if not any(len(v) for v in period_entries.values()):
            if ftype == "prospects":
                try:
                    last_movement_at = await conn.fetchval("""
                        SELECT max(t) FROM (
                            SELECT max(assigned_at) AS t FROM bedrock.jobs_contact_membership
                            UNION ALL SELECT max(first_outreach_at) FROM bedrock.jobs_contact_membership
                            UNION ALL SELECT max(converted_at) FROM bedrock.jobs_contact_membership
                            UNION ALL SELECT max(changed_at) FROM bedrock.jobs_membership_stage_history
                        ) x
                    """)
                except asyncpg.exceptions.InsufficientPrivilegeError:
                    last_movement_at = await conn.fetchval("""
                        SELECT max(t) FROM (
                            SELECT max(assigned_at) AS t FROM bedrock.jobs_contact_membership
                            UNION ALL SELECT max(first_outreach_at) FROM bedrock.jobs_contact_membership
                            UNION ALL SELECT max(converted_at) FROM bedrock.jobs_contact_membership
                        ) x
                    """)
            else:
                last_movement_at = await conn.fetchval("""
                    SELECT max(t) FROM (
                        SELECT max(h.changed_at) AS t
                        FROM bedrock.jobs_stage_history h
                        JOIN bedrock.jobs_opportunity o ON o.id = h.opportunity_id
                        WHERE o.deleted_at IS NULL AND ($1::text IS NULL OR o.deal_type = $1)
                        UNION ALL
                        SELECT max(o.created_at) FROM bedrock.jobs_opportunity o
                        WHERE o.deleted_at IS NULL AND ($1::text IS NULL OR o.deal_type = $1)
                    ) x
                """, dt)

    stages = []
    cohort_conv = locals().get("cohort_conv") or {}
    if period_entries is not None:
        # Period mode supersedes the snapshot entirely: counts, records and
        # conversion all come from entries, so the cohort rate must not leak in.
        by_stage = period_entries
        cohort_conv = {}
    counts = [len(by_stage.get(k, [])) for k, _ in stage_order]
    max_count = max(counts) if counts else 1

    def _conv_to_next(cs: list, i: int, key: str) -> Optional[int]:
        """Share of stage i's arrivals that went on to stage i+1."""
        cnt = cs[i]
        nxt = cs[i + 1] if i + 1 < len(cs) else None
        if nxt is None or cnt <= 0:
            return None
        # On Hold is a parking state, not the step after Converted, so neither it
        # nor Converted advertises a rate into it.
        if key in _NO_CONVERSION_FROM:
            return None
        return round(100 * nxt / cnt)

    # Conversion is reported on the DESTINATION row — "% of the prior stage that
    # reached this one" — which is how the team reads it. That's the same number
    # as the source row's to-next rate, just displayed one row down.
    def _stage_conv(i: int, key: str) -> Optional[int]:
        raw = _conv_to_next(counts, i, key)
        if period_entries is not None:
            # Period flow reports the rate as measured, including >100% (a
            # backlog clearing) — clamping it is what made the old number
            # untrustworthy.
            return raw
        if key in cohort_conv:
            return cohort_conv[key]          # contacts snapshot: cohort rate
        return None if (raw is not None and raw > 100) else raw

    conv_to_next = [_stage_conv(i, k) for i, (k, _) in enumerate(stage_order)]
    prev_conv_to_next = (
        [_conv_to_next(prev_counts, i, k) for i, (k, _) in enumerate(stage_order)]
        if prev_counts is not None else None
    )

    for i, (k, label) in enumerate(stage_order):
        recs = by_stage.get(k, [])
        cnt = len(recs)
        mv = movement_by_stage.get(k, [])
        stages.append({
            "key": k, "label": label, "count": cnt,
            "pct_of_max": round(100 * cnt / max_count) if max_count else 0,
            "conversion_to_next": conv_to_next[i],
            # Same rate, addressed to the row it describes.
            "conversion_in": conv_to_next[i - 1] if i > 0 else None,
            # Prior window of equal length, for the two trend columns.
            "count_prev": prev_counts[i] if prev_counts is not None else None,
            "conversion_in_prev": (
                prev_conv_to_next[i - 1] if (prev_conv_to_next is not None and i > 0) else None
            ),
            "records": recs,
            "movement": mv,
            "advanced_in": sum(1 for m in mv if m["flow"] == "in" and m["direction"] == "advanced"),
            "regressed_in": sum(1 for m in mv if m["flow"] == "in" and m["direction"] == "regressed"),
        })

    return {"success": True, "data": {
        "type": ftype,
        "mode": "period" if period_entries is not None else "snapshot",
        "period": ({"from": period_from, "to": period_to} if period_entries is not None else None),
        # Only set when the window came back empty — the empty state uses it to
        # say where the movement actually is.
        "last_movement_at": last_movement_at.isoformat() if last_movement_at else None,
        "stages": stages,
        "record_columns": record_columns,
    }}


# ── Opportunities weekly overview (Thursday pipeline meeting) ──────────────────
# High-level, read-only view of the employer-deal pipeline. "Time in pipeline" is
# time in the CURRENT stage (from jobs_stage_history), not age since sourced.
# The active "set" = working-stage opps (initial_outreach + active_*).
# 'active_%' covers the two remaining active stages; reviewing_builders is named
# explicitly since the rename dropped its 'active_' prefix. initial_outreach is
# retained only so un-migrated rows still count as in-set.
_OPP_INSET = "(o.stage IN ('initial_outreach','reviewing_builders') OR o.stage LIKE 'active_%')"
# Stage×Time heatmap rows. Initial Outreach is gone as of the 2026-08-05
# simplification (its rows were remapped to In Discussions), so the heatmap's
# column totals now agree with the aging bars instead of sitting below them by
# the count of initial_outreach deals.
_OPP_STAGE_ORDER = [
    "active_in_discussions", "active_opportunity_confirmed", "reviewing_builders",
]
_OPP_AGE_BUCKETS = [
    ("lt2w", "< 2 weeks"), ("2_4w", "2–4 weeks"), ("4_6w", "4–6 weeks"),
    ("6_8w", "6–8 weeks"), ("8w", "8+ weeks"),
]
_OPP_STATUS_LABELS = {"new": "New (<1w)", "active": "Active", "stalled": "Stalled"}


# Read-time stage normalisation to the post-2026-08-05 vocabulary.
#
# The migration that rewrites these values is applied by Jac, so until it lands
# the database still holds the retired names. Canonicalising on READ means the UI
# shows the new vocabulary immediately — 'active_builder_interview' deals appear
# under Reviewing Builders rather than falling into no row at all and quietly
# vanishing from the heatmap and funnel. After the migration this is a no-op.
_STAGE_CANON = {
    "active_builder_interview": "reviewing_builders",
    "initial_outreach": "active_in_discussions",
    "on_hold_not_interested": "closed_lost",
    "on_hold_not_responsive": "closed_lost",
    "on_hold_not_selected": "closed_lost",
}
_MEMBERSHIP_CANON = {"on_hold": "revisit"}


def canon_stage(stage: Optional[str]) -> Optional[str]:
    return _STAGE_CANON.get(stage, stage) if stage else stage


def canon_membership_stage(stage: Optional[str]) -> Optional[str]:
    return _MEMBERSHIP_CANON.get(stage, stage) if stage else stage


def canon_stage_sql(col: str = "o.stage") -> str:
    """The same mapping as an SQL expression, for GROUP BY / filters."""
    whens = " ".join(f"WHEN '{k}' THEN '{v}'" for k, v in _STAGE_CANON.items())
    return f"(CASE {col} {whens} ELSE {col} END)"


def canon_membership_sql(col: str = "m.stage") -> str:
    whens = " ".join(f"WHEN '{k}' THEN '{v}'" for k, v in _MEMBERSHIP_CANON.items())
    return f"(CASE {col} {whens} ELSE {col} END)"


def _opp_age_bucket(days: int) -> int:
    if days < 14: return 0
    if days < 28: return 1
    if days < 42: return 2
    if days < 56: return 3
    return 4


@router.get("/opportunities/overview")
async def opportunities_overview(
    owner: Optional[str] = Query(None),
    deal_type: Optional[str] = Query(None),
    week_end: Optional[str] = Query(None, description="Last day of the window (YYYY-MM-DD), inclusive. Anchors the as-of view. Defaults to now."),
    start: Optional[str] = Query(None, description="First day of the window (YYYY-MM-DD), inclusive. Defaults to week_end - 6 days, i.e. a 7-day window. Lets the Thursday meeting run Thu→Thu."),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Weekly opportunity overview: summary cards, time-in-stage aging, a
    switchable breakdown (status / deal type / segment / stage / owner), the
    Priority×Time and Stage×Time concentration heatmaps, and the needs-attention
    list. Read-only. `owner`/`deal_type` scope the whole view; `week_end` sets the
    as-of reference (Saturday-to-Saturday) — ages and the net-new / moved-to-committed
    windows are measured back from it."""
    owner_f = owner if owner and owner != "all" else None
    dt_f = deal_type if deal_type and deal_type != "all" else None
    # `ref` = the as-of instant: midnight after the selected week-ending Saturday,
    # so the trailing 7-day window is that Sat–Sat week. Defaults to now.
    if week_end:
        try:
            d = date.fromisoformat(week_end)
            ref = datetime(d.year, d.month, d.day, tzinfo=timezone.utc) + timedelta(days=1)
        except ValueError:
            ref = datetime.now(timezone.utc)
    else:
        ref = datetime.now(timezone.utc)
    # Ages ("Nd in stage", stalled/new status, buckets, needs-attention) are
    # measured against the as-of instant but never against the FUTURE: the
    # current in-progress week arrives with the upcoming Saturday's week_end,
    # which would inflate every day-count by up to 7 days — and disagree with
    # Jobs Home, which calls this endpoint without week_end. Past weeks keep
    # their historical anchor. Week windows (net-new, moved, lost) stay on
    # `ref` so they remain calendar Sun–Sat weeks.
    age_ref = min(ref, datetime.now(timezone.utc))
    # Window = [win_start, ref). Defaults to 7 days back from ref so existing
    # callers are unchanged; an explicit `start` gives any span (Thu→Thu, a
    # month, a quarter). `prev_start` is the same span immediately before.
    win_start = ref - timedelta(days=7)
    if start:
        try:
            sd = date.fromisoformat(start)
            cand = datetime(sd.year, sd.month, sd.day, tzinfo=timezone.utc)
            if cand < ref:
                win_start = cand
        except ValueError:
            pass
    span = ref - win_start
    prev_start = win_start - span

    # As-of-`ref` membership: the active set as it stood at the end of the
    # selected week, reconstructed from created_at / closed_at so past weeks show
    # what was open THEN (not the live set). An opp is in the set as-of `ref` when
    # it existed by then and either (a) is still open and in a working stage, or
    # (b) was closed only after `ref` (so it was open at the time). When `ref` is
    # the current week this collapses to the live working set.
    rows = await conn.fetch(f"""
        SELECT o.id, o.account_name, o.stage, o.deal_type, o.segment, o.owner_email,
               o.priority, o.created_at,
               COALESCE((SELECT max(h.changed_at) FROM bedrock.jobs_stage_history h
                         WHERE h.opportunity_id = o.id AND h.to_stage = o.stage),
                        o.created_at) AS entered_stage,
               -- No history row for the current stage (pre-tracking / imported
               -- opps): the age below is time SINCE CREATION, not time in stage.
               (SELECT max(h.changed_at) FROM bedrock.jobs_stage_history h
                 WHERE h.opportunity_id = o.id AND h.to_stage = o.stage) IS NULL
                 AS stage_from_created,
               -- clamp to the as-of instant so activity logged AFTER a selected
               -- past week doesn't retroactively un-stall that week's view
               (SELECT max(a.activity_date) FROM bedrock.activity a
                WHERE a.jobs_opportunity_id = o.id AND a.deleted_at IS NULL
                  AND a.activity_date < $3::timestamptz) AS last_activity
        FROM bedrock.jobs_opportunity o
        WHERE o.deleted_at IS NULL
          AND o.created_at < $3::timestamptz
          AND (o.closed_at IS NULL OR o.closed_at >= $3::timestamptz)
          AND ({_OPP_INSET} OR (o.closed_at IS NOT NULL AND o.closed_at >= $3::timestamptz))
          AND ($1::text IS NULL OR o.owner_email = $1)
          AND ($2::text IS NULL OR o.deal_type = $2)
    """, owner_f, dt_f, ref)

    # Week-over-week, anchored to `ref` (Sat–Sat): net-new = opportunities created
    # in the selected week; prev = the prior 7-day window; moved-to-committed =
    # →closed_won in the week.
    net_new_rows = await conn.fetch(f"""
        SELECT o.id, o.account_name, o.stage, o.owner_email, o.created_at AS at
        FROM bedrock.jobs_opportunity o
        WHERE o.deleted_at IS NULL
          AND o.created_at >= $4::timestamptz AND o.created_at < $3::timestamptz
          AND ($1::text IS NULL OR o.owner_email = $1)
          AND ($2::text IS NULL OR o.deal_type = $2)
        ORDER BY o.created_at DESC
    """, owner_f, dt_f, ref, win_start)
    net_new = len(net_new_rows)
    net_new_prev = await conn.fetchval(f"""
        SELECT count(*) FROM bedrock.jobs_opportunity o
        WHERE o.deleted_at IS NULL
          AND o.created_at >= $4::timestamptz AND o.created_at < $3::timestamptz
          AND ($1::text IS NULL OR o.owner_email = $1)
          AND ($2::text IS NULL OR o.deal_type = $2)
    """, owner_f, dt_f, win_start, prev_start)
    won_rows = await conn.fetch(f"""
        SELECT DISTINCT o.id, o.account_name, o.stage, o.owner_email,
               max(h.changed_at) AS at
        FROM bedrock.jobs_stage_history h
        JOIN bedrock.jobs_opportunity o ON o.id = h.opportunity_id
        WHERE h.to_stage = 'closed_won'
          AND h.changed_at >= $4::timestamptz AND h.changed_at < $3::timestamptz
          AND o.deleted_at IS NULL
          AND ($1::text IS NULL OR o.owner_email = $1)
          AND ($2::text IS NULL OR o.deal_type = $2)
        GROUP BY o.id, o.account_name, o.stage, o.owner_email
        ORDER BY max(h.changed_at) DESC
    """, owner_f, dt_f, ref, win_start)
    moved_committed = len(won_rows)
    lost_rows = await conn.fetch(f"""
        SELECT DISTINCT o.id, o.account_name, o.stage, o.owner_email,
               max(h.changed_at) AS at
        FROM bedrock.jobs_stage_history h
        JOIN bedrock.jobs_opportunity o ON o.id = h.opportunity_id
        WHERE h.to_stage = 'closed_lost'
          AND h.changed_at >= $4::timestamptz AND h.changed_at < $3::timestamptz
          AND o.deleted_at IS NULL
          AND ($1::text IS NULL OR o.owner_email = $1)
          AND ($2::text IS NULL OR o.deal_type = $2)
        GROUP BY o.id, o.account_name, o.stage, o.owner_email
        ORDER BY max(h.changed_at) DESC
    """, owner_f, dt_f, ref, win_start)
    closed_lost = len(lost_rows)

    def _days(dt):
        return None if dt is None else max(0, (age_ref - dt).days)

    def _drill_row(r, at):
        return {
            "opportunity_id": str(r["id"]),
            "account": r["account_name"],
            "stage": canon_stage(r["stage"]),
            "stage_label": STAGE_LABELS.get(canon_stage(r["stage"]), r["stage"]),
            "owner": r["owner_email"],
            "at": at.isoformat() if at else None,
        }

    in_set = len(rows)
    stalled_6wk = 0  # active opps that have been an opportunity for >6 weeks (since created)
    # Every active-set member, flat, carrying the keys each panel groups by
    # (age bucket, status, deal type, segment, stage, owner, priority). The
    # frontend filters THIS array for every drill-down, so a drill can never
    # disagree with the count above it — one array, one source of truth.
    active_set: list = []
    age_counts = [0, 0, 0, 0, 0]
    bd: dict = {"status": {}, "deal_type": {}, "segment": {}, "stage": {}, "owner": {}}
    stage_heat: dict = {}
    prio_heat: dict = {}
    prio_unset = 0
    needs = []

    def _inc(dim: str, key: str):
        bd[dim][key] = bd[dim].get(key, 0) + 1

    for r in rows:
        c_days = _days(r["created_at"])
        if c_days is not None and c_days > 42:
            stalled_6wk += 1
        stage_days = _days(r["entered_stage"]) or 0
        bi = _opp_age_bucket(stage_days)
        age_counts[bi] += 1

        recency = min([d for d in (_days(r["entered_stage"]), _days(r["last_activity"])) if d is not None],
                      default=None)
        if c_days is not None and c_days < 7:
            status = "new"
        elif recency is not None and recency > 14:
            status = "stalled"
        else:
            status = "active"

        # Normalise once, so the drill filters and the counts use identical keys.
        dt_key = r["deal_type"] or "(unset)"
        seg_key = r["segment"] or "(unset)"
        owner_key = r["owner_email"] or "(unassigned)"

        _inc("status", status)
        _inc("deal_type", dt_key)
        _inc("segment", seg_key)
        _inc("stage", r["stage"])
        _inc("owner", owner_key)

        active_set.append({
            "opportunity_id": str(r["id"]),
            "account": r["account_name"],
            "owner": r["owner_email"],
            "stage": canon_stage(r["stage"]),
            "stage_label": STAGE_LABELS.get(canon_stage(r["stage"]), r["stage"]),
            "priority": r["priority"] if r["priority"] in (1, 2, 3, 4, 5) else None,
            "age_bucket": bi,
            "days_in_stage": stage_days,
            "status": status,
            "deal_type": dt_key,
            "segment": seg_key,
            "owner_key": owner_key,
        })

        stage_heat.setdefault(canon_stage(r["stage"]), [0, 0, 0, 0, 0])[bi] += 1
        if r["priority"] in (1, 2, 3, 4, 5):
            prio_heat.setdefault(r["priority"], [0, 0, 0, 0, 0])[bi] += 1
        else:
            prio_unset += 1

        if stage_days >= 21 or status == "stalled":
            act_days = _days(r["last_activity"])
            why_bits = []
            if stage_days >= 21:
                why_bits.append(
                    f"{stage_days}d old, no stage change logged" if r["stage_from_created"]
                    else f"{stage_days}d in {STAGE_LABELS.get(r['stage'], r['stage'])}")
            why_bits.append("no logged activity" if act_days is None else f"last activity {act_days}d ago")
            needs.append({
                "opportunity_id": str(r["id"]),
                "account": r["account_name"], "owner": r["owner_email"],
                "stage": r["stage"], "stage_label": STAGE_LABELS.get(canon_stage(r["stage"]), r["stage"]),
                "days_in_stage": stage_days, "days_since_activity": act_days,
                "stage_from_created": bool(r["stage_from_created"]),
                "why": " · ".join(why_bits),
            })

    def _bd_list(dim: str, labels: Optional[dict] = None):
        return [{"key": k, "label": (labels.get(k, k) if labels else k), "count": v}
                for k, v in sorted(bd[dim].items(), key=lambda kv: -kv[1])]

    # Heatmaps
    buckets = [{"key": k, "label": lbl} for k, lbl in _OPP_AGE_BUCKETS]
    stage_rows = [
        {"key": s, "label": STAGE_LABELS.get(s, s), "cells": stage_heat[s], "total": sum(stage_heat[s])}
        for s in _OPP_STAGE_ORDER if s in stage_heat
    ]
    prio_rows = [
        {"key": f"P{p}", "label": f"P{p}",
         "cells": prio_heat.get(p, [0, 0, 0, 0, 0]), "total": sum(prio_heat.get(p, [0, 0, 0, 0, 0]))}
        for p in (1, 2, 3, 4, 5)
    ]
    def _col_totals(rws):
        return [sum(r["cells"][i] for r in rws) for i in range(5)]

    needs.sort(key=lambda n: -n["days_in_stage"])

    # Recent activity in the selected Sat–Sat week: opportunities added, stage
    # moves (incl. won/lost), and opps crossing the 6-week "stalled" threshold.
    added_rows = await conn.fetch(f"""
        SELECT o.id, o.account_name, o.deal_type, o.stage, o.created_at AS at,
               COALESCE((SELECT h.changed_by FROM bedrock.jobs_stage_history h
                         WHERE h.opportunity_id = o.id AND h.changed_by IS NOT NULL
                         ORDER BY h.changed_at ASC LIMIT 1), o.owner_email) AS actor
        FROM bedrock.jobs_opportunity o
        WHERE o.deleted_at IS NULL
          AND o.created_at >= $4::timestamptz AND o.created_at < $3::timestamptz
          AND ($1::text IS NULL OR o.owner_email = $1)
          AND ($2::text IS NULL OR o.deal_type = $2)
    """, owner_f, dt_f, ref, win_start)
    moved_rows = await conn.fetch(f"""
        SELECT o.id, o.account_name, o.deal_type, h.from_stage, h.to_stage,
               h.changed_at AS at, h.changed_by AS actor
        FROM bedrock.jobs_stage_history h
        JOIN bedrock.jobs_opportunity o ON o.id = h.opportunity_id
        WHERE o.deleted_at IS NULL AND h.from_stage IS NOT NULL
          AND h.changed_at >= $4::timestamptz AND h.changed_at < $3::timestamptz
          AND ($1::text IS NULL OR o.owner_email = $1)
          AND ($2::text IS NULL OR o.deal_type = $2)
    """, owner_f, dt_f, ref, win_start)
    stalled_rows = await conn.fetch(f"""
        SELECT o.id, o.account_name, o.deal_type, o.stage, o.created_at, o.owner_email AS actor
        FROM bedrock.jobs_opportunity o
        WHERE o.deleted_at IS NULL AND {_OPP_INSET}
          AND o.created_at >= $3::timestamptz - interval '49 days' AND o.created_at < $3::timestamptz - interval '42 days'
          AND ($1::text IS NULL OR o.owner_email = $1)
          AND ($2::text IS NULL OR o.deal_type = $2)
    """, owner_f, dt_f, ref)

    recent_activity = []
    for r in added_rows:
        recent_activity.append({
            "type": "added", "opportunity_id": str(r["id"]), "account": r["account_name"],
            "deal_type": r["deal_type"], "stage_label": STAGE_LABELS.get(canon_stage(r["stage"]), r["stage"]),
            "detail": "Added to the set", "at": r["at"].isoformat() if r["at"] else None,
            "actor": r["actor"],
        })
    for r in moved_rows:
        to = canon_stage(r["to_stage"])
        frm = canon_stage(r["from_stage"])
        recent_activity.append({
            "type": "won" if to == "closed_won" else "lost" if to == "closed_lost" else "moved",
            "opportunity_id": str(r["id"]), "account": r["account_name"], "deal_type": r["deal_type"],
            "stage_label": STAGE_LABELS.get(to, to),
            "detail": f"{STAGE_LABELS.get(frm, frm)} → {STAGE_LABELS.get(to, to)}",
            "at": r["at"].isoformat() if r["at"] else None, "actor": r["actor"],
        })
    # 'Stalled' events deliberately NOT in this feed (2026-07-30 review): the
    # feed answers "what MOVED this week", and stalling is the absence of
    # movement — it's covered by the walkthrough's stalled groups instead.
    recent_activity = sorted((e for e in recent_activity if e["at"]), key=lambda e: e["at"], reverse=True)

    return {"success": True, "data": {
        "filters": {"owner": owner_f, "deal_type": dt_f, "week_end": week_end},
        "aging_basis": "time_in_stage",
        "summary": {
            "in_set": in_set, "net_new": net_new, "net_new_prev": net_new_prev,
            "moved_committed": moved_committed, "closed_lost": closed_lost, "stalled_6wk": stalled_6wk,
        },
        "aging": {"buckets": [
            {"key": k, "label": lbl, "count": age_counts[i],
             "pct": round(100 * age_counts[i] / in_set) if in_set else 0}
            for i, (k, lbl) in enumerate(_OPP_AGE_BUCKETS)
        ]},
        "breakdowns": {
            "status": _bd_list("status", _OPP_STATUS_LABELS),
            "deal_type": _bd_list("deal_type"),
            "segment": _bd_list("segment"),
            "stage": _bd_list("stage", STAGE_LABELS),
            "owner": _bd_list("owner"),
        },
        "heatmaps": {
            "buckets": buckets,
            "priority": {"rows": prio_rows, "col_totals": _col_totals(prio_rows),
                         "unset": prio_unset, "populated": prio_unset < in_set},
            "stage": {"rows": stage_rows, "col_totals": _col_totals(stage_rows)},
        },
        "needs_attention": needs[:30],
        # Every active-set member with its grouping keys. Backs the drill-downs
        # on the aging bars, the set distribution and the heatmap cells — all
        # three filter this one array, so no drill can contradict its count.
        "active_set": active_set,
        "recent_activity": recent_activity,
        # Rows behind each summary card — the card count and its drill list are
        # the same query, so they can never disagree (closed-won showed 10 but a
        # closed_at-based client drill found 1: those opps never got closed_at).
        "drills": {
            "in_set":  [_drill_row(r, r["entered_stage"]) for r in rows],
            "stalled": [_drill_row(r, r["created_at"]) for r in rows
                        if (_days(r["created_at"]) or 0) > 42],
            "net_new": [_drill_row(r, r["at"]) for r in net_new_rows],
            "won":     [_drill_row(r, r["at"]) for r in won_rows],
            "lost":    [_drill_row(r, r["at"]) for r in lost_rows],
        },
    }}



@router.get("/roles")
async def get_roles(user=Depends(require_auth), conn=Depends(get_db)):
    """Jobs / roles view — hired counts (from placements) + pipeline rows.

    Hired is counted by DISTINCT BUILDER from paid placements
    (public.employment_records, the single source of truth), matching the North
    Star numbers — NOT by application:
      hired_ft       = distinct builders with any paid FULL-TIME placement
      hired_contract = distinct builders in any paid work that is NOT full-time
                       (so the two partition the "in any paid work" total)
      avg_salary_ft  = avg pay of those FT builders' FT placements
    Interviewing / applied / rejected / withdrawn come from the submission
    funnel (public.job_applications, Pursuit-referred). Builder name comes from
    the 'Name: …' prefix stored in notes on import.
    """
    # ── Hired: placements from employment_records ────────────────────────────
    # COUNTS (cards) are by DISTINCT BUILDER with a recorded amount (payment > 0)
    # — "don't count them unless >0". But the table SHOWS every placement record,
    # including ones with no amount yet ("show all, even with $0"), so the team
    # can see (and backfill) them. Excludes pro_bono (never paid) and pipeline
    # (not an actual placement yet).
    plc = await conn.fetch(f"""
        SELECT * FROM bedrock.secured_jobs()
        WHERE employment_type <> 'pro_bono'
          AND (engagement_stage IS NULL OR engagement_stage NOT IN ('pipeline'))
          AND {_live_placement()}
        ORDER BY (payment_amount > 0) DESC NULLS LAST, payment_amount DESC NULLS LAST, builder
    """)

    out_rows = []
    ft_builders: set = set()
    other_builders: set = set()
    ft_salaries: list[int] = []
    for r in plc:
        is_ft = r["employment_type"] == "full_time"
        salary = int(r["payment_amount"]) if r["payment_amount"] is not None else None
        # distinct-builder counts only credit recorded paid work (payment > 0)
        if salary and salary > 0:
            if is_ft:
                ft_builders.add(r["user_id"])
                ft_salaries.append(salary)
            else:
                other_builders.add(r["user_id"])
        out_rows.append({
            "id": str(r["id"]),
            "builder": r["builder"] or "—",
            "role_title": r["role_title"] or "—",
            "company_name": r["company_name"] or "—",
            "salary": salary,
            "stage": "accepted",
            "segment": "hired_ft" if is_ft else "hired_contract",
        })

    # A builder counted as FT shouldn't also be counted under "other paid".
    other_builders -= ft_builders
    hired_ft = len(ft_builders)
    hired_contract = len(other_builders)

    # ── Pipeline (not-yet-hired): submission funnel from job_applications ─────
    apps = await conn.fetch("""
        SELECT
            ja.job_application_id AS id,
            trim(split_part(ja.notes, ':', 1)) AS builder,
            ja.role_title,
            ja.company_name,
            CASE WHEN ja.salary ~ '^[0-9]+$' THEN ja.salary::int ELSE NULL END AS salary,
            ja.stage
        FROM public.job_applications ja
        WHERE ja.source_type='Pursuit_referred' AND ja.stage <> 'accepted'
        ORDER BY
            CASE ja.stage WHEN 'interview' THEN 1 WHEN 'applied' THEN 2 ELSE 3 END,
            (ja.salary ~ '^[0-9]+$') DESC, ja.company_name
    """)
    seg_map = {"interview": "interviewing", "applied": "applied",
               "rejected": "rejected", "withdrawn": "withdrawn"}
    for r in apps:
        out_rows.append({
            "id": str(r["id"]),
            "builder": r["builder"] or "—",
            "role_title": r["role_title"] or "—",
            "company_name": r["company_name"] or "—",
            "salary": r["salary"],
            "stage": canon_stage(r["stage"]),
            "segment": seg_map.get(r["stage"], "other"),
        })

    interviewing = sum(1 for r in out_rows if r["segment"] == "interviewing")
    avg_ft = round(sum(ft_salaries) / len(ft_salaries)) if ft_salaries else None

    return {
        "success": True,
        "data": {
            "committed":       hired_ft + hired_contract + interviewing,  # hired + interviewing
            "hired_ft":        hired_ft,
            "hired_contract":  hired_contract,  # = builders in any paid work, not FT
            "hired_total":     hired_ft + hired_contract,
            "avg_salary_ft":   avg_ft,
            "rows": out_rows,
        },
    }


@router.get("/contacts/summary")
async def get_contacts_summary(user=Depends(require_auth), conn=Depends(get_db)):
    """Contacts & outreach metrics for the leadership dashboard.

    Mirrors the legacy Airtable "Job Outcomes Dash":
      Contacts & Leads:
        Total Leads   = every contact in the jobs pipeline (is_jobs_contact)
        Engaged Leads = those past the lead stage (initial outreach & beyond),
                        i.e. contact_stage IN (initial_outreach, active, on_hold)
      Employer Outreach (manually-logged engagements; source='manual'):
        All Outreach (last week) = engagements logged in the trailing 7 days
        Calls in total           = all call-type engagements
        Calls in last week       = call-type engagements in the trailing 7 days
    """
    stages = await conn.fetch("""
        SELECT contact_stage, count(*) AS count
        FROM public.contacts WHERE is_jobs_contact = true
        GROUP BY contact_stage ORDER BY count DESC
    """)
    total = sum(r["count"] for r in stages)

    # Engaged = distinct jobs prospects we've actually had activity with (manual
    # or synced gmail/calendar), linked via participant_public_contact_id by the
    # jobs-activity-link pass. Grows as the nightly scrape lands new touches.
    engaged = await conn.fetchval(f"""
        SELECT count(DISTINCT ct.contact_id)
        FROM public.contacts ct
        WHERE ct.is_jobs_contact = true
          AND EXISTS (
            SELECT 1 FROM bedrock.activity a
            WHERE a.participant_public_contact_id = ct.contact_id AND a.deleted_at IS NULL
              AND {_jobs_relevant('a')}
          )
    """)

    # Outreach / Calls = FIRST TOUCHES by the jobs team (JOBS_TEAM_EMAILS).
    # Outreach: each external contact counts once — the week number is contacts
    # whose first-ever outbound email from the team landed in the last 7 days.
    # Calls/Mtgs: same first-touch rule over external attendees of meetings on
    # the team's calendars.
    em = await conn.fetchrow(f"""
        {_first_touch_email_cte()}
        SELECT count(*) FILTER (WHERE first_touch >= now() - interval '7 days')  AS wk,
               count(*) FILTER (WHERE first_touch >= now() - interval '30 days') AS mo,
               count(*)                                                          AS total
        FROM ext
    """)
    mt = await conn.fetchrow(f"""
        {_first_touch_meeting_cte()}
        SELECT count(*) FILTER (WHERE first_touch >= now() - interval '7 days') AS wk,
               count(*)                                                         AS total
        FROM ext
    """)
    activity = {
        "outreach_total":     em["total"],
        "outreach_this_week": em["wk"],
        "outreach_this_month": em["mo"],
        "calls_total":        mt["total"],
        "calls_this_week":    mt["wk"],
        "meetings_total":     mt["total"],
        "active_owners":      len(JOBS_TEAM_EMAILS),
    }

    active_companies = await conn.fetchval("""
        SELECT count(*) FROM bedrock.jobs_opportunity
        WHERE deleted_at IS NULL AND stage LIKE 'active_%'
    """)

    # Account-level leads: distinct companies in the jobs pipeline (jobs contacts'
    # company ∪ opportunity accounts), and how many have ANY activity (engaged).
    acct = await conn.fetchrow(f"""
        WITH companies AS (
            SELECT DISTINCT lower(trim(current_company)) AS company
            FROM public.contacts WHERE is_jobs_contact=true AND coalesce(trim(current_company),'')<>''
            UNION
            SELECT DISTINCT lower(trim(account_name)) FROM bedrock.jobs_opportunity
            WHERE deleted_at IS NULL AND coalesce(trim(account_name),'')<>''
        ),
        engaged AS (
            SELECT DISTINCT lower(trim(c.current_company)) AS company
            FROM bedrock.activity a JOIN public.contacts c ON c.contact_id = a.participant_public_contact_id
            WHERE a.deleted_at IS NULL AND coalesce(trim(c.current_company),'')<>''
              AND {_jobs_relevant('a')}
        )
        SELECT (SELECT count(*) FROM companies) AS total,
               (SELECT count(*) FROM companies WHERE company IN (SELECT company FROM engaged)) AS engaged
    """)

    return {
        "success": True,
        "data": {
            "contacts": {
                "total":    total,
                "engaged":  engaged,
                "by_stage": [{"stage": r["contact_stage"] or "none", "count": r["count"]} for r in stages],
            },
            "accounts": {"total": acct["total"], "engaged": acct["engaged"]},
            "activity": dict(activity),
            "active_companies": active_companies,
        },
    }


def _team_actor(alias: str = "a") -> str:
    """SQL: this activity row was authored BY the jobs team (Avni/Damon) — they
    sent the email, or it's on their synced calendar / a manual log they made."""
    conds = []
    for e in JOBS_TEAM_EMAILS:
        conds.append(f"{alias}.email_from ILIKE '%{e}%'")
        conds.append(f"{alias}.logged_by ILIKE '%{e}%'")
    return "(" + " OR ".join(conds) + ")"


_SAFE_EMAIL = re.compile(r"^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$")


def _engaged_clause(alias: str = "c") -> str:
    """SQL boolean: this contact is 'engaged' — worth showing in the default
    pipeline views — vs a cold LinkedIn import we've never touched. Engaged =
    not a bare linkedin_import, OR has any activity, OR is linked to Salesforce.
    Keeps default lists ~15k instead of ~47k; pass scope=all to drop it."""
    return (
        f"({alias}.source IS DISTINCT FROM 'linkedin_import' "
        f"OR EXISTS(SELECT 1 FROM bedrock.activity a WHERE a.participant_public_contact_id={alias}.contact_id AND a.deleted_at IS NULL AND {_jobs_relevant('a')}) "
        f"OR EXISTS(SELECT 1 FROM bedrock.sf_contact_link l WHERE l.public_contact_id={alias}.contact_id))"
    )


# Subjects/senders that are automated replies or bounces, never real outreach —
# excluded from outreach counts (OOO auto-replies were inflating the numbers).
_AUTOREPLY_SUBJECTS = (
    "out of office", "automatic reply", "auto-reply", "autoreply", "auto reply",
    "ooo:", "ooo -", "away from", "on vacation", "on leave", "maternity leave",
    "thank you for your message", "thank you for your email", "thank you for contacting",
    "undeliverable", "delivery status notification", "mail delivery", "returned mail",
)
_AUTOREPLY_SENDERS = ("mailer-daemon", "postmaster", "no-reply", "noreply", "donotreply", "do-not-reply")


def _not_autoreply(alias: str) -> str:
    """SQL predicate: this activity is NOT an automated reply / bounce."""
    subj = " AND ".join(f"coalesce({alias}.subject,'') NOT ILIKE '%{p}%'" for p in _AUTOREPLY_SUBJECTS)
    frm = " AND ".join(f"coalesce({alias}.email_from,'') NOT ILIKE '%{p}%'" for p in _AUTOREPLY_SENDERS)
    return f"({subj} AND {frm})"


def _jobs_relevant(alias: str) -> str:
    """SQL predicate: this activity counts as jobs outreach. Synced email/meeting
    rows are gated by the content classifier (jobs_relevance='jobs'); manually
    logged touches (call/text/linkedin/note entered in the jobs tool) are
    deliberate jobs outreach and always count. See services/activity_classifier.py."""
    return (f"(coalesce({alias}.jobs_relevance_override, {alias}.jobs_relevance) = 'jobs' "
            f"OR {alias}.type NOT IN ('email','meeting'))")


# ── firmographic derivations (My Network filters) ─────────────────────────────
# Both ladders are PORTED VERBATIM from ~/employer-prospect-ranking/scripts/lib_sql.py,
# which implements the "Prioritizing our employer prospect list" doc
# (1C3BNx4gbEA91BXciq22UP58OoUgAZGbXdzUsC42jshQ). They were reviewed against that
# doc's own tables — do not re-derive them here. If one changes, change both.

def _seniority_case(col: str) -> str:
    """Map a job title to the doc's seniority ladder.

    Evaluated top to bottom, first match wins, so the narrowest patterns come
    first. Two judgment calls, both settled by the doc's "when in doubt, rate it
    down": SVP → Medium (its High is "C-suite or EVP", and SVP is neither) and
    Managing Director → Medium (senior finance, but not C-suite; plain Director
    stays Low)."""
    return f"""
CASE
  WHEN {col} IS NULL OR btrim({col}) = '' THEN 'Lowest'
  WHEN {col} ~* '(chief executive|(^|[^a-z])ceo([^a-z]|$)|founder|co[- ]?founder|(^|[^a-z])owner([^a-z]|$)|proprietor|managing partner|founding partner)'
       THEN 'Highest'
  WHEN {col} ~* '(^|[^a-z])(chief[a-z ]*officer|chief [a-z]+|cto|cfo|coo|chro|cio|cmo|cro|cpo|cso|cdo|cco|ciso|cxo)([^a-z]|$)'
       OR {col} ~* '(^|[^a-z])(evp|executive vice president|president|chairman|chairwoman|chairperson|general partner)([^a-z]|$)'
       THEN 'High'
  WHEN {col} ~* '(^|[^a-z])(svp|senior vice president|vp|vice president|head of|global head|managing director|(^|[^a-z])partner([^a-z]|$))'
       THEN 'Medium'
  WHEN {col} ~* '(^|[^a-z])(director|manager|principal|team lead|tech lead|supervisor)([^a-z]|$)'
       THEN 'Low'
  ELSE 'Lowest'
END"""


def _tristate_case(col: str) -> str:
    """Map a company HQ string to tri-state presence.

    Yields 'Yes' | 'HQ elsewhere' | 'Unknown' and DELIBERATELY NEVER 'No'.
    Presence is not headquarters: Google, Amazon and Microsoft all run large New
    York offices from out-of-state HQs. An HQ can confirm presence; only research
    can refute it. Anything stronger than 'HQ elsewhere' would be a claim this
    column cannot support."""
    return f"""
CASE
  WHEN coalesce({col}, '') = '' OR lower({col}) = 'null' THEN 'Unknown'
  WHEN {col} ~* '(,\\s*(NY|NJ|CT)\\s*$)|(,\\s*(NY|NJ|CT)\\s*,)'
       OR {col} ~* '(new york|brooklyn|bronx|queens|manhattan|staten island|long island city|yonkers|white plains|new rochelle|mount vernon|westchester|hoboken|newark|jersey city|princeton|hackensack|paramus|morristown|stamford|greenwich|norwalk|bridgeport|hartford|new haven|danbury|westport|darien)'
       THEN 'Yes'
  ELSE 'HQ elsewhere'
END"""


# The COMPLETE output domains of the two expressions above. Both CASEs are total
# (every input lands on a branch, NULL included), so these lists need no query to
# discover — which keeps them out of the per-request facet scan.
_SENIORITY_RUNGS = ["Highest", "High", "Medium", "Low", "Lowest"]
_TRISTATE_VALUES = ["Yes", "HQ elsewhere", "Unknown"]


# ── My Network priority banding ──────────────────────────────────────────────
# Jac's rule, 2026-08-05. Three "fits" are scored, then portfolio-company status
# promotes, then an exclusion overrides everything:
#
#   fits = (headcount 51-200) + (tri-state Yes|Unknown) + (seniority High|Highest)
#
#   portco  fits   priority
#   -----------------------
#   no      3      P1
#   no      2      P2
#   no      0-1    NULL (unranked — a badge means "act on this")
#   yes     >=2    P1
#   yes     0-1    P2
#
# Plus an override (Jac, 2026-08-05): seniority 'Highest' at a company of roughly
# 10 to 5,000 people is P1, whatever else is true. That rung is founder /
# co-founder / CEO / owner / proprietor / managing or founding partner — someone who
# can decide to hire without asking anyone.
#
# The headcount window is what makes it useful rather than noise; see
# _PRIORITY_SENIORITY_HEADCOUNT_WINDOW for what each end excludes and why. A NULL
# size_bucket does not qualify either — "10 to 5,000 people" is a positive claim and
# 14% of the network has no headcount on file.
#
# Excluded from P1/P2 entirely: anyone at Pursuit, and anyone carrying an
# alumni_* tag. They aren't employer leads, and the exclusion outranks every rule
# above including the Highest override.
#
# Tri-state counts Unknown as a fit ON PURPOSE: the value is derived from HQ, and
# an unknown HQ is missing data, not an absent office. Treating it as a miss
# would quietly bury leads for the ~40% of companies with no HQ on file.
_PRIORITY_HEADCOUNT = "51-200"
_PRIORITY_TRISTATE = ("Yes", "Unknown")
_PRIORITY_SENIORITY = ("High", "Highest")
# The rung that is P1 on its own, provided the company sits in the headcount window
# below. Must be one of _SENIORITY_RUNGS.
_PRIORITY_SENIORITY_TOP = "Highest"
# The window in which "Highest seniority" alone earns P1: roughly 10 to 5,000 people
# (Jac, 2026-08-05). Both ends are doing work.
#   * 1-10 excluded: every one-person consultancy has a founder. Without this floor
#     the rung banded 4,524 people P1 — more than P2 held.
#   * 5000+ excluded: at that size the seniority ladder's 'owner' and
#     'managing partner' patterns stop meaning ownership. The measured false
#     positives were all of that shape — "Product Portfolio Owner" and "Pittsburgh
#     Office Managing Partner" at PwC — people who cannot decide to hire anyone.
#
# An ALLOW-list, not an exclusion list: if a new size band is added later it will
# not silently start earning P1. Conservative is the right default for a promotion.
_PRIORITY_SENIORITY_HEADCOUNT_WINDOW = ("11-50", "51-200", "201-1000", "1001-5000")


def _net_priority_case(portco_expr: Optional[str] = None,
                       account_floor_expr: Optional[str] = None) -> str:
    """SQL CASE returning 'P1' | 'P2' | NULL for one network row.

    `portco_expr` is a boolean SQL expression, or None when
    bedrock.company_investor hasn't been created yet — in which case nothing is a
    portco and the banding degrades to the headcount/tri-state/seniority rules
    alone. The endpoint checks for the table per request, so applying the
    migration starts the portco clause working with no code change.

    `account_floor_expr` is a scalar subquery yielding a band ('P2') for accounts
    that guarantee a minimum, or None while bedrock.priority_account_floor is
    absent. It is evaluated LAST on purpose — see the comment at the branch.
    """
    tri = ", ".join(f"'{v}'" for v in _PRIORITY_TRISTATE)
    sen = ", ".join(f"'{v}'" for v in _PRIORITY_SENIORITY)
    # Every term is coalesced to false before being summed. Without it, a NULL
    # size_bucket (14% of the network has no headcount on file) makes the whole
    # sum NULL, every comparison against it NULL, and the row falls through to
    # unranked — silently dropping ~1,080 contacts who DID hit the other two
    # criteria. A missing headcount is one criterion unmet, not a void score.
    fits = (
        f"(coalesce(co.size_bucket = '{_PRIORITY_HEADCOUNT}', false)::int"
        f" + coalesce(({_tristate_case('co.hq_location')}) IN ({tri}), false)::int"
        f" + coalesce(({_seniority_case('c.current_title')}) IN ({sen}), false)::int)"
    )
    # Anchored on the "pursuit" prefix, NOT a bare ILIKE '%pursuit%' — the loose
    # form also swallows "In Wild Pursuit", a real founder lead, while the prefix
    # still catches "Pursuit Fellow" and the "Pursuit Company - …" import artifact.
    excluded = (
        "(lower(btrim(coalesce(c.current_company,''))) ~ '^pursuit($|[^a-z])'"
        " OR EXISTS (SELECT 1 FROM unnest(coalesce(c.tags, '{}'::text[])) t WHERE t LIKE 'alumni%'))"
    )
    portco = portco_expr or "false"
    # "Highest seniority at a company of roughly 10 to 5,000 people."
    window = ", ".join(f"'{b}'" for b in _PRIORITY_SENIORITY_HEADCOUNT_WINDOW)
    top_decider = (
        f"(({_seniority_case('c.current_title')}) = '{_PRIORITY_SENIORITY_TOP}'"
        f" AND co.size_bucket IN ({window}))"
    )
    # NULL::text, not "false" — this one yields a band, not a boolean.
    floor = account_floor_expr or "NULL::text"
    return f"""
CASE
  WHEN {excluded} THEN NULL
  WHEN {top_decider} THEN 'P1'
  WHEN ({portco}) AND {fits} >= 2 THEN 'P1'
  WHEN ({portco})                 THEN 'P2'
  WHEN {fits} = 3                 THEN 'P1'
  WHEN {fits} = 2                 THEN 'P2'
  -- Account floor, LAST on purpose. "P2 minimum" may only RAISE a row that would
  -- otherwise be unranked: every P1 path above has already been taken, so this
  -- branch can never demote anyone. It also sits below the exclusion at the top,
  -- so a Pursuit staffer or an alumnus at a Work-now account stays unbanded.
  WHEN ({floor}) IS NOT NULL      THEN ({floor})
  ELSE NULL
END"""


# Portfolio-company test, used only when the table exists. account_key is the
# normalised company name the jobs app already groups on.
_PORTCO_EXISTS = (
    "EXISTS (SELECT 1 FROM bedrock.company_investor ci"
    " WHERE ci.account_key = lower(btrim(coalesce(c.current_company, ''))))"
)

# Account-level floor: "anyone at a Work-now account is P2 minimum" (Jac,
# 2026-08-06). Same name-keyed join as the portco test. Used only when
# bedrock.priority_account_floor exists.
_ACCOUNT_FLOOR_BAND = (
    "(SELECT paf.floor_band FROM bedrock.priority_account_floor paf"
    " WHERE paf.account_key = lower(btrim(coalesce(c.current_company, ''))))"
)


_STAFF_EMAIL_TOKEN = ":staff_email"


def _net_rule_clause(
    rule: dict,
    fields: dict[str, tuple[str, str]],
    params: list,
    staff_email: str = "",
) -> tuple[str, list]:
    """Translate one {field,op,values} filter rule into a SQL clause.

    Returns (clause, params); an empty clause means "no-op, skip this rule".
    Values are always bound, never interpolated, and `fields` is a whitelist —
    an unrecognised field or an op the field's type doesn't support raises 400
    rather than silently matching everything, which would under-filter without
    telling anyone.

    An expression may contain the literal `:staff_email` token; it is bound on
    demand, only for the rules that actually use it. Binding it up front instead
    would hand the count queries a parameter their WHERE never references, and
    asyncpg rejects an argument count that doesn't match the statement.

    list_contacts has its own larger translator (see `_RULE_FIELDS` there) that
    also covers number/date/recency. That one is load-bearing and carries
    filter-fuzzing fixes from 2026-07-16, so it was left alone rather than
    refactored underneath. This helper is the place to consolidate to when
    someone has the appetite to migrate it.
    """
    field, op = rule.get("field"), rule.get("op")
    values = [str(v) for v in (rule.get("values") or [])]
    spec = fields.get(field)
    if spec is None:
        raise HTTPException(400, f"unfilterable field: {field}")
    ftype, expr = spec

    if _STAFF_EMAIL_TOKEN in expr:
        # One binding, substituted into every occurrence in this expression.
        params.append(f"%{staff_email}%")
        expr = expr.replace(_STAFF_EMAIL_TOKEN, f"${len(params)}")

    if op == "is_empty":
        if ftype == "tags":
            return f"coalesce(array_length({expr}, 1), 0) = 0", params
        if ftype == "boolean":
            return f"NOT {expr}", params          # a boolean is never "empty"
        return f"(({expr}) IS NULL OR ({expr})::text = '')", params
    if op == "is_not_empty":
        if ftype == "tags":
            return f"coalesce(array_length({expr}, 1), 0) > 0", params
        if ftype == "boolean":
            return expr, params
        return f"(({expr}) IS NOT NULL AND ({expr})::text <> '')", params

    if ftype == "boolean" and op in ("equals", "not_equals"):
        # "yes"/"no" from the picker, flipped again by not_equals.
        want = ((values[0] if values else "") == "yes") ^ (op == "not_equals")
        return (expr if want else f"NOT {expr}"), params

    if not values:
        return "", params                          # nothing selected yet — no-op

    neg = "NOT " if op == "not_equals" else ""
    if ftype == "tags" and op in ("equals", "not_equals"):
        params.append(values)
        return f"{neg}(coalesce({expr}, '{{}}'::text[]) && ${len(params)}::text[])", params
    if ftype == "select" and op in ("equals", "not_equals"):
        params.append(values)
        return f"{neg}(coalesce(({expr})::text, '') = ANY(${len(params)}::text[]))", params
    if ftype == "text" and op in ("contains", "equals", "not_equals"):
        if op == "contains":
            params.append(f"%{values[0]}%")
            return f"lower(coalesce({expr}, '')) LIKE lower(${len(params)})", params
        params.append(values[0])
        sym = "=" if op == "equals" else "<>"
        return f"lower(coalesce({expr}, '')) {sym} lower(${len(params)})", params

    raise HTTPException(400, f"unsupported op '{op}' for {field}")


# Canonical placement-status derivation for a jobs_role, so every screen (Home,
# Accounts, drills) agrees. Stakeholder vocabulary (JOBS_REVIEW_PLAN.md #9):
#   ft_placed       — a builder is placed full-time (the FT seat is filled)
#   trial_active    — a committed trial with a builder currently in it (converts to a
#                     separate FT role, so the FT number never walks back — e.g. Fowler)
#   committed_open  — signed hiring commitment, seat still open (incl. open trial reqs)
#   open_market     — CVs welcome, no hiring commitment yet
#   cancelled       — role cancelled
def _placement_status_sql(a: str = "r") -> str:
    return f"""CASE
        WHEN {a}.status = 'cancelled' THEN 'cancelled'
        WHEN {a}.filled_by_user_id IS NOT NULL AND coalesce({a}.is_trial,false) = false THEN 'ft_placed'
        WHEN {a}.filled_by_user_id IS NOT NULL AND {a}.is_trial THEN 'trial_active'
        WHEN {a}.commitment = 'committed' THEN 'committed_open'
        ELSE 'open_market' END"""

PLACEMENT_STATUS_LABELS = {
    "ft_placed": "Full-time placed",
    "trial_active": "Committed: trial active",
    "committed_open": "Committed: no placement",
    "open_market": "Open market",
    "cancelled": "Closed",
}


def _staff_actor(alias: str = "a") -> str:
    """SQL: authored by an active staff member OUTSIDE the core jobs team — the
    'staff mobilization' scope. Paired with _jobs_relevant, this surfaces jobs
    outreach the wider staff do on top of their day jobs (kept out of the core
    Outreach & Activation number, which stays Avni/Damon/Devika)."""
    excl = ",".join(f"'{e.lower()}'" for e in JOBS_TEAM_EMAILS)
    return (f"(EXISTS (SELECT 1 FROM public.org_users o WHERE o.is_active "
            f"AND lower(o.email) NOT IN ({excl}) AND o.email IS NOT NULL "
            f"AND ({alias}.email_from ILIKE '%'||o.email||'%' OR {alias}.logged_by ILIKE '%'||o.email||'%')))")


def _actor_sql(alias: str, owner: Optional[str], scope: str = "team") -> str:
    """Actor filter for the outreach trends/detail. With `owner` (a single,
    validated staff email) it scopes to that person; else `scope` picks the core
    jobs team ('team', default), the wider staff ('staff'), or everyone at
    Pursuit ('pursuit'). `owner` is regex-validated so it's safe to interpolate
    into the ILIKE."""
    if owner and _SAFE_EMAIL.match(owner):
        return f"({alias}.email_from ILIKE '%{owner}%' OR {alias}.logged_by ILIKE '%{owner}%')"
    if scope == "staff":
        return _staff_actor(alias)
    if scope == "pursuit":
        return f"({_team_actor(alias)} OR {_staff_actor(alias)})"
    return _team_actor(alias)


@router.get("/activity-trends")
async def activity_trends(
    granularity: str = Query("week", pattern="^(day|week|month)$"),
    channel: str = Query("all", pattern="^(all|email|meeting)$"),
    owner: Optional[str] = Query(None, description="Scope to one staff email (else the scope)"),
    scope: str = Query("team", pattern="^(pursuit|team|staff)$", description="team = Avni/Damon/Devika; staff = everyone else; pursuit = both"),
    date_from: Optional[str] = Query(None, pattern=r"^\d{4}-\d{2}-\d{2}$", description="Custom range start (ISO date); overrides trailing window when both from+to are set"),
    date_to: Optional[str] = Query(None, pattern=r"^\d{4}-\d{2}-\d{2}$", description="Custom range end (ISO date)"),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Account-level outreach over time — one stacked bar per period split into
    touches to NEW accounts (first activated this period) vs EXISTING accounts.

    A "team touch" = an email sent by Avni/Damon, a meeting on their calendars,
    or a manual log, mapped to an account (the counterpart's company) and counted
    once per (activity, account). `channel` = all | email | meeting toggles which
    touches count; new-vs-existing is by the account's first-ever team touch.
    `granularity` = week | month; trailing 12 buckets, zero-filled.
    """
    periods = 14 if granularity == "day" else 12
    actor = _actor_sql("a", owner, scope)
    chan_sql = (
        "CASE WHEN a.source='calendar-sync' OR a.type='meeting' THEN 'meeting' "
        "WHEN a.type IN ('email') OR a.source='gmail-sync' THEN 'email' ELSE 'other' END"
    )

    # Each team activity mapped to an account (company), via the counterpart
    # contact (participant link, email recipient, or meeting attendee), counted
    # once per (activity, account). Then classify each account-touch as NEW (the
    # period the account was first touched) vs EXISTING.
    rows = await conn.fetch(f"""
        WITH team_act AS (
          SELECT a.id, a.activity_date, a.participant_public_contact_id AS cid,
                 a.email_to, a.email_cc, a.meeting_attendees, {chan_sql} AS channel
          FROM bedrock.activity a
          WHERE a.deleted_at IS NULL AND {actor} AND {_not_autoreply('a')} AND {_jobs_relevant('a')}
        ),
        touch_contact AS (
          SELECT id, activity_date, channel, cid AS contact_id FROM team_act WHERE cid IS NOT NULL
          UNION
          SELECT t.id, t.activity_date, t.channel, c.contact_id
          FROM team_act t, unnest(coalesce(t.email_to,'{{}}') || coalesce(t.email_cc,'{{}}')) e
          JOIN public.contacts c ON lower(c.email) = lower(e)
          WHERE t.channel = 'email'
          UNION
          SELECT t.id, t.activity_date, t.channel, c.contact_id
          FROM team_act t, jsonb_array_elements(coalesce(t.meeting_attendees, '[]'::jsonb)) att
          JOIN public.contacts c ON lower(c.email) = lower(att->>'email')
          WHERE t.channel = 'meeting'
        ),
        acct_touch AS (   -- one row per (activity, account, channel)
          SELECT DISTINCT tc.id, tc.activity_date, tc.channel,
                 lower(trim(c.current_company)) AS company
          FROM touch_contact tc
          JOIN public.contacts c ON c.contact_id = tc.contact_id
          WHERE coalesce(trim(c.current_company), '') <> ''
        ),
        acct_first AS (   -- the period each account was first touched (any channel)
          SELECT company, date_trunc('{granularity}', min(activity_date)) AS first_period
          FROM acct_touch GROUP BY company
        )
        SELECT date_trunc('{granularity}', t.activity_date) AS bucket,
               CASE WHEN date_trunc('{granularity}', t.activity_date) = af.first_period
                    THEN 'new' ELSE 'existing' END AS kind,
               count(*) AS n
        FROM acct_touch t JOIN acct_first af USING (company)
        WHERE ($1 = 'all' OR t.channel = $1)
        GROUP BY 1, 2
    """, channel)

    def _add_months(dt, n):
        y, m = dt.year, dt.month + n
        while m <= 0:
            m += 12; y -= 1
        while m > 12:
            m -= 12; y += 1
        return dt.replace(year=y, month=m, day=1)

    buckets = []
    if date_from and date_to:
        # Custom range: enumerate every bucket start between from..to (inclusive),
        # aligned to the granularity via Postgres date_trunc so week starts match
        # the aggregation. Capped at 400 buckets to bound the payload/chart.
        start_b, end_b = await conn.fetchrow(
            f"SELECT date_trunc('{granularity}', $1::timestamp) AS a, "
            f"date_trunc('{granularity}', $2::timestamp) AS b",
            datetime.fromisoformat(date_from), datetime.fromisoformat(date_to),
        )
        cur, guard = start_b, 0
        while cur <= end_b and guard < 400:
            buckets.append({"period": cur.date().isoformat(), "new": 0, "existing": 0})
            if granularity == "day":
                cur = cur + timedelta(days=1)
            elif granularity == "week":
                cur = cur + timedelta(weeks=1)
            else:
                cur = _add_months(cur, 1)
            guard += 1
    else:
        # Assemble trailing `periods` buckets (zero-filled), newest last.
        base = await conn.fetchval(f"SELECT date_trunc('{granularity}', now())")
        for i in range(periods - 1, -1, -1):
            if granularity == "day":
                start = base - timedelta(days=i)
            elif granularity == "week":
                start = base - timedelta(weeks=i)
            else:
                start = _add_months(base, -i)
            buckets.append({"period": start.date().isoformat(), "new": 0, "existing": 0})
    idx = {b["period"]: b for b in buckets}

    def _key(ts):
        return ts.date().isoformat() if ts else None

    for r in rows:
        b = idx.get(_key(r["bucket"]))
        if b:
            b[r["kind"]] = r["n"]

    # Coverage note — Damon's mailbox is only partially synced, so flag low coverage.
    damon = await conn.fetchval(
        "SELECT count(*) FROM bedrock.activity a WHERE a.deleted_at IS NULL "
        "AND (a.email_from ILIKE '%damon.kornhauser@pursuit.org%' OR a.logged_by ILIKE '%damon.kornhauser@pursuit.org%')"
    )
    return {
        "success": True,
        "data": {
            "granularity": granularity,
            "channel": channel,
            "buckets": buckets,
            "totals": {
                "new": sum(b["new"] for b in buckets),
                "existing": sum(b["existing"] for b in buckets),
                "touches": sum(b["new"] + b["existing"] for b in buckets),
            },
            "coverage_note": (
                "Damon's mailbox sync is sparse (and he also sends from non-Pursuit addresses), "
                "so his outreach is undercounted." if damon < 200 else None
            ),
        },
    }


@router.get("/activity-trends/detail")
async def activity_trends_detail(
    period: str = Query(..., description="Bucket start date (ISO, the bar's period)"),
    granularity: str = Query("week", pattern="^(day|week|month)$"),
    channel: str = Query("all", pattern="^(all|email|meeting)$"),
    owner: Optional[str] = Query(None),
    scope: str = Query("team", pattern="^(pursuit|team|staff)$"),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Drill-down for one outreach bar: the individual account-touches in that
    period (account + contact + subject + date + channel), so clicking a bar
    shows exactly who was reached out to. Mirrors /activity-trends' actor +
    account-mapping logic, scoped to the single bucket."""
    try:
        pstart = date.fromisoformat(period)
    except ValueError:
        raise HTTPException(400, "period must be ISO date")
    actor = _actor_sql("a", owner, scope)
    chan_sql = (
        "CASE WHEN a.source='calendar-sync' OR a.type='meeting' THEN 'meeting' "
        "WHEN a.type IN ('email') OR a.source='gmail-sync' THEN 'email' ELSE 'other' END"
    )
    rows = await conn.fetch(f"""
        WITH team_act AS (
          SELECT a.id, a.activity_date, a.subject, a.participant_public_contact_id AS cid,
                 a.email_to, a.email_cc, a.meeting_attendees, {chan_sql} AS channel
          FROM bedrock.activity a
          WHERE a.deleted_at IS NULL AND {actor} AND {_not_autoreply('a')} AND {_jobs_relevant('a')}
            AND date_trunc('{granularity}', a.activity_date) = date_trunc('{granularity}', $1::timestamptz)
        ),
        touch_contact AS (
          SELECT id, activity_date, subject, channel, cid AS contact_id FROM team_act WHERE cid IS NOT NULL
          UNION
          SELECT t.id, t.activity_date, t.subject, t.channel, c.contact_id
          FROM team_act t, unnest(coalesce(t.email_to,'{{}}') || coalesce(t.email_cc,'{{}}')) e
          JOIN public.contacts c ON lower(c.email) = lower(e)
          WHERE t.channel = 'email'
          UNION
          SELECT t.id, t.activity_date, t.subject, t.channel, c.contact_id
          FROM team_act t, jsonb_array_elements(coalesce(t.meeting_attendees, '[]'::jsonb)) att
          JOIN public.contacts c ON lower(c.email) = lower(att->>'email')
          WHERE t.channel = 'meeting'
        )
        SELECT DISTINCT tc.id, tc.activity_date, tc.subject, tc.channel,
               c.contact_id, c.full_name, trim(c.current_company) AS account
        FROM touch_contact tc
        JOIN public.contacts c ON c.contact_id = tc.contact_id
        WHERE coalesce(trim(c.current_company),'') <> ''
          AND ($2 = 'all' OR tc.channel = $2)
        ORDER BY account, tc.activity_date DESC
    """, datetime(pstart.year, pstart.month, pstart.day, tzinfo=timezone.utc), channel)
    # group by account for the drawer
    accounts: dict = {}
    for r in rows:
        g = accounts.setdefault(r["account"], {"account": r["account"], "touches": []})
        g["touches"].append({
            "activity_id": str(r["id"]), "contact_id": r["contact_id"],
            "contact": r["full_name"], "subject": r["subject"], "channel": r["channel"],
            "date": r["activity_date"].isoformat() if r["activity_date"] else None,
        })
    items = sorted(accounts.values(), key=lambda a: -len(a["touches"]))
    return {"success": True, "data": {"period": period, "accounts": items,
            "total_touches": sum(len(a["touches"]) for a in items), "total_accounts": len(items)}}


# ── Outreach Dashboard scorecard ──────────────────────────────────────────────
# Two tables, both split Warm/Cold, this-period vs last-period, with a target:
#   User/Contact Pipeline — contacts ENTERING each funnel stage in the period
#       (flow, keyed by the stage-entry timestamps assigned_at / first_outreach_at
#        / active_at / handed_off_at).
#   Activity Pipeline     — raw activity-row counts (email/linkedin/intro/response).
# Warm/Cold is decided ONCE per contact: warm iff their company already had a
# Bedrock presence (an opportunity, or another jobs-pipeline contact) that
# predates this contact's own first touch; else cold. Every stage-entry and
# activity row for that contact inherits that label (never recomputed per row).

_OUTREACH_STAGE_META = [
    ("assigned",          "Assigned"),
    ("initial_outreach", "Initial Outreach"),
    ("converted_to_opportunity", "Converted to Opportunity"),
]
_OUTREACH_ACTIVITY_META = [
    ("direct_email_sent",      "Direct Email Sent"),
    ("linkedin_message_sent",  "LinkedIn Messages Sent"),
    ("facilitated_intro_sent", "Facilitated Intro"),
    ("engagement",             "Engagements"),
    ("direct_email_response",  "Direct Email Responses"),
]
# Funnel tier per activity metric — the frontend uses this to visually group the
# rows: sent touches (1) → engagements (2) → email replies (3).
_ACTIVITY_TIER = {
    "direct_email_sent": 1, "linkedin_message_sent": 1, "facilitated_intro_sent": 1,
    "engagement": 2, "direct_email_response": 3,
}
_STAGE_ENTERED_COL = {
    "assigned": "assigned_at", "initial_outreach": "first_outreach_at",
    "call_booked": "call_booked_at",
    "converted_to_opportunity": "converted_at",
}


def _shift_months(dt, n):
    m = dt.month - 1 + n
    return dt.replace(year=dt.year + m // 12, month=m % 12 + 1)


def _outreach_windows(granularity, date_from, date_to):
    """(this_start, this_end, last_start, last_end). Reviews happen after a period
    closes (e.g. Monday standup on last week), so `this` is the most recently
    COMPLETED period and `last` the one before it — never an in-progress period.
    Weeks run Sunday–Saturday. A custom date_from/date_to range overrides and
    compares against the immediately preceding equal-length range."""
    now = datetime.now(timezone.utc)
    if date_from and date_to:
        ds = datetime.fromisoformat(date_from).replace(tzinfo=timezone.utc)
        de = datetime.fromisoformat(date_to).replace(tzinfo=timezone.utc) + timedelta(days=1)
        length = de - ds
        return ds, de, ds - length, ds
    midnight = now.replace(hour=0, minute=0, second=0, microsecond=0)
    if granularity == "day":
        this_start = midnight - timedelta(days=1)          # yesterday
        last_start = this_start - timedelta(days=1)
    elif granularity == "week":
        days_since_sun = (midnight.weekday() + 1) % 7       # Mon=0..Sun=6 → days since Sunday
        cur_week_sun = midnight - timedelta(days=days_since_sun)  # Sunday of the in-progress week
        this_start = cur_week_sun - timedelta(days=7)       # last completed Sun–Sat week
        last_start = this_start - timedelta(days=7)
    else:  # month — last completed calendar month
        this_month_first = midnight.replace(day=1)
        this_start = _shift_months(this_month_first, -1)
        last_start = _shift_months(this_start, -1)
    this_end = _shift_months(this_start, 1) if granularity == "month" else this_start + timedelta(days=1 if granularity == "day" else 7)
    return this_start, this_end, last_start, this_start


def _scope_author(alias, scope):
    """Activity authored by the chosen staff scope: core team (Avni/Damon/Devika),
    other active staff, or anyone at Pursuit."""
    if scope == "team":
        return _team_actor(alias)
    if scope == "staff":
        return _staff_actor(alias)
    return f"({_team_actor(alias)} OR {_staff_actor(alias)})"


def _activity_actor(alias, scope, owner):
    """Author filter for the scorecard: a specific sender (owner) overrides the
    scope toggle; otherwise fall back to the scope group. `owner` is regex-checked
    so it's safe to interpolate."""
    if owner and _SAFE_EMAIL.match(owner):
        return f"({alias}.email_from ILIKE '%{owner}%' OR {alias}.logged_by ILIKE '%{owner}%')"
    return _scope_author(alias, scope)


def _message_actor(scope, owner, aem: str = "aem") -> str:
    """Author filter applied to a PARSED per-message sender
    (bedrock.activity_email_message.from_email — bare, lowercased). Mirrors
    _activity_actor's owner/scope semantics at message granularity, so replies
    and follow-ups count on the day they were sent and for the person who sent
    them (thread rows carry only the first message's author/date)."""
    if owner and _SAFE_EMAIL.match(owner):
        return f"{aem}.from_email ILIKE '%{owner}%'"
    team = " OR ".join(f"{aem}.from_email ILIKE '%{e}%'" for e in JOBS_TEAM_EMAILS)
    excl = ",".join(f"'{e.lower()}'" for e in JOBS_TEAM_EMAILS)
    staff = (f"EXISTS (SELECT 1 FROM public.org_users o WHERE o.is_active "
             f"AND o.email IS NOT NULL AND lower(o.email) NOT IN ({excl}) "
             f"AND {aem}.from_email ILIKE '%'||o.email||'%')")
    if scope == "team":
        return f"({team})"
    if scope == "staff":
        return f"({staff})"
    return f"(({team}) OR {staff})"


def _scope_intro_pred(scope, owner):
    """requested_by_email filter for facilitated intros — owner overrides scope."""
    if owner and _SAFE_EMAIL.match(owner):
        return f"lower(ir.requested_by_email) = lower('{owner}')"
    return _scope_email_pred("ir.requested_by_email", scope)


def _email_addr(alias="a"):
    """Bare email address from `{alias}.email_from`, which is often stored as
    'Display Name <addr>' — extract what's inside the angle brackets, else use the
    whole value. Needed to match inbound senders against outbound recipients
    (email_to stores bare addresses)."""
    return f"lower(trim(coalesce(substring({alias}.email_from from '<([^>]+)>'), {alias}.email_from)))"


def _scope_email_pred(col, scope):
    """Predicate: this email column belongs to the chosen staff scope."""
    core = ",".join(f"'{e.lower()}'" for e in JOBS_TEAM_EMAILS)
    if scope == "team":
        return f"lower({col}) IN ({core})"
    staff = (f"EXISTS (SELECT 1 FROM public.org_users o WHERE o.is_active "
             f"AND lower(o.email) = lower({col}) AND lower(o.email) NOT IN ({core}))")
    if scope == "staff":
        return staff
    return f"(lower({col}) IN ({core}) OR {staff})"


# Per-contact Warm/Cold classification, shared by the scorecard + by-sender
# queries (spliced in after WITH). Warm iff the contact's company had a Bedrock
# presence — an opportunity, or another jobs-pipeline contact (membership) —
# predating this contact's first touch; else cold. Decided once per contact.
_OUTREACH_WARMTH_CTES = """
    company_presence AS (
        SELECT lower(trim(account_name)) AS company, min(created_at) AS first_seen
        FROM bedrock.jobs_opportunity
        WHERE deleted_at IS NULL AND coalesce(trim(account_name),'') <> ''
        GROUP BY 1
        UNION ALL
        SELECT lower(trim(c.current_company)) AS company, min(m.assigned_at) AS first_seen
        FROM bedrock.jobs_contact_membership m
        JOIN public.contacts c ON c.contact_id = m.contact_id
        WHERE coalesce(trim(c.current_company),'') <> '' AND m.assigned_at IS NOT NULL
        GROUP BY 1
    ),
    company_first_seen AS (
        SELECT company, min(first_seen) AS first_seen FROM company_presence GROUP BY company
    ),
    contact_activity_first AS (
        SELECT participant_public_contact_id AS contact_id, min(activity_date) AS first_activity
        FROM bedrock.activity
        WHERE deleted_at IS NULL AND participant_public_contact_id IS NOT NULL
        GROUP BY 1
    ),
    contact_universe AS (
        SELECT contact_id FROM bedrock.jobs_contact_membership
        UNION
        SELECT DISTINCT participant_public_contact_id FROM bedrock.activity
          WHERE participant_public_contact_id IS NOT NULL AND deleted_at IS NULL
        UNION
        SELECT contact_id FROM bedrock.intro_request WHERE contact_id IS NOT NULL
    ),
    contact_first_touch AS (
        SELECT u.contact_id,
               lower(trim(c.current_company)) AS company,
               LEAST(
                   coalesce(caf.first_activity,   'infinity'::timestamptz),
                   coalesce(m.first_outreach_at,  'infinity'::timestamptz),
                   coalesce(m.assigned_at,         'infinity'::timestamptz)
               ) AS first_touch
        FROM contact_universe u
        JOIN public.contacts c ON c.contact_id = u.contact_id
        LEFT JOIN contact_activity_first caf ON caf.contact_id = u.contact_id
        LEFT JOIN bedrock.jobs_contact_membership m ON m.contact_id = u.contact_id
    ),
    contact_warmth AS (
        SELECT ct.contact_id,
               CASE WHEN cfs.first_seen IS NOT NULL AND cfs.first_seen < ct.first_touch
                    THEN 'warm' ELSE 'cold' END AS warmth
        FROM contact_first_touch ct
        LEFT JOIN company_first_seen cfs ON cfs.company = ct.company
    )"""


@router.get("/outreach/scorecard")
async def outreach_scorecard(
    granularity: str = Query("week", pattern="^(day|week|month)$"),
    scope: str = Query("team", pattern="^(pursuit|team|staff)$"),
    owner: Optional[str] = Query(None, description="Scope to one staff sender (overrides scope)"),
    date_from: Optional[str] = Query(None, pattern=r"^\d{4}-\d{2}-\d{2}$"),
    date_to: Optional[str] = Query(None, pattern=r"^\d{4}-\d{2}-\d{2}$"),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Outreach Dashboard P0 scorecard — User Pipeline + Activity Pipeline, each
    split Warm/Cold, this period vs last, plus a per-sender breakdown.

    Periods are the most recently COMPLETED period (this week = last full Sun–Sat
    week, this month = last full month, daily = yesterday) vs the one before — so
    a Monday review looks at the week that just closed. A custom date_from/date_to
    range overrides. `scope` (pursuit|team|staff) filters the ACTIVITY side by
    author; `owner` (a specific staff email) overrides scope. The User Pipeline
    stays team-wide because per-contact staff attribution isn't populated yet.

    NOTE: active_at / handed_off_at exist only once the 2026-07-15 migration is
    applied and fill in going forward, so Qualified Lead / Committed read 0 until
    then; first_outreach_at is likewise only stamped on new transitions.
    """
    this_start, this_end, last_start, last_end = _outreach_windows(granularity, date_from, date_to)

    # active_at / handed_off_at are added by the 2026-07-15 migration, applied
    # separately (admin role). Until then they don't exist — detect and omit
    # those stage-entry branches so the query still runs; the Qualified Lead /
    # Committed rows just come back empty (rendered as 0) rather than 500ing.
    have_cols = {r["column_name"] for r in await conn.fetch(
        """SELECT column_name FROM information_schema.columns
           WHERE table_schema='bedrock' AND table_name='jobs_contact_membership'
             AND column_name = ANY($1::text[])""",
        ["qualified_at", "converted_at"],
    )}
    # 'initial_outreach' (Outreached) is NOT taken from the membership stage
    # timestamp (unpopulated) — it's derived from actual outreach email activity
    # below. flagged / active / handed_off stay membership-driven.
    stage_event_parts = [
        "SELECT contact_id, 'assigned' AS stage, assigned_at AS entered_at "
        "FROM bedrock.jobs_contact_membership WHERE assigned_at IS NOT NULL",
    ]
    if "converted_at" in have_cols:
        stage_event_parts.append(
            "SELECT contact_id, 'converted_to_opportunity', converted_at "
            "FROM bedrock.jobs_contact_membership WHERE converted_at IS NOT NULL")
    stage_events_sql = "\n        UNION ALL\n        ".join(stage_event_parts)

    # One query, warmth computed once, two labelled result sets unioned.
    sql = f"""
    WITH {_OUTREACH_WARMTH_CTES},
    stage_events AS (
        {stage_events_sql}
    ),
    -- Message-level sends: every email the scope actually authored, dated by the
    -- message itself (thread rows are dated/attributed to the FIRST message, which
    -- made replies and follow-ups invisible to weekly counts).
    sent_msgs AS (
        SELECT aem.sent_at AS ts, a.participant_public_contact_id AS contact_id
        FROM bedrock.activity a
        JOIN bedrock.activity_email_message aem ON aem.activity_id = a.id
        WHERE a.deleted_at IS NULL AND a.type = 'email' AND {_message_actor(scope, owner)}
          AND {_not_autoreply('a')} AND {_jobs_relevant('a')}
    ),
    activity_events AS (
        SELECT 'direct_email_sent' AS metric, sm.ts, sm.contact_id
        FROM sent_msgs sm
        UNION ALL
        SELECT 'linkedin_message_sent', a.activity_date, a.participant_public_contact_id
        FROM bedrock.activity a
        WHERE a.deleted_at IS NULL AND a.type = 'linkedin' AND {_activity_actor('a', scope, owner)} AND {_jobs_relevant('a')}
        UNION ALL
        -- Facilitated intro that was acted on.
        SELECT 'facilitated_intro_sent', coalesce(ir.responded_at, ir.created_at), ir.contact_id
        FROM bedrock.intro_request ir
        WHERE ir.status IN ('accepted','completed') AND {_scope_intro_pred(scope, owner)}
    ),
    -- Outreached (activity-driven): distinct jobs contacts who RECEIVED an outreach
    -- email from the selected scope in the period (not the empty membership stamp).
    outreach_emails AS (
        SELECT sm.contact_id, sm.ts FROM sent_msgs sm WHERE sm.contact_id IS NOT NULL
    ),
    -- Direct Email Responses: an external address that got a jobs outreach email
    -- from the scope, then sent its FIRST inbound email back afterwards. Counted in
    -- the period of that first reply.
    sent_out AS (
        SELECT lower(e) AS addr, min(aem.sent_at) AS first_out
        FROM bedrock.activity a
        JOIN bedrock.activity_email_message aem ON aem.activity_id = a.id
        CROSS JOIN unnest(coalesce(a.email_to,'{{}}') || coalesce(a.email_cc,'{{}}')) e
        WHERE a.deleted_at IS NULL AND a.type = 'email' AND {_message_actor(scope, owner)}
          AND {_jobs_relevant('a')} AND lower(e) NOT LIKE '%@pursuit.org%'
        GROUP BY 1
    ),
    first_reply AS (
        -- Reply need not be classified (the OUTREACH was jobs-relevant, via sent_out);
        -- just an inbound MESSAGE from that address after we emailed them — including
        -- replies inside threads WE started, which the thread row never surfaces.
        SELECT aem.from_email AS addr, min(aem.sent_at) AS reply_date,
               max(a.participant_public_contact_id) AS contact_id
        FROM bedrock.activity a
        JOIN bedrock.activity_email_message aem ON aem.activity_id = a.id
        JOIN sent_out s ON s.addr = aem.from_email AND aem.sent_at >= s.first_out
        WHERE a.deleted_at IS NULL AND a.type = 'email' AND {_not_autoreply('a')}
          AND aem.from_email NOT LIKE '%@pursuit.org%'
        GROUP BY 1
    ),
    -- Engagements = the counterpart engaging back: a meeting or call, OR a direct
    -- email response. Built so Direct Email Responses always nest inside it.
    engagement_events AS (
        SELECT a.activity_date AS ts, a.participant_public_contact_id AS contact_id
        FROM bedrock.activity a
        WHERE a.deleted_at IS NULL AND a.type IN ('meeting','call')
          AND {_jobs_relevant('a')} AND {_not_autoreply('a')}
        UNION ALL
        SELECT fr.reply_date AS ts, fr.contact_id FROM first_reply fr
    ),
    stage_counts AS (
        SELECT 'user' AS kind, se.stage AS key, coalesce(cw.warmth,'cold') AS warmth,
               count(*) FILTER (WHERE se.entered_at >= $1 AND se.entered_at < $2) AS this_period,
               count(*) FILTER (WHERE se.entered_at >= $3 AND se.entered_at < $4) AS last_period
        FROM stage_events se
        LEFT JOIN contact_warmth cw ON cw.contact_id = se.contact_id
        GROUP BY se.stage, coalesce(cw.warmth,'cold')
    ),
    outreached_counts AS (
        SELECT 'user' AS kind, 'initial_outreach' AS key, coalesce(cw.warmth,'cold') AS warmth,
               count(DISTINCT oe.contact_id) FILTER (WHERE oe.ts >= $1 AND oe.ts < $2) AS this_period,
               count(DISTINCT oe.contact_id) FILTER (WHERE oe.ts >= $3 AND oe.ts < $4) AS last_period
        FROM outreach_emails oe
        LEFT JOIN contact_warmth cw ON cw.contact_id = oe.contact_id
        GROUP BY coalesce(cw.warmth,'cold')
    ),
    activity_counts AS (
        SELECT 'activity' AS kind, ae.metric AS key, coalesce(cw.warmth,'cold') AS warmth,
               count(DISTINCT ae.contact_id) FILTER (WHERE ae.ts >= $1 AND ae.ts < $2) AS this_period,
               count(DISTINCT ae.contact_id) FILTER (WHERE ae.ts >= $3 AND ae.ts < $4) AS last_period
        FROM activity_events ae
        LEFT JOIN contact_warmth cw ON cw.contact_id = ae.contact_id
        GROUP BY ae.metric, coalesce(cw.warmth,'cold')
    ),
    engagement_counts AS (
        SELECT 'activity' AS kind, 'engagement' AS key, coalesce(cw.warmth,'cold') AS warmth,
               count(DISTINCT ee.contact_id) FILTER (WHERE ee.ts >= $1 AND ee.ts < $2) AS this_period,
               count(DISTINCT ee.contact_id) FILTER (WHERE ee.ts >= $3 AND ee.ts < $4) AS last_period
        FROM engagement_events ee
        LEFT JOIN contact_warmth cw ON cw.contact_id = ee.contact_id
        GROUP BY coalesce(cw.warmth,'cold')
    ),
    response_counts AS (
        SELECT 'activity' AS kind, 'direct_email_response' AS key, coalesce(cw.warmth,'cold') AS warmth,
               count(*) FILTER (WHERE fr.reply_date >= $1 AND fr.reply_date < $2) AS this_period,
               count(*) FILTER (WHERE fr.reply_date >= $3 AND fr.reply_date < $4) AS last_period
        FROM first_reply fr
        LEFT JOIN contact_warmth cw ON cw.contact_id = fr.contact_id
        GROUP BY coalesce(cw.warmth,'cold')
    )
    SELECT * FROM stage_counts
    UNION ALL SELECT * FROM outreached_counts
    UNION ALL SELECT * FROM activity_counts
    UNION ALL SELECT * FROM engagement_counts
    UNION ALL SELECT * FROM response_counts
    """
    rows = await conn.fetch(sql, this_start, this_end, last_start, last_end)

    # ── By Sender: per-staff sent volume (email + linkedin) with warm/cold split,
    # this period vs last. Real, author-attributed. Qualified/Committed/response-
    # rate per sender await first_outreach_by attribution (not populated yet).
    core = ",".join(f"'{e.lower()}'" for e in JOBS_TEAM_EMAILS)
    if owner and _SAFE_EMAIL.match(owner):
        sender_scope = f"AND lower(o.email) = lower('{owner}')"
    else:
        sender_scope = {"team": f"AND lower(o.email) IN ({core})",
                        "staff": f"AND lower(o.email) NOT IN ({core})",
                        "pursuit": ""}[scope]
    sender_sql = f"""
    WITH {_OUTREACH_WARMTH_CTES}
    SELECT o.email AS sender,
           count(*) FILTER (WHERE a.activity_date >= $1 AND a.activity_date < $2) AS sent_this,
           count(*) FILTER (WHERE a.activity_date >= $3 AND a.activity_date < $4) AS sent_last,
           count(*) FILTER (WHERE a.activity_date >= $1 AND a.activity_date < $2 AND coalesce(cw.warmth,'cold')='warm') AS warm_this,
           count(*) FILTER (WHERE a.activity_date >= $1 AND a.activity_date < $2 AND coalesce(cw.warmth,'cold')='cold') AS cold_this
    FROM bedrock.activity a
    JOIN public.org_users o ON o.is_active
      AND (a.email_from ILIKE '%'||o.email||'%' OR a.logged_by ILIKE '%'||o.email||'%')
    LEFT JOIN contact_warmth cw ON cw.contact_id = a.participant_public_contact_id
    WHERE a.deleted_at IS NULL AND a.type IN ('email','linkedin') AND {_jobs_relevant('a')}
      AND a.activity_date >= $3 {sender_scope}
    GROUP BY o.email
    HAVING count(*) FILTER (WHERE a.activity_date >= $3 AND a.activity_date < $2) > 0
    ORDER BY sent_this DESC, sent_last DESC
    """
    sender_rows = await conn.fetch(sender_sql, this_start, this_end, last_start, last_end)
    by_sender = [{
        "staff": r["sender"],
        "sent": {"this": r["sent_this"], "last": r["sent_last"]},
        "warm": r["warm_this"], "cold": r["cold_this"],
    } for r in sender_rows]

    # (kind, key) → {"this": {warm,cold}, "last": {warm,cold}}
    agg: dict = {}
    for r in rows:
        bucket = agg.setdefault((r["kind"], r["key"]),
                                {"this": {"warm": 0, "cold": 0}, "last": {"warm": 0, "cold": 0}})
        bucket["this"][r["warmth"]] = r["this_period"]
        bucket["last"][r["warmth"]] = r["last_period"]

    def _row(kind, key, label, target):
        b = agg.get((kind, key), {"this": {"warm": 0, "cold": 0}, "last": {"warm": 0, "cold": 0}})
        tw, tc = b["this"]["warm"], b["this"]["cold"]
        lw, lc = b["last"]["warm"], b["last"]["cold"]
        out = {
            "label": label,
            "this_period": {"warm": tw, "cold": tc, "total": tw + tc},
            "last_period": {"warm": lw, "cold": lc, "total": lw + lc},
            "target": target,
        }
        return out

    user_pipeline = [
        {"stage": s, **_row("user", s, label, user_pipeline_target(s, granularity))}
        for s, label in _OUTREACH_STAGE_META
    ]
    activity_pipeline = [
        {"metric": m, "tier": _ACTIVITY_TIER.get(m), **_row("activity", m, label, activity_pipeline_target(m, granularity))}
        for m, label in _OUTREACH_ACTIVITY_META
    ]
    return {"success": True, "data": {
        "granularity": granularity,
        "scope": scope,
        "period": {
            "this_start": this_start.isoformat(), "this_end": this_end.isoformat(),
            "last_start": last_start.isoformat(), "last_end": last_end.isoformat(),
        },
        "user_pipeline": user_pipeline,
        "activity_pipeline": activity_pipeline,
        "by_sender": by_sender,
    }}


# Zero leads the list and is coloured as a problem: it means "sitting in initial
# outreach and not touched at all in the last four weeks", which is exactly the
# follow-up gap this panel exists to surface.
_TOUCH_BUCKETS = [
    ("0", "0 Touches", 0, 0),
    ("1", "1 Touch", 1, 1),
    ("2", "2 Touches", 2, 2),
    ("3", "3 Touches", 3, 3),
    ("4plus", "4+ Touches", 4, None),
]
_TOUCH_DEPTH_CAP = 40   # contacts listed per bucket for the inline drill


# The touch-counting window, anchored to TODAY (Jac 2026-08-12).
#
# It was previously anchored to the end of the selected period, which made the
# panel answer a reporting question ("how well did we work the cohort we opened
# in July") while sitting on an operational tab. Two things went wrong in
# practice: a contact touched five times last week read as zero when you looked
# at a July period, and on a quiet week the panel was simply empty even though
# 800+ contacts were sitting in initial outreach needing follow-up.
#
# Touch Depth is a WORK QUEUE, not a period metric. Cohort and window are both
# "right now", which is why this no longer takes a period at all. The bounded
# window still does the job it was added for — a lifetime count would make an old
# contact look well-worked purely for being old. The UI reads the number from the
# response, so changing it here changes the label too.
_TOUCH_DEPTH_WEEKS = 4


async def _touch_depth(conn, scope: str, owner: Optional[str]) -> dict:
    """Touch depth: for every contact sitting in initial outreach RIGHT NOW, how
    many jobs touches they received in the trailing `_TOUCH_DEPTH_WEEKS`.

    Deliberately takes no period — see the note on _TOUCH_DEPTH_WEEKS. The cohort
    is the live queue, so the panel is never empty just because it was a quiet
    week, and the window ends today, so a contact worked last week never reads as
    untouched.

    Uses the same _jobs_relevant / _not_autoreply filters as the funnel and the
    drill-downs. The count and `last_touch_at` share those filters, so the two
    can't contradict each other — the previous version left _not_autoreply off
    `last_touch_at`, which could date a contact by an out-of-office reply."""
    owner_pred = "TRUE"
    if owner and _SAFE_EMAIL.match(owner):
        owner_pred = f"lower(m.owner_email) = lower('{owner}')"
    elif scope in ("team", "staff"):
        owner_pred = _scope_email_pred("m.owner_email", scope)

    # One aggregate pass over the window, not two correlated subqueries per
    # contact. At 800+ contacts in the cohort the old shape issued ~1600 index
    # scans and took long enough that the panel looked broken.
    sql = f"""
        WITH cohort AS (
            SELECT m.contact_id, m.owner_email
            FROM bedrock.jobs_contact_membership m
            WHERE m.stage = 'initial_outreach' AND {owner_pred}
        ),
        touched AS (
            SELECT a.participant_public_contact_id AS contact_id,
                   count(*)              AS touches,
                   max(a.activity_date)  AS last_touch_at
            FROM bedrock.activity a
            JOIN cohort c ON c.contact_id = a.participant_public_contact_id
            WHERE a.deleted_at IS NULL
              AND a.activity_date >= $1
              AND {_jobs_relevant('a')} AND {_not_autoreply('a')}
            GROUP BY 1
        )
        SELECT c.contact_id, c.owner_email,
               COALESCE(t.touches, 0) AS touches,
               t.last_touch_at,
               p.full_name, p.current_company
        FROM cohort c
        LEFT JOIN touched t ON t.contact_id = c.contact_id
        LEFT JOIN public.contacts p ON p.contact_id = c.contact_id
        ORDER BY COALESCE(t.touches, 0), p.current_company NULLS LAST
    """
    touch_from = datetime.now(timezone.utc) - timedelta(weeks=_TOUCH_DEPTH_WEEKS)
    rows = await conn.fetch(sql, touch_from)

    # Everyone in the cohort is counted: with a bounded window, zero is a real
    # answer rather than a data gap, so excluding it would understate the problem.
    measured = rows
    total = len(measured)
    buckets = []
    for key, label, lo, hi in _TOUCH_BUCKETS:
        members = [r for r in measured if r["touches"] >= lo and (hi is None or r["touches"] <= hi)]
        buckets.append({
            "key": key, "label": label, "count": len(members),
            "pct": round(100 * len(members) / total) if total else 0,
            "contacts": [{
                "contact_id": r["contact_id"], "name": r["full_name"],
                "company": r["current_company"], "owner": r["owner_email"],
                "touches": r["touches"],
                "last_touch_at": r["last_touch_at"].isoformat() if r["last_touch_at"] else None,
            } for r in members[:_TOUCH_DEPTH_CAP]],
            "truncated": max(0, len(members) - _TOUCH_DEPTH_CAP),
        })
    # `undated` is gone: it counted contacts whose stage entry had no timestamp,
    # which mattered only while the cohort was date-filtered. Nobody is excluded
    # for a missing stamp now, so reporting it would imply a gap that isn't there.
    return {"total": total,
            "weeks": _TOUCH_DEPTH_WEEKS,
            "touch_from": touch_from.date().isoformat(),
            "touch_to": datetime.now(timezone.utc).date().isoformat(),
            "buckets": buckets}


@router.get("/outreach/touch-depth")
async def outreach_touch_depth(
    scope: str = Query("team", pattern="^(pursuit|team|staff)$"),
    owner: Optional[str] = Query(None, description="Scope to one staff owner (overrides scope)"),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Who in initial outreach is under-touched right now.

    Split off the scorecard deliberately. The scorecard is period-scoped, so the
    Outreach tab fetches it once per period to build its trend columns — which
    meant this recomputed several times on every page load while answering a
    question that doesn't depend on the period at all. On its own endpoint it is
    fetched once and cached independently, and the scorecard gets lighter."""
    return {"success": True, "data": await _touch_depth(conn, scope, owner)}


def _touch_direction(type_: str, email_from: Optional[str]) -> str:
    """Label a touch as sent / received / meeting for the drill-down."""
    if type_ == "email":
        return "sent" if "@pursuit.org" in (email_from or "").lower() else "received"
    if type_ == "meeting":
        return "meeting"
    return "sent"


def _touch_actor(alias: str = "a") -> str:
    """The Pursuit person behind a touch, for the drill's Owner column.

    Which side of the row that is depends on direction: when we sent it the
    actor is the sender, when the contact replied it's the Pursuit address the
    reply landed on, and for a manually logged call/LinkedIn note it's whoever
    logged it. Returned as a bare lowercase email so the frontend resolves it to
    a staff name through the same map every other owner label uses.

    logged_by is only trusted when it looks like an email. Checked against
    production 2026-08-04: 2,958 of 18,627 activity rows since January carry a
    raw Salesforce user id there instead ('0051U0000017rDxQAI'), and only 1 of
    the 16 distinct ids resolves through bedrock.app_user. Printing the rest
    would put a Salesforce id in a column headed "Owner", so they come back NULL
    and the drill shows an explicit dash."""
    frm = _email_addr(alias)
    return f"""coalesce(
        CASE WHEN {frm} LIKE '%@pursuit.org' THEN {frm} END,
        (SELECT lower(e)
           FROM unnest(coalesce({alias}.email_to,'{{}}') || coalesce({alias}.email_cc,'{{}}')) e
          WHERE lower(e) LIKE '%@pursuit.org' LIMIT 1),
        CASE WHEN {alias}.logged_by LIKE '%@%'
             THEN lower(nullif(trim({alias}.logged_by),'')) END)"""


@router.get("/outreach/scorecard/detail")
async def outreach_scorecard_detail(
    kind: str = Query(..., pattern="^(user|activity)$"),
    key: str = Query(..., description="stage (user) or metric (activity)"),
    period: str = Query("this", pattern="^(this|last)$"),
    granularity: str = Query("week", pattern="^(day|week|month)$"),
    scope: str = Query("team", pattern="^(pursuit|team|staff)$"),
    owner: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None, pattern=r"^\d{4}-\d{2}-\d{2}$"),
    date_to: Optional[str] = Query(None, pattern=r"^\d{4}-\d{2}-\d{2}$"),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Drill-down behind one scorecard row: the contacts in that stage / behind
    that activity count for the period, each with the actual touches that went
    out (or came in). Powers the click-to-expand on the Outreach tab."""
    this_start, this_end, last_start, last_end = _outreach_windows(granularity, date_from, date_to)
    start, end = (this_start, this_end) if period == "this" else (last_start, last_end)

    contacts: dict = {}   # contact_id → {contact_id, name, company, entered_at, touches:[]}

    def _contact(cid, name, company):
        return contacts.setdefault(cid, {
            "contact_id": cid, "name": name, "company": company,
            "entered_at": None, "touches": []})

    if kind == "activity":
        if key == "facilitated_intro_sent":
            rows = await conn.fetch(f"""
                SELECT ir.contact_id, c.full_name, c.current_company,
                       coalesce(ir.responded_at, ir.created_at) AS activity_date,
                       ir.specific_ask AS subject, ir.context AS snippet,
                       lower(ir.requested_by_email) AS actor
                FROM bedrock.intro_request ir
                JOIN public.contacts c ON c.contact_id = ir.contact_id
                WHERE ir.status IN ('accepted','completed')
                  AND {_scope_intro_pred(scope, owner)}
                  AND coalesce(ir.responded_at, ir.created_at) >= $1
                  AND coalesce(ir.responded_at, ir.created_at) < $2
                ORDER BY activity_date DESC LIMIT 500
            """, start, end)
            for r in rows:
                g = _contact(r["contact_id"], r["full_name"], r["current_company"])
                g["touches"].append({
                    "date": r["activity_date"].isoformat() if r["activity_date"] else None,
                    "type": "intro", "subject": r["subject"], "snippet": r["snippet"],
                    "direction": "sent", "actor": r["actor"]})
        elif key == "direct_email_response":
            rows = await conn.fetch(f"""
                WITH sent_out AS (
                    -- `sender` = who sent the outreach this is a reply TO, so the
                    -- drill can credit the reply to the person who earned it.
                    SELECT lower(e) AS addr, min(aem.sent_at) AS first_out,
                           (array_agg(aem.from_email ORDER BY aem.sent_at))[1] AS sender
                    FROM bedrock.activity a
                    JOIN bedrock.activity_email_message aem ON aem.activity_id = a.id
                    CROSS JOIN unnest(coalesce(a.email_to,'{{}}') || coalesce(a.email_cc,'{{}}')) e
                    WHERE a.deleted_at IS NULL AND a.type = 'email' AND {_message_actor(scope, owner)}
                      AND {_jobs_relevant('a')} AND lower(e) NOT LIKE '%@pursuit.org%'
                    GROUP BY 1
                ),
                first_reply AS (
                    SELECT aem.from_email AS addr, min(aem.sent_at) AS reply_date,
                           max(a.participant_public_contact_id) AS contact_id
                    FROM bedrock.activity a
                    JOIN bedrock.activity_email_message aem ON aem.activity_id = a.id
                    JOIN sent_out s ON s.addr = aem.from_email AND aem.sent_at >= s.first_out
                    WHERE a.deleted_at IS NULL AND a.type = 'email' AND {_not_autoreply('a')}
                      AND aem.from_email NOT LIKE '%@pursuit.org%'
                    GROUP BY 1
                )
                SELECT fr.addr, fr.reply_date, fr.contact_id, c.full_name, c.current_company,
                       lower(so.sender) AS actor
                FROM first_reply fr
                JOIN sent_out so ON so.addr = fr.addr
                LEFT JOIN public.contacts c ON c.contact_id = fr.contact_id
                WHERE fr.reply_date >= $1 AND fr.reply_date < $2
                ORDER BY fr.reply_date DESC LIMIT 500
            """, start, end)
            for idx, r in enumerate(rows):
                cid = r["contact_id"] if r["contact_id"] is not None else -(idx + 1)
                g = _contact(cid, r["full_name"] or r["addr"], r["current_company"])
                g["touches"].append({
                    "date": r["reply_date"].isoformat() if r["reply_date"] else None,
                    "type": "email", "subject": "First reply to outreach", "snippet": r["addr"],
                    "direction": "received", "actor": r["actor"]})
            rows = []  # already consumed
        else:
            where = {
                # Window + actor applied per MESSAGE — a follow-up sent this week in an
                # old thread must appear in this week's drill.
                "direct_email_sent":
                    f"a.type = 'email' AND {_not_autoreply('a')} AND {_jobs_relevant('a')} AND EXISTS ("
                    f"  SELECT 1 FROM bedrock.activity_email_message aem WHERE aem.activity_id = a.id"
                    f"  AND {_message_actor(scope, owner)} AND aem.sent_at >= $1 AND aem.sent_at < $2)",
                "linkedin_message_sent":
                    f"a.type = 'linkedin' AND {_activity_actor('a', scope, owner)} AND {_jobs_relevant('a')}",
                "engagement":  # meetings/calls only here; email replies appended below
                    f"a.type IN ('meeting','call') AND {_jobs_relevant('a')} AND {_not_autoreply('a')}",
            }.get(key)
            if where is None:
                raise HTTPException(400, "invalid activity key")
            # direct_email_sent windows on the message timestamps inside `where`;
            # the other keys window on the activity row itself.
            date_pred = ("TRUE" if key == "direct_email_sent"
                         else "a.activity_date >= $1 AND a.activity_date < $2")
            rows = await conn.fetch(f"""
                SELECT a.participant_public_contact_id AS contact_id, c.full_name, c.current_company,
                       a.activity_date, a.type, a.subject, a.email_snippet, a.email_from,
                       -- Prefer the per-message sender: a thread row carries only the
                       -- FIRST message's author, so a follow-up sent this week by
                       -- someone else would otherwise be credited to the wrong person.
                       coalesce(lower(msg.from_email), {_touch_actor('a')}) AS actor
                FROM bedrock.activity a
                JOIN public.contacts c ON c.contact_id = a.participant_public_contact_id
                LEFT JOIN LATERAL (
                    SELECT aem.from_email
                    FROM bedrock.activity_email_message aem
                    WHERE aem.activity_id = a.id AND {_message_actor(scope, owner)}
                      AND aem.sent_at >= $1 AND aem.sent_at < $2
                    ORDER BY aem.sent_at LIMIT 1
                ) msg ON true
                WHERE a.deleted_at IS NULL AND {date_pred} AND {where}
                ORDER BY c.current_company NULLS LAST, a.activity_date DESC LIMIT 500
            """, start, end)
            for r in rows:
                g = _contact(r["contact_id"], r["full_name"], r["current_company"])
                g["touches"].append({
                    "date": r["activity_date"].isoformat() if r["activity_date"] else None,
                    "type": r["type"], "subject": r["subject"], "snippet": r["email_snippet"],
                    "direction": _touch_direction(r["type"], r["email_from"]),
                    "actor": r["actor"]})
            if key == "engagement":
                # Engagements also include direct email responses — append them.
                reps = await conn.fetch(f"""
                    WITH sent_out AS (
                        SELECT lower(e) AS addr, min(aem.sent_at) AS first_out
                        FROM bedrock.activity a
                        JOIN bedrock.activity_email_message aem ON aem.activity_id = a.id
                        CROSS JOIN unnest(coalesce(a.email_to,'{{}}') || coalesce(a.email_cc,'{{}}')) e
                        WHERE a.deleted_at IS NULL AND a.type='email' AND {_message_actor(scope, owner)}
                          AND {_jobs_relevant('a')} AND lower(e) NOT LIKE '%@pursuit.org%'
                        GROUP BY 1
                    ),
                    first_reply AS (
                        SELECT aem.from_email AS addr, min(aem.sent_at) AS reply_date, max(a.participant_public_contact_id) AS contact_id
                        FROM bedrock.activity a
                        JOIN bedrock.activity_email_message aem ON aem.activity_id = a.id
                        JOIN sent_out s ON s.addr = aem.from_email AND aem.sent_at >= s.first_out
                        WHERE a.deleted_at IS NULL AND a.type='email' AND {_not_autoreply('a')} AND aem.from_email NOT LIKE '%@pursuit.org%'
                        GROUP BY 1
                    )
                    SELECT fr.addr, fr.reply_date, fr.contact_id, c.full_name, c.current_company,
                           lower(so.sender) AS actor
                    FROM first_reply fr
                    JOIN sent_out so ON so.addr = fr.addr
                    LEFT JOIN public.contacts c ON c.contact_id = fr.contact_id
                    WHERE fr.reply_date >= $1 AND fr.reply_date < $2 ORDER BY fr.reply_date DESC LIMIT 500
                """, start, end)
                for idx, r in enumerate(reps):
                    cid = r["contact_id"] if r["contact_id"] is not None else -(idx + 1)
                    g = _contact(cid, r["full_name"] or r["addr"], r["current_company"])
                    g["touches"].append({
                        "date": r["reply_date"].isoformat() if r["reply_date"] else None,
                        "type": "email", "subject": "Email reply", "snippet": r["addr"],
                        "direction": "received", "actor": r["actor"]})
    else:  # kind == "user"
        if key == "initial_outreach":
            # Activity-driven: distinct contacts emailed by the scope this period.
            crows = await conn.fetch(f"""
                SELECT a.participant_public_contact_id AS contact_id, c.full_name, c.current_company,
                       min(a.activity_date) AS entered_at
                FROM bedrock.activity a
                JOIN public.contacts c ON c.contact_id = a.participant_public_contact_id
                WHERE a.deleted_at IS NULL AND a.type = 'email' AND {_activity_actor('a', scope, owner)}
                  AND {_not_autoreply('a')} AND {_jobs_relevant('a')}
                  AND a.activity_date >= $1 AND a.activity_date < $2
                GROUP BY a.participant_public_contact_id, c.full_name, c.current_company
                ORDER BY entered_at DESC LIMIT 300
            """, start, end)
        else:
            col = _STAGE_ENTERED_COL.get(key)
            if col is None:
                raise HTTPException(400, "invalid stage key")
            # active_at / handed_off_at may not exist pre-migration → empty drill.
            exists = await conn.fetchval(
                """SELECT true FROM information_schema.columns
                   WHERE table_schema='bedrock' AND table_name='jobs_contact_membership'
                     AND column_name=$1""", col)
            crows = await conn.fetch(f"""
                SELECT m.contact_id, c.full_name, c.current_company, m.{col} AS entered_at
                FROM bedrock.jobs_contact_membership m
                JOIN public.contacts c ON c.contact_id = m.contact_id
                WHERE m.{col} IS NOT NULL AND m.{col} >= $1 AND m.{col} < $2
                ORDER BY m.{col} DESC LIMIT 300
            """, start, end) if exists else []
        for r in crows:
            g = _contact(r["contact_id"], r["full_name"], r["current_company"])
            g["entered_at"] = r["entered_at"].isoformat() if r["entered_at"] else None
        if contacts:
            touch_rows = await conn.fetch(f"""
                SELECT a.participant_public_contact_id AS contact_id, a.activity_date,
                       a.type, a.subject, a.email_snippet, a.email_from,
                       {_touch_actor('a')} AS actor
                FROM bedrock.activity a
                WHERE a.deleted_at IS NULL AND a.participant_public_contact_id = ANY($1::int[])
                  AND a.activity_date >= $2 AND a.activity_date < $3 AND {_jobs_relevant('a')}
                ORDER BY a.activity_date DESC
            """, list(contacts.keys()), start, end)
            for r in touch_rows:
                g = contacts.get(r["contact_id"])
                if g is not None:
                    g["touches"].append({
                        "date": r["activity_date"].isoformat() if r["activity_date"] else None,
                        "type": r["type"], "subject": r["subject"], "snippet": r["email_snippet"],
                        "direction": _touch_direction(r["type"], r["email_from"]),
                        "actor": r["actor"]})

    items = sorted(contacts.values(), key=lambda g: (-(len(g["touches"])), g["company"] or ""))
    # Distinct actors per contact, so the collapsed row can name who worked it
    # without expanding to read the per-touch column.
    for g in items:
        g["actors"] = sorted({t["actor"] for t in g["touches"] if t.get("actor")})
    return {"success": True, "data": {"kind": kind, "key": key, "period": period,
            "count": len(items), "contacts": items}}


_TARGETING_DIMS = [
    # Tag replaced lead source (2026-07-30): source is an import artifact
    # ("linkedin_import"), while the curated campaign tags are how the team
    # actually chooses who to work.
    ("tag",           "By Tag"),
    ("industry",      "By Industry"),
    ("size_bucket",   "By Company Size"),
    ("company_stage", "By Company Stage"),
]


@router.get("/outreach/targeting-mix")
async def outreach_targeting_mix(
    granularity: str = Query("week", pattern="^(day|week|month)$"),
    scope: str = Query("team", pattern="^(pursuit|team|staff)$"),
    owner: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None, pattern=r"^\d{4}-\d{2}-\d{2}$"),
    date_to: Optional[str] = Query(None, pattern=r"^\d{4}-\d{2}-\d{2}$"),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Targeting Mix — outreach volume + email replies for the period, cut by the
    contact's lead source and their company's industry / size / stage. Answers
    'where is our outreach going and which segments reply'. Contact-linked subset
    only (needs participant_public_contact_id + company_id)."""
    this_start, this_end, _ls, _le = _outreach_windows(granularity, date_from, date_to)

    sql = f"""
    WITH outreach AS (
        SELECT a.participant_public_contact_id AS cid
        FROM bedrock.activity a
        WHERE a.deleted_at IS NULL AND a.type='email' AND {_activity_actor('a', scope, owner)}
          AND {_not_autoreply('a')} AND {_jobs_relevant('a')}
          AND a.participant_public_contact_id IS NOT NULL
          AND a.activity_date >= $1 AND a.activity_date < $2
    ),
    sent_out AS (
        SELECT lower(e) AS addr, min(a.activity_date) AS first_out
        FROM bedrock.activity a,
             unnest(coalesce(a.email_to,'{{}}') || coalesce(a.email_cc,'{{}}')) e
        WHERE a.deleted_at IS NULL AND a.type='email' AND {_activity_actor('a', scope, owner)}
          AND {_jobs_relevant('a')} AND lower(e) NOT LIKE '%@pursuit.org%'
        GROUP BY 1
    ),
    first_reply AS (
        SELECT {_email_addr('a')} AS addr, min(a.activity_date) AS reply_date,
               max(a.participant_public_contact_id) AS cid
        FROM bedrock.activity a
        JOIN sent_out s ON s.addr = {_email_addr('a')} AND a.activity_date >= s.first_out
        WHERE a.deleted_at IS NULL AND a.type='email' AND {_not_autoreply('a')}
          AND a.email_from NOT ILIKE '%@pursuit.org%'
        GROUP BY 1
    ),
    replies AS (SELECT cid FROM first_reply WHERE reply_date >= $1 AND reply_date < $2 AND cid IS NOT NULL),
    contact_dims AS (
        SELECT c.contact_id,
               coalesce(co.industry, '(unknown)')                    AS industry,
               coalesce(co.size_bucket, '(unknown)')                 AS size_bucket,
               coalesce(co.stage, '(unknown)')                       AS company_stage
        FROM public.contacts c
        LEFT JOIN public.companies co ON co.company_id = c.company_id
    ),
    -- One row per (contact, curated tag); a contact with 3 tags counts in all
    -- three buckets, and untagged contacts land in '(untagged)'.
    contact_tags AS (
        SELECT c.contact_id,
               coalesce((SELECT tc.label FROM bedrock.contact_tag_catalog tc WHERE tc.slug = t), t) AS tag
        FROM public.contacts c
        LEFT JOIN LATERAL unnest(
            CASE WHEN EXISTS (SELECT 1 FROM unnest(coalesce(c.tags,'{{}}'::text[])) x
                              WHERE x IN (SELECT slug FROM bedrock.contact_tag_catalog))
                 THEN ARRAY(SELECT x FROM unnest(c.tags) x
                            WHERE x IN (SELECT slug FROM bedrock.contact_tag_catalog))
                 ELSE ARRAY['(untagged)'] END) AS t ON true
    ),
    agg AS (
        SELECT 'tag' AS dim, ct.tag AS bucket, count(*) AS sent, 0 AS resp FROM outreach o JOIN contact_tags ct ON ct.contact_id=o.cid GROUP BY 2
        UNION ALL SELECT 'industry', cd.industry, count(*), 0 FROM outreach o JOIN contact_dims cd ON cd.contact_id=o.cid GROUP BY 2
        UNION ALL SELECT 'size_bucket', cd.size_bucket, count(*), 0 FROM outreach o JOIN contact_dims cd ON cd.contact_id=o.cid GROUP BY 2
        UNION ALL SELECT 'company_stage', cd.company_stage, count(*), 0 FROM outreach o JOIN contact_dims cd ON cd.contact_id=o.cid GROUP BY 2
        UNION ALL SELECT 'tag', ct.tag, 0, count(*) FROM replies r JOIN contact_tags ct ON ct.contact_id=r.cid GROUP BY 2
        UNION ALL SELECT 'industry', cd.industry, 0, count(*) FROM replies r JOIN contact_dims cd ON cd.contact_id=r.cid GROUP BY 2
        UNION ALL SELECT 'size_bucket', cd.size_bucket, 0, count(*) FROM replies r JOIN contact_dims cd ON cd.contact_id=r.cid GROUP BY 2
        UNION ALL SELECT 'company_stage', cd.company_stage, 0, count(*) FROM replies r JOIN contact_dims cd ON cd.contact_id=r.cid GROUP BY 2
    )
    SELECT dim, bucket, sum(sent)::int AS sent, sum(resp)::int AS responses
    FROM agg GROUP BY dim, bucket HAVING sum(sent) + sum(resp) > 0
    ORDER BY dim, sent DESC, responses DESC
    """
    rows = await conn.fetch(sql, this_start, this_end)
    by_dim: dict = {d: [] for d, _ in _TARGETING_DIMS}
    for r in rows:
        by_dim.setdefault(r["dim"], []).append(
            {"bucket": r["bucket"], "sent": r["sent"], "responses": r["responses"]})
    return {"success": True, "data": {
        "dims": [{"key": d, "label": lbl, "rows": by_dim.get(d, [])} for d, lbl in _TARGETING_DIMS],
    }}


@router.get("/outreach/responded-contacts")
async def outreach_responded_contacts(
    owner: Optional[str] = Query(None),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Contacts still sitting in initial_outreach who HAVE replied — the queue
    that needs a human decision. A positive reply belongs in
    converted_to_opportunity, a neutral/negative one in on_hold / not_a_fit, but
    nothing here moves automatically: the owner reads the reply and picks. Sorted
    by how long the reply has gone un-actioned."""
    team_aem = " OR ".join(f"aem.from_email ILIKE '%{e}%'" for e in JOBS_TEAM_EMAILS)
    params: list = []
    if owner:
        owner_where = "AND lower(coalesce(c.owner_email,'')) = lower($1)"
        params.append(owner)
    else:
        # Default to the jobs team's own contacts. Prospects owned by PBD folks
        # (Nick/Greg/David) surfaced replies that aren't this team's queue to
        # triage; the per-sender select can still target anyone explicitly.
        owner_list = ", ".join(f"'{e}'" for e in JOBS_TEAM_EMAILS)
        owner_where = f"AND lower(coalesce(c.owner_email,'')) IN ({owner_list})"
    rows = await conn.fetch(f"""
        WITH sends AS (
            SELECT a.participant_public_contact_id AS cid, aem.sent_at AS ts, a.id AS activity_id
            FROM bedrock.activity a
            JOIN bedrock.activity_email_message aem ON aem.activity_id = a.id
            WHERE a.deleted_at IS NULL AND a.type = 'email'
              AND a.participant_public_contact_id IS NOT NULL
              AND coalesce(a.jobs_relevance_override, a.jobs_relevance) = 'jobs'
              AND ({team_aem})
        ),
        touch_counts AS (
            SELECT cid, count(*) AS touches, min(ts) AS first_send FROM sends GROUP BY 1),
        replies AS (
            -- inbound MESSAGES on threads linked to the contact (their reply or a
            -- colleague's), and only replies that came AFTER this outreach cycle
            -- started — otherwise a 2020 thread reads as an un-actioned reply.
            SELECT a.participant_public_contact_id AS cid,
                   max(aem.sent_at) AS last_reply,
                   count(*) AS reply_count,
                   (array_agg(aem.from_email ORDER BY aem.sent_at DESC))[1] AS last_reply_from,
                   (array_agg(coalesce(a.email_snippet, a.subject) ORDER BY aem.sent_at DESC))[1] AS snippet
            FROM bedrock.activity a
            JOIN bedrock.activity_email_message aem ON aem.activity_id = a.id
            JOIN touch_counts tc ON tc.cid = a.participant_public_contact_id
            JOIN bedrock.jobs_contact_membership mm
              ON mm.contact_id = a.participant_public_contact_id
            WHERE a.deleted_at IS NULL AND a.participant_public_contact_id IS NOT NULL
              AND aem.from_email NOT ILIKE '%@pursuit.org%'
              AND aem.from_email NOT ILIKE '%@pursuit.com%'
              -- the reply has to be on a jobs-classified thread, not any old
              -- correspondence that happens to involve this person
              AND coalesce(a.jobs_relevance_override, a.jobs_relevance) = 'jobs'
              AND aem.sent_at >= coalesce(mm.first_outreach_at, tc.first_send)
            GROUP BY 1
        )
        SELECT c.contact_id, c.full_name, c.current_title, c.current_company, c.owner_email,
               coalesce(tc.touches, 0)::int AS touches,
               r.last_reply, r.reply_count::int AS reply_count, r.last_reply_from,
               left(coalesce(r.snippet, ''), 240) AS snippet
        FROM replies r
        JOIN public.contacts c ON c.contact_id = r.cid
        JOIN bedrock.jobs_contact_membership m
          ON m.contact_id = c.contact_id AND m.stage = 'initial_outreach'
        LEFT JOIN touch_counts tc ON tc.cid = r.cid
        WHERE true {owner_where}
        ORDER BY r.last_reply ASC
        LIMIT 200
    """, *params)
    return {"success": True, "data": [dict(r) for r in rows]}


@router.get("/outreach/stuck-contacts")
async def outreach_stuck_contacts(
    min_touches: int = Query(3, ge=1, le=20),
    owner: Optional[str] = Query(None),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Contacts stuck in initial outreach: N+ touches from the jobs team and no
    reply. The signal that this person isn't the way in and the team should work
    a different contact at that account — so each row carries how many OTHER
    jobs prospects exist at the same company. Replaces the account working list."""
    team_aem = " OR ".join(f"aem.from_email ILIKE '%{e}%'" for e in JOBS_TEAM_EMAILS)
    owner_where = ""
    params: list = [min_touches]
    if owner:
        owner_where = "AND lower(coalesce(c.owner_email,'')) = lower($2)"
        params.append(owner)
    rows = await conn.fetch(f"""
        WITH sends AS (
            -- one row per outbound MESSAGE (thread-level rows undercount follow-ups)
            SELECT a.participant_public_contact_id AS cid, aem.sent_at AS ts, TRUE AS is_email
            FROM bedrock.activity a
            JOIN bedrock.activity_email_message aem ON aem.activity_id = a.id
            WHERE a.deleted_at IS NULL AND a.type = 'email'
              AND coalesce(a.jobs_relevance_override, a.jobs_relevance) = 'jobs'
              AND a.participant_public_contact_id IS NOT NULL
              AND ({team_aem})
            UNION ALL
            SELECT a.participant_public_contact_id, a.activity_date, FALSE
            FROM bedrock.activity a
            WHERE a.deleted_at IS NULL AND a.type IN ('call','text','linkedin')
              AND a.participant_public_contact_id IS NOT NULL
        ),
        agg AS (
            SELECT cid, count(*) AS touches, max(ts) AS last_touch,
                   -- email-only floor, to match /outreach/responded-contacts exactly
                   min(ts) FILTER (WHERE is_email) AS first_email_send
            FROM sends GROUP BY 1),
        replied AS (
            -- a meeting at all, OR any inbound MESSAGE on a thread linked to
            -- the contact (their own reply or a colleague's — either way the
            -- account engaged).
            -- 'call' is deliberately NOT here: bedrock.activity has no direction
            -- column and the only writer of type='call' is log_jobs_activity, i.e.
            -- OUR outbound touch — already counted in `sends` above. Treating it as
            -- engagement disqualified every call-only contact from this queue, which
            -- is exactly the population the panel exists to surface.
            SELECT DISTINCT a.participant_public_contact_id AS cid
            FROM bedrock.activity a
            WHERE a.deleted_at IS NULL AND a.participant_public_contact_id IS NOT NULL
              AND a.type = 'meeting'
            UNION
            SELECT DISTINCT a.participant_public_contact_id AS cid
            FROM bedrock.activity a
            JOIN bedrock.activity_email_message aem ON aem.activity_id = a.id
            JOIN agg ON agg.cid = a.participant_public_contact_id
            JOIN bedrock.jobs_contact_membership mm
              ON mm.contact_id = a.participant_public_contact_id
            WHERE a.deleted_at IS NULL AND a.participant_public_contact_id IS NOT NULL
              AND aem.from_email NOT ILIKE '%@pursuit.org%'
              AND aem.from_email NOT ILIKE '%@pursuit.com%'
              -- These two predicates MUST match /outreach/responded-contacts
              -- (search: "the reply has to be on a jobs-classified thread").
              -- Without them a stale or non-jobs inbound message excluded a
              -- contact from Stuck without admitting them to Replied, so they
              -- appeared in neither queue on the same page.
              AND coalesce(a.jobs_relevance_override, a.jobs_relevance) = 'jobs'
              AND aem.sent_at >= coalesce(mm.first_outreach_at, agg.first_email_send)
        )
        SELECT c.contact_id, c.full_name, c.current_title, c.current_company,
               c.owner_email, agg.touches::int AS touches, agg.last_touch,
               m.first_outreach_at,
               (SELECT count(*) FROM public.contacts oc
                 WHERE oc.is_jobs_contact AND oc.contact_id <> c.contact_id
                   AND coalesce(trim(oc.current_company),'') <> ''
                   AND lower(trim(oc.current_company)) = lower(trim(c.current_company)))::int
                 AS other_contacts_at_account
        FROM agg
        JOIN public.contacts c ON c.contact_id = agg.cid
        JOIN bedrock.jobs_contact_membership m
          ON m.contact_id = c.contact_id AND m.stage = 'initial_outreach'
        LEFT JOIN replied r ON r.cid = agg.cid
        WHERE agg.touches >= $1 AND r.cid IS NULL {owner_where}
        ORDER BY agg.touches DESC, agg.last_touch DESC NULLS LAST
        LIMIT 200
    """, *params)
    return {"success": True, "data": [dict(r) for r in rows]}


@router.get("/outreach/accounts")
async def outreach_accounts(
    owner: Optional[str] = Query(None, description="Filter to accounts this staff sender is involved with"),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Account working list for the deep-dive: every account with a comment or an
    open task, with its owner, open tasks, and comments — rolled up from the
    opportunity- and prospect-level jobs_comment / jobs_task rows. Ordered by most
    recent note/task so freshly-worked accounts surface first. Not period-scoped —
    it's the live discussion queue.

    With `owner` set, restricts to accounts that staffer is involved with: they
    outreached to it (jobs email/linkedin to a contact there), own the opportunity,
    or authored a comment/task. Accounts with no notes/tasks simply don't appear."""
    # owner is regex-validated → safe to interpolate. When set, add a filter CTE.
    owner_cte, owner_where = "", ""
    if owner and _SAFE_EMAIL.match(owner):
        owner_cte = f""",
    owner_accounts AS (
        SELECT DISTINCT lower(trim(c.current_company)) AS acct
        FROM bedrock.activity a JOIN public.contacts c ON c.contact_id = a.participant_public_contact_id
        WHERE a.deleted_at IS NULL AND a.type IN ('email','linkedin')
          AND (a.email_from ILIKE '%{owner}%' OR a.logged_by ILIKE '%{owner}%')
          AND {_jobs_relevant('a')} AND coalesce(trim(c.current_company),'') <> ''
        UNION
        SELECT lower(trim(account_name)) FROM bedrock.jobs_opportunity
          WHERE deleted_at IS NULL AND lower(owner_email) = lower('{owner}') AND coalesce(trim(account_name),'') <> ''
        UNION SELECT acct FROM comments WHERE lower(author) = lower('{owner}')
        UNION SELECT acct FROM tasks WHERE lower(owner) = lower('{owner}')
    )"""
        owner_where = "WHERE a.acct IN (SELECT acct FROM owner_accounts)"
    sql = f"""
    WITH comments AS (
        SELECT lower(trim(o.account_name)) AS acct, o.account_name AS display,
               jc.author_email AS author, jc.content, jc.created_at
        FROM bedrock.jobs_comment jc
        JOIN bedrock.jobs_opportunity o ON o.id::text = jc.parent_id
        WHERE jc.parent_type='opportunity' AND o.deleted_at IS NULL AND coalesce(trim(o.account_name),'') <> ''
        UNION ALL
        SELECT lower(trim(c.current_company)), c.current_company, jc.author_email, jc.content, jc.created_at
        FROM bedrock.jobs_comment jc
        JOIN public.contacts c ON jc.parent_id ~ '^[0-9]+$' AND c.contact_id = jc.parent_id::int
        WHERE jc.parent_type='prospect' AND coalesce(trim(c.current_company),'') <> '' AND lower(trim(c.current_company)) <> 'n/a'
    ),
    tasks AS (
        SELECT lower(trim(o.account_name)) AS acct, o.account_name AS display,
               jt.title, jt.status, jt.deadline, jt.owner, jt.created_at
        FROM bedrock.jobs_task jt
        JOIN bedrock.jobs_opportunity o ON o.id::text = jt.parent_id
        WHERE jt.parent_type='opportunity' AND jt.deleted_at IS NULL AND jt.status <> 'Completed'
          AND o.deleted_at IS NULL AND coalesce(trim(o.account_name),'') <> ''
        UNION ALL
        SELECT lower(trim(c.current_company)), c.current_company, jt.title, jt.status, jt.deadline, jt.owner, jt.created_at
        FROM bedrock.jobs_task jt
        JOIN public.contacts c ON jt.parent_id ~ '^[0-9]+$' AND c.contact_id = jt.parent_id::int
        WHERE jt.parent_type='prospect' AND jt.deleted_at IS NULL AND jt.status <> 'Completed'
          AND coalesce(trim(c.current_company),'') <> '' AND lower(trim(c.current_company)) <> 'n/a'
    ),
    accts AS (SELECT acct FROM comments UNION SELECT acct FROM tasks),
    owners AS (
        SELECT lower(trim(account_name)) AS acct,
               (array_agg(owner_email ORDER BY updated_at DESC) FILTER (WHERE owner_email IS NOT NULL))[1] AS owner
        FROM bedrock.jobs_opportunity WHERE deleted_at IS NULL AND coalesce(trim(account_name),'') <> '' GROUP BY 1
    ){owner_cte}
    SELECT a.acct,
           coalesce((SELECT display FROM comments c WHERE c.acct=a.acct AND display IS NOT NULL LIMIT 1),
                    (SELECT display FROM tasks t WHERE t.acct=a.acct AND display IS NOT NULL LIMIT 1), a.acct) AS display,
           coalesce(ja.owner_email, ow.owner) AS owner,
           greatest(
             (SELECT max(created_at) FROM comments c WHERE c.acct=a.acct),
             (SELECT max(created_at) FROM tasks t WHERE t.acct=a.acct)
           ) AS last_activity,
           (SELECT json_agg(json_build_object('author', cc.author, 'content', cc.content, 'date', cc.created_at) ORDER BY cc.created_at DESC)
            FROM comments cc WHERE cc.acct=a.acct) AS comments,
           (SELECT json_agg(json_build_object('title', tt.title, 'status', tt.status, 'deadline', tt.deadline, 'owner', tt.owner) ORDER BY tt.deadline NULLS LAST)
            FROM tasks tt WHERE tt.acct=a.acct) AS open_tasks,
           (SELECT count(*) FROM public.contacts pc WHERE lower(trim(pc.current_company))=a.acct AND pc.is_jobs_contact = true) AS contact_count,
           (SELECT json_agg(json_build_object('name', x.full_name, 'title', x.current_title))
            FROM (SELECT full_name, current_title FROM public.contacts pc2
                  WHERE lower(trim(pc2.current_company))=a.acct AND pc2.is_jobs_contact = true
                  ORDER BY pc2.updated_at DESC LIMIT 30) x) AS contacts
    FROM accts a
    LEFT JOIN bedrock.jobs_account ja ON ja.account_key = a.acct
    LEFT JOIN owners ow ON ow.acct = a.acct
    {owner_where}
    ORDER BY last_activity DESC NULLS LAST
    """
    rows = await conn.fetch(sql)
    items = []
    for r in rows:
        comments = _jsonb(r["comments"]) or []
        tasks = _jsonb(r["open_tasks"]) or []
        contacts_l = _jsonb(r["contacts"]) or []
        items.append({
            "account": r["display"], "owner": r["owner"],
            "last_activity": r["last_activity"].isoformat() if r["last_activity"] else None,
            "comment_count": len(comments), "open_task_count": len(tasks),
            "contact_count": r["contact_count"] or 0,
            "comments": comments, "open_tasks": tasks, "contacts": contacts_l,
        })
    return {"success": True, "data": {"accounts": items}}


@router.get("/daily-digest")
async def jobs_daily_digest(date_q: Optional[str] = Query(None, alias="date"),
                            user=Depends(require_auth), conn=Depends(get_db)):
    """The morning-Slack digest, computed: a day's outreach split into
    new vs existing accounts (message-level emails + manual touches +
    jobs-relevant meetings), plus builder submissions to roles. "New account"
    = no activity mapped to that account before the day. Defaults to
    yesterday (UTC). Builder OUTREACH (staff→builder emails) isn't captured
    in the activity model yet — that line is a future addition."""
    if date_q:
        try:
            d = date.fromisoformat(date_q)
        except ValueError:
            raise HTTPException(400, "invalid date; use YYYY-MM-DD")
    else:
        d = (datetime.now(timezone.utc) - timedelta(days=1)).date()
    day_start = datetime(d.year, d.month, d.day, tzinfo=timezone.utc)
    day_end = day_start + timedelta(days=1)

    # Scope to the JOBS TEAM, not all of Pursuit: the digest mirrors Avni's
    # morning post, which counts her team's outreach. Counting every active
    # mailbox pulled in Rebecca/Kirstie/Nick/Jukay's unrelated email and
    # overstated the day (7/27: 15 touches vs her 11).
    team_aem = " OR ".join(f"aem.from_email ILIKE '%{e}%'" for e in JOBS_TEAM_EMAILS)
    team_act = " OR ".join(
        f"a.email_from ILIKE '%{e}%' OR a.logged_by ILIKE '%{e}%'" for e in JOBS_TEAM_EMAILS)
    outreach = await conn.fetchrow(f"""
        WITH touches AS (
            -- message-level email sends by the jobs team, jobs-relevant
            SELECT aem.sent_at AS ts, 'email' AS kind,
                   a.account_id, a.participant_public_contact_id AS pid
            FROM bedrock.activity a
            JOIN bedrock.activity_email_message aem ON aem.activity_id = a.id
            WHERE a.type = 'email' AND a.deleted_at IS NULL
              AND coalesce(a.jobs_relevance_override, a.jobs_relevance) = 'jobs'
              AND aem.sent_at >= $1 AND aem.sent_at < $2
              AND ({team_aem})
            UNION ALL
            -- manual touches by the team; a meeting counts when it's classified
            -- jobs-relevant OR the person sitting in it is a jobs prospect (the
            -- classifier misses meetings, which cost Avni a meeting on 7/27).
            SELECT a.activity_date, a.type, a.account_id, a.participant_public_contact_id
            FROM bedrock.activity a
            WHERE a.type IN ('call','text','linkedin','meeting') AND a.deleted_at IS NULL
              AND a.activity_date >= $1 AND a.activity_date < $2
              AND ({team_act})
              AND (coalesce(a.jobs_relevance_override, a.jobs_relevance) = 'jobs'
                   OR a.type <> 'meeting'
                   OR EXISTS (SELECT 1 FROM public.contacts jc
                              WHERE jc.contact_id = a.participant_public_contact_id
                                AND jc.is_jobs_contact))
        ),
        keyed AS (
            SELECT t.*, COALESCE(t.account_id,
                (SELECT lower(trim(c.current_company)) FROM public.contacts c
                 WHERE c.contact_id = t.pid AND coalesce(trim(c.current_company),'') <> '')) AS akey
            FROM touches t
        ),
        flagged AS (
            SELECT k.*, (k.akey IS NOT NULL AND NOT EXISTS (
                SELECT 1 FROM bedrock.activity p
                LEFT JOIN public.contacts pc ON pc.contact_id = p.participant_public_contact_id
                WHERE p.deleted_at IS NULL AND p.activity_date < $1
                  AND COALESCE(p.account_id, lower(trim(pc.current_company))) = k.akey
            )) AS is_new
            FROM keyed k
        )
        SELECT count(*) FILTER (WHERE is_new)                                   AS new_touches,
               count(*) FILTER (WHERE NOT is_new)                               AS existing_touches,
               count(DISTINCT akey) FILTER (WHERE is_new)                       AS new_accounts,
               count(DISTINCT akey) FILTER (WHERE NOT is_new AND akey IS NOT NULL) AS existing_accounts,
               count(*) FILTER (WHERE kind = 'meeting')                         AS meetings
        FROM flagged
    """, day_start, day_end)

    submissions = await conn.fetch("""
        SELECT coalesce(nullif(trim(ja.company_name),''),'(unknown)') AS company,
               count(*) AS builders,
               count(DISTINCT coalesce(ja.jobs_role_id::text, lower(coalesce(ja.role_title,'')))) AS roles
        FROM public.job_applications ja
        WHERE ja.date_applied = $1 AND ja.source_type = 'Pursuit_referred'
        GROUP BY 1 ORDER BY 2 DESC
    """, d)

    return {"success": True, "data": {
        "date": d.isoformat(),
        "outreach": dict(outreach) if outreach else {},
        "submissions": [dict(r) for r in submissions],
    }}


@router.get("/this-week-summary")
async def this_week_summary(user=Depends(require_auth), conn=Depends(get_db)):
    """A narrative of the jobs team's last 7 days: who Avni/Damon emailed &
    met (first touches this week), and which opportunities progressed."""
    emailed = await conn.fetch(f"""
        {_first_touch_email_cte()}
        SELECT ext.counterpart AS email, ext.first_touch,
               p.full_name, p.current_company
        FROM ext
        LEFT JOIN LATERAL (
            SELECT full_name, current_company FROM public.contacts c
            WHERE lower(c.email) = ext.counterpart AND c.is_jobs_contact = true LIMIT 1
        ) p ON true
        WHERE ext.first_touch >= now() - interval '7 days'
        ORDER BY ext.first_touch DESC
    """)
    met = await conn.fetch(f"""
        {_first_touch_meeting_cte()}
        SELECT ext.counterpart AS email, ext.first_touch,
               p.full_name, p.current_company
        FROM ext
        LEFT JOIN LATERAL (
            SELECT full_name, current_company FROM public.contacts c
            WHERE lower(c.email) = ext.counterpart AND c.is_jobs_contact = true LIMIT 1
        ) p ON true
        WHERE ext.first_touch >= now() - interval '7 days'
        ORDER BY ext.first_touch DESC
    """)
    progressed = await conn.fetch("""
        SELECT o.account_name, h.from_stage, h.to_stage, h.changed_at
        FROM bedrock.jobs_stage_history h
        JOIN bedrock.jobs_opportunity o ON o.id = h.opportunity_id
        WHERE h.changed_at >= now() - interval '7 days' AND o.deleted_at IS NULL
          AND h.from_stage IS NOT NULL
        ORDER BY h.changed_at DESC
    """)

    def _person(r):
        return {"email": r["email"], "name": r["full_name"],
                "company": r["current_company"], "when": r["first_touch"]}

    return {"success": True, "data": {
        "emailed":  [_person(r) for r in emailed],
        "met":      [_person(r) for r in met],
        "progressed": [{
            "account": r["account_name"],
            "from_stage": STAGE_LABELS.get(r["from_stage"], r["from_stage"]),
            "to_stage": STAGE_LABELS.get(r["to_stage"], r["to_stage"]),
            "when": r["changed_at"],
        } for r in progressed],
        "counts": {"emailed": len(emailed), "met": len(met), "progressed": len(progressed)},
    }}


class ContactCreate(BaseModel):
    full_name:       str
    email:           Optional[str] = None
    current_title:   Optional[str] = None
    current_company: Optional[str] = None
    contact_stage:   str = "lead"
    linkedin_url:    Optional[str] = None


@router.post("/contacts")
async def create_contact(
    body: ContactCreate,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    import uuid as _uuid
    at_id = f"manual-{_uuid.uuid4().hex[:8]}"
    parts = body.full_name.split(" ", 1)
    first = parts[0]
    last  = parts[1] if len(parts) > 1 else ""
    # TKT-135: email/linkedin have UNIQUE indexes that treat '' as a value —
    # one ''-email row made EVERY email-less create explode with a raw 500.
    # Blanks become NULL, and duplicates come back as a friendly 409 naming
    # the existing contact instead of an unhandled 500.
    email = (body.email or "").strip() or None
    linkedin = (body.linkedin_url or "").strip() or None
    try:
        cid = await conn.fetchval("""
            INSERT INTO public.contacts
                (first_name, last_name, full_name, email, current_title,
                 current_company, linkedin_url, source, airtable_id, contact_stage)
            VALUES ($1,$2,$3,$4,$5,$6,$7,'manual',$8,$9)
            RETURNING contact_id
        """, first, last, body.full_name, email, body.current_title,
            body.current_company, linkedin, at_id, body.contact_stage)
    except asyncpg.exceptions.UniqueViolationError:
        dupe = await conn.fetchrow(
            "SELECT contact_id, full_name FROM public.contacts "
            "WHERE ($1::text IS NOT NULL AND email = $1) OR ($2::text IS NOT NULL AND linkedin_url = $2) LIMIT 1",
            email, linkedin)
        detail = (f"A contact with that email or LinkedIn already exists: "
                  f"{dupe['full_name']} (#{dupe['contact_id']})" if dupe
                  else "A contact with that email or LinkedIn already exists.")
        raise HTTPException(409, detail)
    row = await conn.fetchrow("SELECT * FROM public.contacts WHERE contact_id=$1", cid)
    return {"success": True, "data": dict(row)}


@router.get("/contacts/search")
async def search_all_contacts(
    q: str = Query(..., min_length=2),
    limit: int = Query(20, le=50),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Search ALL contacts in the DB (LinkedIn, SF-linked, Airtable) for the picker UI."""
    rows = await conn.fetch(
        """
        SELECT
            c.contact_id,
            c.full_name,
            c.email,
            c.current_title,
            c.current_company,
            c.source,
            c.airtable_id,
            c.contact_stage,
            scl.sf_contact_id IS NOT NULL AS in_sf,
            -- Reference key to use when linking to a deal
            CASE
                WHEN c.airtable_id IS NOT NULL THEN 'airtable:' || c.airtable_id
                WHEN scl.sf_contact_id IS NOT NULL THEN scl.sf_contact_id
                ELSE 'pub:' || c.contact_id::text
            END AS contact_ref
        FROM public.contacts c
        LEFT JOIN bedrock.sf_contact_link scl ON scl.public_contact_id = c.contact_id
        WHERE
            lower(c.full_name) LIKE lower($1)
            OR lower(c.email) LIKE lower($1)
            OR lower(c.current_company) LIKE lower($1)
        ORDER BY
            CASE WHEN c.airtable_id IS NOT NULL THEN 0
                 WHEN scl.sf_contact_id IS NOT NULL THEN 1
                 ELSE 2 END,
            c.full_name
        LIMIT $2
        """,
        f"%{q}%",
        limit,
    )
    return {
        "success": True,
        "data": [dict(r) for r in rows],
    }


@router.get("/contacts/by-account")
async def contacts_by_account(
    deal_type: Optional[str] = Query(None),
    scope: str = Query("engaged", pattern="^(engaged|all)$"),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Jobs prospects grouped into account rows by company name.

    No hard account_id on contacts, so we group by the company text
    (COALESCE(NULLIF(trim(current_company),''),'(no company)')). Accounts are
    ordered by contact_count desc, then name.

    `deal_type` narrows to prospects at companies that have a deal of that type.
    `scope` defaults to engaged (excludes ~32k cold linkedin imports); scope=all
    includes everything.
    """
    dt = deal_type if deal_type and deal_type != "all" else None
    eng = "" if scope == "all" else f"AND {_engaged_clause('c')}"
    rows = await conn.fetch(
        f"""
        SELECT
            COALESCE(NULLIF(trim(c.current_company), ''), '(no company)') AS account,
            c.contact_id, c.full_name, c.email, c.current_title, c.contact_stage, c.linkedin_url
        FROM public.contacts c
        WHERE c.is_jobs_contact = true {eng}
          AND ($1::text IS NULL OR lower(c.current_company) IN (
                SELECT lower(account_name) FROM bedrock.jobs_opportunity
                WHERE deleted_at IS NULL AND deal_type = $1 AND account_name IS NOT NULL))
        ORDER BY account, full_name
        """,
        dt,
    )
    accounts: dict = {}
    for r in rows:
        acct = r["account"]
        g = accounts.setdefault(acct, {"account": acct, "contact_count": 0, "contacts": []})
        g["contact_count"] += 1
        g["contacts"].append({
            "contact_id":    r["contact_id"],
            "full_name":     r["full_name"],
            "email":         r["email"],
            "current_title": r["current_title"],
            "contact_stage": r["contact_stage"],
            "linkedin_url":  r["linkedin_url"],
        })

    # Attach each account's current opportunity (matched by company name). When a
    # company has more than one deal, prefer an active one, then won, on-hold, lost,
    # and break ties by most-recently-updated — so the account row shows the deal
    # the team is actually working.
    deal_rows = await conn.fetch(
        """
        SELECT DISTINCT ON (lower(account_name))
            account_name, id, stage, deal_type, owner_email
        FROM bedrock.jobs_opportunity
        WHERE deleted_at IS NULL AND account_name IS NOT NULL
        ORDER BY lower(account_name),
            CASE
                WHEN stage LIKE 'active%'  THEN 0
                WHEN stage = 'closed_won'  THEN 1
                WHEN stage LIKE 'on_hold%' THEN 2
                WHEN stage = 'closed_lost' THEN 3
                ELSE 4
            END,
            updated_at DESC NULLS LAST
        """,
    )
    deal_by_company = {r["account_name"].strip().lower(): r for r in deal_rows}
    for acct, g in accounts.items():
        d = deal_by_company.get(acct.strip().lower())
        g["deal"] = (
            {
                "id":          str(d["id"]),
                "stage":       d["stage"],
                "deal_type":   d["deal_type"],
                "owner_email": d["owner_email"],
            }
            if d
            else None
        )

    out = sorted(accounts.values(), key=lambda a: (-a["contact_count"], a["account"]))
    return {"success": True, "data": out}


# Account status vocabulary mirrors the portfolio Accounts tab so the two read
# the same. Status is DERIVED from the account's jobs data (not stored):
#   Stewarding   – a won deal (delivery / placed relationship)
#   Pursuing     – an active open opportunity
#   Re-activating– only stale opps (on-hold/lost) but touched in the last 90 days
#   Dormant      – only stale opps, no recent touch
#   Activating    – no opportunity yet, but we've made contact / done outreach
#   Prospect     – no opportunity AND no activity (untouched)
_ACCOUNT_STATUS_RANK = {
    "Pursuing": 0, "Stewarding": 1, "Re-activating": 2,
    "Activating": 3, "Prospect": 4, "Dormant": 5,
}


@router.get("/accounts/names")
async def jobs_account_names(
    scope: str = Query("engaged", pattern="^(engaged|all)$"),
    user=Depends(require_auth), conn=Depends(get_db),
):
    """Lightweight account picker: distinct account key + display name only.
    Used by dropdowns (home quick-add, candidate company field) so they don't
    pull the full /accounts payload. Defaults to engaged accounts; scope=all
    includes cold linkedin-import companies too."""
    eng = "" if scope == "all" else f"AND {_engaged_clause('c')}"
    rows = await conn.fetch(
        f"""
        SELECT key, (array_agg(name ORDER BY length(name)))[1] AS name FROM (
          SELECT lower(trim(account_name)) AS key, trim(account_name) AS name
          FROM bedrock.jobs_opportunity
          WHERE deleted_at IS NULL AND coalesce(trim(account_name),'') <> ''
          UNION ALL
          SELECT lower(trim(c.current_company)) AS key, trim(c.current_company) AS name
          FROM public.contacts c
          WHERE c.is_jobs_contact = true AND coalesce(trim(c.current_company),'') <> '' {eng}
        ) s
        GROUP BY key ORDER BY name
        """,
    )
    return {"success": True, "data": [{"account_key": r["key"], "account": r["name"]} for r in rows]}


# ── Schema probes ─────────────────────────────────────────────────────────────
# Migrations are applied by Jac, not by the app, so a feature can ship before its
# column exists. Probing lets the same build work on both sides of a migration
# and light up on apply — referencing a missing column would 42703 the request.
# Cached: columns are only ever ADDED, so a True is permanent; a False is
# re-checked, which is what makes the feature appear without a restart.
_COLUMN_CACHE: dict[tuple[str, str, str], bool] = {}


async def _has_column(schema: str, table: str, column: str) -> bool:
    key = (schema, table, column)
    if _COLUMN_CACHE.get(key):
        return True
    pool = get_pool()
    found = bool(await pool.fetchval(
        "SELECT 1 FROM information_schema.columns "
        "WHERE table_schema = $1 AND table_name = $2 AND column_name = $3",
        schema, table, column))
    _COLUMN_CACHE[key] = found
    return found


# ── Stage vocabulary probe ────────────────────────────────────────────────────
# The 2026-08-05 stage migration is applied by Jac, not by the app, so at any
# moment the database may be on the old vocabulary or the new one. Rather than
# guess, read the CHECK constraint and report what it actually accepts. The UI
# builds its stage pickers from this, so it can never offer a value the database
# will reject — which is the failure mode that would break saves for the whole
# team during the window between migration and deploy.
_STAGE_VOCAB_CACHE: dict[str, list[str]] = {}


async def _writable_stages(table: str, constraint: str, fallback: list[str],
                           settles_when: Optional[str] = None) -> list[str]:
    """Stage values this table's CHECK constraint currently allows, in the order
    given by `fallback` (constraint order is not meaningful). Falls back to the
    full list if there's no constraint to read.

    Only caches the POST-migration answer. Caching the pre-migration list would
    pin it for the life of the process: Jac applies the migration, the constraint
    changes underneath us, and every request keeps serving the stale vocabulary
    until someone restarts the API — which defeats the entire point of probing.
    `settles_when` names a value that only the new constraint contains; until it
    shows up we re-read each time (one indexed catalog lookup)."""
    cached = _STAGE_VOCAB_CACHE.get(table)
    if cached is not None:
        return cached
    pool = get_pool()
    try:
        definition = await pool.fetchval(
            "SELECT pg_get_constraintdef(oid) FROM pg_constraint WHERE conname = $1",
            constraint)
    except Exception:            # nosec - probe only; never block a request on it
        definition = None
    if not definition:
        return fallback
    allowed = [v for v in fallback if f"'{v}'" in definition]
    if not allowed:              # unrecognised shape — don't lock the UI out
        return fallback
    if settles_when is None or settles_when in allowed:
        _STAGE_VOCAB_CACHE[table] = allowed
    return allowed


@router.get("/stage-vocabulary")
async def stage_vocabulary(user=Depends(require_auth)):
    """What the database will accept for each stage column right now, plus the
    labels. Lets the UI stay correct across the stage migration without a
    coordinated deploy: pre-migration it offers the old values, post-migration
    the new ones, with no code change in between."""
    opp = await _writable_stages(
        "jobs_opportunity", "jobs_opportunity_stage_check",
        OPPORTUNITY_STAGES_NEW + ["initial_outreach", "active_builder_interview",
                                  "on_hold_not_selected", "on_hold_not_interested",
                                  "on_hold_not_responsive"],
        settles_when="reviewing_builders")
    mem = await _writable_stages(
        "jobs_contact_membership", "jobs_contact_membership_stage_vals",
        MEMBERSHIP_STAGES_NEW + ["on_hold"],
        settles_when="call_booked")
    # Report the TARGET vocabulary with an `available` flag, not just what's
    # writable. Filtering the unavailable ones out entirely made Call Booked and
    # Revisit simply missing from the dropdown, which reads as "not built" rather
    # than "waiting on the migration". Showing them disabled, with a reason on
    # hover, tells the truth and still can't produce a failing save.
    def _opt(v: str, label: str, allowed: list[str]):
        ok = v in allowed
        return {"value": v, "label": label, "available": ok,
                "unavailable_reason": None if ok else
                "Available once the 2026-08-05 stage migration is applied"}

    # on_hold is deliberately NOT appended when the database still accepts it:
    # Revisit replaces it outright (Kwame 2026-08-05), so offering both would let
    # someone park a contact in the stage we're retiring. Until the migration
    # lands Revisit shows disabled, which means parking is briefly unavailable —
    # the alternative is writing more rows into a stage about to be renamed.
    membership_target = MEMBERSHIP_STAGES_NEW
    opp_target = OPPORTUNITY_STAGES_NEW

    return {"success": True, "data": {
        "opportunity_stages": [_opt(v, STAGE_LABELS.get(v, v), opp) for v in opp_target],
        "membership_stages": [_opt(v, MEMBERSHIP_STAGE_LABELS.get(v, v), mem) for v in membership_target],
        "closed_lost_reasons": [{"value": v, "label": l, "available": True} for v, l in CLOSED_LOST_REASONS],
        # True once the stage migration has landed.
        "migrated": "call_booked" in mem,
    }}


@router.get("/accounts")
async def jobs_accounts(
    deal_type: Optional[str] = Query(None),
    scope: str = Query("engaged", pattern="^(engaged|all)$"),
    user=Depends(require_auth),
):
    """Account-level hub: every company with an opportunity OR a jobs prospect,
    keyed by normalized company name, with its opportunities and prospects nested
    and a derived account status (same vocabulary as the portfolio Accounts tab).

    `account_name`/`current_company` is the canonical key — SF `account_id` is
    carried through when it's a real Account Id but is too sparse to group on.
    `deal_type` narrows to accounts that have an opportunity of that type.
    `scope` defaults to engaged (~15k contacts) so the hub doesn't nest ~38k
    rows; scope=all includes cold linkedin imports.
    """
    eng = "" if scope == "all" else f"AND {_engaged_clause('c')}"

    # investor_account_key arrives with the 2026-08-05 migration. Selected only
    # once it exists — naming a missing column would 42703 the whole accounts
    # list, and this endpoint is the Accounts page. Named columns rather than
    # SELECT *: `notes` is free text across every account and has no business
    # riding along in a list payload.
    investor_select = (
        ", investor_account_key"
        if await _has_column("bedrock", "jobs_account", "investor_account_key") else "")

    # Every input below is an independent read. Run sequentially on one
    # connection they cost ~2.6s; gather them across the pool so wall-time ≈ the
    # single slowest query (~0.8s). (Was 3s+ end-to-end for the whole endpoint.)
    team_sender = " OR ".join(f"a.email_from ILIKE '%{e}%'" for e in JOBS_TEAM_EMAILS)
    # SQL expr → the jobs-team member who authored a row (NULL if none), so we can
    # aggregate the distinct set of team members who've touched each account.
    actor_case = "CASE " + " ".join(
        f"WHEN a.email_from ILIKE '%{e}%' OR a.logged_by ILIKE '%{e}%' THEN '{e}'"
        for e in JOBS_TEAM_EMAILS
    ) + " END"
    pool = get_pool()
    (
        opp_rows, prospect_rows, ja_rows, task_rows,
        act1_rows, act2_rows, role_resp_rows, er_resp_rows,
        hires_rows, name_sf_rows, flagged_rows, listings_rows,
        industry_rows, contact_total_rows, domain_rows,
    ) = await asyncio.gather(
        pool.fetch(
            """
            SELECT id, account_id, account_name, stage, deal_type, title,
                   owner_email, priority, num_roles, likelihood, updated_at
            FROM bedrock.jobs_opportunity
            WHERE deleted_at IS NULL AND coalesce(trim(account_name), '') <> ''
            ORDER BY updated_at DESC NULLS LAST
            """),
        # Prospects are ~38k rows across ~21k companies — nesting them all into the
        # account list produced a 2.5MB (16MB on scope=all) payload and a 3s+ render.
        # The list only needs a COUNT per account; the actual prospect rows load
        # lazily when a row is expanded (GET /account-prospects). So aggregate here.
        pool.fetch(
            f"""
            SELECT lower(trim(c.current_company)) AS key,
                   min(c.current_company)          AS display,
                   count(*)                        AS n,
                   max(c.updated_at)               AS last_updated
            FROM public.contacts c
            WHERE c.is_jobs_contact = true AND coalesce(trim(c.current_company), '') <> '' {eng}
            GROUP BY 1
            """),
        # Persistent account record (owner, manual status override, SF link).
        pool.fetch("SELECT account_key, display_name, owner_email, status_override, sf_account_id"
                   f"{investor_select} FROM bedrock.jobs_account"),
        # Open account-level tasks per account_key.
        pool.fetch(
            "SELECT parent_id, count(*) AS n FROM bedrock.jobs_account_task "
            "WHERE parent_type='account' AND deleted_at IS NULL AND status <> 'Completed' "
            "GROUP BY parent_id"),
        # Per-account activity for warmth: recent volume (90d), recency (last
        # touch), whether the account has RESPONDED (a meeting/call, or an inbound
        # email — not just our outbound), and which team members touched it.
        pool.fetch(
            f"""
            SELECT lower(trim(c.current_company)) AS company,
                   max(a.activity_date) AS last_act,
                   min(a.activity_date) AS first_act,
                   count(*) FILTER (WHERE a.activity_date >= now() - interval '90 days') AS recent,
                   bool_or(a.type IN ('call','meeting')
                           OR (a.type = 'email' AND NOT ({team_sender}))) AS responded,
                   array_agg(DISTINCT {actor_case}) AS actors
            FROM bedrock.activity a JOIN public.contacts c ON c.contact_id = a.participant_public_contact_id
            WHERE a.deleted_at IS NULL AND coalesce(trim(c.current_company),'') <> ''
              AND {_jobs_relevant('a')}
            GROUP BY 1"""),
        # Meeting attendees (calendar rows have no participant link) so
        # meetings/manual aren't undercounted.
        pool.fetch(
            f"""
            SELECT lower(trim(c.current_company)) AS company,
                   max(a.activity_date) AS last_act,
                   min(a.activity_date) AS first_act,
                   count(*) FILTER (WHERE a.activity_date >= now() - interval '90 days') AS recent,
                   true AS responded,
                   array_agg(DISTINCT {actor_case}) AS actors
            FROM bedrock.activity a, jsonb_array_elements(coalesce(a.meeting_attendees, '[]'::jsonb)) att
            JOIN public.contacts c ON lower(c.email) = lower(att->>'email')
            WHERE a.deleted_at IS NULL AND a.source = 'calendar-sync'
              AND coalesce(trim(c.current_company),'') <> ''
              AND {_jobs_relevant('a')}
            GROUP BY 1"""),
        # A role created or a hire tagged is real momentum — count it as responded.
        pool.fetch(
            """SELECT lower(trim(o.account_name)) AS company, max(r.created_at) AS last_act,
                  count(*) FILTER (WHERE r.created_at >= now() - interval '90 days') AS recent, true AS responded
           FROM bedrock.jobs_role r JOIN bedrock.jobs_opportunity o ON o.id = r.opportunity_id
           WHERE o.deleted_at IS NULL AND coalesce(trim(o.account_name),'') <> '' GROUP BY 1"""),
        pool.fetch(
            """SELECT lower(trim(o.account_name)) AS company, max(er.created_at) AS last_act,
                  count(*) FILTER (WHERE er.created_at >= now() - interval '90 days') AS recent, true AS responded
           FROM public.employment_records er JOIN bedrock.jobs_opportunity o ON o.id = er.opportunity_id
           WHERE o.deleted_at IS NULL AND coalesce(trim(o.account_name),'') <> '' GROUP BY 1"""),
        # Builders hired per account (distinct builders with an employment_record
        # at the company), via the jobs opp OR the record's own company_name.
        # Own-venture records aren't "hires" at an external account.
        pool.fetch(
            """
            SELECT company, count(DISTINCT user_id) AS n FROM (
                SELECT er.user_id, lower(trim(o.account_name)) AS company
                FROM public.employment_records er
                JOIN bedrock.jobs_opportunity o ON o.id = er.opportunity_id
                WHERE o.deleted_at IS NULL AND coalesce(trim(o.account_name),'') <> ''
                  AND coalesce(er.is_own_venture, false) = false
                UNION
                SELECT er.user_id, lower(trim(er.company_name)) AS company
                FROM public.employment_records er
                WHERE coalesce(trim(er.company_name),'') <> ''
                  AND coalesce(er.is_own_venture, false) = false
            ) s GROUP BY company
            """),
        # Every Salesforce account id that maps to each account (by company name →
        # public.companies → sf_account_company_map), so SF fellow counts attach to
        # more than just the directly-SF-linked accounts.
        pool.fetch(
            """
            SELECT lower(trim(co.name)) AS key, array_agg(DISTINCT m.sf_account_id) AS ids
            FROM public.companies co
            JOIN bedrock.sf_account_company_map m ON m.public_company_id = co.company_id
            WHERE coalesce(trim(co.name), '') <> '' AND m.sf_account_id IS NOT NULL
            GROUP BY 1
            """),
        # Accounts (by normalized company name) with ≥1 contact flagged for jobs
        # activation and actively in the funnel → derive to "Activating".
        pool.fetch(
            """
            SELECT lower(trim(c.current_company)) AS key, count(*) AS n
            FROM bedrock.jobs_contact_membership m
            JOIN public.contacts c ON c.contact_id = m.contact_id
            WHERE m.stage IN ('assigned','initial_outreach','converted_to_opportunity')
              AND coalesce(trim(c.current_company), '') <> ''
            GROUP BY 1
            """),
        # Job listings per account (by normalized company name): roles the team
        # sourced (job_postings) + distinct roles builders applied to on the open
        # market (job_applications).
        pool.fetch(
            """
            SELECT k, sum(sourced)::int AS sourced, sum(applied)::int AS applied FROM (
              SELECT lower(trim(company_name)) k, count(*) sourced, 0 applied
              FROM public.job_postings WHERE coalesce(trim(company_name),'') <> '' GROUP BY 1
              UNION ALL
              SELECT lower(trim(company_name)) k, 0 sourced, count(DISTINCT lower(trim(role_title))) applied
              FROM public.job_applications
              WHERE coalesce(trim(company_name),'') <> '' AND coalesce(trim(role_title),'') <> '' GROUP BY 1
            ) s GROUP BY k
            """),
        # Company firmographics by normalized name (enriched via claude_ai / the
        # SERP-verified PE-VC pass) — powers the accounts Industry filter and the
        # Size / HQ / Industry columns. The WHERE deliberately admits a row that
        # has ANY of the four: gating on industry alone hid the size band for
        # every company whose industry was never enriched.
        pool.fetch(
            """
            SELECT lower(trim(name)) AS key,
                   max(industry)     AS industry,
                   max(size_bucket)  AS size_bucket,
                   max(hq_location)  AS hq_location,
                   max(stage)        AS company_stage
            FROM public.companies
            WHERE coalesce(trim(name),'') <> ''
              AND (coalesce(industry,'') <> '' OR coalesce(size_bucket,'') <> ''
                   OR coalesce(hq_location,'') <> '' OR coalesce(stage,'') <> '')
            GROUP BY 1
            """),
        # ALL contacts per company (not just flagged prospects). Lookup only —
        # it never seeds an account. Lets the UI distinguish "nobody exists
        # here" from "11 people on file, none activated into the pipeline"
        # (PFNYC read as empty while 11 contacts sat under it).
        pool.fetch(
            """
            SELECT lower(trim(current_company)) AS key, count(*) AS n
            FROM public.contacts
            WHERE coalesce(trim(current_company), '') <> ''
            GROUP BY 1
            """),
        # Company-name variants that share an email domain — the same real
        # company filed under 2-3 spellings ("PFNYC" holds the opportunity while
        # the flagged prospects sit under "Partnership For NYC"). Lets the UI say
        # where the prospects actually are instead of "nobody identified".
        pool.fetch(
            """
            SELECT lower(trim(c.current_company)) AS key,
                   lower(split_part(c.email, '@', 2)) AS domain,
                   count(*) AS n,
                   count(*) FILTER (WHERE c.is_jobs_contact) AS prospects,
                   min(c.current_company) AS display
            FROM public.contacts c
            WHERE coalesce(c.email,'') LIKE '%@%' AND coalesce(trim(c.current_company),'') <> ''
              AND lower(split_part(c.email, '@', 2)) NOT IN
                  ('gmail.com','yahoo.com','hotmail.com','outlook.com','icloud.com','aol.com','me.com','msn.com','comcast.net','pursuit.org')
            GROUP BY 1, 2
            """),
    )

    accounts: dict = {}

    def bucket(key: str, display: str) -> dict:
        return accounts.setdefault(
            key,
            {
                "account": display,
                "account_id": None,
                "owner_email": None,
                "opportunities": [],
                "_last": None,
            },
        )

    def touch(g: dict, ts) -> None:
        # contacts.updated_at is tz-naive while jobs_opportunity.updated_at is
        # tz-aware — normalize to UTC so they're comparable.
        if ts is None:
            return
        if ts.tzinfo is None:
            ts = ts.replace(tzinfo=timezone.utc)
        if g["_last"] is None or ts > g["_last"]:
            g["_last"] = ts

    for r in opp_rows:
        key = r["account_name"].strip().lower()
        g = bucket(key, r["account_name"].strip())
        if r["account_id"] and str(r["account_id"]).startswith("001"):
            g["account_id"] = r["account_id"]
        if not g["owner_email"] and r["owner_email"]:
            g["owner_email"] = r["owner_email"]
        touch(g, r["updated_at"])
        g["opportunities"].append({
            "id":         str(r["id"]),
            "title":      r["title"],
            "stage":      r["stage"],
            "deal_type":  r["deal_type"],
            "owner_email": r["owner_email"],
            "priority":   r["priority"],
            "num_roles":  r["num_roles"],
            "likelihood": r["likelihood"],
            "updated_at": r["updated_at"].isoformat() if r["updated_at"] else None,
        })

    prospect_counts: dict = {}
    for r in prospect_rows:
        key = r["key"]
        g = bucket(key, (r["display"] or "").strip())
        touch(g, r["last_updated"])
        prospect_counts[key] = r["n"]

    ja = {r["account_key"]: r for r in ja_rows}
    # Pretty name for an investor key, and the reverse count (how many companies
    # name this account as their investor). Derived from the rows already
    # fetched — no extra query, and no-op before the migration adds the column.
    display_by_key = {r["account_key"]: (r["display_name"] or r["account_key"]) for r in ja_rows}
    portfolio_counts: dict[str, int] = {}
    for r in ja_rows:
        inv = dict(r).get("investor_account_key")
        if inv:
            portfolio_counts[inv] = portfolio_counts.get(inv, 0) + 1
    company_by_key = {r["key"]: dict(r) for r in industry_rows}
    contact_totals = {r["key"]: r["n"] for r in contact_total_rows}
    # domain → the name variant that actually holds flagged prospects
    _by_domain: dict = {}
    for r in domain_rows:
        _by_domain.setdefault(r["domain"], []).append(r)
    sibling_hint: dict = {}
    for variants in _by_domain.values():
        if len(variants) < 2:
            continue
        best = max(variants, key=lambda v: v["prospects"])
        domain_total = sum(int(v["n"]) for v in variants)
        # High-precision only: the sibling must hold 3+ flagged prospects on the
        # shared domain. Looser rules produced junk hints from single
        # miscategorized rows (one "@uber.com but company=Pursuit" contact made
        # Uber look like a duplicate of Pursuit).
        if int(best["prospects"]) < 3 or int(best["n"]) < 3:
            continue
        for v in variants:
            if v["key"] != best["key"] and not v["prospects"]:
                sibling_hint[v["key"]] = {"account": best["display"], "prospects": int(best["prospects"])}
    # Manually-created accounts (a jobs_account row with no opps/prospects yet)
    # still need to appear on the hub, so seed a bucket for each.
    for k, rec in ja.items():
        if k not in accounts:
            bucket(k, (rec["display_name"] or k).strip())
    open_tasks = {r["parent_id"]: r["n"] for r in task_rows}
    acct_act: dict = {}

    def _merge(company, last_act, recent, responded, actors=None, first_act=None):
        cur = acct_act.setdefault(
            company, {"recent": 0, "last": None, "first": None, "responded": False, "actors": set()})
        cur["recent"] += recent or 0
        if last_act and (cur["last"] is None or last_act > cur["last"]):
            cur["last"] = last_act
        if first_act and (cur["first"] is None or first_act < cur["first"]):
            cur["first"] = first_act
        cur["responded"] = cur["responded"] or bool(responded)
        if actors:
            cur["actors"].update(a for a in actors if a)

    for r in act1_rows:
        _merge(r["company"], r["last_act"], r["recent"], r["responded"], r["actors"], r["first_act"])
    for r in act2_rows:
        _merge(r["company"], r["last_act"], r["recent"], r["responded"], r["actors"], r["first_act"])
    # A role created or a hire tagged is real momentum — count it as responded.
    for r in role_resp_rows:
        _merge(r["company"], r["last_act"], r["recent"], r["responded"])
    for r in er_resp_rows:
        _merge(r["company"], r["last_act"], r["recent"], r["responded"])

    # Builders hired per account (distinct builders with an employment_record at
    # the company), linked via the jobs opp OR the record's own company_name.
    # Own-venture records aren't "hires" at an external account. Keyed on the
    # normalized account name = account_key. (SF "fellows hired" is merged
    # separately — historical Pursuit placements live only in Salesforce.)
    hires_by_account: dict[str, int] = {r["company"]: r["n"] for r in hires_rows}

    # One company name can map to several SF accounts — keep them all so fellow
    # counts (keyed by sf_account_id) can be summed across every match.
    name_sf_ids: dict[str, list[str]] = {
        r["key"]: [x for x in (r["ids"] or []) if x] for r in name_sf_rows
    }
    flagged_keys = {r["key"] for r in flagged_rows}
    listings_by_key = {r["k"]: (r["sourced"] or 0, r["applied"] or 0) for r in listings_rows}

    dt = deal_type if deal_type and deal_type != "all" else None
    now = datetime.now(timezone.utc)
    out = []
    for key, g in accounts.items():
        opps = g["opportunities"]
        if dt and not any(o["deal_type"] == dt for o in opps):
            continue
        rec = ja.get(key)
        # A stored owner wins over the one derived from the opportunities.
        if rec and rec["owner_email"]:
            g["owner_email"] = rec["owner_email"]
        # An explicit SF link wins over the account_id derived from opps.
        if rec and rec["sf_account_id"]:
            g["account_id"] = rec["sf_account_id"]
        stages = {o["stage"] for o in opps}
        # An opportunity is "open" while it's anywhere in the live funnel —
        # initial_outreach through builder interview (NOT closed/on-hold). These
        # accounts are being actively pursued. Re-activating/Dormant is only for
        # accounts whose opps are ALL closed-lost or on-hold.
        has_open = any(s and (s == "initial_outreach" or s.startswith("active")) for s in stages)
        has_won = "closed_won" in stages
        last = g.pop("_last")
        recent = bool(last and (now - last).days <= 90)
        a = acct_act.get(key)
        has_activity = bool(a and (a["last"] or a["recent"]))
        has_flagged = key in flagged_keys
        if rec and rec["status_override"]:
            status = rec["status_override"]
        elif has_open:
            status = "Pursuing"
        elif has_won:
            status = "Stewarding"
        elif opps:
            status = "Re-activating" if recent else "Dormant"
        elif has_activity or has_flagged:
            # No opportunity yet, but we're working it — activity, or a
            # contact here flagged for jobs activation.
            status = "Activating"
        else:
            status = "Prospect"
        g["account_key"] = key
        g["account_status"] = status
        g["opp_count"] = len(opps)
        g["prospect_count"] = prospect_counts.get(key, 0)
        # Everyone on file at this company, flagged or not — "no prospects" and
        # "no people at all" are different problems with different fixes.
        g["contact_count"] = contact_totals.get(key, 0)
        g["prospect_sibling"] = sibling_hint.get(key)
        # Read-only firmographics from public.companies — Bedrock displays them,
        # other systems own them, so nothing here is editable.
        co = company_by_key.get(key) or {}
        g["industry"] = co.get("industry")
        g["size_bucket"] = co.get("size_bucket")
        g["hq_location"] = co.get("hq_location")
        g["company_stage"] = co.get("company_stage")
        # Investor link both ways: who owns this company, and (for an investor
        # account) which companies it owns. Present only after the 2026-08-05
        # migration; `ja` rows simply won't carry the key before then.
        _ja = ja.get(key)
        inv_key = dict(_ja).get("investor_account_key") if _ja else None
        g["investor_account_key"] = inv_key
        g["investor_name"] = display_by_key.get(inv_key) if inv_key else None
        g["portfolio_count"] = portfolio_counts.get(key, 0)
        _src, _app = listings_by_key.get(key, (0, 0))
        g["roles_sourced"] = _src
        g["roles_applied"] = _app
        g["job_listings"] = _src + _app
        g["last_activity"] = last.isoformat() if last else None
        g["open_tasks"] = open_tasks.get(key, 0)
        g["recent_activity_count"] = a["recent"] if a else 0
        g["last_activity_at"] = a["last"].isoformat() if (a and a["last"]) else None
        g["first_activity_at"] = a["first"].isoformat() if (a and a.get("first")) else None
        g["responded"] = bool(a and a["responded"])
        # Which jobs-team members have touched this account (for the team filter).
        g["activity_actors"] = sorted(a["actors"]) if a else []
        # # hired: builders we placed (our DB) + SF fellows hired (historical,
        # merged from Salesforce by the caller; null until that enrichment runs).
        g["builders_hired"] = hires_by_account.get(key, 0)
        g["fellows_hired"] = None
        # Every SF account id this account resolves to (direct opp link + explicit
        # jobs_account link + company-name bridge), so the SF fellow counts can
        # attach to far more than just the directly-SF-linked accounts.
        sf_ids = set(name_sf_ids.get(key, []))
        if g["account_id"] and str(g["account_id"]).startswith("001"):
            sf_ids.add(g["account_id"])
        if rec and rec["sf_account_id"]:
            sf_ids.add(rec["sf_account_id"])
        g["sf_account_ids"] = sorted(sf_ids)
        out.append(g)

    out.sort(key=lambda a: (
        _ACCOUNT_STATUS_RANK.get(a["account_status"], 9),
        -(a["opp_count"] + a["prospect_count"]),
        a["account"].lower(),
    ))
    return {"success": True, "data": out}


_VALID_ACCOUNT_STATUS = {"Prospect", "Activating", "Pursuing", "Stewarding", "Re-activating", "Dormant"}


class JobsAccountUpdate(BaseModel):
    account: str                              # the account display/company name
    owner_email: Optional[str] = None
    status_override: Optional[str] = None     # "" clears the override (back to derived)
    notes: Optional[str] = None
    # account_key of the investor that owns this company. "" clears it. The
    # investor is itself a jobs_account, so the link is navigable both ways.
    # Column arrives with the 2026-08-05 migration.
    investor_account_key: Optional[str] = None


@router.patch("/accounts")
async def update_jobs_account(
    body: JobsAccountUpdate,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Upsert the persistent account record (owner / manual status / notes).

    Keyed by the normalized company name so it lines up with GET /accounts.
    Only the provided fields are written; status_override="" clears it.
    """
    key = body.account.strip().lower()
    if not key:
        raise HTTPException(status_code=400, detail="account is required")
    if body.status_override and body.status_override not in _VALID_ACCOUNT_STATUS:
        raise HTTPException(status_code=400, detail=f"invalid status_override: {body.status_override}")

    # Fixed positional params; only the provided columns are touched on UPDATE
    # (EXCLUDED = the attempted-insert values), so a partial PATCH never nulls
    # the fields it didn't send. A new row inserts NULL for anything omitted.
    sets = ["display_name = EXCLUDED.display_name", "updated_at = now()"]
    if body.owner_email is not None:
        sets.append("owner_email = EXCLUDED.owner_email")
    if body.status_override is not None:
        sets.append("status_override = EXCLUDED.status_override")
    if body.notes is not None:
        sets.append("notes = EXCLUDED.notes")

    # The investor column ships with the 2026-08-05 migration. Before it lands,
    # accept the request and skip the write rather than 42703-ing an otherwise
    # valid owner/status/notes save that happened to travel with it.
    has_investor = await _has_column("bedrock", "jobs_account", "investor_account_key")
    cols, vals = "", ""
    if body.investor_account_key is not None and has_investor:
        # Self-reference guard: an account that is its own investor renders as an
        # infinite portfolio loop in the UI.
        inv = body.investor_account_key.strip().lower() or None
        if inv == key:
            raise HTTPException(400, "an account cannot be its own investor")
        cols, vals = ", investor_account_key", ", $6"
        sets.append("investor_account_key = EXCLUDED.investor_account_key")

    params = [key, body.account.strip(), body.owner_email or None,
              body.status_override or None, body.notes or None]
    if cols:
        params.append(inv)

    await conn.execute(
        f"""
        INSERT INTO bedrock.jobs_account (account_key, display_name, owner_email, status_override, notes{cols})
        VALUES ($1, $2, $3, $4, $5{vals})
        ON CONFLICT (account_key) DO UPDATE SET {', '.join(sets)}
        """,
        *params,
    )
    return {"success": True}


class JobsAccountCreate(BaseModel):
    name: str
    sf_account_id: Optional[str] = None   # link to an existing SF account; never creates one in SF


def _acct_nkey(name: str) -> str:
    """Normalized account-name key for de-duplication (case/punct/legal-suffix
    insensitive). Keep in step with the dedupe tooling."""
    s = (name or "").strip().lower()
    s = re.sub(r"\([^)]*\)", " ", s)                 # drop parenthetical acronyms
    s = re.sub(r"[,\.\&\'\"()]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    for _ in range(3):
        s = re.sub(r"\s+(inc|llc|ltd|limited|corp|corporation|co|company|the|group|"
                   r"holdings|holding|plc|gmbh|sa|ag|incorporated)$", "", s).strip()
    return s


@router.get("/accounts/resolve")
async def resolve_account(
    name: str = Query(..., min_length=1),
    user=Depends(require_auth),
    conn=Depends(get_db),
    client=Depends(get_mcp_client),
):
    """Does an account matching this name already exist? Returns ONE de-duplicated
    list of accounts, merging our pipeline and Salesforce by normalized name — the
    caller never sees which system a match came from. Each match carries whatever
    ids it has (local key and/or SF id) so selecting it can reconcile silently.
    `exact` = an account with this exact (normalized) name already exists."""
    from security import escape_soql_string
    q = name.strip()
    like = f"%{q.lower()}%"
    local_rows = await conn.fetch(
        """
        SELECT k, min(disp) AS display, sum(n) AS n FROM (
          SELECT lower(trim(account_name)) k, min(account_name) disp, count(*) n
          FROM bedrock.jobs_opportunity WHERE deleted_at IS NULL AND lower(trim(account_name)) LIKE $1 GROUP BY 1
          UNION ALL
          SELECT lower(trim(current_company)) k, min(current_company) disp, count(*) n
          FROM public.contacts WHERE is_jobs_contact AND lower(trim(current_company)) LIKE $1 GROUP BY 1
          UNION ALL
          SELECT account_key k, min(display_name) disp, 0 n
          FROM bedrock.jobs_account WHERE account_key LIKE $1 GROUP BY 1
        ) s GROUP BY k
        """,
        like,
    )
    sf_rows = []
    try:
        safe = escape_soql_string(q)
        res = await client.salesforce.query(
            "SELECT Id, Name FROM Account WHERE RecordType.Name='Organization' "
            f"AND Name LIKE '%{safe}%' ORDER BY Name LIMIT 25")
        sf_rows = res.get("records", [])
    except Exception as e:
        logger.warning(f"resolve_account SF lookup failed: {e}")

    # Merge both sources by normalized name → one row per real company.
    merged: dict = {}
    for r in local_rows:
        label = (r["display"] or r["k"]).strip()
        e = merged.setdefault(_acct_nkey(label), {"key": None, "label": label, "sf_account_id": None, "record_count": 0})
        if e["key"] is None:
            e["key"] = r["k"]
        e["record_count"] += int(r["n"] or 0)
    for a in sf_rows:
        e = merged.setdefault(_acct_nkey(a["Name"]), {"key": None, "label": a["Name"], "sf_account_id": None, "record_count": 0})
        if not e["sf_account_id"]:
            e["sf_account_id"] = a["Id"]
    matches = sorted(merged.values(), key=lambda e: (-e["record_count"], (e["label"] or "").lower()))
    return {"success": True, "data": {"matches": matches[:12], "exact": _acct_nkey(q) in merged}}


@router.post("/accounts")
async def create_jobs_account(
    body: JobsAccountCreate,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Create a net-new local jobs account. Optionally linked to an existing SF
    account (sf_account_id) — we NEVER create the account in Salesforce here.
    Idempotent on account_key; won't clobber an existing SF link with null."""
    key = body.name.strip().lower()
    if not key:
        raise HTTPException(status_code=400, detail="name is required")
    await conn.execute(
        """
        INSERT INTO bedrock.jobs_account (account_key, display_name, sf_account_id)
        VALUES ($1, $2, $3)
        ON CONFLICT (account_key) DO UPDATE
          SET display_name  = EXCLUDED.display_name,
              sf_account_id = COALESCE(jobs_account.sf_account_id, EXCLUDED.sf_account_id),
              updated_at    = now()
        """,
        key, body.name.strip(), body.sf_account_id,
    )
    return {"success": True, "data": {"account_key": key, "display": body.name.strip()}}


@router.get("/account-prospects")
async def account_prospects(
    key: str = Query(...),
    scope: str = Query("engaged", pattern="^(engaged|all)$"),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Prospects (jobs contacts) at one account, keyed by normalized company name.
    Split out of GET /accounts so the account list doesn't ship ~38k prospect
    rows — these load lazily when an account row is expanded. `scope` mirrors the
    list's engaged/all toggle so the drill-in matches the row's prospect_count."""
    eng = "" if scope == "all" else f"AND {_engaged_clause('c')}"
    rows = await conn.fetch(
        f"""
        SELECT c.contact_id, c.full_name, c.email, c.current_title,
               c.contact_stage, c.linkedin_url
        FROM public.contacts c
        WHERE c.is_jobs_contact = true AND lower(trim(c.current_company)) = $1 {eng}
        ORDER BY c.full_name
        """,
        key.strip().lower(),
    )
    return {"success": True, "data": [dict(r) for r in rows]}


# ── Jobs contact activation (flag + funnel membership) ────────────────────────
# Accepts both vocabularies: 'on_hold' until the 2026-08-05 migration lands,
# 'call_booked'/'revisit' after. The database CHECK is the real gate — this only
# catches typos, so being permissive here costs nothing and avoids rejecting a
# valid stage in whichever direction the schema currently sits.
_MEMBERSHIP_STAGES = ('assigned', 'initial_outreach', 'call_booked',
                      'converted_to_opportunity', 'revisit', 'on_hold', 'not_a_fit')


# TKT-161: placements whose linked opportunity was soft-deleted (data-entry
# errors) must not count anywhere. Self-sourced placements (no opp link) stay.
def _live_placement(a: str = "") -> str:
    col = f"{a}.opportunity_id" if a else "opportunity_id"
    return (f"({col} IS NULL OR NOT EXISTS (SELECT 1 FROM bedrock.jobs_opportunity dop "
            f"WHERE dop.id = {col} AND dop.deleted_at IS NOT NULL))")


def _user_email(user) -> Optional[str]:
    return user.get("email") if isinstance(user, dict) else getattr(user, "email", None)


async def _flag_contacts(conn, contact_ids: list[int], owner_email: Optional[str],
                         reason: str, note: Optional[str], by: Optional[str],
                         stage: Optional[str] = None) -> int:
    """Create/keep a membership per contact. With no stage → 'assigned' and never
    downgrades an existing further stage (plain flag). With an explicit stage →
    set it (bulk advance). Writes through is_jobs_contact=true for legacy views."""
    if not contact_ids:
        return 0
    stg = stage or "assigned"
    # When bulk-advancing to an explicit stage, stamp the stage-entry timestamp
    # for qualified/converted too (mirrors the single-contact PATCH path) so the
    # scorecard flow count sees these transitions. On a fresh insert the CASE on
    # $6 stamps directly; on conflict it stamps only on a genuine stage change.
    if stage:
        conflict_stage = "stage = EXCLUDED.stage,"
        stamp_cols = ", converted_at"
        stamp_vals = ", CASE WHEN $6 = 'converted_to_opportunity' THEN now() END"
        conflict_stamp = (
            "converted_at = CASE WHEN EXCLUDED.stage = 'converted_to_opportunity' AND jobs_contact_membership.stage "
            "IS DISTINCT FROM 'converted_to_opportunity' THEN now() ELSE jobs_contact_membership.converted_at END,")
        # call_booked_at arrives with the 2026-08-05 migration. Referencing a
        # column that doesn't exist yet would 42703 every flag/advance call, so
        # it's added only once the column is there.
        if await _has_column("bedrock", "jobs_contact_membership", "call_booked_at"):
            stamp_cols += ", call_booked_at"
            stamp_vals += ", CASE WHEN $6 = 'call_booked' THEN now() END"
            conflict_stamp += (
                "call_booked_at = CASE WHEN EXCLUDED.stage = 'call_booked' AND jobs_contact_membership.stage "
                "IS DISTINCT FROM 'call_booked' THEN now() ELSE jobs_contact_membership.call_booked_at END,")
    else:
        conflict_stage = stamp_cols = stamp_vals = conflict_stamp = ""
    async with conn.transaction():
        # Snapshot pre-upsert stages so we can record only genuine transitions.
        old = {r["contact_id"]: r["stage"] for r in await conn.fetch(
            "SELECT contact_id, stage FROM bedrock.jobs_contact_membership WHERE contact_id = ANY($1::int[])",
            contact_ids)}
        await conn.execute(
            f"""
            INSERT INTO bedrock.jobs_contact_membership
                (contact_id, stage, owner_email, activation_reason, activation_note, assigned_by{stamp_cols})
            SELECT cid, $6, $2, $3, $4, $5{stamp_vals} FROM unnest($1::int[]) AS cid
            ON CONFLICT (contact_id) DO UPDATE SET
                {conflict_stage}
                {conflict_stamp}
                owner_email       = COALESCE(jobs_contact_membership.owner_email, EXCLUDED.owner_email),
                activation_reason = COALESCE(jobs_contact_membership.activation_reason, EXCLUDED.activation_reason),
                updated_at        = now()
            """,
            contact_ids, owner_email, reason, note, by, stg,
        )
        if await has_membership_history(conn):
            # Fresh inserts always enter `stg`; conflicts change stage only when
            # an explicit stage was passed (the no-stage upsert never touches it).
            hist = [(cid, None, stg, by) for cid in contact_ids if cid not in old]
            if stage:
                hist += [(cid, old[cid], stage, by) for cid in contact_ids
                         if cid in old and old[cid] != stage]
            if hist:
                await conn.executemany(
                    "INSERT INTO bedrock.jobs_membership_stage_history "
                    "(contact_id, from_stage, to_stage, changed_by) VALUES ($1, $2, $3, $4)",
                    hist)
        await conn.execute(
            "UPDATE public.contacts SET is_jobs_contact = true WHERE contact_id = ANY($1::int[]) AND NOT is_jobs_contact",
            contact_ids)
    return len(contact_ids)


class FlagJobsBody(BaseModel):
    contact_ids: list[int]
    owner_email: Optional[str] = None
    activation_reason: str = "manual"
    note: Optional[str] = None
    stage: Optional[str] = None   # bulk-advance to this funnel stage; None = 'assigned'


@router.post("/contacts/flag-jobs")
async def flag_contacts_for_jobs(body: FlagJobsBody, user=Depends(require_auth), conn=Depends(get_db)):
    """Bulk 'flag for jobs activation' — the contacts-page carve action. An
    optional `stage` bulk-advances the funnel for the selected contacts."""
    if body.activation_reason not in ("manual", "scraper_job", "strategic", "algorithm"):
        raise HTTPException(400, "invalid activation_reason")
    if body.stage is not None and body.stage not in _MEMBERSHIP_STAGES:
        raise HTTPException(400, f"invalid stage; one of {_MEMBERSHIP_STAGES}")
    n = await _flag_contacts(conn, body.contact_ids, body.owner_email,
                             body.activation_reason, body.note, _user_email(user), body.stage)
    return {"success": True, "data": {"flagged": n}}


class MembershipPatch(BaseModel):
    stage: Optional[str] = None
    owner_email: Optional[str] = None
    first_outreach_by: Optional[str] = None
    opportunity_id: Optional[str] = None
    not_a_fit_reason: Optional[str] = None
    # Set with stage='revisit': when to pick this contact back up. Stored on the
    # membership AND turned into a jobs_task so it surfaces in the owner's Jobs
    # Home widget on the day, which is the whole point of Revisit over On Hold.
    revisit_date: Optional[str] = None


@router.patch("/contacts/{contact_id}/jobs-membership")
async def update_jobs_membership(contact_id: int, body: MembershipPatch,
                                 user=Depends(require_auth), conn=Depends(get_db)):
    """Advance the funnel / reassign owner. Stamps first_outreach_by (who actually
    reached out — may differ from owner) on the → initial_outreach transition."""
    old_stage = await conn.fetchval(
        "SELECT stage FROM bedrock.jobs_contact_membership WHERE contact_id = $1", contact_id)
    if old_stage is None:
        raise HTTPException(404, "contact is not flagged for jobs")
    sets, params, i = [], [contact_id], 2
    if body.stage is not None:
        if body.stage not in _MEMBERSHIP_STAGES:
            raise HTTPException(400, f"invalid stage; one of {_MEMBERSHIP_STAGES}")
        sets.append(f"stage = ${i}"); params.append(body.stage); i += 1
        if body.stage == "initial_outreach":
            sets.append(f"first_outreach_by = COALESCE(first_outreach_by, ${i})")
            params.append(body.first_outreach_by or _user_email(user)); i += 1
            sets.append("first_outreach_at = COALESCE(first_outreach_at, now())")
        # Stamp the stage-entry timestamp ONLY when this call actually moves the
        # contact INTO that stage — SET expressions see the pre-update row, so
        # `jobs_contact_membership.stage` here is the OLD stage. Re-stamps on a
        # genuine re-entry (e.g. active → on_hold → active), which is what the
        # scorecard's "entered this period" flow count wants.
        if body.stage == "converted_to_opportunity":
            sets.append("converted_at = CASE WHEN jobs_contact_membership.stage "
                        "IS DISTINCT FROM 'converted_to_opportunity' THEN now() ELSE jobs_contact_membership.converted_at END")
        if body.stage == "call_booked" and await _has_column(
                "bedrock", "jobs_contact_membership", "call_booked_at"):
            sets.append("call_booked_at = CASE WHEN jobs_contact_membership.stage "
                        "IS DISTINCT FROM 'call_booked' THEN now() ELSE jobs_contact_membership.call_booked_at END")
    if body.revisit_date is not None and await _has_column(
            "bedrock", "jobs_contact_membership", "revisit_date"):
        try:
            _rd = date.fromisoformat(body.revisit_date)
        except ValueError:
            raise HTTPException(400, "invalid revisit_date; use YYYY-MM-DD")
        sets.append(f"revisit_date = ${i}"); params.append(_rd); i += 1
    if body.owner_email is not None:
        sets.append(f"owner_email = ${i}"); params.append(body.owner_email); i += 1
    if body.opportunity_id is not None:
        sets.append(f"opportunity_id = ${i}"); params.append(body.opportunity_id); i += 1
    if body.not_a_fit_reason is not None:
        sets.append(f"not_a_fit_reason = ${i}"); params.append(body.not_a_fit_reason); i += 1
    if not sets:
        return {"success": True}
    sets.append("updated_at = now()")
    async with conn.transaction():
        try:
            res = await conn.execute(
                f"UPDATE bedrock.jobs_contact_membership SET {', '.join(sets)} WHERE contact_id = $1", *params)
        except asyncpg.exceptions.CheckViolationError:
            # call_booked / revisit until the 2026-08-05 migration enables them.
            # The pickers already grey these out, so this is the belt to that
            # braces — but a raw 500 here told the user nothing.
            raise HTTPException(
                409, f"'{MEMBERSHIP_STAGE_LABELS.get(body.stage, body.stage)}' isn't accepted by "
                     "the database yet — it needs the 2026-08-05 pipeline-stage migration.")
        # Row can vanish between the pre-fetch and the UPDATE (unflag race) —
        # bail before recording a phantom transition.
        if res == "UPDATE 0":
            raise HTTPException(404, "contact is not flagged for jobs")
        if body.stage is not None and body.stage != old_stage and await has_membership_history(conn):
            await conn.execute(
                "INSERT INTO bedrock.jobs_membership_stage_history "
                "(contact_id, from_stage, to_stage, changed_by) VALUES ($1, $2, $3, $4)",
                contact_id, old_stage, body.stage, _user_email(user))
        if body.revisit_date:
            await _ensure_revisit_task(conn, contact_id, body.revisit_date, _user_email(user))
    return {"success": True}


async def _ensure_revisit_task(conn, contact_id: int, when: str, actor: Optional[str]) -> None:
    """Put the revisit on the owner's task list for that date.

    Reuses bedrock.jobs_task, which already drives the Jobs Home Overdue/Today/
    Upcoming widget — so Revisit needs no new surface, it just files a row the
    existing widget already renders. Idempotent per contact: re-setting the date
    moves the open task rather than stacking a second one, otherwise changing your
    mind three times leaves three reminders."""
    row = await conn.fetchrow(
        "SELECT m.owner_email, c.full_name, c.current_company "
        "FROM bedrock.jobs_contact_membership m "
        "JOIN public.contacts c ON c.contact_id = m.contact_id "
        "WHERE m.contact_id = $1", contact_id)
    if row is None:
        return
    who = row["full_name"] or f"contact {contact_id}"
    where = f", {row['current_company']}" if row["current_company"] else ""
    title = f"Revisit: {who}{where}"
    owner = row["owner_email"] or actor
    existing = await conn.fetchval(
        "SELECT id FROM bedrock.jobs_task "
        "WHERE parent_type = 'prospect' AND parent_id = $1 AND deleted_at IS NULL "
        "  AND status <> 'Completed' AND title LIKE 'Revisit:%' LIMIT 1",
        str(contact_id))
    if existing:
        await conn.execute(
            "UPDATE bedrock.jobs_task SET deadline = $2::date, title = $3, updated_at = now() "
            "WHERE id = $1", existing, when, title)
        return
    # status must be one of jobs_task_status_check's values ('todo' is rejected),
    # and owner is NOT NULL DEFAULT '' — an unowned contact would otherwise
    # violate the column rather than simply having no assignee.
    await conn.execute(
        "INSERT INTO bedrock.jobs_task (parent_type, parent_id, title, status, owner, deadline) "
        "VALUES ('prospect', $1, $2, 'Not Started', $3, $4::date)",
        str(contact_id), title, owner or "", when)


@router.delete("/contacts/{contact_id}/jobs-membership")
async def unflag_jobs_contact(contact_id: int, user=Depends(require_auth), conn=Depends(get_db)):
    """Unflag — remove the jobs membership (leaves the legacy is_jobs_contact as-is)."""
    async with conn.transaction():
        old_stage = await conn.fetchval(
            "DELETE FROM bedrock.jobs_contact_membership WHERE contact_id = $1 RETURNING stage", contact_id)
        if old_stage is not None and await has_membership_history(conn):
            await conn.execute(
                "INSERT INTO bedrock.jobs_membership_stage_history "
                "(contact_id, from_stage, to_stage, changed_by) VALUES ($1, $2, 'unflagged', $3)",
                contact_id, old_stage, _user_email(user))
    return {"success": True}


class FlagAccountBody(BaseModel):
    owner_email: Optional[str] = None
    activation_reason: str = "manual"


@router.post("/accounts/{account_key}/flag-contacts")
async def flag_account_contacts(account_key: str, body: FlagAccountBody,
                                user=Depends(require_auth), conn=Depends(get_db)):
    """Flag every (not-yet-flagged) contact at an account — the account-driven side
    of the bi-directional flag."""
    rows = await conn.fetch(
        "SELECT contact_id FROM public.contacts WHERE lower(trim(current_company)) = $1",
        account_key.strip().lower())
    ids = [r["contact_id"] for r in rows]
    n = await _flag_contacts(conn, ids, body.owner_email, body.activation_reason, None, _user_email(user))
    return {"success": True, "data": {"flagged": n}}


@router.get("/account-activity")
async def account_activity(
    key: str = Query(...),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """All engagement activity across an account's opportunities + contacts."""
    k = key.strip().lower()
    opps = await conn.fetch(
        "SELECT id, account_id FROM bedrock.jobs_opportunity WHERE deleted_at IS NULL AND lower(trim(account_name)) = $1",
        k,
    )
    opp_ids = [r["id"] for r in opps]
    account_ids = [r["account_id"] for r in opps if r["account_id"] and str(r["account_id"]).startswith("001")]
    contact_ids = [
        r["contact_id"] for r in await conn.fetch(
            # Include unconfirmed candidates matching the account so their activity
            # rolls up immediately, before anyone reviews them.
            "SELECT contact_id FROM public.contacts "
            "WHERE (is_jobs_contact = true OR contact_stage = 'candidate') "
            "AND lower(trim(current_company)) = $1",
            k,
        )
    ]
    rows = await conn.fetch(
        """
        SELECT a.id, a.type, a.subject, a.description, a.activity_date, a.source, a.logged_by,
               a.synced_at, a.email_from, a.email_to, a.email_snippet,
               left(a.email_body_text, 2000) AS email_body_text,  -- cap body: the rollup of 250 rows was ~1.9MB
               a.meeting_duration_minutes,
               a.jobs_relevance, a.jobs_relevance_override,
               a.jobs_relevance, a.jobs_relevance_override,
            a.jobs_relevance, a.jobs_relevance_override,
            """ + _jobs_activity_flag("a") + """ AS is_jobs,
               a.deleted_at
        FROM bedrock.activity a
        WHERE a.deleted_at IS NULL AND (
            ($1::uuid[] <> '{}' AND a.jobs_opportunity_id = ANY($1::uuid[]))
            OR ($2::text[] <> '{}' AND a.account_id = ANY($2::text[]))
            OR ($3::int[] <> '{}' AND a.participant_public_contact_id = ANY($3::int[]))
        )
        -- Activity should be real mail + calendar only. Exclude Salesforce ToDo
        -- tasks (sf_type='Task') that were imported as meeting/note — they show up
        -- as bogus "calendar" entries. SF Tasks logged as emails/calls are kept
        -- (those are real logged comms); genuine calendar events are sf_type='Event'.
        AND NOT (a.sf_type = 'Task' AND a.type IN ('meeting', 'note'))
        ORDER BY a.activity_date DESC NULLS LAST
        LIMIT 250
        """,
        opp_ids, account_ids, contact_ids,
    )
    return {
        "success": True,
        "data": [
            {**dict(r), "activity_date": r["activity_date"].isoformat() if r["activity_date"] else None,
             "synced_at": r["synced_at"].isoformat() if r["synced_at"] else None,
             "id": str(r["id"]), "email_to": list(r["email_to"]) if r["email_to"] else None}
            for r in rows
        ],
    }


async def _account_opp_contact_ids(conn, key: str):
    """(opp rows [id,title], contact rows [contact_id,full_name]) for an account."""
    k = key.strip().lower()
    opps = await conn.fetch(
        "SELECT id, title FROM bedrock.jobs_opportunity WHERE deleted_at IS NULL AND lower(trim(account_name)) = $1",
        k,
    )
    contacts = await conn.fetch(
        "SELECT contact_id, full_name FROM public.contacts "
        "WHERE (is_jobs_contact = true OR contact_stage = 'candidate') AND lower(trim(current_company)) = $1",
        k,
    )
    return opps, contacts


@router.get("/account-tasks")
async def account_tasks(key: str = Query(...), user=Depends(require_auth), conn=Depends(get_db)):
    """Tasks tagged to any of an account's opportunities or contacts, each
    annotated with what it's tagged to (scope + label)."""
    opps, contacts = await _account_opp_contact_ids(conn, key)
    opp_label = {str(o["id"]): (o["title"] or "Opportunity") for o in opps}
    contact_label = {str(c["contact_id"]): (c["full_name"] or "Contact") for c in contacts}
    rows = await conn.fetch(
        """
        SELECT id, parent_type, parent_id, title, status, deadline, created_at
        FROM bedrock.jobs_task
        WHERE deleted_at IS NULL AND (
            (parent_type = 'opportunity' AND parent_id = ANY($1::text[]))
            OR (parent_type = 'prospect' AND parent_id = ANY($2::text[]))
        )
        ORDER BY (status = 'done'), deadline ASC NULLS LAST, created_at DESC
        """,
        list(opp_label.keys()), list(contact_label.keys()),
    )
    out = []
    for r in rows:
        scope = "opportunity" if r["parent_type"] == "opportunity" else "contact"
        label = (opp_label if scope == "opportunity" else contact_label).get(r["parent_id"], "")
        out.append({
            "id": str(r["id"]), "title": r["title"], "status": r["status"],
            "deadline": r["deadline"].isoformat() if r["deadline"] else None,
            "scope": scope, "parent_id": r["parent_id"], "scope_label": label,
        })
    return {"success": True, "data": out}


@router.get("/account-comments")
async def account_comments(key: str = Query(...), user=Depends(require_auth), conn=Depends(get_db)):
    """Comments across an account's opportunities + contacts (read rollup)."""
    opps, contacts = await _account_opp_contact_ids(conn, key)
    opp_label = {str(o["id"]): (o["title"] or "Opportunity") for o in opps}
    contact_label = {str(c["contact_id"]): (c["full_name"] or "Contact") for c in contacts}
    rows = await conn.fetch(
        """
        SELECT id, parent_type, parent_id, author_email, content, created_at
        FROM bedrock.jobs_comment
        WHERE (parent_type = 'opportunity' AND parent_id = ANY($1::text[]))
           OR (parent_type = 'prospect' AND parent_id = ANY($2::text[]))
        ORDER BY created_at DESC
        """,
        list(opp_label.keys()), list(contact_label.keys()),
    )
    out = []
    for r in rows:
        scope = "opportunity" if r["parent_type"] == "opportunity" else "contact"
        label = (opp_label if scope == "opportunity" else contact_label).get(r["parent_id"], "")
        out.append({
            "id": str(r["id"]), "author_email": r["author_email"], "content": r["content"],
            "created_at": r["created_at"].isoformat() if r["created_at"] else None,
            "scope": scope, "scope_label": label,
        })
    return {"success": True, "data": out}


@router.get("/account-builders")
async def account_builders(key: str = Query(...), user=Depends(require_auth), conn=Depends(get_db)):
    """Builder applications across all of an account's opportunities."""
    k = key.strip().lower()
    rows = await conn.fetch(
        """
        SELECT ja.job_application_id, trim(split_part(ja.notes, ':', 1)) AS builder,
               ja.company_name, ja.role_title, ja.stage, ja.jobs_role_id, ja.date_applied,
               ja.jobs_opportunity_id, o.title AS opp_title
        FROM public.job_applications ja
        JOIN bedrock.jobs_opportunity o ON o.id = ja.jobs_opportunity_id
        WHERE o.deleted_at IS NULL AND lower(trim(o.account_name)) = $1
        ORDER BY ja.date_applied DESC NULLS LAST
        """,
        k,
    )
    summary = {"applied": 0, "interview": 0, "accepted": 0}
    out = []
    for r in rows:
        if r["stage"] in summary:
            summary[r["stage"]] += 1
        out.append({
            "job_application_id": r["job_application_id"], "builder": r["builder"],
            "company_name": r["company_name"], "role_title": r["role_title"], "stage": canon_stage(r["stage"]),
            "jobs_role_id": str(r["jobs_role_id"]) if r["jobs_role_id"] else None,
            "date_applied": r["date_applied"].isoformat() if r["date_applied"] else None,
            "opportunity_id": str(r["jobs_opportunity_id"]) if r["jobs_opportunity_id"] else None,
            "opp_title": r["opp_title"],
        })
    return {"success": True, "data": {"rows": out, "summary": summary}}


@router.get("/account-roles")
async def account_roles(key: str = Query(...), user=Depends(require_auth), conn=Depends(get_db)):
    """Committed roles across all of an account's opportunities."""
    k = key.strip().lower()
    rows = await conn.fetch(
        """
        SELECT r.id, r.opportunity_id, r.title, r.status, r.employment_type,
               r.approx_salary, r.commitment, r.is_trial, r.filled_by_user_id, o.title AS opp_title
        FROM bedrock.jobs_role r
        JOIN bedrock.jobs_opportunity o ON o.id = r.opportunity_id
        WHERE o.deleted_at IS NULL AND lower(trim(o.account_name)) = $1
        ORDER BY r.created_at DESC
        """,
        k,
    )
    return {
        "success": True,
        "data": [
            {"id": str(r["id"]), "opportunity_id": str(r["opportunity_id"]), "opp_title": r["opp_title"],
             "title": r["title"], "status": r["status"], "employment_type": r["employment_type"],
             "approx_salary": r["approx_salary"], "commitment": r["commitment"], "is_trial": r["is_trial"],
             "filled_by_user_id": r["filled_by_user_id"]}
            for r in rows
        ],
    }


@router.get("/contacts")
async def list_contacts(
    stage: Optional[str] = Query(None),
    company: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    flagged: Optional[bool] = Query(None, description="filter to (un)flagged-for-jobs-activation"),
    membership_stage: Optional[str] = Query(None, description="funnel stage on the jobs membership"),
    industry: Optional[str] = Query(None),
    has_open_roles: Optional[bool] = Query(None, description="only contacts whose company has open job postings"),
    scope: str = Query("jobs", pattern="^(jobs|all)$", description="'jobs' = jobs prospects only (default); 'all' = every contact"),
    filter_rules: Optional[str] = Query(None, alias="filters", description="JSON array of {field,op,values} rules — applied in SQL so filters see the full universe, not the loaded page"),
    limit: int = Query(200, le=5000),
    offset: int = Query(0),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Contacts list. Default scope='jobs' keeps today's working view
    (is_jobs_contact=true — kept in sync at opp create/link time so the flag
    alone is filterable without the index-defeating EXISTS). scope='all' opens
    the full contact universe so any contact can be promoted to a jobs
    prospect from the table.
    """
    filters = ["c.is_jobs_contact = true"] if scope == "jobs" else ["true"]
    params: list = []
    i = 1

    if stage:
        filters.append(f"c.contact_stage = ${i}"); params.append(stage); i += 1
    if company:
        filters.append(f"(lower(c.current_company) LIKE lower(${i}) OR lower(c.full_name) LIKE lower(${i}))")
        params.append(f"%{company}%"); i += 1
    if search:
        filters.append(f"(lower(c.full_name) LIKE lower(${i}) OR lower(c.email) LIKE lower(${i}) OR lower(c.current_company) LIKE lower(${i}) OR lower(c.current_title) LIKE lower(${i}))")
        params.append(f"%{search}%"); i += 1
    # Flag / funnel / signal filters — EXISTS subqueries so the count query needs
    # no extra joins.
    if flagged is not None:
        filters.append(("EXISTS" if flagged else "NOT EXISTS")
                       + " (SELECT 1 FROM bedrock.jobs_contact_membership m WHERE m.contact_id = c.contact_id)")
    if membership_stage:
        # Compare canonically so ?membership_stage=revisit still finds the rows
        # stored as on_hold before the migration rewrites them.
        filters.append(
            f"EXISTS (SELECT 1 FROM bedrock.jobs_contact_membership m "
            f"WHERE m.contact_id = c.contact_id AND {canon_membership_sql('m.stage')} = ${i})")
        params.append(canon_membership_stage(membership_stage)); i += 1
    if industry:
        filters.append(f"EXISTS (SELECT 1 FROM public.companies co WHERE co.company_id = c.company_id AND co.industry ILIKE ${i})")
        params.append(f"%{industry}%"); i += 1
    if has_open_roles:
        filters.append("EXISTS (SELECT 1 FROM public.job_postings jp WHERE coalesce(trim(jp.company_name),'') <> '' "
                       "AND lower(trim(jp.company_name)) = lower(trim(c.current_company)))")

    # ── Generic rule filters (the page's Filter menu) ─────────────────────
    # Filtering MUST constrain the SQL: the old client-side approach sifted
    # only the loaded page, so "connected staff contains X" matched 9 of a
    # few hundred. Each UI field maps to a whitelisted SQL expression that
    # mirrors the page's column semantics exactly; unknown fields/ops 400
    # instead of silently under-filtering. The frontend normalizes date
    # presets (in_range → between) before sending.
    _ACT_GATE = "(coalesce(a.jobs_relevance_override, a.jobs_relevance) = 'jobs' OR a.type NOT IN ('email','meeting'))"
    _ACT_LAST = ("(SELECT max(a.activity_date) FROM bedrock.activity a "
                 "WHERE a.participant_public_contact_id = c.contact_id "
                 f"AND a.deleted_at IS NULL AND {_ACT_GATE})")
    _ACT_FIRST = _ACT_LAST.replace("max(", "min(", 1)
    _RULE_FIELDS: dict[str, tuple[str, str]] = {
        "name":     ("text", "c.full_name"),
        "title":    ("text", "c.current_title"),
        "company":  ("text", "c.current_company"),
        "industry": ("text", "(SELECT co2.industry FROM public.companies co2 WHERE co2.company_id = c.company_id)"),
        "stage":    ("select", "c.contact_stage"),
        "owner":    ("select", "c.owner_email"),
        "tags":     ("tags",   "c.tags"),
        "is_jobs":  ("boolean", "(c.is_jobs_contact = true)"),
        "flag":     ("select", "(SELECT m2.stage FROM bedrock.jobs_contact_membership m2 WHERE m2.contact_id = c.contact_id)"),
        "connected":        ("text", "(SELECT string_agg(sm.display_name, ', ') FROM public.staff_contact_relationships scr "
                                     "JOIN bedrock.staff_user_id_map sm ON sm.staff_user_id = scr.staff_user_id "
                                     "WHERE scr.contact_id = c.contact_id)"),
        "connection_count": ("number", "(SELECT count(*) FROM public.staff_contact_relationships scr2 WHERE scr2.contact_id = c.contact_id)"),
        # listings + activity dates resolve through hash JOINs (see _RULE_JOINS):
        # aggregate the source table once, join by key. The correlated
        # per-contact versions re-executed per row across 47k contacts in the
        # COUNT query and wedged the pool (found by filter fuzzing 2026-07-16).
        "listings": ("number", "(coalesce(f_jp.cnt, 0) + coalesce(f_ja.cnt, 0))"),
        "has_deal": ("boolean", "EXISTS (SELECT 1 FROM bedrock.jobs_opportunity o2 WHERE o2.deleted_at IS NULL AND ("
                                "(c.airtable_id IS NOT NULL AND ('airtable:' || c.airtable_id) = ANY(o2.sf_contact_ids)) "
                                "OR ('pub:' || c.contact_id::text) = ANY(o2.sf_contact_ids) "
                                "OR lower(trim(o2.account_name)) = lower(trim(c.current_company))))"),
        "last_activity":      ("recency", "f_act.last_act"),
        "first_contact_date": ("date", "f_act.first_act"),
        "last_contact_date":  ("date", "f_act.last_act"),
    }
    # Joins backing the expressions above — added to BOTH the rows and count
    # queries only when a rule references them.
    _RULE_JOINS: dict[str, str] = {
        "listings": (
            " LEFT JOIN (SELECT lower(trim(company_name)) AS comp, count(*) AS cnt"
            " FROM public.job_postings WHERE coalesce(trim(company_name),'') <> '' GROUP BY 1) f_jp"
            "   ON f_jp.comp = lower(trim(c.current_company))"
            " LEFT JOIN (SELECT lower(trim(company_name)) AS comp, count(DISTINCT lower(trim(role_title))) AS cnt"
            " FROM public.job_applications WHERE coalesce(trim(company_name),'') <> ''"
            " AND coalesce(trim(role_title),'') <> '' GROUP BY 1) f_ja"
            "   ON f_ja.comp = lower(trim(c.current_company))"
        ),
        "last_activity": (
            " LEFT JOIN (SELECT a.participant_public_contact_id AS pid,"
            " min(a.activity_date) AS first_act, max(a.activity_date) AS last_act"
            " FROM bedrock.activity a WHERE a.deleted_at IS NULL AND " + _ACT_GATE +
            " GROUP BY 1) f_act ON f_act.pid = c.contact_id"
        ),
    }
    _RULE_JOINS["first_contact_date"] = _RULE_JOINS["last_activity"]
    _RULE_JOINS["last_contact_date"] = _RULE_JOINS["last_activity"]
    filter_joins: dict[str, str] = {}
    if filter_rules:
        try:
            parsed_rules = json.loads(filter_rules)
            assert isinstance(parsed_rules, list)
        except Exception:
            raise HTTPException(400, "filters must be a JSON array of {field,op,values}")
        for rule in parsed_rules[:20]:
            field, op = rule.get("field"), rule.get("op")
            values = [str(v) for v in (rule.get("values") or [])]
            spec = _RULE_FIELDS.get(field)
            if not spec:
                raise HTTPException(400, f"unfilterable field: {field}")
            if field in _RULE_JOINS:
                filter_joins[_RULE_JOINS[field]] = field
            ftype, expr = spec
            first = values[0] if values else ""
            _TAGS_ANY = ("EXISTS (SELECT 1 FROM unnest(coalesce(c.tags, '{}'::text[])) t "
                         "WHERE t IN (SELECT slug FROM bedrock.contact_tag_catalog))")
            if op == "is_empty":
                if ftype == "tags":
                    filters.append(f"NOT {_TAGS_ANY}")
                else:
                    filters.append(f"({expr} IS NULL OR ({expr})::text = '')" if ftype != "number" else f"coalesce({expr}, 0) = 0")
            elif op == "is_not_empty":
                if ftype == "tags":
                    filters.append(_TAGS_ANY)
                else:
                    filters.append(f"({expr} IS NOT NULL AND ({expr})::text <> '')" if ftype != "number" else f"coalesce({expr}, 0) > 0")
            elif ftype == "tags" and op in ("equals", "contains", "has_any", "not_equals"):
                if not values:
                    continue
                neg = "NOT " if op == "not_equals" else ""
                filters.append(f"{neg}(coalesce(c.tags, '{{}}'::text[]) && ${i}::text[])")
                params.append(values); i += 1
            elif ftype == "select" and op in ("equals", "not_equals"):
                if not values:
                    continue
                neg = "NOT " if op == "not_equals" else ""
                filters.append(f"{neg}(coalesce({expr}, '') = ANY(${i}::text[]))")
                params.append(values); i += 1
            elif ftype == "text" and op in ("contains", "equals", "not_equals"):
                if op == "contains":
                    filters.append(f"lower(coalesce({expr}, '')) LIKE lower(${i})")
                    params.append(f"%{first}%")
                else:
                    filters.append(f"lower(coalesce({expr}, '')) {'=' if op == 'equals' else '<>'} lower(${i})")
                    params.append(first)
                i += 1
            elif ftype == "number" and op in ("gt", "lt", "equals", "not_equals"):
                sym = {"gt": ">", "lt": "<", "equals": "=", "not_equals": "<>"}[op]
                try:
                    num = float(first)
                except ValueError:
                    raise HTTPException(400, f"non-numeric value for {field}")
                filters.append(f"coalesce({expr}, 0) {sym} ${i}")
                params.append(num); i += 1
            elif ftype == "boolean" and op in ("equals", "not_equals"):
                want = (first == "yes") ^ (op == "not_equals")
                filters.append(expr if want else f"NOT {expr}")
            elif ftype == "recency" and op == "within":
                if first == "none":
                    filters.append(f"({expr} IS NULL OR {expr} < now() - interval '90 days')")
                else:
                    try:
                        days = int(first)
                    except ValueError:
                        raise HTTPException(400, "recency window must be a number of days or 'none'")
                    filters.append(f"{expr} >= now() - make_interval(days => ${i})")
                    params.append(days); i += 1
            elif ftype == "date" and op in ("before", "after", "equals", "not_equals", "between"):
                # asyncpg binds ::date params as Python dates, not strings.
                def _d(s: str):
                    try:
                        return date.fromisoformat(s)
                    except ValueError:
                        raise HTTPException(400, f"invalid date for {field}: {s!r}")
                if op == "between":
                    filters.append(f"({expr} >= ${i}::date AND {expr} < ${i+1}::date + 1)")
                    params.append(_d(first)); params.append(_d(values[1] if len(values) > 1 else first)); i += 2
                elif op in ("equals", "not_equals"):
                    filters.append(f"({expr})::date {'=' if op == 'equals' else '<>'} ${i}::date")
                    params.append(_d(first)); i += 1
                else:
                    filters.append(f"{expr} {'<' if op == 'before' else '>'} ${i}::date")
                    params.append(_d(first)); i += 1
            else:
                raise HTTPException(400, f"unsupported op '{op}' for {field}")

    where = " AND ".join(filters)
    filter_join_sql = "".join(filter_joins.keys()) if filter_rules else ""
    # A pathological filter must never wedge the shared pool: cap this
    # request's statement time (the pool resets session state on release).
    if filter_rules:
        await conn.execute("SET statement_timeout = '15000'")
    # When the contact entered its CURRENT membership stage. Real answer comes
    # from the stage-history table; until that migration lands, degrade to
    # assigned_at (row creation — accurate only for never-restaged contacts).
    # Per-stage entry time. Falls back to the stage's OWN stamp before
    # assigned_at: a contact in initial_outreach entered that stage at
    # first_outreach_at, and using assigned_at made "contacted this week" read 0
    # while 6 contacts had actually been reached (2026-07-30).
    _stamp_fallback = ("CASE m.stage"
                       " WHEN 'initial_outreach' THEN m.first_outreach_at"
                       " WHEN 'converted_to_opportunity' THEN m.converted_at"
                       " END")   # call_booked_at joins this once the column exists
    entered_expr = (
        f"COALESCE((SELECT max(h.changed_at) FROM bedrock.jobs_membership_stage_history h"
        f" WHERE h.contact_id = c.contact_id AND h.to_stage = m.stage), {_stamp_fallback}, m.assigned_at)"
        if await has_membership_history(conn)
        else f"COALESCE({_stamp_fallback}, m.assigned_at)")
    rows = await conn.fetch(
        f"""
        SELECT
            c.contact_id,
            c.full_name,
            c.first_name,
            c.last_name,
            c.email,
            c.current_title,
            c.current_company,
            c.contact_stage,
            c.linkedin_url,
            c.airtable_id,
            c.is_jobs_contact,
            c.owner_email,
            -- user-facing tags only: intersection with the curated catalog, so
            -- system markers (email_review) never reach the UI
            ARRAY(SELECT t FROM unnest(coalesce(c.tags, '{{}}'::text[])) t
                  WHERE t IN (SELECT slug FROM bedrock.contact_tag_catalog)) AS crm_tags,
            -- linked deal via sf_contact_ids
            jo.id           AS deal_id,
            jo.account_name AS deal_account,
            jo.stage        AS deal_stage,
            -- OR via company name match (fallback)
            jo2.id           AS deal_id_by_company,
            jo2.account_name AS deal_account_by_company,
            jo2.stage        AS deal_stage_by_company,
            -- jobs activation membership (the flag + funnel)
            -- Canonical: 'on_hold' rows read as 'revisit' until the migration
            -- rewrites them, so the UI never shows a stage it no longer offers.
            {canon_membership_sql('m.stage')} AS membership_stage,
            m.owner_email    AS membership_owner,
            m.first_outreach_by AS first_outreach_by,
            {entered_expr}   AS membership_stage_entered_at,
            -- signals for triage
            co.industry      AS company_industry,
            (SELECT count(*) FROM public.job_postings jp
               WHERE coalesce(trim(jp.company_name),'') <> ''
                 AND lower(trim(jp.company_name)) = lower(trim(c.current_company))) AS open_roles,
            -- DISTINCT roles builders have applied to at this company (not one per
            -- applicant — 8 builders on the same role is 1 listing).
            (SELECT count(DISTINCT lower(trim(ja.role_title))) FROM public.job_applications ja
               WHERE coalesce(trim(ja.company_name),'') <> ''
                 AND lower(trim(ja.company_name)) = lower(trim(c.current_company))
                 AND coalesce(trim(ja.role_title),'') <> '') AS builder_apps
        FROM public.contacts c
        LEFT JOIN bedrock.jobs_contact_membership m ON m.contact_id = c.contact_id
        LEFT JOIN public.companies co ON co.company_id = c.company_id
        -- direct link via sf_contact_ids (airtable: or pub: ref).
        -- LATERAL + LIMIT 1: a contact tied to N opportunities must still be
        -- ONE row — pick the best deal (active first, then freshest) instead
        -- of fanning out (Emily Zhao appeared 8× for 8 deals at her company).
        LEFT JOIN LATERAL (
            SELECT o.id, o.account_name, o.stage
            FROM bedrock.jobs_opportunity o
            WHERE o.deleted_at IS NULL
              AND (
                (c.airtable_id IS NOT NULL AND ('airtable:' || c.airtable_id) = ANY(o.sf_contact_ids))
                OR ('pub:' || c.contact_id::text) = ANY(o.sf_contact_ids)
              )
            ORDER BY (o.stage LIKE 'active%') DESC, o.updated_at DESC NULLS LAST
            LIMIT 1
        ) jo ON true
        -- company fallback (only when no direct link): EXACT current-company
        -- match only. Substring matching linked contacts to every deal whose
        -- account name overlapped their company string — a contact's implied
        -- deal must come from their current employer, nothing looser.
        LEFT JOIN LATERAL (
            SELECT o.id, o.account_name, o.stage
            FROM bedrock.jobs_opportunity o
            WHERE jo.id IS NULL
              AND o.deleted_at IS NULL
              AND lower(trim(o.account_name)) = lower(trim(c.current_company))
            ORDER BY (o.stage LIKE 'active%') DESC, o.updated_at DESC NULLS LAST
            LIMIT 1
        ) jo2 ON true
        {filter_join_sql}
        WHERE {where}
        -- Pipeline-first: contacts with a jobs stage sort above the (large)
        -- flagged-but-unstaged population, so every pipeline stage is visible
        -- without paging through ~47k blank prospects. Alphabetical within each.
        ORDER BY (m.stage IS NOT NULL) DESC, c.full_name NULLS LAST
        LIMIT ${i} OFFSET ${i+1}
        """,
        *params, limit, offset,
    )
    total = await conn.fetchval(
        f"SELECT count(*) FROM public.contacts c{filter_join_sql} WHERE {where}", *params
    )
    if filter_rules:
        await conn.execute("RESET statement_timeout")

    # Batch-resolve LinkedIn-connected staff names for all returned contacts so
    # the contacts list can show who on the team knows each person.
    contact_ids = [r["contact_id"] for r in rows]
    staff_by_contact: dict[int, list[str]] = {}
    if contact_ids:
        srows = await conn.fetch(
            """
            SELECT scr.contact_id, m.display_name
            FROM public.staff_contact_relationships scr
            JOIN bedrock.staff_user_id_map m ON m.staff_user_id = scr.staff_user_id
            WHERE scr.contact_id = ANY($1::int[]) AND m.display_name IS NOT NULL
            ORDER BY m.display_name
            """,
            contact_ids,
        )
        for s in srows:
            staff_by_contact.setdefault(s["contact_id"], []).append(s["display_name"])

    # Per-contact activity for warmth: recent volume (90d), recency (last touch),
    # and whether they've RESPONDED (a meeting/call, or an inbound email — not
    # just our outbound).
    team_sender = " OR ".join(f"a.email_from ILIKE '%{e}%'" for e in JOBS_TEAM_EMAILS)
    actor_case = "CASE " + " ".join(
        f"WHEN a.email_from ILIKE '%{e}%' OR a.logged_by ILIKE '%{e}%' THEN '{e}'"
        for e in JOBS_TEAM_EMAILS
    ) + " END"
    activity_by_contact: dict[int, dict] = {}
    if contact_ids:
        arows = await conn.fetch(
            f"""
            SELECT a.participant_public_contact_id AS cid,
                   count(*) FILTER (WHERE a.activity_date >= now() - interval '90 days') AS recent,
                   max(a.activity_date) AS last_act,
                   min(a.activity_date) AS first_act,
                   bool_or(a.type IN ('call','meeting')
                           OR (a.type = 'email' AND NOT ({team_sender}))) AS responded,
                   array_agg(DISTINCT {actor_case}) AS actors
            FROM bedrock.activity a
            WHERE a.deleted_at IS NULL AND a.participant_public_contact_id = ANY($1::int[])
              AND {_jobs_relevant('a')}
            GROUP BY a.participant_public_contact_id
            """,
            contact_ids,
        )
        activity_by_contact = {a["cid"]: {"recent": a["recent"], "last": a["last_act"],
                                          "first": a["first_act"],
                                          "responded": a["responded"],
                                          "actors": sorted(x for x in (a["actors"] or []) if x)} for a in arows}

    # Open tasks per contact (parent_type='prospect').
    tasks_by_contact: dict[int, int] = {}
    if contact_ids:
        trows = await conn.fetch(
            "SELECT parent_id, count(*) AS n FROM bedrock.jobs_task "
            "WHERE parent_type='prospect' AND deleted_at IS NULL AND status <> 'Completed' "
            "AND parent_id = ANY($1::text[]) GROUP BY parent_id",
            [str(cid) for cid in contact_ids],
        )
        tasks_by_contact = {int(t["parent_id"]): t["n"] for t in trows}

    return {
        "success": True,
        "total": total,
        "data": [
            {
                **{k: v for k, v in dict(r).items() if not k.startswith("deal_")},
                "deal": (
                    {"id": str(r["deal_id"]), "account_name": r["deal_account"], "stage": r["deal_stage"]}
                    if r["deal_id"] else
                    {"id": str(r["deal_id_by_company"]), "account_name": r["deal_account_by_company"], "stage": r["deal_stage_by_company"]}
                    if r["deal_id_by_company"] else None
                ),
                "connected_staff_names": staff_by_contact.get(r["contact_id"], []),
                "recent_activity_count": (activity_by_contact.get(r["contact_id"]) or {}).get("recent", 0),
                "last_activity_at": (lambda v: v.isoformat() if v else None)((activity_by_contact.get(r["contact_id"]) or {}).get("last")),
                "first_activity_at": (lambda v: v.isoformat() if v else None)((activity_by_contact.get(r["contact_id"]) or {}).get("first")),
                "responded": bool((activity_by_contact.get(r["contact_id"]) or {}).get("responded")),
                "activity_actors": (activity_by_contact.get(r["contact_id"]) or {}).get("actors", []),
                "open_tasks": tasks_by_contact.get(r["contact_id"], 0),
            }
            for r in rows
        ],
    }


@router.get("/contacts/{contact_id}/opportunities")
async def contact_opportunities(
    contact_id: int,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Opportunities this contact is attached to — directly (sf_contact_ids) or
    via a company-name match to the opportunity's account."""
    rows = await conn.fetch(
        """
        WITH c AS (
            SELECT contact_id, airtable_id, current_company
            FROM public.contacts WHERE contact_id = $1
        )
        SELECT o.id, o.account_name, o.title, o.stage, o.deal_type,
               o.owner_email, o.num_roles, o.priority, o.updated_at
        FROM bedrock.jobs_opportunity o, c
        WHERE o.deleted_at IS NULL AND (
            ('pub:' || c.contact_id::text) = ANY(o.sf_contact_ids)
            OR (c.airtable_id IS NOT NULL AND ('airtable:' || c.airtable_id) = ANY(o.sf_contact_ids))
            OR (coalesce(trim(c.current_company), '') <> ''
                AND lower(trim(o.account_name)) = lower(trim(c.current_company)))
        )
        ORDER BY o.updated_at DESC NULLS LAST
        """,
        contact_id,
    )
    return {
        "success": True,
        "data": [
            {
                "id": str(r["id"]),
                "account_name": r["account_name"],
                "title": r["title"],
                "stage": canon_stage(r["stage"]),
                "deal_type": r["deal_type"],
                "owner_email": r["owner_email"],
                "num_roles": r["num_roles"],
                "priority": r["priority"],
                "updated_at": r["updated_at"].isoformat() if r["updated_at"] else None,
            }
            for r in rows
        ],
    }


@router.get("/contacts/{contact_id}")
async def get_contact(
    contact_id: int,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Full contact detail — info, linked deal, and all engagement activity."""
    row = await conn.fetchrow(
        """
        SELECT c.*,
            jo.id           AS deal_id,
            jo.account_name AS deal_account,
            jo.stage        AS deal_stage,
            jo.owner_email  AS deal_owner,
            jo2.id           AS deal_id2,
            jo2.account_name AS deal_account2,
            jo2.stage        AS deal_stage2
        FROM public.contacts c
        LEFT JOIN bedrock.jobs_opportunity jo
            ON jo.deleted_at IS NULL
            AND ('airtable:' || c.airtable_id) = ANY(jo.sf_contact_ids)
        LEFT JOIN bedrock.jobs_opportunity jo2
            ON jo2.deleted_at IS NULL AND jo.id IS NULL
            -- exact current-company match only (see list_contacts note)
            AND lower(trim(jo2.account_name)) = lower(trim(c.current_company))
        WHERE c.contact_id = $1
        """,
        contact_id,
    )
    if not row:
        raise HTTPException(404, "Contact not found")

    deal_id    = row["deal_id"] or row["deal_id2"]
    deal_id2   = row["deal_id2"]
    contact_email = row["email"]
    contact_name  = row["full_name"] or ""
    first_name    = contact_name.split()[0] if contact_name else ""

    # Activity the contact was ACTUALLY a participant in — set by the nightly
    # relink (participant_public_contact_id) or their email being on the thread
    # (from / to / cc). We deliberately do NOT pull deal- or account-level
    # activity: a contact at a company must not inherit calls/emails they
    # weren't part of (that previously showed e.g. every Adonis call on one
    # Adonis contact). first_name fuzzy-matching is dropped for the same reason.
    rows_act = await conn.fetch(
        """
        SELECT a.id, a.type, a.subject, a.description, a.activity_date,
               a.logged_by, a.source, a.email_from, a.email_snippet,
               a.meeting_duration_minutes, a.deleted_at,
               a.jobs_relevance, a.jobs_relevance_override,
               """ + _jobs_activity_flag("a") + """ AS is_jobs
        FROM bedrock.activity a
        WHERE a.deleted_at IS NULL
          AND (
            a.participant_public_contact_id = $1
            OR (
              $2 <> '' AND (
                lower(a.email_from) LIKE '%' || lower($2) || '%'
                OR EXISTS (
                  SELECT 1 FROM unnest(coalesce(a.email_to, '{}') || coalesce(a.email_cc, '{}')) e
                  WHERE lower(e) = lower($2)
                )
              )
            )
          )
        ORDER BY a.activity_date DESC NULLS LAST LIMIT 100
        """,
        contact_id,
        contact_email or "",
    )
    all_activity: list = [dict(r) for r in rows_act]

    all_activity.sort(key=lambda x: x.get("activity_date") or "", reverse=True)
    activity = all_activity[:150]

    # Who on staff is connected to this contact — from LinkedIn connections
    # (public.staff_contact_relationships), resolved to names via
    # bedrock.staff_user_id_map. (The map may be sparsely populated; unresolved
    # staff are returned with name=null so the UI can choose to hide them.)
    conn_rows = await conn.fetch(
        """
        SELECT scr.staff_user_id,
               m.display_name,
               m.email,
               scr.source,
               scr.relationship_strength,
               scr.connected_date
        FROM public.staff_contact_relationships scr
        LEFT JOIN bedrock.staff_user_id_map m ON m.staff_user_id = scr.staff_user_id
        WHERE scr.contact_id = $1
        ORDER BY (m.display_name IS NULL), m.display_name
        """,
        contact_id,
    )
    connected_staff = [
        {
            "staff_user_id":  r["staff_user_id"],
            "name":           r["display_name"],
            "email":          r["email"],
            "source":         r["source"],
            "strength":       r["relationship_strength"],
            "connected_date": r["connected_date"].isoformat() if r["connected_date"] else None,
        }
        for r in conn_rows
    ]

    # Open roles the team has sourced at this contact's company (what the
    # "open roles" signal on the list counts — here we return the actual rows).
    open_roles: list = []
    if row["current_company"]:
        pr = await conn.fetch(
            """
            SELECT id, job_title, job_url, status, source, salary_range, location,
                   aligned_sector, builder_interest_count, created_at
            FROM public.job_postings
            WHERE coalesce(trim(company_name), '') <> ''
              AND lower(trim(company_name)) = lower(trim($1))
            ORDER BY created_at DESC NULLS LAST
            """,
            row["current_company"],
        )
        open_roles = [dict(r) for r in pr]

    # Jobs builders have applied to at this company — self-directed (no team opp)
    # or team-linked. A strong signal the company is already a builder target.
    builder_apps: list = []
    if row["current_company"]:
        ba = await conn.fetch(
            """
            SELECT coalesce(nullif(trim(role_title), ''), 'Unspecified role') AS role_title,
                   count(DISTINCT builder_id)                                  AS applicant_count,
                   array_agg(DISTINCT stage) FILTER (WHERE stage IS NOT NULL)  AS stages,
                   bool_or(jobs_opportunity_id IS NOT NULL)                    AS team_linked,
                   max(date_applied)                                          AS last_applied,
                   (array_agg(source_type ORDER BY date_applied DESC NULLS LAST))[1] AS source_type
            FROM public.job_applications
            WHERE coalesce(trim(company_name), '') <> ''
              AND lower(trim(company_name)) = lower(trim($1))
            GROUP BY 1
            ORDER BY count(DISTINCT builder_id) DESC, max(date_applied) DESC NULLS LAST
            """,
            row["current_company"],
        )
        builder_apps = [dict(r) for r in ba]

    # Membership (jobs activation flag + funnel), if any.
    mem = await conn.fetchrow(
        "SELECT stage, owner_email, first_outreach_by FROM bedrock.jobs_contact_membership WHERE contact_id = $1",
        contact_id)

    deal = None
    if row["deal_id"]:
        deal = {"id": str(row["deal_id"]), "account_name": row["deal_account"], "stage": row["deal_stage"], "owner_email": row["deal_owner"]}
    elif row["deal_id2"]:
        deal = {"id": str(row["deal_id2"]), "account_name": row["deal_account2"], "stage": row["deal_stage2"], "owner_email": None}

    return {
        "success": True,
        "data": {
            "contact_id":      row["contact_id"],
            "full_name":       row["full_name"],
            "first_name":      row["first_name"],
            "last_name":       row["last_name"],
            "email":           row["email"],
            "current_title":   row["current_title"],
            "current_company": row["current_company"],
            "contact_stage":   row["contact_stage"],
            "linkedin_url":    row["linkedin_url"],
            "airtable_id":     row["airtable_id"],
            "deal":            deal,
            "activity":        [dict(a) for a in activity],
            "connected_staff": connected_staff,
            "open_roles_list": open_roles,
            "builder_applications": builder_apps,
            "membership_stage": canon_membership_stage(mem["stage"]) if mem else None,
            "membership_owner": mem["owner_email"] if mem else None,
            "first_outreach_by": mem["first_outreach_by"] if mem else None,
        },
    }


CONTACT_SELECT = """
    SELECT contact_id, first_name, last_name, full_name, email,
           current_title, current_company, contact_stage, linkedin_url, source, airtable_id
"""

async def _resolve_contacts(conn, sf_contact_ids: list[str]) -> list[dict]:
    """Resolve contact refs to public.contacts records.

    Supported ref formats:
      airtable:{airtable_id}  — from Airtable employer contacts
      pub:{contact_id}        — any public.contacts row by PK
      {sf_contact_id}         — 15/18-char SF contact ID via sf_contact_link
    """
    if not sf_contact_ids:
        return []

    airtable_ids = [r.replace("airtable:", "") for r in sf_contact_ids if r.startswith("airtable:")]
    # skip malformed pub refs (e.g. 'pub:abc') so one bad ref can't 500 the whole load
    pub_ids      = [int(r[4:]) for r in sf_contact_ids if r.startswith("pub:") and r[4:].isdigit()]
    sf_ids       = [r for r in sf_contact_ids if not r.startswith("airtable:") and not r.startswith("pub:")]

    contacts = []
    if airtable_ids:
        rows = await conn.fetch(
            CONTACT_SELECT + "FROM public.contacts WHERE airtable_id = ANY($1::text[])",
            airtable_ids,
        )
        contacts.extend([dict(r) for r in rows])
    if pub_ids:
        rows = await conn.fetch(
            CONTACT_SELECT + "FROM public.contacts WHERE contact_id = ANY($1::int[])",
            pub_ids,
        )
        contacts.extend([dict(r) for r in rows])
    if sf_ids:
        rows = await conn.fetch(
            CONTACT_SELECT + """FROM public.contacts c
               JOIN bedrock.sf_contact_link scl ON scl.public_contact_id = c.contact_id
               WHERE scl.sf_contact_id = ANY($1::text[])""",
            sf_ids,
        )
        contacts.extend([dict(r) for r in rows])
    return contacts


async def _flag_jobs_contacts(conn, sf_contact_ids: list[str]) -> None:
    """Mark contacts linked to an opp as is_jobs_contact=true so they keep
    showing up in GET /contacts (which now filters on the flag alone)."""
    if not sf_contact_ids:
        return
    airtable_ids = [r[len("airtable:"):] for r in sf_contact_ids if r.startswith("airtable:")]
    pub_ids      = [int(r[4:]) for r in sf_contact_ids if r.startswith("pub:") and r[4:].isdigit()]
    if airtable_ids:
        await conn.execute(
            "UPDATE public.contacts SET is_jobs_contact=true, updated_at=now() "
            "WHERE airtable_id = ANY($1::text[]) AND NOT is_jobs_contact",
            airtable_ids,
        )
    if pub_ids:
        await conn.execute(
            "UPDATE public.contacts SET is_jobs_contact=true, updated_at=now() "
            "WHERE contact_id = ANY($1::int[]) AND NOT is_jobs_contact",
            pub_ids,
        )


@router.get("/opportunities/pipeline")
async def get_pipeline_summary(user=Depends(require_auth), conn=Depends(get_db)):
    """Stage counts + deal-type breakdown for the pipeline dashboard."""
    rows = await conn.fetch(
        """
        SELECT
            stage,
            deal_type,
            count(*) AS count,
            avg(salary_expected) FILTER (WHERE salary_expected IS NOT NULL) AS avg_salary
        FROM bedrock.jobs_opportunity
        WHERE deleted_at IS NULL
        GROUP BY stage, deal_type
        ORDER BY stage, deal_type
        """
    )
    # Group into stage buckets
    pipeline = {}
    for r in rows:
        s = r["stage"]
        if s not in pipeline:
            pipeline[s] = {
                "stage": s,
                "label": STAGE_LABELS.get(s, s),
                "group": s.split("_")[0],
                "total": 0,
                "by_type": {},
                "avg_salary": None,
            }
        pipeline[s]["total"] += r["count"]
        if r["deal_type"]:
            pipeline[s]["by_type"][r["deal_type"]] = r["count"]
        if r["avg_salary"]:
            pipeline[s]["avg_salary"] = round(float(r["avg_salary"]))

    return {"success": True, "data": list(pipeline.values())}


def _norm_opp(d: dict) -> dict:
    """Guarantee array fields are never None (the columns allow NULL, which
    would crash the frontend pickers that call .map/.length on them), and hand
    the frontend a CANONICAL stage.

    Every opportunity the API returns passes through here, so canonicalising in
    one place is what keeps a row's stage chip, its board column and its status
    roll-up agreeing with each other before the migration lands."""
    d["builder_ids"] = d.get("builder_ids") or []
    d["sf_contact_ids"] = d.get("sf_contact_ids") or []
    d["tags"] = d.get("tags") or []
    if d.get("stage"):
        d["stage"] = canon_stage(d["stage"])
    return d


@router.get("/opportunities")
async def list_opportunities(
    stage: Optional[str] = Query(None),
    stage_group: Optional[str] = Query(None, description="active | on_hold | closed"),
    owner_email: Optional[str] = Query(None),
    account_id: Optional[str] = Query(None),
    deal_type: Optional[str] = Query(None),
    limit: int = Query(200, le=500),
    offset: int = Query(0),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    filters = ["o.deleted_at IS NULL"]
    params: list = []
    i = 1

    if stage:
        filters.append(f"o.stage = ${i}"); params.append(stage); i += 1
    elif stage_group:
        filters.append(f"o.stage LIKE ${i}"); params.append(f"{stage_group}%"); i += 1

    if owner_email:
        filters.append(f"o.owner_email = ${i}"); params.append(owner_email); i += 1
    if account_id:
        filters.append(f"o.account_id = ${i}"); params.append(account_id); i += 1
    if deal_type:
        filters.append(f"o.deal_type = ${i}"); params.append(deal_type); i += 1

    where = " AND ".join(filters)
    rows = await conn.fetch(
        f"""
        SELECT
            o.*,
            act.activity_count,
            act.last_activity_at,
            act.recent_activity_count,
            (SELECT count(*) FROM bedrock.jobs_task t
               WHERE t.parent_type='opportunity' AND t.parent_id = o.id::text
                 AND t.deleted_at IS NULL AND t.status <> 'Completed') AS open_tasks,
            -- Suggested priority (1–5, 5 = highest) the team can override. Bumped by
            -- signals: committed roles, multiple contacts, recent activity, and
            -- builders already applying. AI-first scoring can replace this later.
            LEAST(5, 1
                + (COALESCE(o.num_roles, 0) > 0)::int
                + (COALESCE(array_length(o.sf_contact_ids, 1), 0) > 1)::int
                + (act.recent_activity_count > 0)::int
                + (EXISTS (SELECT 1 FROM public.job_applications ja WHERE ja.jobs_opportunity_id = o.id))::int
            ) AS priority_suggested
        FROM bedrock.jobs_opportunity o
        LEFT JOIN LATERAL (
            -- Same scope as the detail Activity tab: deal-tagged OR company-level
            -- activity. Powers the row "recent activity" indicator so the team can
            -- see at a glance which accounts moved this week.
            SELECT
                count(*)                                              AS activity_count,
                max(a.activity_date)                                  AS last_activity_at,
                count(*) FILTER (
                    WHERE a.activity_date >= now() - interval '7 days'
                )                                                     AS recent_activity_count
            FROM bedrock.activity a
            WHERE a.deleted_at IS NULL
              AND {_jobs_relevant('a')}
              AND (
                a.jobs_opportunity_id = o.id
                OR (o.account_id <> 'UNKNOWN' AND a.account_id = o.account_id)
              )
        ) act ON true
        WHERE {where}
        ORDER BY o.updated_at DESC
        LIMIT ${i} OFFSET ${i+1}
        """,
        *params, limit, offset,
    )
    total = await conn.fetchval(
        f"SELECT count(*) FROM bedrock.jobs_opportunity o WHERE {where}", *params
    )
    return {"success": True, "data": [_norm_opp(dict(r)) for r in rows], "total": total}


@router.get("/opportunities/search")
async def search_opportunities(
    q: str = Query(..., min_length=1),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Lightweight opportunity picker (by employer/account name or deal title) —
    used by the Roles board's Add Role flow to attach a new role to an existing
    opportunity instead of always creating one from scratch."""
    # Escape LIKE metacharacters in the raw query — unescaped, a search for
    # e.g. "50%" or "a_b" is interpreted as a wildcard pattern instead of a
    # literal string, matching far more (or differently) than intended.
    escaped = q.strip().lower().replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
    like = f"%{escaped}%"
    rows = await conn.fetch(
        """
        SELECT id, account_name, title, stage
        FROM bedrock.jobs_opportunity
        WHERE deleted_at IS NULL
          AND (lower(trim(account_name)) LIKE $1 OR lower(trim(title)) LIKE $1)
        ORDER BY updated_at DESC
        LIMIT 20
        """,
        like,
    )
    return {"success": True, "data": [
        {"id": str(r["id"]), "account_name": r["account_name"], "title": r["title"], "stage": r["stage"]}
        for r in rows
    ]}


@router.post("/opportunities")
async def create_opportunity(
    body: OpportunityCreate,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    if body.stage not in VALID_STAGES:
        raise HTTPException(400, f"Invalid stage: {body.stage}")
    if body.deal_type and body.deal_type not in VALID_DEAL_TYPES:
        raise HTTPException(400, f"Invalid deal_type: {body.deal_type}")
    if body.likelihood and body.likelihood not in VALID_LIKELIHOODS:
        raise HTTPException(400, f"Invalid likelihood: {body.likelihood}")

    user_email = user.get("email") if isinstance(user, dict) else getattr(user, "email", None)

    async with conn.transaction():
        row_id = await conn.fetchval(
            """
            INSERT INTO bedrock.jobs_opportunity (
                account_id, account_name, stage, deal_type,
                title, description, salary_expected, num_roles, likelihood,
                source, owner_email, relationship_owner, sf_contact_ids, builder_ids, follow_up_date
            ) VALUES ($1,$2,$3,$4,$5,$6,$7,$8,$9,$10,$11,$12,$13,$14,$15)
            RETURNING id
            """,
            body.account_id, body.account_name, body.stage, body.deal_type,
            body.title, body.description, body.salary_expected, body.num_roles, body.likelihood,
            body.source, body.owner_email, body.relationship_owner,
            body.sf_contact_ids, body.builder_ids, body.follow_up_date,
        )
        await conn.execute(
            """
            INSERT INTO bedrock.jobs_stage_history
                (opportunity_id, from_stage, to_stage, changed_by, note)
            VALUES ($1, NULL, $2, $3, $4)
            """,
            row_id, body.stage, user_email, body.note,
        )
        await _flag_jobs_contacts(conn, list(body.sf_contact_ids or []))

    row = await conn.fetchrow(
        "SELECT * FROM bedrock.jobs_opportunity WHERE id=$1", row_id
    )
    return {"success": True, "data": dict(row)}


@router.get("/opportunities/{opp_id}")
async def get_opportunity(
    opp_id: UUID,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    row = await conn.fetchrow(
        "SELECT * FROM bedrock.jobs_opportunity WHERE id=$1 AND deleted_at IS NULL",
        opp_id,
    )
    if not row:
        raise HTTPException(404, "Opportunity not found")

    history = await conn.fetch(
        """
        SELECT * FROM bedrock.jobs_stage_history
        WHERE opportunity_id=$1 ORDER BY changed_at DESC
        """,
        opp_id,
    )
    contacts = await _resolve_contacts(conn, list(row["sf_contact_ids"] or []))

    account_id = row["account_id"]

    # All activity for this account (Gmail, Calendar, SF, manual) + jobs-tagged
    activity = await conn.fetch(
        """
        SELECT
            a.id, a.type, a.subject, a.description, a.activity_date,
            a.source, a.logged_by, a.synced_at, a.email_from, a.email_to,
            a.email_snippet, a.email_body_text,
            a.meeting_duration_minutes, a.meeting_attendees, a.deleted_at,
            """ + _jobs_activity_flag("a") + """ AS is_jobs
        FROM bedrock.activity a
        WHERE a.deleted_at IS NULL
          AND (
            a.jobs_opportunity_id = $1
            OR (a.account_id = $2 AND $2 != 'UNKNOWN')
          )
        ORDER BY a.activity_date DESC NULLS LAST
        LIMIT 250
        """,
        opp_id,
        account_id,
    )
    return {
        "success": True,
        "data": {
            **_norm_opp(dict(row)),
            "stage_history": [dict(h) for h in history],
            "activity":      [dict(a) for a in activity],
            "contacts":      contacts,
        },
    }


@router.patch("/opportunities/{opp_id}")
async def update_opportunity(
    opp_id: UUID,
    body: OpportunityUpdate,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    existing = await conn.fetchrow(
        "SELECT * FROM bedrock.jobs_opportunity WHERE id=$1 AND deleted_at IS NULL", opp_id
    )
    if not existing:
        raise HTTPException(404, "Opportunity not found")

    if body.stage and body.stage not in VALID_STAGES:
        raise HTTPException(400, f"Invalid stage: {body.stage}")
    if body.deal_type and body.deal_type not in VALID_DEAL_TYPES:
        raise HTTPException(400, f"Invalid deal_type: {body.deal_type}")
    if body.likelihood and body.likelihood not in VALID_LIKELIHOODS:
        raise HTTPException(400, f"Invalid likelihood: {body.likelihood}")
    if body.priority is not None and not (1 <= body.priority <= 5):
        raise HTTPException(400, "priority must be between 1 and 5")

    user_email = user.get("email") if isinstance(user, dict) else getattr(user, "email", None)
    stage_changed = body.stage and body.stage != existing["stage"]

    sets = []
    params: list = []
    i = 1

    # An explicitly-sent null must CLEAR the column (e.g. unassigning an owner),
    # so key off model_fields_set rather than `val is not None` — otherwise a
    # null is indistinguishable from "field omitted" and is silently dropped.
    # `stage` is NOT NULL, so it can never be cleared this way.
    fields_set = body.model_fields_set
    for field in ("stage", "deal_type", "title", "description", "salary_expected",
                  "num_roles", "likelihood",
                  "source", "owner_email", "relationship_owner", "sf_contact_ids", "builder_ids",
                  "follow_up_date", "target_close_date", "touch_count", "sf_opportunity_id",
                  "closed_lost_reason", "closed_lost_note", "priority", "segment", "intro_by",
                  "tags"):
        if field not in fields_set:
            continue
        if field == "tags" and not await _has_column("bedrock", "jobs_opportunity", "tags"):
            continue   # pre-migration: silently no-op rather than fail the save
        val = getattr(body, field, None)
        if val is None and field == "stage":
            continue
        sets.append(f"{field} = ${i}"); params.append(val); i += 1

    # Auto-set closed_at
    if body.stage in ("closed_won", "closed_lost") and not existing["closed_at"]:
        sets.append(f"closed_at = ${i}"); params.append(datetime.now(timezone.utc)); i += 1
    elif body.stage and not body.stage.startswith("closed_") and existing["closed_at"]:
        sets.append(f"closed_at = ${i}"); params.append(None); i += 1

    if not sets:
        return {"success": True, "data": dict(existing)}

    params.append(opp_id)
    async with conn.transaction():
        try:
            await conn.execute(
                f"UPDATE bedrock.jobs_opportunity SET {', '.join(sets)} WHERE id=${i}",
                *params,
            )
        except asyncpg.exceptions.CheckViolationError as e:
            # Almost always a stage the 2026-08-05 migration hasn't enabled yet
            # (reviewing_builders is the only new value). Raw, this surfaced as a
            # 500 with no hint of what to do about it.
            if "stage" in str(e):
                raise HTTPException(
                    409, f"'{STAGE_LABELS.get(body.stage, body.stage)}' isn't accepted by the "
                         "database yet — it needs the 2026-08-05 pipeline-stage migration.")
            raise HTTPException(400, f"That value isn't allowed: {e}")
        if stage_changed:
            await conn.execute(
                """
                INSERT INTO bedrock.jobs_stage_history
                    (opportunity_id, from_stage, to_stage, changed_by, note)
                VALUES ($1,$2,$3,$4,$5)
                """,
                opp_id, existing["stage"], body.stage, user_email, body.note,
            )
        if body.sf_contact_ids is not None:
            await _flag_jobs_contacts(conn, list(body.sf_contact_ids or []))
        # closed_lost + reason=revisit is "come back to this", not an ending, so
        # it has to leave something behind that resurfaces. The date rides on the
        # existing follow_up_date column; this turns it into a task for the
        # deal's owner, which the Jobs Home widget already renders.
        if body.closed_lost_reason == "revisit" and body.follow_up_date:
            await _ensure_opp_revisit_task(
                conn, str(opp_id), existing["account_name"],
                body.owner_email or existing["owner_email"] or user_email,
                body.follow_up_date)

    row = await conn.fetchrow("SELECT * FROM bedrock.jobs_opportunity WHERE id=$1", opp_id)
    return {"success": True, "data": dict(row)}


async def _ensure_opp_revisit_task(conn, opp_id: str, account: Optional[str],
                                   owner: Optional[str], when) -> None:
    """File (or move) the revisit reminder for a closed-lost-but-come-back deal.

    Mirrors _ensure_revisit_task on the contact side: idempotent per opportunity,
    so changing the date moves the open task rather than stacking reminders.
    status/owner match jobs_task's constraints — 'Not Started' is a valid status
    where 'todo' is not, and owner is NOT NULL."""
    title = f"Revisit: {account or 'opportunity'}"
    # follow_up_date is a timestamp on the model but jobs_task.deadline is a date.
    day = when.date() if hasattr(when, "date") else when
    existing = await conn.fetchval(
        "SELECT id FROM bedrock.jobs_task "
        "WHERE parent_type = 'opportunity' AND parent_id = $1 AND deleted_at IS NULL "
        "  AND status <> 'Completed' AND title LIKE 'Revisit:%' LIMIT 1",
        opp_id)
    if existing:
        await conn.execute(
            "UPDATE bedrock.jobs_task SET deadline = $2::date, title = $3, updated_at = now() "
            "WHERE id = $1", existing, day, title)
        return
    await conn.execute(
        "INSERT INTO bedrock.jobs_task (parent_type, parent_id, title, status, owner, deadline) "
        "VALUES ('opportunity', $1, $2, 'Not Started', $3, $4::date)",
        opp_id, title, owner or "", day)


@router.delete("/opportunities/{opp_id}")
async def delete_opportunity(
    opp_id: UUID,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    user_email = user.get("email") if isinstance(user, dict) else getattr(user, "email", None)
    result = await conn.execute(
        """
        UPDATE bedrock.jobs_opportunity
        SET deleted_at=now(), deleted_by=$2
        WHERE id=$1 AND deleted_at IS NULL
        """,
        opp_id, user_email,
    )
    if result == "UPDATE 0":
        raise HTTPException(404, "Opportunity not found")
    return {"success": True, "data": {"deleted": True}}


# ── Activity delete ──────────────────────────────────────────────────────────

@router.delete("/activity/{activity_id}")
async def delete_activity(
    activity_id: UUID,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    result = await conn.execute(
        "UPDATE bedrock.activity SET deleted_at=now() WHERE id=$1 AND deleted_at IS NULL",
        activity_id,
    )
    if result == "UPDATE 0":
        raise HTTPException(404, "Activity not found")
    return {"success": True, "data": {"deleted": True}}


# ── Builders search ───────────────────────────────────────────────────────────

@router.get("/builders")
async def list_builders(
    search: Optional[str] = Query(None, min_length=1),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Search all platform builders via SECURITY DEFINER function (bypasses RLS)."""
    rows = await conn.fetch(
        "SELECT user_id, full_name, email, cohort FROM bedrock.search_builders($1)",
        search or "",
    )
    return {
        "success": True,
        "data": [
            {
                "user_id": r["user_id"],
                "email":   r["email"] or "",
                "name":    r["full_name"] or r["email"] or "",
                "cohort":  r["cohort"] or "",
            }
            for r in rows
            if r["email"]
        ],
    }


class BulkOwner(BaseModel):
    contact_ids: list[int]
    owner_email: Optional[str] = None   # null/empty clears the owner


@router.post("/contacts/bulk-owner")
async def bulk_set_contact_owner(
    body: BulkOwner,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Set (or clear) the org-wide owner on many contacts at once."""
    ids = [int(i) for i in body.contact_ids]
    if not ids:
        return {"success": True, "data": {"updated": 0}}
    owner = (body.owner_email or "").strip().lower() or None
    res = await conn.execute(
        "UPDATE public.contacts SET owner_email = $2, updated_at = now() WHERE contact_id = ANY($1::int[])",
        ids, owner)
    n = int(res.split()[-1]) if res and res.split()[-1].isdigit() else 0
    return {"success": True, "data": {"updated": n}}


class BulkProspect(BaseModel):
    contact_ids: list[int]
    value: bool = True   # true = mark as jobs prospect; false = remove the flag


@router.post("/contacts/bulk-prospect")
async def bulk_set_prospect(
    body: BulkProspect,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Bulk mark (or unmark) contacts as jobs prospects — sets is_jobs_contact
    without creating a pipeline stage (that's what /flag-jobs does)."""
    ids = [int(i) for i in body.contact_ids]
    if not ids:
        return {"success": True, "data": {"updated": 0}}
    res = await conn.execute(
        "UPDATE public.contacts SET is_jobs_contact = $2, updated_at = now() "
        "WHERE contact_id = ANY($1::int[]) AND is_jobs_contact IS DISTINCT FROM $2",
        ids, body.value)
    n = int(res.split()[-1]) if res and res.split()[-1].isdigit() else 0
    return {"success": True, "data": {"updated": n}}


@router.get("/contact-tags")
async def contact_tag_catalog(
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Curated contact-tag vocabulary (drives the tags picker + filter).
    Slugs are what's stored on contacts.tags; labels are display-only.
    sort_order is the campaign priority (lower = higher priority)."""
    rows = await conn.fetch(
        "SELECT slug, label, sort_order FROM bedrock.contact_tag_catalog WHERE active ORDER BY sort_order, label")
    return {"success": True, "data": [{"slug": r["slug"], "label": r["label"], "sort_order": r["sort_order"]} for r in rows]}


# ── Tag campaigns (Performance → prioritize outreach) ─────────────────────────
# A "campaign" is one prioritizable tag; all alumni cohorts collapse into a
# single "Fellow Alumni" campaign (key 'alumni'). Priority = catalog.sort_order.
def _campaign_key(slug: str) -> str:
    return "alumni" if slug.startswith("alumni") else slug


@router.get("/tag-campaigns")
async def tag_campaigns(user=Depends(require_auth), conn=Depends(get_db)):
    """Each tag as a prioritizable campaign with contact + account counts,
    ordered by priority. Alumni cohorts are merged into one 'Fellow Alumni' row."""
    rows = await conn.fetch("""
        WITH tagged AS (
          SELECT c.contact_id, c.is_jobs_contact,
                 CASE WHEN t LIKE 'alumni%' THEN 'alumni' ELSE t END AS campaign,
                 nullif(lower(trim(c.current_company)), '') AS company,
                 (SELECT m.stage FROM bedrock.jobs_contact_membership m WHERE m.contact_id = c.contact_id) AS stage
          FROM public.contacts c, unnest(c.tags) t
          WHERE t IN (SELECT slug FROM bedrock.contact_tag_catalog WHERE active)
        )
        SELECT campaign,
               count(DISTINCT contact_id) AS contacts,
               -- "in pipeline" = the jobs-prospect flag (is_jobs_contact), NOT a
               -- membership. accounts/in_pipeline reflect that population.
               count(DISTINCT company) FILTER (WHERE is_jobs_contact) AS accounts,
               count(DISTINCT contact_id) FILTER (WHERE is_jobs_contact) AS in_pipeline,
               -- disjoint funnel stages; 'assigned' is a real stage a user sets
               -- (no stage at all = blank/not-yet, computed on the client):
               count(DISTINCT contact_id) FILTER (WHERE is_jobs_contact AND stage = 'assigned') AS assigned,
               count(DISTINCT contact_id) FILTER (WHERE is_jobs_contact AND stage = 'initial_outreach') AS contacted,
               count(DISTINCT contact_id) FILTER (WHERE is_jobs_contact AND stage = 'converted_to_opportunity') AS converted,
               count(DISTINCT contact_id) FILTER (WHERE is_jobs_contact AND stage = 'on_hold') AS on_hold
        FROM tagged GROUP BY campaign
    """)
    counts = {r["campaign"]: r for r in rows}
    cat = await conn.fetch(
        "SELECT slug, label, sort_order, owner_email FROM bedrock.contact_tag_catalog WHERE active")
    camps: dict = {}
    for r in cat:
        key = _campaign_key(r["slug"])
        c = camps.setdefault(key, {"key": key, "label": "Fellow Alumni" if key == "alumni" else r["label"],
                                   "slugs": [], "sort_order": r["sort_order"], "owner_email": r["owner_email"]})
        c["slugs"].append(r["slug"])
        c["sort_order"] = min(c["sort_order"], r["sort_order"])
        c["owner_email"] = c["owner_email"] or r["owner_email"]
    out = []
    for c in camps.values():
        r = counts.get(c["key"])
        contacts = r["contacts"] if r else 0
        in_pipeline = r["in_pipeline"] if r else 0
        assigned = r["assigned"] if r else 0
        contacted = r["contacted"] if r else 0
        converted = r["converted"] if r else 0
        on_hold = r["on_hold"] if r else 0
        out.append({**c,
                    "contacts": contacts,
                    "accounts": r["accounts"] if r else 0,
                    "in_pipeline": in_pipeline,
                    # funnel over the in-pipeline (jobs-prospect) population, DISJOINT:
                    # not_yet (no stage) + assigned + contacted + converted + on_hold
                    # = in_pipeline.
                    "funnel": {
                        "not_yet":   max(0, in_pipeline - assigned - contacted - converted - on_hold),
                        "assigned":  assigned,
                        "contacted": contacted,
                        "converted": converted,
                        "on_hold":   on_hold,
                    }})
    out.sort(key=lambda x: (x["sort_order"], x["label"]))
    return {"success": True, "data": out}


@router.get("/tag-campaigns/{key}/records")
async def tag_campaign_records(
    key: str,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """The contacts and accounts behind a campaign's counts.

    Deliberately re-derives the slug set from the catalog via `_campaign_key` and
    applies the same `is_jobs_contact` gate as /tag-campaigns, so a drill list can
    never disagree with the number that opened it.
    """
    cat = await conn.fetch(
        "SELECT slug FROM bedrock.contact_tag_catalog WHERE active")
    slugs = [r["slug"] for r in cat if _campaign_key(r["slug"]) == key]
    if not slugs:
        raise HTTPException(404, f"Unknown campaign: {key}")

    rows = await conn.fetch("""
        WITH tagged AS (
          SELECT DISTINCT c.contact_id
          FROM public.contacts c
          CROSS JOIN LATERAL unnest(c.tags) AS t
          WHERE t = ANY($1::text[])
        ),
        act AS (
          SELECT a.participant_public_contact_id AS cid,
                 count(*) AS touches,
                 max(a.activity_date) AS last_touch
          FROM bedrock.activity a
          WHERE a.deleted_at IS NULL
            AND a.participant_public_contact_id IN (SELECT contact_id FROM tagged)
          GROUP BY 1
        )
        SELECT c.contact_id, c.full_name, c.current_company AS company,
               c.is_jobs_contact, m.stage, m.owner_email,
               -- Latest stage stamp the contact reached, so "when" lines up with
               -- the stage shown next to it. The label lives in the frontend's
               -- MEMBERSHIP_STAGE_LABELS, which is the single source for it.
               COALESCE(m.converted_at, m.first_outreach_at, m.assigned_at) AS stage_entered_at,
               COALESCE(act.touches, 0) AS touches, act.last_touch
        FROM tagged tg
        JOIN public.contacts c ON c.contact_id = tg.contact_id
        LEFT JOIN bedrock.jobs_contact_membership m ON m.contact_id = c.contact_id
        LEFT JOIN act ON act.cid = c.contact_id
        ORDER BY c.full_name
    """, slugs)

    # in_pipeline / accounts count only jobs prospects, so the drills must too.
    in_pipe = [r for r in rows if r["is_jobs_contact"]]
    contacts = [{
        "contact_id": r["contact_id"],
        "full_name": r["full_name"],
        "company": r["company"],
        "stage": canon_stage(r["stage"]),
        "stage_entered_at": r["stage_entered_at"].isoformat() if r["stage_entered_at"] else None,
        "owner": r["owner_email"],
        "touches": int(r["touches"] or 0),
        "last_touch": r["last_touch"].isoformat() if r["last_touch"] else None,
    } for r in in_pipe]

    by_company: dict = {}
    for r in in_pipe:
        name = r["company"] or "(no company)"
        a = by_company.setdefault(name, {"company": name, "contacts": 0, "contacted": 0})
        a["contacts"] += 1
        if r["stage"] in ("initial_outreach", "converted_to_opportunity"):
            a["contacted"] += 1
    accounts = sorted(by_company.values(), key=lambda a: (-a["contacts"], a["company"]))

    return {"success": True, "data": {"contacts": contacts, "accounts": accounts}}


class CampaignOrder(BaseModel):
    order: list[str]   # campaign keys, top = highest priority


@router.put("/tag-campaigns/order")
async def set_tag_campaign_order(body: CampaignOrder, user=Depends(require_auth), conn=Depends(get_db)):
    """Persist a new campaign priority order → catalog.sort_order (10, 20, …).
    A campaign key maps to its slug (or all alumni_* slugs for 'alumni')."""
    async with conn.transaction():
        for i, key in enumerate(body.order):
            so = (i + 1) * 10
            if key == "alumni":
                await conn.execute("UPDATE bedrock.contact_tag_catalog SET sort_order=$1 WHERE slug LIKE 'alumni%'", so)
            else:
                await conn.execute("UPDATE bedrock.contact_tag_catalog SET sort_order=$1 WHERE slug=$2", so, key)
    return {"success": True, "data": {"updated": len(body.order)}}


class CampaignOwner(BaseModel):
    key: str
    owner_email: Optional[str] = None   # null/empty clears


@router.put("/tag-campaigns/owner")
async def set_tag_campaign_owner(body: CampaignOwner, user=Depends(require_auth), conn=Depends(get_db)):
    """Assign (or clear) a staff owner on a campaign. 'alumni' → all alumni_* slugs."""
    owner = (body.owner_email or "").strip().lower() or None
    if body.key == "alumni":
        await conn.execute("UPDATE bedrock.contact_tag_catalog SET owner_email=$1 WHERE slug LIKE 'alumni%'", owner)
    else:
        await conn.execute("UPDATE bedrock.contact_tag_catalog SET owner_email=$1 WHERE slug=$2", owner, body.key)
    return {"success": True, "data": {"key": body.key, "owner_email": owner}}


@router.post("/contacts/{contact_id}/add-to-jobs")
async def add_contact_to_jobs(
    contact_id: int,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Flag any contact (SF, LinkedIn, manual) as a jobs pipeline contact."""
    result = await conn.execute(
        "UPDATE public.contacts SET is_jobs_contact=true, updated_at=now() WHERE contact_id=$1",
        contact_id,
    )
    if result == "UPDATE 0":
        raise HTTPException(404, "Contact not found")
    return {"success": True, "data": {"contact_id": contact_id, "is_jobs_contact": True}}


@router.delete("/contacts/{contact_id}/add-to-jobs")
async def remove_contact_from_jobs(
    contact_id: int,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Remove a contact from the jobs pipeline (unflag)."""
    await conn.execute(
        "UPDATE public.contacts SET is_jobs_contact=false, updated_at=now() WHERE contact_id=$1",
        contact_id,
    )
    return {"success": True, "data": {"contact_id": contact_id, "is_jobs_contact": False}}


# ── Contact PATCH ────────────────────────────────────────────────────────────

class ContactUpdate(BaseModel):
    full_name:       Optional[str] = None
    email:           Optional[str] = None
    current_title:   Optional[str] = None
    current_company: Optional[str] = None
    contact_stage:   Optional[str] = None
    linkedin_url:    Optional[str] = None
    owner_email:     Optional[str] = None
    tags:            Optional[list[str]] = None   # curated slugs only; system markers preserved server-side


class ContactsMerge(BaseModel):
    canonical_id: int
    loser_ids: list[int]


@router.post("/contacts/merge")
async def merge_contacts(
    body: ContactsMerge,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Merge duplicate contacts into a canonical one — repoints activity,
    aliases, SF links, statuses, tasks/comments, relationships, and candidate
    rows via the SECURITY DEFINER bedrock.merge_contacts, then marks losers
    merged (audited). Returns how many were merged."""
    losers = [int(i) for i in body.loser_ids if int(i) != body.canonical_id]
    if not losers:
        return {"success": True, "data": {"merged": 0, "canonical_id": body.canonical_id}}
    for lid in losers:
        await conn.execute("SELECT bedrock.merge_contacts($1, $2, $3)",
                           body.canonical_id, lid, "in-app duplicate merge")
    return {"success": True, "data": {"merged": len(losers), "canonical_id": body.canonical_id}}


@router.patch("/contacts/{contact_id}")
async def update_contact(
    contact_id: int,
    body: ContactUpdate,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    # Any live contact is editable — the old `airtable_id IS NOT NULL` guard
    # dated from the Airtable-only era and silently 404'd edits on every
    # linkedin/outreach/candidate/manual contact (most of the table).
    existing = await conn.fetchrow(
        "SELECT contact_id FROM public.contacts WHERE contact_id=$1 AND coalesce(contact_stage,'') <> 'merged'",
        contact_id,
    )
    if not existing:
        raise HTTPException(404, "Contact not found")

    sets, params = [], []
    i = 1
    # Key off model_fields_set (like update_opportunity) so an explicit null
    # CLEARS a field — `is not None` made clearing email/title/linkedin a
    # silent no-op.
    fields_set = body.model_fields_set
    for field in ("full_name", "email", "current_title", "current_company",
                  "contact_stage", "linkedin_url", "owner_email"):
        if field not in fields_set:
            continue
        val = getattr(body, field, None)
        # UI clears send "" — store NULL (email '' would collide with the
        # unique constraint). Never blank a name: skip empty full_name.
        if isinstance(val, str) and not val.strip():
            val = None
        if field == "full_name" and val is None:
            continue
        sets.append(f"{field} = ${i}"); params.append(val); i += 1

    if "tags" in fields_set:
        new_tags = sorted({t.strip() for t in (body.tags or []) if t and t.strip()})
        valid = {r["slug"] for r in await conn.fetch(
            "SELECT slug FROM bedrock.contact_tag_catalog WHERE active")}
        unknown = [t for t in new_tags if t not in valid]
        if unknown:
            raise HTTPException(400, f"Unknown tags (not in catalog): {unknown}")
        # Replace only the curated tags; system markers (email_review) survive.
        sets.append(
            "tags = (SELECT coalesce(array_agg(t), '{}'::text[]) "
            "        FROM unnest(coalesce(tags, '{}'::text[])) t "
            "        WHERE t NOT IN (SELECT slug FROM bedrock.contact_tag_catalog)) "
            f"|| ${i}::text[]")
        params.append(new_tags); i += 1

    if not sets:
        row = await conn.fetchrow("SELECT * FROM public.contacts WHERE contact_id=$1", contact_id)
        return {"success": True, "data": dict(row)}

    params.append(contact_id)
    await conn.execute(
        f"UPDATE public.contacts SET {', '.join(sets)}, updated_at=now() WHERE contact_id=${i}",
        *params,
    )
    row = await conn.fetchrow("SELECT * FROM public.contacts WHERE contact_id=$1", contact_id)
    return {"success": True, "data": dict(row)}


# ── My Network / Pursuit Network ──────────────────────────────────────────────
# Two scopes on one endpoint, because everything except the row universe is shared
# — the filters, the priority banding, the vote/fit/note cells:
#
#   scope=mine     the caller's own LinkedIn connections (staff_contact_relationships)
#   scope=pursuit  every tagged contact at Pursuit, leadership only
#
# The Pursuit scope's universe is Jac's definition (2026-08-05): a contact carrying
# at least one CURATED tag that is neither an alumni_* cohort nor 'influence'.
#   - Curated means present in bedrock.contact_tag_catalog. That deliberately
#     excludes 'email_review', which is not a catalog tag — it marks 2,624
#     auto-created contacts from unidentified email addresses, a triage queue
#     rather than a network.
#   - 'influence' is excluded as too broad (it alone carried 2,646 rows).
#   - An alumni tag does NOT disqualify a contact that also carries a qualifying
#     tag: an alumnus who is now a hiring partner is precisely who this view is
#     for. 148 contacts land that way.
# Currently 4,658 contacts, 2,853 of them via volunteer_historical.
_PURSUIT_TAG_EXCLUDED = ("influence",)
_PURSUIT_UNIVERSE = f"""EXISTS (
    SELECT 1 FROM unnest(c.tags) t
    JOIN bedrock.contact_tag_catalog cat ON cat.slug = t
    WHERE t NOT LIKE 'alumni%'
      AND t NOT IN ({', '.join(f"'{x}'" for x in _PURSUIT_TAG_EXCLUDED)})
)"""

# Profiles allowed into the Pursuit scope. Enforced on the ENDPOINT, not only by
# hiding the tab — a hidden tab is not access control.
_PURSUIT_SCOPE_PROFILES = ("Executive", "Admin")


async def _ensure_staff_mapping(email: str, conn) -> Optional[int]:
    """This caller's staff_user_id, self-provisioning on first use.

    Every write on these pages is keyed by staff_user_id (public.users.user_id).
    Without a bedrock.staff_user_id_map row a staff member can open the page and
    have every save refused, which is how six of thirteen leadership accounts —
    including four of five Executives — ended up unable to do anything.

    So the mapping is resolved lazily rather than waiting on a hand-written row.
    Lazy rather than at login on purpose: it also repairs accounts that existed
    before this code, not only new joiners.

    Degrades safely. bedrock.resolve_staff_user_id is a postgres-owned SECURITY
    DEFINER function (public.users has RLS that hides every row from bedrock_user);
    until its migration is applied this returns None exactly as before, and the
    endpoint reports the account unmapped instead of erroring.
    """
    sid = await conn.fetchval(
        "SELECT staff_user_id FROM bedrock.staff_user_id_map WHERE lower(email) = $1", email)
    if sid is not None or not email:
        return sid

    if not await conn.fetchval(
            "SELECT to_regprocedure('bedrock.resolve_staff_user_id(text)') IS NOT NULL"):
        return None
    resolved = await conn.fetchval("SELECT bedrock.resolve_staff_user_id($1)", email)
    if resolved is None:
        # Not staff in the learning platform, or no record there at all. Two active
        # logins are in that position (angielausche@, yoshiyuki.minami@) and no
        # amount of retrying will fix them — they need a public.users row first.
        return None

    await conn.execute(
        """INSERT INTO bedrock.staff_user_id_map (staff_user_id, email, display_name, notes)
           VALUES ($1, $2, NULL, 'auto-mapped on first use')
           ON CONFLICT (staff_user_id) DO NOTHING""", resolved, email)
    sid = await conn.fetchval(
        "SELECT staff_user_id FROM bedrock.staff_user_id_map WHERE lower(email) = $1", email)
    if sid is not None:
        return sid

    # The id is already claimed by a DIFFERENT login — the same human with two
    # addresses (firstname@ vs firstname.lastname@). Take it over only if that
    # login has been retired; if both are live, leave it alone and let the account
    # read as unmapped. Silently moving one active user's votes to another
    # would be worse than the visible gap.
    claimed_by_active = await conn.fetchval(
        """SELECT coalesce(bool_or(coalesce(ou.is_active, true)), false)
             FROM bedrock.staff_user_id_map m
             LEFT JOIN public.org_users ou ON lower(ou.email) = lower(m.email)
            WHERE m.staff_user_id = $1""", resolved)
    if claimed_by_active:
        logger.warning("staff_user_id %s already mapped to an active login; %s left unmapped",
                       resolved, email)
        return None
    await conn.execute(
        """UPDATE bedrock.staff_user_id_map
              SET email = $2, notes = coalesce(notes,'') || ' | repointed from a retired login 2026-08-05',
                  updated_at = now()
            WHERE staff_user_id = $1""", resolved, email)
    return resolved


async def _assert_pursuit_scope_allowed(email: str, conn) -> None:
    from routes.permissions import get_user_permissions
    profile = (await get_user_permissions(email, conn) or {}).get("profile_name")
    if profile not in _PURSUIT_SCOPE_PROFILES:
        raise HTTPException(
            403, "The Pursuit network is limited to leadership "
                 f"({' or '.join(_PURSUIT_SCOPE_PROFILES)} profile)")



# The vote vocabulary behind the 👍/👎 column. Renamed 2026-08-05 to match what
# the column actually asks ("Expect a response") rather than what it used to
# ("Willing to reach out").
#
# NOTE: bedrock.intro_request ALSO has a 'declined' status. It is a separate
# column on a separate table meaning a connector turned down an intro request —
# unrelated to this, and untouched by the rename.
CONNECTION_STATUSES = ("new", "expect_response", "dont_expect_response")
# Pre-rename spellings, still accepted on write and normalised on read, so the
# code and the data migration can land in EITHER order without a window where the
# 123 recorded votes read as "no vote". Safe to delete once
# db/migrations/2026-08-05-rename-connection-status.sql has been applied and no
# older client is still deployed.
LEGACY_CONNECTION_STATUSES = {
    "will_reach_out": "expect_response",
    "declined": "dont_expect_response",
}


def _normalize_connection_status(value: Optional[str]) -> str:
    """Legacy spelling -> current one. Anything unrecognised reads as 'new'."""
    v = LEGACY_CONNECTION_STATUSES.get(value or "new", value or "new")
    return v if v in CONNECTION_STATUSES else "new"


HIRING_FIT_VALUES = ("yes", "no")

# The My Network Note cell is stored as a real team comment
# (bedrock.jobs_comment, parent_type='prospect') rather than a private field, so
# it shows up in the contact's Comments thread like any other note. This prefix is
# what makes it round-trippable: it marks which comment the cell owns, and it
# labels the note in the thread for anyone reading it there.
RELATIONSHIP_CONTEXT_PREFIX = "relationship context:"


def _strip_relationship_context(content: Optional[str]) -> Optional[str]:
    """Comment body -> what the cell should show (the prefix removed)."""
    if not content:
        return None
    body = content[len(RELATIONSHIP_CONTEXT_PREFIX):] \
        if content[:len(RELATIONSHIP_CONTEXT_PREFIX)].lower() == RELATIONSHIP_CONTEXT_PREFIX \
        else content
    return body.strip() or None


async def _connection_status_has_hiring_fit(conn) -> bool:
    """Whether db/migrations/2026-08-05-connection-hiring-fit.sql has been applied.

    Checked per request, not cached at import, so the column starts working the
    moment the migration lands — no restart. to_regclass can't see columns, hence
    information_schema."""
    return bool(await conn.fetchval(
        """SELECT EXISTS (SELECT 1 FROM information_schema.columns
             WHERE table_schema = 'bedrock' AND table_name = 'connection_status'
               AND column_name = 'hiring_fit')"""))



@router.get("/my-network")
async def my_network(
    q: Optional[str] = Query(None, description="Search name/company/title"),
    limit: int = Query(500, le=2000),
    staff_email: Optional[str] = Query(None, description="Admin override; else the caller"),
    filter_rules: Optional[str] = Query(None, alias="filters", description="JSON array of {field,op,values} rules — applied in SQL so filters see the whole network, not the loaded page"),
    prioritized: bool = Query(True, description="Order P1 then P2 then the rest — ranked in SQL so the top band is drawn from the whole network, not just the page. Defaults ON (Jac, 2026-08-05): pass prioritized=false for the older warmth/recency order"),
    scope: str = Query("mine", pattern="^(mine|pursuit)$", description="'mine' = the caller's LinkedIn connections; 'pursuit' = every curated-tagged contact (leadership only)"),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """The logged-in staff member's LinkedIn connections (from
    staff_contact_relationships), joined to contacts + company, with flags for
    whether we've had activity with them (warm) and whether they're already a
    jobs prospect. Drives the "My Network" page. Sputnik staff ids are
    mapped to emails via bedrock.staff_user_id_map.

    Returns `total` (the whole network) alongside `matched` (the count under the
    active filters) so the UI can say "412 of 5,121" and warn honestly when a
    filtered result exceeds `limit`. Seven staff have more than 2,000
    connections, so a filter that only sifted the loaded page would under-report
    without saying so."""
    email = (staff_email or (user.get("email") if isinstance(user, dict) else getattr(user, "email", None)) or "").lower()
    if scope == "pursuit":
        await _assert_pursuit_scope_allowed(email, conn)
    sid = await _ensure_staff_mapping(email, conn)
    if sid is None:
        # The Pursuit scope doesn't depend on the caller having LinkedIn imports —
        # it's every tagged contact — but the vote/fit/note cells are stored per
        # staff_user_id, so without a mapping there is nowhere to write.
        return {"success": True, "data": {"mapped": False, "connections": [], "total": 0, "matched": 0,
                "hiring_fit_available": False,
                # Mapping self-provisions on first use (see _ensure_staff_mapping), so
                # reaching here means the lookup genuinely found nothing: no active
                # staff record in the learning platform under this address.
                "message": "This account has no staff record in the learning platform, "
                           "so there is nowhere to save your answers. Ask an admin to "
                           "check the email on your Pursuit staff record."}}

    # Scope decides the row universe; everything downstream is identical, which is
    # why both FROMs expose the same r / c / co / cs aliases.
    #
    # `mine` inner-joins the relationship table. `pursuit` starts from contacts and
    # LEFT joins the caller's own relationship, so warmth and connected-date still
    # light up for the tagged contacts an exec happens to know, and are simply NULL
    # for the rest.
    #
    # The companies AND connection_status joins are unconditional on every query
    # below: the shared `where` string may reference `co.` or `cs.`, so both must
    # resolve identically in the count, the rows, and the facet queries. Both are
    # keyed joins, so the cost is negligible.
    if scope == "pursuit":
        NET_FROM = ("public.contacts c "
                    "LEFT JOIN public.staff_contact_relationships r "
                    "  ON r.contact_id = c.contact_id AND r.staff_user_id = $1 "
                    "LEFT JOIN public.companies co ON co.company_id = c.company_id "
                    "LEFT JOIN bedrock.connection_status cs "
                    "  ON cs.contact_id = c.contact_id AND cs.staff_user_id = $1")
    else:
        NET_FROM = ("public.staff_contact_relationships r "
                    "JOIN public.contacts c ON c.contact_id = r.contact_id "
                    "LEFT JOIN public.companies co ON co.company_id = c.company_id "
                    "LEFT JOIN bedrock.connection_status cs "
                    "  ON cs.contact_id = c.contact_id AND cs.staff_user_id = $1")

    # bedrock.company_investor may not be created yet (its migration is applied by
    # hand). Checked per request rather than cached at import, so the portco half
    # of the priority rule starts working the moment the migration lands — no
    # restart, no code change. A missing table means "nothing is a portco", never
    # a 500.
    has_portco = await conn.fetchval("SELECT to_regclass('bedrock.company_investor') IS NOT NULL")
    # Same story for the account floor ("Work now" accounts are P2 minimum) — its
    # table is loaded from a snapshot and may not exist yet.
    has_floor = await conn.fetchval("SELECT to_regclass('bedrock.priority_account_floor') IS NOT NULL")
    priority_case = _net_priority_case(
        _PORTCO_EXISTS if has_portco else None,
        _ACCOUNT_FLOOR_BAND if has_floor else None)
    # Same story for the hiring_fit column — checked per request so the page works
    # either side of its migration. Reads as unset until then; the PATCH refuses
    # rather than silently discarding a write (see set_connection_status).
    has_hiring_fit = await _connection_status_has_hiring_fit(conn)
    hiring_fit_sel = "cs.hiring_fit" if has_hiring_fit else "NULL::text"
    # And again for the LinkedIn re-enrichment landing table, so the page works
    # either side of db/migrations/2026-08-06-contact-enrichment.sql.
    has_enrichment = await conn.fetchval(
        "SELECT to_regclass('bedrock.contact_enrichment') IS NOT NULL")

    params: list = [sid]
    # $1 is consumed by the joins in NET_FROM for both scopes, so it is always bound
    # even where the WHERE clause doesn't mention it.
    where = ("coalesce(c.contact_stage,'') <> 'merged' AND " + _PURSUIT_UNIVERSE) \
        if scope == "pursuit" else \
        "r.staff_user_id = $1 AND coalesce(c.contact_stage,'') <> 'merged'"
    # `base_where` excludes the search box and the filter rules — the facet lists
    # describe the WHOLE universe, so the filter menu offers a stable set of
    # options that doesn't shrink as you narrow.
    base_where = where
    if q:
        params.append(f"%{q}%")
        where += f" AND (c.full_name ILIKE ${len(params)} OR c.current_company ILIKE ${len(params)} OR c.current_title ILIKE ${len(params)})"

    # ── Filter rules (the page's Filter menu) ─────────────────────────────────
    # Each UI field maps to a whitelisted (type, SQL expression) pair mirroring
    # the column the page displays. An unknown field or op is a 400, never a
    # silent pass-through that would under-filter.
    _NET_RULE_FIELDS: dict[str, tuple[str, str]] = {
        # firmographics
        "headcount": ("select", "co.size_bucket"),
        "industry":  ("select", "co.industry"),
        "tristate":  ("select", _tristate_case("co.hq_location")),
        "seniority": ("select", _seniority_case("c.current_title")),
        # the contact themselves
        "company":   ("text", "c.current_company"),
        "title":     ("text", "c.current_title"),
        "tags":      ("tags", "c.tags"),
        # signals — the same three chips the row renders, so the filter and the
        # badge can never disagree.
        "is_jobs":   ("boolean", "(c.is_jobs_contact = true)"),
        "has_open_opp": ("boolean",
            "EXISTS (SELECT 1 FROM bedrock.jobs_opportunity o2 WHERE o2.deleted_at IS NULL"
            " AND o2.stage LIKE 'active%'"
            " AND lower(trim(o2.account_name)) = lower(trim(c.current_company)))"),
        "hired_before": ("boolean",
            "EXISTS (SELECT 1 FROM public.employment_records e2 WHERE c.current_company IS NOT NULL"
            " AND lower(e2.company_name) = lower(c.current_company))"),
        # "warm" = THIS staff member has touched them; "touched" = anyone has.
        # Mirrors the two LATERALs in the rows query, including the relevance
        # gate. :staff_email is bound on demand — see _net_rule_clause.
        "warm": ("boolean",
            f"EXISTS (SELECT 1 FROM bedrock.activity a2 WHERE a2.participant_public_contact_id = c.contact_id"
            f" AND a2.deleted_at IS NULL AND {_jobs_relevant('a2')}"
            f" AND (a2.email_from ILIKE {_STAFF_EMAIL_TOKEN} OR a2.logged_by ILIKE {_STAFF_EMAIL_TOKEN}))"),
        "touched": ("boolean",
            f"EXISTS (SELECT 1 FROM bedrock.activity a2 WHERE a2.participant_public_contact_id = c.contact_id"
            f" AND a2.deleted_at IS NULL AND {_jobs_relevant('a2')})"),
        # The caller's own two answers. Normalised to yes/no/'' so the filter reads
        # the same as the thumbs regardless of which vocabulary the row was stored
        # under, and so is_empty means "not answered yet".
        "expect_response": ("select",
            "CASE"
            f" WHEN cs.status IN ('{CONNECTION_STATUSES[1]}', 'will_reach_out') THEN 'yes'"
            f" WHEN cs.status IN ('{CONNECTION_STATUSES[2]}', 'declined') THEN 'no'"
            " ELSE '' END"),
        "hiring_fit": ("select", hiring_fit_sel),
    }
    if filter_rules:
        try:
            parsed_rules = json.loads(filter_rules)
            assert isinstance(parsed_rules, list)
        except Exception:
            raise HTTPException(400, "filters must be a JSON array of {field,op,values}")
        for rule in parsed_rules[:20]:
            clause, params = _net_rule_clause(rule, _NET_RULE_FIELDS, params, email)
            if clause:
                where += f" AND {clause}"

    # `total` is the whole network; `matched` is what the filters + search leave.
    total = await conn.fetchval(f"SELECT count(*) FROM {NET_FROM} WHERE {base_where}", sid)
    matched = await conn.fetchval(f"SELECT count(*) FROM {NET_FROM} WHERE {where}", *params)

    params.append(f"%{email}%"); p_email = len(params)   # staff-specific activity match
    params.append(RELATIONSHIP_CONTEXT_PREFIX); p_ctx_prefix = len(params)
    params.append(email);        p_author = len(params)   # the caller's own note
    params.append(limit);        p_lim = len(params)
    # Ranking in SQL, not in the browser: with 5,000+ connections and a 2,000-row
    # page, sorting client-side would only rank the window and could leave P1s
    # off-screen entirely. NULLS LAST keeps unranked rows below the bands.
    priority_order = f"({priority_case}) ASC NULLS LAST," if prioritized else ""
    rows = await conn.fetch(
        f"""
        SELECT c.contact_id, c.full_name, c.current_title, c.current_company, c.email,
               c.linkedin_url, c.is_jobs_contact, r.relationship_strength, r.connected_date,
               act.n AS activity_count, act.last AS last_activity, act.last_type AS last_channel,
               coalesce(mine.n, 0) AS my_activity_count, mine.last AS my_last_activity,
               coalesce(cc.n, 0) AS co_connections,
               (hire.hired IS NOT NULL) AS company_hired_before,
               (opp.has_open IS NOT NULL) AS has_open_opp,
               cs.status AS status, cs.reason AS status_reason,
               {hiring_fit_sel} AS hiring_fit,
               rc.id AS rc_id, rc.content AS rc_content,
               coalesce(rco.n, 0) AS rc_others,
               -- firmographics behind the page's filters
               coalesce(c.tags, '{{}}'::text[]) AS tags,
               co.size_bucket AS headcount_band, co.industry AS industry, co.hq_location,
               ({_tristate_case('co.hq_location')}) AS tristate,
               ({_seniority_case('c.current_title')}) AS seniority,
               ({priority_case}) AS priority,
               {'(' + _PORTCO_EXISTS + ')' if has_portco else 'false'} AS is_portco,
               coalesce(cmt.n, 0) AS comment_count,
               -- Live LinkedIn values, when this contact has been re-enriched.
               -- The row keeps rendering c.current_* as the authoritative value;
               -- these ride alongside as "there is something fresher". Nothing
               -- here feeds the priority banding or the filters — promoting a
               -- change into contacts is a separate reviewed step, because
               -- current_company is the join key to the firmographics.
               {'ce.live_title'   if has_enrichment else 'NULL::text'} AS live_title,
               {'ce.live_company' if has_enrichment else 'NULL::text'} AS live_company,
               {'ce.enriched_at'  if has_enrichment else 'NULL::timestamptz'} AS enriched_at,
               -- Compared against c.current_* as it is NOW, not against the
               -- prior_* snapshot the loader took: if someone has since fixed the
               -- contact by hand, the row must stop claiming a difference.
               {'''(ce.live_title IS NOT NULL
                    AND lower(btrim(ce.live_title)) IS DISTINCT FROM lower(btrim(coalesce(c.current_title,''))))'''
                 if has_enrichment else 'false'} AS live_title_differs,
               {'''(ce.live_company IS NOT NULL
                    AND lower(btrim(ce.live_company)) IS DISTINCT FROM lower(btrim(coalesce(c.current_company,''))))'''
                 if has_enrichment else 'false'} AS live_company_differs
        FROM {NET_FROM}
        {"LEFT JOIN bedrock.contact_enrichment ce ON ce.contact_id = c.contact_id AND ce.status = 'ok'"
         if has_enrichment else ""}
        LEFT JOIN LATERAL (
            SELECT count(*) n, max(activity_date) last,
                   (array_agg(type ORDER BY activity_date DESC))[1] AS last_type
            FROM bedrock.activity a
            WHERE a.participant_public_contact_id = c.contact_id AND a.deleted_at IS NULL
              AND {_jobs_relevant('a')}
        ) act ON true
        LEFT JOIN LATERAL (
            SELECT count(*) n, max(activity_date) last FROM bedrock.activity a
            WHERE a.participant_public_contact_id = c.contact_id AND a.deleted_at IS NULL
              AND {_jobs_relevant('a')}
              AND (a.email_from ILIKE ${p_email} OR a.logged_by ILIKE ${p_email})
        ) mine ON true
        LEFT JOIN LATERAL (
            SELECT count(*) - 1 AS n FROM public.staff_contact_relationships r2
            WHERE r2.contact_id = c.contact_id
        ) cc ON true
        LEFT JOIN LATERAL (
            SELECT 1 AS hired FROM public.employment_records e
            WHERE c.current_company IS NOT NULL AND lower(e.company_name) = lower(c.current_company) LIMIT 1
        ) hire ON true
        LEFT JOIN LATERAL (
            SELECT 1 AS has_open FROM bedrock.jobs_opportunity o
            WHERE o.deleted_at IS NULL AND o.stage LIKE 'active%'
              AND lower(trim(o.account_name)) = lower(trim(c.current_company)) LIMIT 1
        ) opp ON true
        LEFT JOIN LATERAL (
            SELECT count(*) n FROM bedrock.jobs_comment jc
            WHERE jc.parent_type = 'prospect' AND jc.parent_id = c.contact_id::text
        ) cmt ON true
        -- The row's Note cell: THE CALLER'S OWN note, blank until they write one.
        -- Stored as a real team comment (so it shows up in the contact's thread and
        -- colleagues can read it), but scoped to author_email here — with several
        -- people working the same Pursuit list, a cell showing whoever wrote last
        -- would overwrite itself under them and never be theirs to fill in.
        LEFT JOIN LATERAL (
            SELECT jc.id, jc.content
            FROM bedrock.jobs_comment jc
            WHERE jc.parent_type = 'prospect' AND jc.parent_id = c.contact_id::text
              AND jc.content ILIKE ${p_ctx_prefix} || '%'
              AND lower(jc.author_email) = ${p_author}
            ORDER BY jc.created_at DESC, jc.id DESC
            LIMIT 1
        ) rc ON true
        -- How many OTHER people have left context here. Not shown in the cell —
        -- just a hint so 13 execs don't each write the same thing blind.
        LEFT JOIN LATERAL (
            SELECT count(*) n
            FROM bedrock.jobs_comment jc
            WHERE jc.parent_type = 'prospect' AND jc.parent_id = c.contact_id::text
              AND jc.content ILIKE ${p_ctx_prefix} || '%'
              AND lower(coalesce(jc.author_email,'')) <> ${p_author}
        ) rco ON true
        WHERE {where}
        ORDER BY {priority_order} (mine.n > 0) DESC, (act.n > 0) DESC, coalesce(mine.last, act.last) DESC NULLS LAST, c.full_name
        LIMIT ${p_lim}
        """, *params)
    return {"success": True, "data": {
        "mapped": True, "total": total, "matched": matched,
        # Lets the UI grey out the Hiring fit cells rather than offer a control
        # whose writes would be refused.
        "hiring_fit_available": has_hiring_fit,
        "connections": [{
            "contact_id": r["contact_id"], "full_name": r["full_name"], "current_title": r["current_title"],
            "current_company": r["current_company"], "email": r["email"], "linkedin_url": r["linkedin_url"],
            "is_jobs_contact": r["is_jobs_contact"], "relationship_strength": r["relationship_strength"],
            "connected_date": r["connected_date"].isoformat() if r["connected_date"] else None,
            "activity_count": r["activity_count"] or 0,
            "last_activity": r["last_activity"].isoformat() if r["last_activity"] else None,
            "last_channel": r["last_channel"],
            "my_activity_count": r["my_activity_count"] or 0,
            "my_last_activity": r["my_last_activity"].isoformat() if r["my_last_activity"] else None,
            # "warm" = THIS staff member has been in touch with their connection
            "warm": (r["my_activity_count"] or 0) > 0,
            # "touched" = anyone at Pursuit has activity with them
            "touched": (r["activity_count"] or 0) > 0,
            "co_connections": r["co_connections"] or 0,
            "company_hired_before": r["company_hired_before"],
            "has_open_opp": r["has_open_opp"],
            # Normalised so a pre-rename row still renders as a real vote.
            "status": _normalize_connection_status(r["status"]),
            "status_reason": r["status_reason"],
            "hiring_fit": r["hiring_fit"],
            # The Note cell — this staff member's OWN note, blank until they write
            # one. Others' notes are only counted, never shown in the cell.
            "relationship_context": _strip_relationship_context(r["rc_content"]),
            "relationship_context_id": str(r["rc_id"]) if r["rc_id"] else None,
            "relationship_context_others": r["rc_others"] or 0,
            "tags": list(r["tags"] or []),
            "priority": r["priority"],
            "is_portco": r["is_portco"],
            "comment_count": r["comment_count"] or 0,
            # Re-enrichment: what LinkedIn says today, only when it disagrees with
            # what we hold. Sending the value only when it differs keeps the row
            # from rendering a "fresher" marker that says the same thing twice.
            "live_title": r["live_title"] if r["live_title_differs"] else None,
            "live_company": r["live_company"] if r["live_company_differs"] else None,
            "enriched_at": r["enriched_at"].isoformat() if r["enriched_at"] else None,
            "headcount_band": r["headcount_band"],
            "industry": r["industry"],
            "hq_location": r["hq_location"],
            "tristate": r["tristate"],
            "seniority": r["seniority"],
        } for r in rows]}}


@router.get("/my-network/facets")
async def my_network_facets(
    staff_email: Optional[str] = Query(None, description="Admin override; else the caller"),
    scope: str = Query("mine", pattern="^(mine|pursuit)$"),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """The option lists for the My Network filter menu.

    Separate from /my-network on purpose: these depend only on WHOSE network it
    is, never on the search box or the active filters, but the scan costs about
    as much as the row query itself. Folded into the main response it re-ran on
    every keystroke; on its own key the client fetches it once.

    Headcount and industry are discovered from the data so a new value can't go
    missing from the menu. Tri-state and seniority are the complete output
    domains of two total CASE expressions, so they need no scan at all."""
    email = (staff_email or (user.get("email") if isinstance(user, dict) else getattr(user, "email", None)) or "").lower()
    if scope == "pursuit":
        await _assert_pursuit_scope_allowed(email, conn)
    sid = await _ensure_staff_mapping(email, conn)
    if sid is None:
        return {"success": True, "data": {"headcount": [], "industry": [], "tristate": [], "seniority": []}}
    if scope == "pursuit":
        src = ("public.contacts c LEFT JOIN public.companies co ON co.company_id = c.company_id "
               f"WHERE coalesce(c.contact_stage,'') <> 'merged' AND {_PURSUIT_UNIVERSE}")
        args: tuple = ()
    else:
        src = ("public.staff_contact_relationships r "
               "JOIN public.contacts c ON c.contact_id = r.contact_id "
               "LEFT JOIN public.companies co ON co.company_id = c.company_id "
               "WHERE r.staff_user_id = $1 AND coalesce(c.contact_stage,'') <> 'merged'")
        args = (sid,)
    row = await conn.fetchrow(
        f"""SELECT array_agg(DISTINCT co.size_bucket) FILTER (WHERE co.size_bucket IS NOT NULL) AS headcount,
                   array_agg(DISTINCT co.industry)    FILTER (WHERE co.industry    IS NOT NULL) AS industry
            FROM {src}""", *args)
    return {"success": True, "data": {
        # Ordering is the client's job — it holds the size/seniority ladders. Tag
        # options aren't here: the client already has the full curated catalog from
        # useContactTagCatalog(), so discovering them per scope would be redundant.
        "headcount": sorted(row["headcount"] or []),
        "industry": sorted(row["industry"] or []),
        "tristate": _TRISTATE_VALUES,
        "seniority": _SENIORITY_RUNGS,
    }}


class ConnectionStatusUpdate(BaseModel):
    contact_id: int
    # PARTIAL: None means "leave alone", so editing one cell of the row can't
    # blank the other. Clearing is explicit — status='new', hiring_fit=''.
    #
    # The Note cell is deliberately NOT here: it writes a real team comment via
    # /api/jobs/jobs-comments (see RELATIONSHIP_CONTEXT_PREFIX), not a private
    # column on this row.
    status: Optional[str] = None          # new | expect_response | dont_expect_response
    hiring_fit: Optional[str] = None      # yes | no | '' to clear
    note: Optional[str] = None            # legacy field, still honoured if sent
    reason: Optional[str] = None


@router.patch("/my-network/status")
async def set_connection_status(
    body: ConnectionStatusUpdate,
    staff_email: Optional[str] = Query(None),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Update a staff member's row for one of their connections — the vote
    ("expect a response"), the hiring-fit call, and the inline note. Stored per
    (staff, contact) in bedrock.connection_status.

    A partial update by design: the page saves one cell at a time, and the
    previous version wrote all columns on every call, so toggling a vote silently
    erased the note beside it. Only fields present in the body are touched.

    Legacy status spellings are accepted and rewritten, so a client that hasn't
    been redeployed still records a usable vote."""
    email = (staff_email or (user.get("email") if isinstance(user, dict) else getattr(user, "email", None)) or "").lower()
    sid = await _ensure_staff_mapping(email, conn)
    if sid is None:
        raise HTTPException(400, "No connection mapping for this account")

    sets: dict[str, object] = {}
    if body.status is not None:
        if body.status not in CONNECTION_STATUSES and body.status not in LEGACY_CONNECTION_STATUSES:
            raise HTTPException(400, "invalid status")
        sets["status"] = _normalize_connection_status(body.status)
    if body.hiring_fit is not None:
        if body.hiring_fit not in HIRING_FIT_VALUES and body.hiring_fit != "":
            raise HTTPException(400, "hiring_fit must be 'yes', 'no', or '' to clear")
        if not await _connection_status_has_hiring_fit(conn):
            # Refuse rather than accept-and-drop: a 200 that quietly loses the
            # answer is worse than a clear error naming the missing migration.
            raise HTTPException(
                503, "hiring_fit is not available yet — apply "
                     "db/migrations/2026-08-05-connection-hiring-fit.sql")
        sets["hiring_fit"] = body.hiring_fit or None
    if body.note is not None:
        sets["note"] = body.note.strip() or None
    if body.reason is not None:
        sets["reason"] = body.reason or None
    if not sets:
        raise HTTPException(400, "nothing to update")

    # INSERT needs a status for the NOT NULL column; the table defaults it to
    # 'new', which is right for a row created by a note or hiring-fit edit alone.
    cols = list(sets)
    insert_cols = ", ".join(["staff_user_id", "contact_id", *cols, "updated_by", "updated_at"])
    insert_vals = ", ".join(["$1", "$2", *(f"${i + 3}" for i in range(len(cols))),
                             f"${len(cols) + 3}", "now()"])
    updates = ", ".join(f"{c}=EXCLUDED.{c}" for c in cols)
    await conn.execute(
        f"""INSERT INTO bedrock.connection_status ({insert_cols})
            VALUES ({insert_vals})
            ON CONFLICT (staff_user_id, contact_id) DO UPDATE
              SET {updates}, updated_by=EXCLUDED.updated_by, updated_at=now()""",
        sid, body.contact_id, *(sets[c] for c in cols), email)
    return {"success": True, "data": {"contact_id": body.contact_id, **sets}}


# ── Candidate review queue ────────────────────────────────────────────────────
# Email recipients we auto-created but couldn't confidently identify (personal
# domains / no full name) land as 'candidate' contacts tagged 'email_review'.
# Avni/Damon review them on the Home page: fill name/company then promote into
# the pipeline, or dismiss. Promote flips is_jobs_contact=true + drops the tag.

@router.get("/candidates")
async def list_candidates(
    owner: Optional[str] = Query(None, description="Filter to candidates whose activity involved this staff email"),
    status: str = Query("candidate", pattern="^(candidate|dismissed)$"),
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Pending review candidates, newest-emailed first. Includes the email
    domain + an exact-domain account suggestion (cheap, set-based) so the list
    can show a linkage chip; the detail endpoint adds fuzzy matching + AI.
    `owner` filters to candidates a given staff member corresponded with
    (from bedrock.email_candidate.owners). Owners/channels are returned per row
    so the UI can show + filter by who reached out."""
    params: list = []
    where = [f"c.contact_stage = '{status}'",
             "('email_review' = ANY(c.tags) OR c.source = 'email_candidate')"]
    if owner:
        params.append(owner)
        where.append(f"${len(params)} = ANY(ec.owners)")
    rows = await conn.fetch(
        f"""
        SELECT c.contact_id, c.full_name, c.email, c.current_company, c.current_title,
               lower(split_part(c.email,'@',2)) AS domain,
               count(a.id) AS email_count,
               max(a.activity_date) AS last_email,
               (array_agg(a.subject ORDER BY a.activity_date DESC))[1] AS last_subject,
               aed.sf_account_name AS domain_account,
               e.full_name AS ai_name, e.company AS ai_company, e.confidence AS ai_confidence,
               e.is_employer_contact,
               e.account_suggestion->>'account_name' AS ai_account,
               coalesce(array_length(e.possible_duplicate_ids, 1), 0) AS dup_count,
               md.top_dup_id AS top_dup_id,
               md.top_dup_company AS top_dup_company,
               md.top_dup_title AS top_dup_title,
               md.dup_n AS dup_n,
               (c.company_id IS NOT NULL) AS account_linked,
               (e.contact_id IS NOT NULL) AS enriched,
               ec.owners AS owners, ec.channels AS channels, ec.tier AS tier
        FROM public.contacts c
        LEFT JOIN bedrock.activity a
          ON a.participant_public_contact_id = c.contact_id AND a.deleted_at IS NULL
        LEFT JOIN bedrock.account_email_domain aed
          ON aed.domain = lower(split_part(c.email,'@',2))
        LEFT JOIN bedrock.candidate_enrichment e ON e.contact_id = c.contact_id
        LEFT JOIN bedrock.email_candidate ec ON ec.contact_id = c.contact_id
        LEFT JOIN LATERAL (
            -- best one-click merge target: a real contact with the same name
            -- (exclude self, merged, and other queue rows). Unique only.
            SELECT (array_agg(d.contact_id ORDER BY d.contact_id))[1] AS top_dup_id,
                   (array_agg(d.current_company ORDER BY d.contact_id))[1] AS top_dup_company,
                   (array_agg(d.current_title ORDER BY d.contact_id))[1] AS top_dup_title,
                   count(*) AS dup_n
            FROM public.contacts d
            WHERE c.full_name IS NOT NULL AND position('@' in c.full_name) = 0
              AND lower(d.full_name) = lower(c.full_name) AND d.contact_id <> c.contact_id
              AND coalesce(d.contact_stage,'') NOT IN ('merged','candidate','dismissed')
        ) md ON true
        WHERE {' AND '.join(where)}
        GROUP BY c.contact_id, c.company_id, aed.sf_account_name, md.top_dup_id, md.top_dup_company, md.top_dup_title, md.dup_n, e.full_name, e.company, e.confidence,
                 e.is_employer_contact, e.account_suggestion, e.possible_duplicate_ids, e.contact_id,
                 ec.owners, ec.channels, ec.tier
        ORDER BY (e.contact_id IS NOT NULL) DESC, max(a.activity_date) DESC NULLS LAST, c.email
        """,
        *params,
    )
    return {"success": True, "data": [
        {"contact_id": r["contact_id"], "full_name": r["full_name"], "email": r["email"],
         "current_company": r["current_company"], "current_title": r["current_title"],
         "domain": r["domain"],
         # Prefer AI account, then exact-domain map. Best name = stored AI name.
         "suggested_account": r["ai_account"] or r["domain_account"],
         "top_dup_id": r["top_dup_id"] if r["dup_n"] == 1 else None,
         "top_dup_company": r["top_dup_company"], "top_dup_title": r["top_dup_title"],
         "dup_match_count": r["dup_n"],
         "account_linked": r["account_linked"],
         "ai_name": r["ai_name"], "ai_company": r["ai_company"], "ai_confidence": r["ai_confidence"],
         "is_employer_contact": r["is_employer_contact"],
         "dup_count": r["dup_count"], "enriched": r["enriched"],
         "email_count": r["email_count"],
         "owners": list(r["owners"]) if r["owners"] else [],
         "channels": list(r["channels"]) if r["channels"] else [],
         "tier": r["tier"],
         "last_email": r["last_email"].isoformat() if r["last_email"] else None,
         "last_subject": r["last_subject"]}
        for r in rows
    ]}


@router.get("/candidates/owners")
async def list_candidate_owners(user=Depends(require_auth), conn=Depends(get_db)):
    """Distinct staff owners across review candidates, with counts — drives the
    person filter dropdown on the candidates zone."""
    rows = await conn.fetch(
        """
        SELECT o AS owner, count(*) AS n
        FROM bedrock.email_candidate ec
        JOIN public.contacts c ON c.contact_id = ec.contact_id AND c.contact_stage='candidate'
        CROSS JOIN LATERAL unnest(ec.owners) AS o
        JOIN bedrock.sync_staff ss ON lower(ss.email) = lower(o) AND ss.enabled
        GROUP BY o ORDER BY o
        """,
    )
    return {"success": True, "data": [{"owner": r["owner"], "count": r["n"]} for r in rows]}


@router.get("/candidates/{contact_id}")
async def candidate_detail(
    contact_id: int,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Full candidate: stored AI enrichment (instant — no live call), the best
    account linkage suggestion, resolved duplicate contacts, and the emails."""
    import json as _json
    from services.candidate_enrich import suggest_account
    c = await conn.fetchrow(
        "SELECT contact_id, full_name, email, current_company, current_title, linkedin_url, company_id "
        "FROM public.contacts WHERE contact_id=$1", contact_id)
    if not c:
        raise HTTPException(404, "Candidate not found")
    enr = await conn.fetchrow("SELECT * FROM bedrock.candidate_enrichment WHERE contact_id=$1", contact_id)

    enrichment = None
    suggestion = None
    dup_ids: list[int] = []
    if enr:
        enrichment = {
            "full_name": enr["full_name"], "title": enr["title"], "company": enr["company"],
            "linkedin_url": enr["linkedin_url"], "is_employer_contact": enr["is_employer_contact"],
            "confidence": enr["confidence"], "reasoning": enr["reasoning"],
            "enriched_at": enr["enriched_at"].isoformat() if enr["enriched_at"] else None,
        }
        if enr["account_suggestion"]:
            suggestion = enr["account_suggestion"] if isinstance(enr["account_suggestion"], dict) else _json.loads(enr["account_suggestion"])
        dup_ids = list(enr["possible_duplicate_ids"] or [])
    if c["company_id"] is not None:
        suggestion = None            # already linked to an account — don't propose another
    elif suggestion is None:
        suggestion = await suggest_account(conn, c["email"])
    # Only ever surface suggestions that link to an EXISTING account — drop any
    # stale "create a new account" enrichment (no account_key / not in pipeline).
    if suggestion and not (suggestion.get("account_key") or suggestion.get("sf_account_id") or suggestion.get("in_pipeline")):
        suggestion = None

    # Resolve duplicate ids to contacts, deduped by (name, company).
    possible_duplicates = []
    seen = set()
    if dup_ids:
        drows = await conn.fetch(
            "SELECT contact_id, full_name, current_company, current_title FROM public.contacts WHERE contact_id = ANY($1::int[])",
            dup_ids)
        for d in drows:
            k = (d["full_name"], d["current_company"])
            if k in seen:
                continue
            seen.add(k)
            possible_duplicates.append({"contact_id": d["contact_id"], "full_name": d["full_name"],
                                        "current_company": d["current_company"], "current_title": d["current_title"]})
    # Live exact-name probe as well — AI enrichment can be missing or stale
    # (Alan Joos / Jumi Barnes sat unlinked next to obvious matches). Email-ish
    # display names can't match here; the drawer's manual link search covers those.
    nm = (c["full_name"] or "").strip()
    if nm and "@" not in nm:
        nrows = await conn.fetch(
            """SELECT contact_id, full_name, current_company, current_title FROM public.contacts
               WHERE lower(full_name) = lower($1) AND contact_id <> $2
                 AND coalesce(contact_stage,'') NOT IN ('merged', 'candidate', 'dismissed')
               ORDER BY (current_company IS NOT NULL) DESC LIMIT 5""", nm, contact_id)
        for d in nrows:
            k = (d["full_name"], d["current_company"])
            if k in seen:
                continue
            seen.add(k)
            possible_duplicates.append({"contact_id": d["contact_id"], "full_name": d["full_name"],
                                        "current_company": d["current_company"], "current_title": d["current_title"]})

    emails = await conn.fetch(
        """
        SELECT a.id, a.subject, a.email_from, a.email_to, a.email_snippet,
               left(a.email_body_text, 4000) AS body, a.type, a.source, a.activity_date
        FROM bedrock.activity a
        WHERE a.participant_public_contact_id = $1 AND a.deleted_at IS NULL
        ORDER BY a.activity_date DESC LIMIT 50
        """, contact_id)
    return {"success": True, "data": {
        "contact": {"contact_id": c["contact_id"], "full_name": c["full_name"], "email": c["email"],
                    "current_company": c["current_company"], "current_title": c["current_title"],
                    "linkedin_url": c["linkedin_url"]},
        "enrichment": enrichment,
        "suggested_account": suggestion,
        "possible_duplicates": possible_duplicates,
        "emails": [
            {"id": str(e["id"]), "subject": e["subject"], "email_from": e["email_from"],
             "email_to": list(e["email_to"]) if e["email_to"] else None,
             "snippet": e["email_snippet"], "body": e["body"], "type": e["type"], "source": e["source"],
             "activity_date": e["activity_date"].isoformat() if e["activity_date"] else None}
            for e in emails
        ],
    }}


@router.post("/candidates/{contact_id}/enrich")
async def enrich_candidate_endpoint(
    contact_id: int,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Re-run AI enrichment for one candidate and PERSIST it (manual refresh).
    Normally enrichment is pre-computed in batch; this forces a fresh pass."""
    from services.candidate_enrich import enrich_and_store
    c = await conn.fetchrow("SELECT contact_id FROM public.contacts WHERE contact_id=$1", contact_id)
    if not c:
        raise HTTPException(404, "Candidate not found")
    result = await enrich_and_store(conn, contact_id)
    return {"success": True, "data": result}


class CandidateLink(BaseModel):
    target_contact_id: int


@router.post("/candidates/{contact_id}/link")
async def link_candidate(
    contact_id: int,
    body: CandidateLink,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Approve a duplicate match: re-point this candidate's emails onto the
    existing contact and retire the candidate. One-click merge."""
    target = await conn.fetchrow(
        "SELECT contact_id FROM public.contacts WHERE contact_id=$1", body.target_contact_id)
    if not target:
        raise HTTPException(404, "Target contact not found")
    async with conn.transaction():
        await conn.execute(
            "UPDATE bedrock.activity SET participant_public_contact_id=$1 "
            "WHERE participant_public_contact_id=$2", body.target_contact_id, contact_id)
        await conn.execute(
            "UPDATE public.contacts SET contact_stage='merged', "
            "tags=array_remove(coalesce(tags,'{}'), 'email_review'), updated_at=now() "
            "WHERE contact_id=$1", contact_id)
        await conn.execute("DELETE FROM bedrock.candidate_enrichment WHERE contact_id=$1", contact_id)
        await conn.execute(
            "UPDATE public.contacts SET is_jobs_contact=true WHERE contact_id=$1", body.target_contact_id)
    return {"success": True, "data": {"linked_to": body.target_contact_id}}


# ── Salesforce matching (MECE: search DB + SF by email; SF hit → link + pipeline)

async def _upsert_sf_contact(conn, sf: dict, source_email: str, company: str | None,
                             employer_account_id: str | None = None) -> int:
    """Ensure the SF-matched contact exists in public.contacts (jobs pipeline)
    and return its contact_id. `company` is the resolved EMPLOYER (from the
    primary affiliation), never the NPSP Household. Reuses an existing
    link/email row when present; else inserts a new pipeline contact + SF link."""
    sf_id = sf.get("sf_contact_id")
    pid = await conn.fetchval(
        "SELECT public_contact_id FROM bedrock.sf_contact_link WHERE sf_contact_id=$1", sf_id)
    if not pid:
        pid = await conn.fetchval("SELECT contact_id FROM public.contacts WHERE lower(email)=lower($1) LIMIT 1", source_email)
    if not pid:
        pid = await conn.fetchval(
            """INSERT INTO public.contacts
               (full_name, email, current_company, current_title, dedup_key, source, contact_stage, is_jobs_contact, created_at, updated_at)
               VALUES ($1,$2,$3,$4,$5,'salesforce','lead',true,now(),now()) RETURNING contact_id""",
            sf.get("name"), source_email, company, sf.get("title"), f"sf:{sf_id}")
    else:
        # backfill company on the existing record if we have one and it's blank
        if company:
            await conn.execute(
                "UPDATE public.contacts SET current_company=coalesce(nullif(current_company,''),$2) WHERE contact_id=$1",
                pid, company)
    await conn.execute("UPDATE public.contacts SET is_jobs_contact=true WHERE contact_id=$1", pid)
    if sf_id:
        await conn.execute(
            """INSERT INTO bedrock.sf_contact_link (sf_contact_id, public_contact_id, confidence, matched_by, sf_account_id)
               VALUES ($1,$2,'email','candidate_resolve',$3)
               ON CONFLICT DO NOTHING""", sf_id, pid, employer_account_id)
    return pid


async def _link_candidate(conn, candidate_id: int, target_pid: int) -> None:
    """Re-point a candidate's emails onto an existing contact + retire it."""
    await conn.execute(
        "UPDATE bedrock.activity SET participant_public_contact_id=$1 WHERE participant_public_contact_id=$2",
        target_pid, candidate_id)
    await conn.execute(
        "UPDATE public.contacts SET contact_stage='merged', is_jobs_contact=false, "
        "tags=array_remove(coalesce(tags,'{}'),'email_review'), updated_at=now() WHERE contact_id=$1", candidate_id)
    await conn.execute("DELETE FROM bedrock.candidate_enrichment WHERE contact_id=$1", candidate_id)


@router.post("/candidates/resolve-sf")
async def resolve_candidates_sf(
    user=Depends(require_auth),
    client=Depends(require_sf_mcp_client),
    conn=Depends(get_db),
):
    """MECE batch: for every candidate, look the email up in Salesforce
    (Email / HomeEmail / WorkEmail). A match is definitive → import the SF
    contact into the pipeline, re-point the candidate's emails onto it, retire
    the candidate. Unmatched candidates stay for human review."""
    from services.candidate_enrich import (
        sf_contact_match_soql, index_sf_matches, sf_affiliation_employer_soql, resolve_employer_company)
    cands = await conn.fetch(
        "SELECT contact_id, email FROM public.contacts WHERE contact_stage='candidate' AND 'email_review'=ANY(tags)")
    emails = [c["email"] for c in cands if c["email"]]
    matched_index: dict = {}
    for i in range(0, len(emails), 150):   # batch SOQL: contacts by email
        try:
            res = await client.salesforce.query(sf_contact_match_soql(emails[i:i + 150]))
            matched_index.update(index_sf_matches(res.get("records", [])))
        except Exception as e:
            logger.warning("resolve-sf contact batch failed: %s", e)
    # Employer per matched contact, from the primary affiliation (not Household).
    sf_ids = list({m["sf_contact_id"] for m in matched_index.values() if m.get("sf_contact_id")})
    employer_by_contact: dict = {}
    for i in range(0, len(sf_ids), 150):
        try:
            res = await client.salesforce.query(sf_affiliation_employer_soql(sf_ids[i:i + 150]))
            for r in res.get("records", []):
                employer_by_contact[r.get("npe5__Contact__c")] = r.get("Account_ForFellowsOnly__c")
        except Exception as e:
            logger.warning("resolve-sf affiliation batch failed: %s", e)
    linked = 0
    async with conn.transaction():
        for c in cands:
            sf = matched_index.get((c["email"] or "").lower().strip())
            if not sf:
                continue
            emp_acct = employer_by_contact.get(sf["sf_contact_id"])
            company = await resolve_employer_company(conn, emp_acct) or sf.get("company")
            pid = await _upsert_sf_contact(conn, sf, c["email"], company, emp_acct)
            if pid and pid != c["contact_id"]:
                await _link_candidate(conn, c["contact_id"], pid)
                linked += 1
    return {"success": True, "data": {"candidates": len(cands), "salesforce_linked": linked,
                                      "remaining": len(cands) - linked}}


@router.get("/candidates/{contact_id}/sf-match")
async def candidate_sf_match(
    contact_id: int,
    user=Depends(require_auth),
    client=Depends(require_sf_mcp_client),
    conn=Depends(get_db),
):
    """Salesforce contact matching this candidate's email, with the resolved
    EMPLOYER (primary affiliation → jobs company) — for the drawer."""
    from services.candidate_enrich import (
        sf_contact_match_soql, index_sf_matches, sf_affiliation_employer_soql, resolve_employer_company)
    c = await conn.fetchrow("SELECT email FROM public.contacts WHERE contact_id=$1", contact_id)
    if not c:
        raise HTTPException(404, "Candidate not found")
    try:
        res = await client.salesforce.query(sf_contact_match_soql([c["email"]]))
    except Exception as e:
        return {"success": True, "data": {"match": None, "error": str(e)[:120]}}
    match = index_sf_matches(res.get("records", [])).get((c["email"] or "").lower().strip())
    if match and match.get("sf_contact_id"):
        try:
            aff = await client.salesforce.query(sf_affiliation_employer_soql([match["sf_contact_id"]]))
            recs = aff.get("records", [])
            emp_acct = recs[0].get("Account_ForFellowsOnly__c") if recs else None
            company = await resolve_employer_company(conn, emp_acct) or match.get("company")
            match["employer_account_id"] = emp_acct
            match["account_name"] = company
        except Exception as e:
            logger.warning("sf-match affiliation lookup failed: %s", e)
            match["account_name"] = match.get("company")
    return {"success": True, "data": {"match": match}}


class CandidateLinkSf(BaseModel):
    sf_contact_id: str
    name: Optional[str] = None
    account_name: Optional[str] = None
    account_id: Optional[str] = None
    title: Optional[str] = None


@router.post("/candidates/{contact_id}/link-sf")
async def link_candidate_sf(
    contact_id: int,
    body: CandidateLinkSf,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Approve a Salesforce match: import the SF contact into the pipeline and
    re-point this candidate's emails onto it."""
    c = await conn.fetchrow("SELECT email FROM public.contacts WHERE contact_id=$1", contact_id)
    if not c:
        raise HTTPException(404, "Candidate not found")
    async with conn.transaction():
        # body.account_name is the resolved EMPLOYER company (from sf-match).
        pid = await _upsert_sf_contact(conn,
            {"sf_contact_id": body.sf_contact_id, "name": body.name, "title": body.title},
            c["email"], body.account_name, body.account_id)
        if pid != contact_id:
            await _link_candidate(conn, contact_id, pid)
    return {"success": True, "data": {"linked_to": pid}}


async def _resolve_or_create_company(conn, name: str, email: Optional[str] = None) -> Optional[int]:
    """Resolve a free-typed company name to a public.companies id, creating the row
    if it's new. Matches case-insensitively on name; on create, sets the domain from
    a non-freemail email so future activity auto-links. Returns company_id (or None
    if name is blank). Requires INSERT on public.companies."""
    nm = (name or "").strip()
    if not nm:
        return None
    existing = await conn.fetchval("SELECT company_id FROM public.companies WHERE lower(name) = lower($1) LIMIT 1", nm)
    if existing:
        return existing
    FREEMAIL = {"gmail.com","yahoo.com","hotmail.com","outlook.com","icloud.com","aol.com","me.com",
                "msn.com","live.com","proton.me","protonmail.com"}
    dom = None
    if email and "@" in email:
        d = email.split("@", 1)[1].lower().strip()
        if d and d not in FREEMAIL:
            dom = d
    try:
        return await conn.fetchval(
            "INSERT INTO public.companies (name, domain, source, created_at, updated_at) "
            "VALUES ($1, $2, 'candidate_promote', now(), now()) RETURNING company_id", nm, dom)
    except Exception as e:  # unique-name race or domain collision — fall back to lookup
        logger.warning("company create fell back to lookup for %r: %s", nm, e)
        return await conn.fetchval("SELECT company_id FROM public.companies WHERE lower(name)=lower($1) LIMIT 1", nm)


class CandidatePromote(BaseModel):
    full_name: Optional[str] = None
    current_company: Optional[str] = None
    current_title: Optional[str] = None
    contact_stage: str = "lead"


@router.post("/candidates/{contact_id}/promote")
async def promote_candidate(
    contact_id: int,
    body: CandidatePromote,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Promote a candidate into the pipeline: set the filled-in fields, flip
    is_jobs_contact=true, and drop the 'email_review' tag so it leaves the queue."""
    async with conn.transaction():
        # Link to an EXISTING account only (match by name). Account *creation* is
        # intentionally not done here — a dedicated create-account flow (with a
        # domain prompt etc.) will own that. A new/unmatched name is kept as text.
        company_id = None
        if body.current_company:
            company_id = await conn.fetchval(
                "SELECT company_id FROM public.companies WHERE lower(name) = lower($1) LIMIT 1",
                body.current_company.strip())
        sets = ["is_jobs_contact = true",
                "contact_stage = $2",
                "tags = array_remove(coalesce(tags,'{}'), 'email_review')",
                "updated_at = now()"]
        params = [contact_id, body.contact_stage]
        i = 3
        for field in ("full_name", "current_company", "current_title"):
            val = getattr(body, field)
            if val:
                sets.append(f"{field} = ${i}"); params.append(val); i += 1
        if company_id:
            sets.append(f"company_id = ${i}"); params.append(company_id); i += 1
        res = await conn.execute(
            f"UPDATE public.contacts SET {', '.join(sets)} "
            f"WHERE contact_id = $1 AND contact_stage = 'candidate'", *params)
        if res == "UPDATE 0":
            raise HTTPException(404, "Candidate not found")
    return {"success": True, "data": {"contact_id": contact_id, "promoted": True, "company_id": company_id}}


class CandidateSetAccount(BaseModel):
    company: str


@router.post("/candidates/{contact_id}/set-account")
async def set_candidate_account(
    contact_id: int,
    body: CandidateSetAccount,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Attach an account to a candidate WITHOUT promoting it — sets
    current_company (+ company_id when the name matches an existing account),
    but keeps it in the review queue so the name/details can still be edited
    before a deliberate promote."""
    company = (body.company or "").strip()
    if not company:
        raise HTTPException(400, "company required")
    company_id = await conn.fetchval(
        "SELECT company_id FROM public.companies WHERE lower(name) = lower($1) LIMIT 1", company)
    sets = ["current_company = $2", "updated_at = now()"]
    params = [contact_id, company]
    if company_id:
        sets.append(f"company_id = ${len(params)+1}"); params.append(company_id)
    res = await conn.execute(
        f"UPDATE public.contacts SET {', '.join(sets)} "
        f"WHERE contact_id = $1 AND contact_stage = 'candidate'", *params)
    if res == "UPDATE 0":
        raise HTTPException(404, "Candidate not found")
    return {"success": True, "data": {"contact_id": contact_id, "company": company,
                                      "company_id": company_id, "matched": company_id is not None}}


@router.post("/candidates/{contact_id}/dismiss")
async def dismiss_candidate(
    contact_id: int,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Dismiss a candidate: move it out of the queue (public.contacts has no
    soft-delete) by clearing the stage + email_review tag. Activity stays linked."""
    res = await conn.execute(
        "UPDATE public.contacts SET contact_stage='dismissed', "
        "tags=array_remove(coalesce(tags,'{}'), 'email_review'), updated_at=now() "
        "WHERE contact_id=$1 AND contact_stage='candidate'", contact_id)
    if res == "UPDATE 0":
        # Idempotent: only 404 if the contact truly doesn't exist. If it's already
        # dismissed/promoted (e.g. absorbed by the builder sweep), treat as success
        # so the UI's dismiss click never errors on an already-handled row.
        exists = await conn.fetchval("SELECT 1 FROM public.contacts WHERE contact_id=$1", contact_id)
        if not exists:
            raise HTTPException(404, "Candidate not found")
    return {"success": True, "data": {"contact_id": contact_id, "dismissed": True}}


class BulkDismiss(BaseModel):
    contact_ids: list[int]


@router.post("/candidates/bulk-dismiss")
async def bulk_dismiss_candidates(
    body: BulkDismiss,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Dismiss many candidates in one atomic statement (only rows still in the
    candidate stage are touched)."""
    ids = [int(i) for i in (body.contact_ids or [])]
    if not ids:
        return {"success": True, "data": {"dismissed": 0}}
    res = await conn.execute(
        "UPDATE public.contacts SET contact_stage='dismissed', "
        "tags=array_remove(coalesce(tags,'{}'), 'email_review'), updated_at=now() "
        "WHERE contact_id = ANY($1::int[]) AND contact_stage='candidate'", ids)
    n = int(res.split()[-1]) if res and res.split()[-1].isdigit() else 0
    return {"success": True, "data": {"dismissed": n}}


@router.post("/candidates/bulk-restore")
async def bulk_restore_candidates(
    body: BulkDismiss,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Undo dismissal — move rows back into the review queue. Only affects rows
    currently dismissed that still carry the candidate markers."""
    ids = [int(i) for i in (body.contact_ids or [])]
    if not ids:
        return {"success": True, "data": {"restored": 0}}
    res = await conn.execute(
        "UPDATE public.contacts SET contact_stage='candidate', "
        "tags=(CASE WHEN 'email_review' = ANY(coalesce(tags,'{}')) THEN tags "
        "      ELSE array_append(coalesce(tags,'{}'), 'email_review') END), updated_at=now() "
        "WHERE contact_id = ANY($1::int[]) AND contact_stage='dismissed' "
        "  AND ('email_review' = ANY(coalesce(tags,'{}')) OR source='email_candidate')", ids)
    n = int(res.split()[-1]) if res and res.split()[-1].isdigit() else 0
    return {"success": True, "data": {"restored": n}}


@router.post("/candidates/match-builders")
async def match_builder_candidates(
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Sweep the queue for candidates who are actually Pursuit builders — save
    their personal email onto the builder record and drop them from review."""
    from services.builder_match import sweep_builder_candidates
    return {"success": True, "data": await sweep_builder_candidates(conn)}


# ── Activity logging ─────────────────────────────────────────────────────────

class ActivityCreate(BaseModel):
    jobs_opportunity_id: Optional[str] = None
    contact_id:          Optional[int] = None   # log against a prospect instead of a deal
    type:                str                     # call | text | linkedin
    description:         str
    activity_date:       Optional[datetime] = None
    subject:             Optional[str] = None


@router.post("/activity")
async def log_activity(
    body: ActivityCreate,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    user_email = user.get("email") if isinstance(user, dict) else getattr(user, "email", None)
    if body.type not in ("call", "text", "linkedin"):
        raise HTTPException(400, f"Invalid type: {body.type}")
    if not body.jobs_opportunity_id and not body.contact_id:
        raise HTTPException(400, "Provide jobs_opportunity_id or contact_id")

    import uuid as _uuid

    if body.jobs_opportunity_id:
        opp_id = _uuid.UUID(body.jobs_opportunity_id)
        row_id = await conn.fetchval(
            """
            INSERT INTO bedrock.activity
                (type, subject, description, activity_date, source, jobs_opportunity_id, logged_by)
            VALUES ($1, $2, $3, COALESCE($4, now()), 'manual', $5, $6)
            RETURNING id
            """,
            body.type,
            body.subject or f"{body.type.capitalize()} — {user_email}",
            body.description,
            body.activity_date,
            opp_id,
            user_email,
        )
    else:
        # Prospect-scoped log: tie to the public.contacts row, leave the deal null.
        row_id = await conn.fetchval(
            """
            INSERT INTO bedrock.activity
                (type, subject, description, activity_date, source,
                 participant_public_contact_id, logged_by)
            VALUES ($1, $2, $3, COALESCE($4, now()), 'manual', $5, $6)
            RETURNING id
            """,
            body.type,
            body.subject or f"{body.type.capitalize()} — {user_email}",
            body.description,
            body.activity_date,
            body.contact_id,
            user_email,
        )
        # A logged touch IS outreach — a still-flagged contact advances to
        # initial_outreach immediately (the nightly pass covers synced email).
        await conn.execute(
            """
            UPDATE bedrock.jobs_contact_membership
            SET stage = 'initial_outreach',
                first_outreach_at = COALESCE(first_outreach_at, $2),
                first_outreach_by = COALESCE(first_outreach_by, $3),
                updated_at = now()
            WHERE contact_id = $1 AND stage = 'assigned'
            """,
            body.contact_id, body.activity_date or datetime.now(timezone.utc), user_email,
        )

    row = await conn.fetchrow("SELECT * FROM bedrock.activity WHERE id=$1", row_id)
    return {"success": True, "data": dict(row)}


# ============================================================================
# Builders tab — per-builder job-search view (L3 population)
# ============================================================================
# Reads identity/apps/placements/intake/learning from platform+bedrock sources;
# the editable coach/readiness/competency overlay lives in
# bedrock.builder_job_profile. Status is auto-derived with a manual override.

BUILDER_STATUSES = {"not_started", "actively_applying", "interviewing", "placed", "paused"}
_READY_KEYS = ("ready_lookbook", "ready_linkedin", "ready_github", "ready_cv", "ready_mock")


def _derive_builder_status(placed: bool, interviewing: bool, applying: bool) -> str:
    if placed:
        return "placed"
    if interviewing:
        return "interviewing"
    if applying:
        return "actively_applying"
    return "not_started"


def _is_placed(payment_amount, engagement_stage) -> bool:
    # "Placed" = a PAID placement, consistent with the placements / North-Star
    # definition. Unpaid active/completed freelance projects don't count — those
    # builders are still job-searching.
    return (payment_amount or 0) > 0


@router.get("/builders/board")
async def builders_board(user=Depends(require_auth), conn=Depends(get_db)):
    """One row per L3 builder: derived status, counts, readiness, coach."""
    builders = await conn.fetch("SELECT * FROM bedrock.l3_builders() ORDER BY full_name")

    apps = await conn.fetch("""
        SELECT builder_id,
               count(*)                                                    AS total,
               count(*) FILTER (WHERE stage = 'interview')                 AS interviews,
               bool_or(date_applied >= CURRENT_DATE - INTERVAL '60 days')  AS recent
        FROM public.job_applications
        WHERE source_type = 'Pursuit_referred'
        GROUP BY builder_id
    """)
    apps_by = {r["builder_id"]: r for r in apps}

    plc_by: dict = {}
    for r in await conn.fetch(f"SELECT user_id, payment_amount, engagement_stage, opportunity_id FROM bedrock.secured_jobs() WHERE {_live_placement()}"):
        b = plc_by.setdefault(r["user_id"], {"count": 0, "placed": False, "deals": set()})
        b["count"] += 1
        if _is_placed(r["payment_amount"], r["engagement_stage"]):
            b["placed"] = True
        if r["opportunity_id"]:
            b["deals"].add(r["opportunity_id"])

    enrolled = {r["builder_id"] for r in await conn.fetch("SELECT DISTINCT builder_id FROM public.job_strategy_enrollments")}
    profs = {r["user_id"]: r for r in await conn.fetch("SELECT * FROM bedrock.builder_job_profile")}

    out = []
    status_counts = {s: 0 for s in BUILDER_STATUSES}
    for b in builders:
        uid = b["user_id"]
        a, p, prof = apps_by.get(uid), plc_by.get(uid), profs.get(uid)
        interviews = a["interviews"] if a else 0
        placed = bool(p and p["placed"])
        applying = (uid in enrolled) or bool(a and a["recent"])
        derived = _derive_builder_status(placed, interviews > 0, applying)
        overridden = bool(prof and prof["status_overridden"])
        status = prof["job_search_status"] if (overridden and prof and prof["job_search_status"]) else derived
        status_counts[status] = status_counts.get(status, 0) + 1
        ready = {k: bool(prof[k]) if prof else False for k in _READY_KEYS}
        out.append({
            "user_id": uid, "name": b["full_name"], "email": b["email"],
            "cohort": b["cohort"], "cohort_completed": b["cohort_completed"],
            "status": status, "status_overridden": overridden,
            "coach": prof["pursuit_coach"] if prof else None,
            "counts": {
                "applications": (a["total"] if a else 0),
                "interviews": interviews,
                "placements": (p["count"] if p else 0),
                "deal_matches": (len(p["deals"]) if p else 0),
            },
            "readiness": {"complete": sum(ready.values()), "total": len(_READY_KEYS),
                          **{k.replace("ready_", ""): v for k, v in ready.items()}},
            "prof_strength": prof["prof_strength"] if prof else None,
            "technical_strength": prof["technical_strength"] if prof else None,
            "has_profile": prof is not None,
        })
    return {"success": True, "data": {"builders": out, "status_counts": status_counts}}


def _jsonb(v):
    import json as _json
    if v is None:
        return None
    return _json.loads(v) if isinstance(v, str) else v


@router.get("/builders/{user_id}")
async def builder_detail(user_id: int, user=Depends(require_auth), conn=Depends(get_db)):
    """Full per-builder detail: identity + apps/interviews/placements/deals +
    platform intake + learning model + editable job profile + derived status."""
    ident = await conn.fetchrow("SELECT * FROM bedrock.l3_builders() WHERE user_id = $1", user_id)
    if not ident:
        raise HTTPException(404, "Builder not found in the L3 population")

    apps = await conn.fetch("""
        SELECT job_application_id AS id, company_name, role_title, stage,
               date_applied, salary, job_url, response_date
        FROM public.job_applications
        WHERE builder_id = $1 AND source_type = 'Pursuit_referred'
        ORDER BY date_applied DESC NULLS LAST
    """, user_id)
    placements = await conn.fetch(f"""
        SELECT id, role_title, company_name, employment_type, payment_amount,
               engagement_stage, influenced, opportunity_id, start_date
        FROM bedrock.secured_jobs() WHERE user_id = $1 AND {_live_placement()}
        ORDER BY (payment_amount > 0) DESC NULLS LAST, payment_amount DESC NULLS LAST
    """, user_id)
    deal_ids = [r["opportunity_id"] for r in placements if r["opportunity_id"]]
    deals = []
    if deal_ids:
        deals = await conn.fetch("""
            SELECT id, account_name, stage, deal_type, owner_email
            FROM bedrock.jobs_opportunity WHERE id = ANY($1::uuid[]) AND deleted_at IS NULL
        """, deal_ids)

    quiz = await conn.fetch("""
        SELECT question_key, response_text, response_structured
        FROM public.job_strategy_quiz_responses WHERE builder_id = $1
    """, user_id)
    enrollment = await conn.fetchrow("""
        SELECT current_profile, onboarding_completed_at, has_coach
        FROM public.job_strategy_enrollments WHERE builder_id = $1
        ORDER BY updated_at DESC NULLS LAST LIMIT 1
    """, user_id)
    learning = await conn.fetchrow("""
        SELECT skill_levels, competencies, interview_readiness
        FROM public.builder_profiles WHERE user_id = $1
    """, user_id)
    prof = await conn.fetchrow("SELECT * FROM bedrock.builder_job_profile WHERE user_id = $1", user_id)

    placed = any(_is_placed(r["payment_amount"], r["engagement_stage"]) for r in placements)
    interviewing = any(r["stage"] == "interview" for r in apps)
    recent = any(r["date_applied"] and (datetime.now().date() - r["date_applied"]).days <= 60 for r in apps)
    applying = enrollment is not None or recent
    derived = _derive_builder_status(placed, interviewing, applying)
    overridden = bool(prof and prof["status_overridden"])
    status = prof["job_search_status"] if (overridden and prof and prof["job_search_status"]) else derived

    profile = None
    if prof:
        profile = dict(prof)
        profile["intake"] = _jsonb(profile.get("intake"))

    return {"success": True, "data": {
        "identity": {
            "user_id": ident["user_id"], "name": ident["full_name"], "email": ident["email"],
            "cohort": ident["cohort"], "cohort_completed": ident["cohort_completed"],
            "linkedin_url": ident["linkedin_url"], "github_url": ident["github_url"],
        },
        "status": status, "status_overridden": overridden, "derived_status": derived,
        "applications": [dict(r) for r in apps],
        "placements": [dict(r) for r in placements],
        "deal_matches": [dict(r) for r in deals],
        "intake_quiz": [{"question_key": r["question_key"], "response_text": r["response_text"],
                         "response_structured": _jsonb(r["response_structured"])} for r in quiz],
        "enrollment": dict(enrollment) if enrollment else None,
        "learning": ({"skill_levels": _jsonb(learning["skill_levels"]),
                      "competencies": _jsonb(learning["competencies"]),
                      "interview_readiness": learning["interview_readiness"]} if learning else None),
        "profile": profile,
    }}


class BuilderProfileUpdate(BaseModel):
    job_search_status:      Optional[str] = None
    status_overridden:      Optional[bool] = None
    pursuit_coach:          Optional[str] = None
    gen_notes:              Optional[str] = None
    coach_notes:            Optional[str] = None
    coach_flags:            Optional[list[str]] = None
    improvement_tags:       Optional[list[str]] = None
    ready_lookbook:         Optional[bool] = None
    ready_linkedin:         Optional[bool] = None
    ready_github:           Optional[bool] = None
    ready_cv:               Optional[bool] = None
    ready_mock:             Optional[bool] = None
    technical_capability:   Optional[str] = None
    ai_reasoning:           Optional[str] = None
    problem_solving:        Optional[str] = None
    presentation:           Optional[str] = None
    professional_behaviors: Optional[str] = None
    prof_strength:          Optional[str] = None
    technical_strength:     Optional[str] = None
    target_industries:      Optional[list[str]] = None
    preferred_modes:        Optional[list[str]] = None
    certifications:         Optional[list[str]] = None
    resume_url:             Optional[str] = None
    lookbook_url:           Optional[str] = None
    university:             Optional[str] = None
    degree:                 Optional[str] = None
    graduation_year:        Optional[int] = None
    languages:              Optional[list[str]] = None
    applying_regularly:     Optional[bool] = None
    networking_regularly:   Optional[bool] = None
    intake:                 Optional[dict] = None


@router.patch("/builders/{user_id}")
async def update_builder_profile(user_id: int, body: BuilderProfileUpdate,
                                 user=Depends(require_auth), conn=Depends(get_db)):
    """Upsert the builder's job-profile overlay. Setting job_search_status flips
    status_overridden on; pass status_overridden=false to revert to derived.
    intake is merged (||) rather than replaced."""
    fields = body.dict(exclude_unset=True)
    if "job_search_status" in fields and "status_overridden" not in fields:
        fields["status_overridden"] = True
    if not fields:
        raise HTTPException(400, "No fields to update")

    import json as _json
    cols, vals = ["user_id"], [user_id]
    for k, v in fields.items():
        cols.append(k)
        vals.append(_json.dumps(v) if k == "intake" else v)
    ph = [f"${i+1}" for i in range(len(cols))]
    sets = []
    for k in cols[1:]:
        if k == "intake":
            sets.append("intake = bedrock.builder_job_profile.intake || EXCLUDED.intake")
        else:
            sets.append(f"{k} = EXCLUDED.{k}")
    sets.append("updated_at = now()")
    sql = (f"INSERT INTO bedrock.builder_job_profile ({', '.join(cols)}) "
           f"VALUES ({', '.join(ph)}) "
           f"ON CONFLICT (user_id) DO UPDATE SET {', '.join(sets)} RETURNING *")
    row = await conn.fetchrow(sql, *vals)
    d = dict(row)
    d["intake"] = _jsonb(d.get("intake"))
    return {"success": True, "data": d}


# ── Export ────────────────────────────────────────────────────────────────────
# Phase 1 per the build plan: server-generated .xlsx returned as a download. No
# Drive write — that needs a new OAuth scope or a delegated service account, which
# is an access conversation with Jac rather than a coding task.
#
# openpyxl is already in requirements.txt (3.1+), so this adds no dependency.

class ExportRequest(BaseModel):
    """Rows to export.

    `ids` are the selected records and are required — see the endpoint for why
    "export everything I'm looking at" isn't a thing the server can honestly do.
    """
    ids: Optional[list[str]] = None
    # Column keys to include, in order. Omitted → the entity's default set, so a
    # caller that doesn't care doesn't have to enumerate them.
    columns: Optional[list[str]] = None


# entity → (label, id column, source SQL, default column list).
# Each SELECT is read-only and already-scoped to non-deleted rows. The id column
# is cast to text so contacts (int) and opportunities (uuid) share one code path.
_EXPORT_SPECS: dict[str, dict] = {
    "contacts": {
        "label": "Contacts",
        # Mirrors what the Contacts table can show, so "export what I'm looking
        # at" means the same fields on screen and in the file. `default` is the
        # subset sent when the caller doesn't pick; `columns` is everything
        # selectable. Post-migration columns (call_booked_at, revisit_date) are
        # deliberately absent — this file has to run before that migration too.
        "sql": """
            SELECT c.contact_id::text AS id, c.full_name, c.email, c.current_title,
                   c.current_company, c.linkedin_url, c.location, c.source,
                   c.contact_stage, c.owner_email AS contact_owner, c.notes,
                   c.created_at,
                   m.stage AS jobs_stage, m.owner_email AS jobs_owner,
                   m.assigned_at, m.first_outreach_at, m.first_outreach_by,
                   m.converted_at, m.not_a_fit_reason,
                   co.industry, co.size_bucket, co.hq_location,
                   (SELECT string_agg(t, ', ') FROM unnest(coalesce(c.tags,'{}')) t
                     WHERE t IN (SELECT slug FROM bedrock.contact_tag_catalog)) AS campaign_tags,
                   (SELECT max(a.activity_date) FROM bedrock.activity a
                     WHERE a.participant_public_contact_id = c.contact_id
                       AND a.deleted_at IS NULL) AS last_activity_at
            FROM public.contacts c
            LEFT JOIN bedrock.jobs_contact_membership m ON m.contact_id = c.contact_id
            LEFT JOIN LATERAL (
                SELECT industry, size_bucket, hq_location
                FROM public.companies
                WHERE lower(trim(name)) = lower(trim(c.current_company)) LIMIT 1
            ) co ON true
        """,
        "default": ["full_name", "email", "current_title", "current_company",
                    "linkedin_url", "jobs_stage", "jobs_owner", "campaign_tags",
                    "assigned_at", "first_outreach_at"],
        "columns": ["full_name", "email", "current_title", "current_company",
                    "linkedin_url", "location", "industry", "size_bucket",
                    "hq_location", "jobs_stage", "jobs_owner", "campaign_tags",
                    "contact_stage", "contact_owner", "source",
                    "assigned_at", "first_outreach_at", "first_outreach_by",
                    "converted_at", "not_a_fit_reason", "last_activity_at",
                    "created_at", "notes"],
        "headers": {"full_name": "Name", "email": "Email", "current_title": "Title",
                    "current_company": "Company", "linkedin_url": "LinkedIn",
                    "location": "Location", "industry": "Industry",
                    "size_bucket": "Size", "hq_location": "HQ",
                    "jobs_stage": "Jobs stage", "jobs_owner": "Jobs owner",
                    "campaign_tags": "Campaigns", "contact_stage": "Contact stage",
                    "contact_owner": "Contact owner", "source": "Source",
                    "assigned_at": "Assigned", "first_outreach_at": "First outreach",
                    "first_outreach_by": "First outreach by",
                    "converted_at": "Converted", "not_a_fit_reason": "Not-a-fit reason",
                    "last_activity_at": "Last activity", "created_at": "Created",
                    "notes": "Notes"},
    },
    "opportunities": {
        "label": "Opportunities",
        "sql": """
            SELECT o.id::text AS id, o.account_name, o.title, o.stage, o.deal_type,
                   o.owner_email, o.salary_expected, o.num_roles, o.likelihood,
                   o.priority, o.segment, o.touch_count, o.follow_up_date,
                   o.target_close_date, o.created_at, o.closed_at,
                   o.closed_lost_reason, o.closed_lost_note
            FROM bedrock.jobs_opportunity o
            WHERE o.deleted_at IS NULL
        """,
        "columns": ["account_name", "title", "stage", "deal_type", "owner_email",
                    "salary_expected", "num_roles", "likelihood", "priority",
                    "segment", "touch_count", "follow_up_date", "target_close_date",
                    "created_at", "closed_at", "closed_lost_reason", "closed_lost_note"],
        "headers": {"account_name": "Account", "title": "Title", "stage": "Stage",
                    "deal_type": "Deal type", "owner_email": "Owner",
                    "salary_expected": "Salary", "num_roles": "Roles",
                    "likelihood": "Likelihood", "priority": "Priority",
                    "segment": "Segment", "touch_count": "Touches",
                    "follow_up_date": "Follow up", "target_close_date": "Target close",
                    "created_at": "Created", "closed_at": "Closed",
                    "closed_lost_reason": "Lost reason", "closed_lost_note": "Lost note"},
    },
    "accounts": {
        "label": "Accounts",
        # Accounts are keyed by normalized company name, and the account list is
        # assembled in Python from several sources. Export the durable core here
        # (the jobs_account row plus firmographics) rather than trying to
        # reproduce the whole derived hub payload in one query.
        "sql": """
            SELECT ja.account_key AS id, ja.display_name, ja.owner_email,
                   ja.status_override, ja.notes,
                   co.industry, co.size_bucket, co.hq_location, co.stage AS company_stage,
                   (SELECT count(*) FROM bedrock.jobs_opportunity o
                     WHERE o.deleted_at IS NULL
                       AND lower(trim(o.account_name)) = ja.account_key) AS opportunities
            FROM bedrock.jobs_account ja
            LEFT JOIN LATERAL (
                SELECT industry, size_bucket, hq_location, stage
                FROM public.companies
                WHERE lower(trim(name)) = ja.account_key LIMIT 1
            ) co ON true
        """,
        "columns": ["display_name", "owner_email", "status_override", "industry",
                    "size_bucket", "hq_location", "company_stage", "opportunities", "notes"],
        "headers": {"display_name": "Account", "owner_email": "Jobs owner",
                    "status_override": "Status", "industry": "Industry",
                    "size_bucket": "Size", "hq_location": "HQ",
                    "company_stage": "Company stage", "opportunities": "Opps",
                    "notes": "Notes"},
    },
}

# Ceiling on an unselected ("export everything I'm looking at") request. Excel
# handles far more, but a 40k-row build holds a pool connection and a few hundred
# MB of memory; past this the answer is to filter first.
_EXPORT_CAP = 5000


@router.get("/export/{entity}/columns")
async def jobs_export_columns(entity: str, user=Depends(require_auth)):
    """The columns this entity can export, and which are on by default.

    The picker builds from this rather than from a list hard-coded in the
    frontend: the allowlist that filters the export request lives here, so a
    second copy in TypeScript would be a drift bug waiting to happen."""
    spec = _EXPORT_SPECS.get(entity)
    if spec is None:
        raise HTTPException(404, f"Nothing to export for '{entity}'")
    default = set(spec.get("default") or spec["columns"])
    return {"success": True, "data": {
        "columns": [
            {"key": c,
             "label": spec["headers"].get(c, c.replace("_", " ").title()),
             "default": c in default}
            for c in spec["columns"]
        ],
    }}


@router.post("/export/{entity}")
async def jobs_export(
    entity: str,
    body: ExportRequest,
    user=Depends(require_auth),
    conn=Depends(get_db),
):
    """Export contacts / accounts / opportunities as .xlsx.

    Pass `ids` for the selected rows; omit it to export the whole set (capped).
    Read-only: this touches nothing, it just reads and formats."""
    spec = _EXPORT_SPECS.get(entity)
    if spec is None:
        raise HTTPException(404, f"Nothing to export for '{entity}'")

    # Unknown keys are dropped rather than 400'd — a stale picker in an open tab
    # should still produce a file. An all-unknown list falls back to the default
    # set, so the sheet is never empty of columns.
    allowed = spec["columns"]
    cols = [c for c in (body.columns or spec.get("default") or allowed) if c in allowed]
    if not cols:
        cols = spec.get("default") or allowed

    # Filter OUTSIDE the spec's SQL, never by appending to it. Appending needed a
    # "does this already have a WHERE" guess, and substring-matching WHERE is
    # wrong twice over: the contacts SQL has a WHERE inside a subquery (so the
    # guess produced `LEFT JOIN ... AND id = ANY(...)`, a syntax error), and the
    # accounts SQL has one inside a LATERAL (so the filter attached to a LEFT JOIN
    # ON clause and silently exported every account instead of the selection).
    # Every spec aliases its key as `id`, so one wrapper covers all three.
    # `ids` is REQUIRED. The old fallback ("no ids = export the whole set") could
    # not keep that promise: nothing here knows the caller's filters, so it ran
    # an unordered `LIMIT 5000` over the raw table — an arbitrary 9% of the
    # 54,919 contacts, merged duplicates included, in a file that looks
    # complete. Silently handing someone the wrong 5,000 rows is worse than
    # refusing, and every caller today sends a selection.
    if not body.ids:
        raise HTTPException(
            400,
            "Select the rows to export. Exporting a whole filtered view isn't "
            "supported yet — the server can't see the filters you applied, so it "
            "would return an arbitrary slice rather than what's on your screen.",
        )
    if len(body.ids) > _EXPORT_CAP:
        raise HTTPException(
            400,
            f"That's {len(body.ids):,} rows; the export caps at {_EXPORT_CAP:,}. "
            "Narrow the selection and try again.",
        )
    params: list = [[str(i) for i in body.ids]]
    sql = f"SELECT * FROM ({spec['sql']}) src WHERE src.id::text = ANY($1::text[])"

    try:
        rows = await conn.fetch(sql, *params)
    except asyncpg.exceptions.InsufficientPrivilegeError as e:
        raise HTTPException(503, f"Export needs a read this role doesn't have: {e}")

    from openpyxl import Workbook
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = spec["label"][:31]     # Excel caps sheet names at 31 chars

    headers = [spec["headers"].get(c, c.replace("_", " ").title()) for c in cols]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for r in rows:
        out = []
        for c in cols:
            v = r.get(c)
            if isinstance(v, (list, tuple)):
                v = ", ".join(str(x) for x in v)
            elif hasattr(v, "tzinfo") and v is not None:
                # Excel has no timezone concept; a tz-aware datetime raises on
                # write, so normalise to naive UTC rather than losing the row.
                v = v.replace(tzinfo=None)
            elif isinstance(v, (dict,)):
                v = json.dumps(v)
            out.append(v)
        ws.append(out)

    # Width from the widest cell in each column, clamped: unbounded widths make a
    # notes column 600px and push everything else off screen.
    for i, header in enumerate(headers, start=1):
        longest = max([len(str(header))] + [len(str(r.get(cols[i - 1]) or "")) for r in rows[:200]])
        ws.column_dimensions[get_column_letter(i)].width = min(48, max(10, longest + 2))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    stamp = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    filename = f"bedrock-{entity}-{stamp}.xlsx"
    logger.info("export: %s rows=%d by=%s", entity, len(rows), user.get("email"))
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
